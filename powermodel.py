#!/usr/bin/python3
#
#  Copyright (C) 2015-2024 Sustainable Energy Now Inc., Angus King
#
#  powermodel.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

from copy import copy
from math import asin, ceil, cos, fabs, floor, log10, pow, radians, sin, sqrt
from PyQt5 import QtCore, QtGui, QtWidgets
import matplotlib
if matplotlib.__version__ > '3.5.1':
    matplotlib.use('Qt5Agg')
else:
    matplotlib.use('TkAgg')
from matplotlib.font_manager import FontProperties
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as oxl
import os
import sys
import ssc
import time
import xlwt

import configparser  # decode .ini file

from senutils import getParents, getUser, ssCol, techClean
import displayobject
import displaytable
from editini import SaveIni
from getmodels import getModelFile
from grid import Grid
from powerclasses import *
from superpower import SuperPower
from sirenicons import Icons
# import Station
from turbine import Turbine
from visualise import Visualise
from zoompan import ZoomPanX

the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


class PowerModel():
    powerExit = QtCore.pyqtSignal(str)
   # barProgress = QtCore.pyqtSignal(int, str)
   # barRange = QtCore.pyqtSignal(int, int)

    def showGraphs(self, ydata, x):
        def shrinkKey(key):
            remove = ['Biomass', 'Community', 'Farm', 'Fixed', 'Geothermal', 'Hydro', 'Pumped',
                      'PV', 'Rooftop', 'Solar', 'Station', 'Thermal', 'Tracking', 'Wave', 'Wind']
            oukey = key
            for i in range(len(remove)):
                oukey = oukey.replace(remove[i], '')
            oukey = ' '.join(oukey.split())
            if oukey == '' or oukey == 'Existing':
                return key
            else:
                return oukey

        def stepPlot(self, period, data, x_labels=None):
            k1 = list(data.keys())[0]
            if self.plots['cumulative']:
                pc = 1
            else:
                pc = 0
            if self.plots['gross_load']:
                pc += 1
            if self.plots['shortfall']:
                pc += 1
            fig = plt.figure(self.hdrs['by_' + period].title() + self.suffix)
            plt.grid(True)
            bbdx = plt.subplot(111)
            plt.title(self.hdrs['by_' + period].title() + self.suffix)
            maxy = 0
            miny = 0
            xs = []
            for i in range(len(data[k1]) + 1):
                xs.append(i)
            if self.plots['save_plot']:
                sp_data = []
                sp_data.append(xs[1:])
                if period == 'day':
                    sp_vals = [period, 'Date']
                    sp_data.append([])
                    mm = 0
                    dy = 1
                    for d in range(1, len(xs)):
                        sp_data[-1].append('%s-%s-%s' % (self.load_year,
                                           str(mm + 1).zfill(2), str(dy).zfill(2)))
                        dy += 1
                        if dy > the_days[mm]:
                            mm += 1
                            dy = 1
                else:
                    sp_vals = ['No.', period]
                    sp_data.append(x_labels)
            if self.plots['cumulative']:
                cumulative = [0.]
                for i in range(len(data[k1])):
                    cumulative.append(0.)
            load = []
            i = -1
            storage = None
            if self.plots['show_pct']:
                load_sum = 0.
                gen_sum = 0.
            for key, value in iter(sorted(ydata.items())):
                if key == 'Generation':
                    continue
                dval = [0.]
                if self.plots['show_pct']:
                    for d in range(len(data[key])):
                        if key[:4] == 'Load':
                            for k in range(len(data[key][d])):
                                load_sum += data[key][d][k]
                        elif key == 'Storage':
                            pass
                        else:
                            for k in range(len(data[key][d])):
                                gen_sum += data[key][d][k]
                                if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                    load_sum += data[key][d][k]
                for d in range(len(data[key])):
                   dval.append(0.)
                   for k in range(len(data[key][0])):
                       dval[-1] += data[key][d][k] / 1000
                maxy = max(maxy, max(dval))
                lw = self.other_width
                if self.plots['cumulative'] and key[:4] != 'Load' and key != 'Storage':
                    lw = 1.0
                    for j in range(len(xs)):
                        cumulative[j] += dval[j]
                bbdx.step(xs, dval, linewidth=lw, label=shrinkKey(key), color=self.colours[key],
                          linestyle=self.linestyle[key])
                if self.plots['save_plot']:
                    sp_vals.append(shrinkKey(key))
                    sp_data.append(dval[1:])
                if (self.plots['shortfall'] or self.plots['show_load']) and key[:4] == 'Load':
                    load = dval[:]
                if self.plots['shortfall'] and key == 'Storage':
                    storage = dval[:]
            if self.plots['show_pct']:
                self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                plt.title(self.hdrs['by_' + period].title() + self.suffix + self.gen_pct)
            if self.plots['cumulative']:
                bbdx.step(xs, cumulative, linewidth=self.other_width, label='Tot. Generation', color=self.colours['cumulative'])
                maxy = max(maxy, max(cumulative))
                if self.plots['save_plot']:
                    sp_vals.append('Tot. Generation')
                    sp_data.append(cumulative[1:])
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                pass
            if (self.plots['shortfall'] and self.do_load):
                load2 = []
                if storage is None:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] - load[i])
                        if load2[-1] < miny:
                            miny = load2[-1]
                else:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] + storage[i] - load[i])
                        if load2[-1] < miny:
                            miny = load2[-1]
                bbdx.step(xs, load2, linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                plt.axhline(0, color='black')
                if rndup != 0 and miny < 0:
                    miny = -ceil(-miny / rndup) * rndup
                if self.plots['save_plot']:
                    sp_vals.append('Shortfall')
                    sp_data.append(load2[1:])
            else:
                miny = 0
            if self.plots['save_plot']:
                titl = 'By_' + period
                decpts = [3] * len(sp_vals)
                decpts[0] = decpts[1] = 0
                dialog = displaytable.Table(list(map(list, list(zip(*sp_data)))), title=titl, fields=sp_vals,
                                            save_folder=self.scenarios, decpts=decpts)
                dialog.exec_()
                del dialog, sp_data, sp_vals
            plt.ylim([miny, maxy])
            plt.xlim([0, len(data[k1])])
            if (len(ydata) + pc) > 9:
                 # Shrink current axis by 5%
                box = bbdx.get_position()
                bbdx.set_position([box.x0, box.y0, box.width * 0.95, box.height])
                 # Put a legend to the right of the current axis
                bbdx.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop=lbl_font)
            else:
                bbdx.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=(len(ydata) + pc),
                            prop=lbl_font)
            rotn = 'horizontal'
            if len(data[k1]) > 12:
                stp = 7
                rotn = 'vertical'
            else:
                stp = 1
            plt.xticks(list(range(0, len(data[k1]), stp)))
            tick_spot = []
            for i in range(0, len(data[k1]), stp):
                tick_spot.append(i + .5)
            bbdx.set_xticks(tick_spot)
            bbdx.set_xticklabels(x_labels, rotation=rotn)
            bbdx.set_xlabel(period.title() + ' of the year')
            bbdx.set_ylabel('Energy (MWh)')
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(bbdx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()

        def dayPlot(self, period, data, per_labels=None, x_labels=None):
            k1 = list(data.keys())[0]
            if self.plots['cumulative']:
                pc = 1
            else:
                pc = 0
            if self.plots['gross_load']:
                pc += 1
            if self.plots['shortfall']:
                pc += 1
            hdr = self.hdrs[period].replace('Power - ', '')
            plt.figure(hdr + self.suffix)
            plt.suptitle(self.hdrs[period] + self.suffix, fontsize=16)
            maxy = 0
            miny = 0
            if len(data[k1]) > 9:
                p1 = 3
                p2 = 4
                xl = 8
                yl = [0, 4, 8]
            elif len(data[k1]) > 6:
                p1 = 3
                p2 = 3
                xl = 6
                yl = [0, 3, 6]
            elif len(data[k1]) > 4:
                p1 = 2
                p2 = 3
                xl = 3
                yl = [0, 3]
            elif len(data[k1]) > 2:
                p1 = 2
                p2 = 2
                xl = 2
                yl = [0, 2]
            else:
                p1 = 1
                p2 = 2
                xl = 0
                yl = [0, 1]
            for key in list(data.keys()):
                for p in range(len(data[key])):
                    maxy = max(maxy, max(data[key][p]))
            if self.plots['show_pct']:
                load_sum = []
                gen_sum = []
                for p in range(len(data[k1])):
                    load_sum.append(0.)
                    gen_sum.append(0.)
            for p in range(len(data[k1])):
                if self.plots['cumulative']:
                    cumulative = []
                    for i in range(len(x24)):
                        cumulative.append(0.)
                if self.plots['gross_load']:
                    gross_load = []
                    for i in range(len(x24)):
                        gross_load.append(0.)
                px = plt.subplot(p1, p2, p + 1)
                l_k = ''
                for key, value in iter(sorted(ydata.items())):
                    if key == 'Generation':
                        continue
                    if key[:4] == 'Load':
                        l_k = key
                    if self.plots['show_pct']:
                        for d in range(len(data[key][p])):
                            if key[:4] == 'Load':
                                load_sum[p] += data[key][p][d]
                            elif key == 'Storage':
                                pass
                            else:
                                gen_sum[p] += data[key][p][d]
                                if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                    load_sum[p] += data[key][p][d]
                    lw = self.other_width
                    if self.plots['cumulative'] and key[:4] != 'Load' and key != 'Storage':
                        lw = 1.0
                        for j in range(len(x24)):
                            cumulative[j] += data[key][p][j]
                    if self.plots['gross_load'] and (key[:4] == 'Load' or
                      key == 'Existing Rooftop PV'):
                        for j in range(len(x24)):
                            gross_load[j] += data[key][p][j]
                    px.plot(x24, data[key][p], linewidth=lw, label=shrinkKey(key),
                            color=self.colours[key], linestyle=self.linestyle[key])
                    plt.title(per_labels[p])
                if self.plots['cumulative']:
                    px.plot(x24, cumulative, linewidth=self.other_width, label='Tot. Generation',
                            color=self.colours['cumulative'])
                    maxy = max(maxy, max(cumulative))
                if self.plots['gross_load'] and 'Existing Rooftop PV' in list(ydata.keys()):
                    px.plot(x24, gross_load, linewidth=1.0, label='Gross Load', color=self.colours['gross_load'])
                    maxy = max(maxy, max(gross_load))
                if self.plots['shortfall'] and self.do_load:
                    load2 = []
                    for i in range(len(x24)):
                        load2.append(cumulative[i] - data[l_k][p][i])
                    miny = min(miny, min(load2))
                    px.plot(x24, load2, linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                    plt.axhline(0, color='black')
                plt.xticks(list(range(4, 25, 4)))
                px.set_xticklabels(x_labels[1:])
                plt.xlim([1, 24])
                if p >= xl:
                    px.set_xlabel('Hour of the Day')
                if p in yl:
                    px.set_ylabel('Power (MW)')
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                pass
            if self.plots['shortfall']:
                if rndup != 0 and miny < 0:
                    miny = -ceil(-miny / rndup) * rndup
            for p in range(len(data[k1])):
                px = plt.subplot(p1, p2, p + 1)
                plt.ylim([miny, maxy])
                plt.xlim([1, 24])
                if self.plots['show_pct']:
                    pct = ' (%s%%)' % '{:0,.1f}'.format(gen_sum[p] * 100. / load_sum[p])
                    titl = px.get_title()
                    px.set_title(titl + pct)
                    #  px.annotate(pct, xy=(1.0, 3.0))
            px = plt.subplot(p1, p2, len(data[k1]))
         #    px.legend(bbox_to_anchor=[1., -0.15], loc='best', ncol=min((len(ly) + pc), 9),
         # prop=lbl_font)
            if (len(ydata) + pc) > 9:
                if len(data[k1]) > 9:
                    do_in = [1, 5, 9, 2, 6, 10, 3, 7, 11, 4, 8, 12]
                elif len(data[k1]) > 6:
                    do_in = [1, 4, 7, 2, 5, 8, 3, 6, 9]
                elif len(data[k1]) > 4:
                    do_in = [1, 4, 2, 5, 3, 6]
                elif len(data[k1]) > 2:
                    do_in = [1, 3, 2, 4]
                else:
                    do_in = [1, 2]
                do_in = do_in[:len(data[k1])]
                for i in range(len(do_in)):
                    px = plt.subplot(p1, p2, do_in[i])
                 # Shrink current axis by 5%
                    box = px.get_position()
                    px.set_position([box.x0, box.y0, box.width * 0.95, box.height])
                 # Put a legend to the right of the current axis
                px.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop=lbl_font)
            else:
                px.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=(len(ydata) + pc),
                          prop=lbl_font)
            if self.plots['show_pct']:
                for p in range(1, len(gen_sum)):
                    load_sum[0] += load_sum[p]
                    gen_sum[0] += gen_sum[p]
                self.gen_pct = ' (%s%%) of load)' % '{:0,.1f}'.format(gen_sum[0] * 100. / load_sum[0])
                titl = px.get_title()
                plt.suptitle(self.hdrs[period] + self.suffix + self.gen_pct, fontsize=16)
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(px, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()

        def saveBalance(self, shortstuff):
            data_file = 'Powerbalance_data_%s.xls' % (
                        QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                        'yyyy-MM-dd_hhmm'))
            data_file = QtWidgets.QFileDialog.getSaveFileName(None, 'Save Powerbalance data file',
                        self.scenarios + data_file, 'Excel Files (*.xls*);;CSV Files (*.csv)')[0]
            if data_file != '':
                if data_file[-4:] == '.csv' or data_file[-4:] == '.xls' \
                  or data_file[-5:] == '.xlsx':
                    pass
                else:
                    data_file += '.xls'
                if os.path.exists(data_file):
                    if os.path.exists(data_file + '~'):
                        os.remove(data_file + '~')
                    os.rename(data_file, data_file + '~')
                stns = {}
                techs = {}
                for i in range(len(self.power_summary)):
                    stns[self.power_summary[i].name] = i
                    techs[self.power_summary[i].technology] = [0., 0., 0.]
                if data_file[-4:] == '.csv':
                    tf = open(data_file, 'w')
                    line = 'Generation Summary Table'
                    tf.write(line + '\n')
                    line = 'Name,Technology,Capacity (MW),CF,Generation'
                    if getattr(self.power_summary[0], 'transmitted') != None:
                        line += ',Transmitted'
                    tf.write(line + '\n')
                    for key, value in iter(sorted(stns.items())):
                        if self.power_summary[value].generation > 0:
                            cf = '{:0.2f}'.format(self.power_summary[value].generation /
                                 (self.power_summary[value].capacity * 8760))
                        else:
                            cf = ''
                        if self.power_summary[value].transmitted is not None:
                            ts = '{:0.2f}'.format(self.power_summary[value].transmitted)
                            techs[self.power_summary[value].technology][2] += self.power_summary[value].transmitted
                        else:
                            ts = ''
                        line = '"%s",%s,%s,%s,%s,%s' % (self.power_summary[value].name,
                                               self.power_summary[value].technology,
                                               '{:0.2f}'.format(self.power_summary[value].capacity),
                                               cf,
                                               '{:0.0f}'.format(self.power_summary[value].generation),
                                               ts)
                        techs[self.power_summary[value].technology][0] += self.power_summary[value].capacity
                        techs[self.power_summary[value].technology][1] += self.power_summary[value].generation
                        tf.write(line + '\n')
                    total = [0., 0., 0.]
                    for key, value in iter(sorted(techs.items())):
                        total[0] += value[0]
                        total[1] += value[1]
                        if value[2] > 0:
                            v2 = ',{:0.0f}'.format(value[2])
                            total[2] += value[2]
                        else:
                            v2 = ''
                        line = ',%s,%s,,%s%s' % (key,
                                               '{:0.2f}'.format(value[0]),
                                               '{:0.0f}'.format(value[1]),
                                               v2)
                        tf.write(line + '\n')
                    if total[2] > 0:
                        v2 = ',{:0.0f}'.format(total[2])
                        total[2] += value[2]
                    else:
                        v2 = ''
                    line = ',Total,%s,,%s%s' % ('{:0.2f}'.format(total[0]),
                                               '{:0.0f}'.format(total[1]),
                                               v2)
                    tf.write(line + '\n')
                    line = '\nHourly Shortfall Table'
                    tf.write(line + '\n')
                    line = 'Hour,Period,Shortfall'
                    tf.write(line + '\n')
                    for i in range(len(shortstuff)):
                        line = '%s,%s,%s' % (str(shortstuff[i].hour), shortstuff[i].period,
                                             '{:0.2f}'.format(shortstuff[i].shortfall))
                        tf.write(line + '\n')
                    tf.close()
                else:
                    wb = xlwt.Workbook()
                    fnt = xlwt.Font()
                    fnt.bold = True
                    styleb = xlwt.XFStyle()
                    styleb.font = fnt
                    style2d = xlwt.XFStyle()
                    style2d.num_format_str = '#,##0.00'
                    style0d = xlwt.XFStyle()
                    style0d.num_format_str = '#,##0'
                    pattern = xlwt.Pattern()
                    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                    pattern.pattern_fore_colour = xlwt.Style.colour_map['ice_blue']
                    style2db = xlwt.XFStyle()
                    style2db.num_format_str = '#,##0.00'
                    style2db.pattern = pattern
                    style0db = xlwt.XFStyle()
                    style0db.num_format_str = '#,##0'
                    style0db.pattern = pattern
                    ws = wb.add_sheet('Powermatch')
                    xl_lens = {}
                    row = 0
                    col = 0
                    ws.write(row, col, 'Hourly Shortfall Table', styleb)
                    row += 1
                    shrt_cols = ['Hour', 'Period', 'Shortfall']
                    for i in range(len(shrt_cols)):
                        ws.write(row, col + i, shrt_cols[i], styleb)
                        xl_lens[col + i] = 0
                    row += 1
                    for i in range(len(shortstuff)):
                        ws.write(row, col, shortstuff[i].hour)
                        ws.write(row, col + 1, shortstuff[i].period)
                        xl_lens[col + 1] = max(xl_lens[col + 1], len(shortstuff[i].period))
                        ws.write(row, col + 2, shortstuff[i].shortfall, style2db)
                        row += 1
                    row = 0
                    col = len(shrt_cols) + 1
                    ws.write(row, col, 'Generation Summary Table', styleb)
                    sum_cols = ['Name', 'Technology', 'Capacity (MW)', 'CF', 'Generated\n(to be\ncosted)']
                    if getattr(self.power_summary[0], 'transmitted') != None:
                        sum_cols.append('Transmitted\n(reduces\nShortfall)')
                    for i in range(len(sum_cols)):
                        ws.write(1, col + i, sum_cols[i], styleb)
                        j = sum_cols[i].find('\n') - 1
                        if j < 0:
                            j = len(sum_cols[i])
                        xl_lens[col + i] = j
                    for key, value in iter(stns.items()):
                        techs[self.power_summary[value].technology][0] += self.power_summary[value].capacity
                        techs[self.power_summary[value].technology][1] += self.power_summary[value].generation
                        if self.power_summary[value].transmitted is not None:
                            techs[self.power_summary[value].technology][2] += self.power_summary[value].transmitted
                    total = [0., 0., 0.]
                    row = 2
                    ws.write(row, col, 'Totals', styleb)
                    row += 1
                    for key, value in iter(sorted(techs.items())):
                        ws.write(row, col + 1, key)
                        ws.write(row, col + 2, value[0], style2db)
                        total[0] += value[0]
                        ws.write(row, col + 4, value[1], style0db)
                        total[1] += value[1]
                        if value[2] > 0:
                            ws.write(row, col + 5, value[2], style0d)
                            total[2] += value[2]
                        row += 1
                    ws.write(row, col + 1, 'Total', styleb)
                    ws.write(row, col + 2, total[0], style2db)
                    ws.write(row, col + 4, total[1], style0db)
                    if total[2] > 0:
                        ws.write(row, col + 5, total[2], style0d)
                    row += 1
                    ws.write(row, col, 'Stations', styleb)
                    row += 1
                    for key, value in iter(sorted(stns.items())):
                        ws.write(row, col, self.power_summary[value].name)
                        xl_lens[col] = max(xl_lens[col], len(self.power_summary[value].name))
                        ws.write(row, col + 1, self.power_summary[value].technology)
                        xl_lens[col + 1] = max(xl_lens[col + 1], len(self.power_summary[value].technology))
                        ws.write(row, col + 2, self.power_summary[value].capacity, style2d)
                        if self.power_summary[value].generation > 0:
                            ws.write(row, col + 3, self.power_summary[value].generation /
                                     (self.power_summary[value].capacity * 8760), style2d)
                        ws.write(row, col + 4, self.power_summary[value].generation, style0d)
                        if self.power_summary[value].transmitted is not None:
                            ws.write(row, col + 5, self.power_summary[value].transmitted, style0d)
                        row += 1
                    for key in xl_lens:
                        if xl_lens[key] * 275 > ws.col(key).width:
                            ws.col(key).width = xl_lens[key] * 275
                    ws.row(1).height_mismatch = True
                    ws.row(1).height = 256 * 3
                    ws.set_panes_frozen(True)
                    ws.set_horz_split_pos(2)
                    ws.set_remove_splits(True)
                    wb.save(data_file)

        def saveMatch(self, shortstuff):
            def cell_format(cell, new_cell):
                if cell.has_style:
                    new_cell.number_format = cell.number_format

            def cell_style(cell, new_cell, value=False):
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                    if value:
                        new_cell.value = copy(cell.value)

            def stn_rows(ss):
                if not self.plots['save_zone']:
                    for col in range(1, ss.max_column + 1):
                        try:
                            if ss.cell(row=1, column=col).value.lower() == 'zone':
                                ss.delete_cols(col)
                                break
                        except:
                            pass
                st_row = []
                st_col = []
                for i in range(len(type_tags)):
                    st_row.append(0)
                    st_col.append(0)
                for row in range(1, ss.max_row + 1):
                    for col in range(1, ss.max_column + 1):
                        try:
                            if ss.cell(row=row, column=col).value[0] != '<':
                                continue
                            elif ss.cell(row=row, column=col).value[:5] == '<stn_':
                                bit = ss.cell(row=row, column=col).value[:-1].split('_')
                                ty = type_tags.index(bit[-1])
                                st_row[ty] = row
                                st_col[ty] = col
                                ss.cell(row=row, column=col).value = None
                        except:
                            pass
                st = 0
                for key, value in iter(sorted(stns.items())):
                    try:
                        ss.cell(row=st_row[type_tags.index('name')] + st, column=st_col[type_tags.index('name')]).value = self.power_summary[value].name
                    except:
                        pass
                    if self.plots['save_zone']:
                        try:
                            ss.cell(row=st_row[type_tags.index('zone')] + st, column=st_col[type_tags.index('zone')]).value = self.power_summary[value].zone
                        except:
                            pass
                    try:
                        ss.cell(row=st_row[type_tags.index('tech')] + st, column=st_col[type_tags.index('tech')]).value = self.power_summary[value].technology
                    except:
                        pass
                    try:
                        ss.cell(row=st_row[type_tags.index('cap')] + st, column=st_col[type_tags.index('cap')]).value = self.power_summary[value].capacity
                        ss.cell(row=st_row[type_tags.index('cap')] + st, column=st_col[type_tags.index('cap')]).number_format = '#,##0.00'
                    except:
                        pass
                    if self.power_summary[value].generation > 0:
                        try:
                            ss.cell(row=st_row[type_tags.index('cf')] + st, column=st_col[type_tags.index('cf')]).value = self.power_summary[value].generation / \
                                 (self.power_summary[value].capacity * 8760)
                            ss.cell(row=st_row[type_tags.index('cf')] + st, column=st_col[type_tags.index('cf')]).number_format = '#,##0.00'
                        except:
                            pass
                    try:
                        ss.cell(row=st_row[type_tags.index('gen')] + st, column=st_col[type_tags.index('gen')]).value = self.power_summary[value].generation
                        ss.cell(row=st_row[type_tags.index('gen')] + st, column=st_col[type_tags.index('gen')]).number_format = '#,##0'
                    except:
                        pass
                    if self.power_summary[value].transmitted is not None:
                        try:
                            ss.cell(row=st_row[type_tags.index('tmit')] + st, column=st_col[type_tags.index('tmit')]).value = self.power_summary[value].transmitted
                            ss.cell(row=st_row[type_tags.index('tmit')] + st, column=st_col[type_tags.index('tmit')]).number_format = '#,##0'
                        except:
                            pass
                    st += 1

            ts = oxl.load_workbook(self.pm_template)
            ws = ts.worksheets[0]
            type_tags = ['name', 'zone', 'tech', 'cap', 'cf', 'gen', 'tmit', 'hrly']
            tech_tags = ['load', 'wind', 'offw', 'roof', 'fixed', 'single', 'dual', 'biomass', 'geotherm', 'other1', 'cst']
            tech_names = ['Load', 'Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV', 'Dual Axis PV',
                          'Biomass', 'Geothermal', 'Other1', 'CST']
            tech_names2 = [''] * len(tech_names)
            tech_names2[tech_names.index('Wind')] = 'Onshore Wind'
            tech_names2[tech_names.index('CST')] = 'Solar Thermal'
            tech_names2[tech_names.index('Dual Axis PV')] = 'Tracking PV'
            if self.plots['save_zone']:
                zone_techs = []
                for stn in self.stations:
                    try:
                        te = tech_names.index(stn.technology)
                    except:
                        try:
                            te = tech_names2.index(stn.technology)
                        except:
                            continue
                    zone_tech = '%s.%02d' % (stn.zone, te)
                    if zone_tech not in zone_techs:
                        zone_techs.append(zone_tech)
                zone_techs.sort()
                zone_techs.insert(0, 'load')
                tech_col = list(range(len(zone_techs)))
            else:
                i = type_tags.index('zone')
                del type_tags[i]
                heights = []
                # some code to delete the zone row and reapply cell formats after the delete
                del_zone = False
                for row in range(ws.max_row, 0, -1):
                    try:
                        if ws.cell(row=row, column=1).value.lower() == 'zone':
                            del_zone = True
                            ws.delete_rows(row)
                            break
                    except:
                        pass
                    heights.append(ws.row_dimensions[row].height)
                if del_zone:
                    for hght in reversed(heights):
                        ws.row_dimensions[row].height = hght
                        cell2 = ws.cell(row=row, column=2)
                        if cell2.value == 'Period':
                            cell = ws.cell(row=row, column=1)
                            ws.unmerge_cells('A' + str(row) + ':B' + str(row))
                            ws.cell(row=row, column=2).value = 'Period'
                            cell_style(cell, ws.cell(row=row, column=2))
                        row += 1
                tech_col = [0] * len(tech_tags)
            type_row = [0] * len(type_tags)
            per_row = [0, 0]
            per_col= [0, 0]
            # setup header rows
            dte = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                                  'yyyy-MM-dd hh:mm')
            if self.plots['save_zone']:
                have_zone = False
                for row in range(1, ws.max_row + 1):
                    try:
                        if ws.cell(row=row, column=1).value.lower() == 'zone':
                            have_zone = True
                            type_row[type_tags.index('zone')] = row
                        elif ws.cell(row=row, column=1).value.lower() == 'technology':
                            if not have_zone:
                                heights = []
                                # some code to insert the zone row and reapply cell formats after the insert
                                for row2 in range(row - 1, ws.max_row + 1):
                                    heights.append(ws.row_dimensions[row2].height)
                                ws.insert_rows(row)
                                ws.cell(row=row, column=1).value = 'Zone'
                                row2 = row
                                for hght in heights:
                                    ws.row_dimensions[row2].height = hght
                                    row2 += 1
                                for col in range(1, ws.max_column + 1):
                                    cell = ws.cell(row=row + 1, column=col)
                                    cell_style(ws.cell(row=row + 1, column=col), ws.cell(row=row, column=col))
                                for row2 in range(ws.max_row, row, -1):
                                    if ws.cell(row=row2, column=1).value == 'Hour':
                                        ws.merge_cells('A' + str(row2 - 1) + ':B' + str(row2 - 1))
                                        break
                                        try:
                                            if ws.cell(row=row + 1, column=1).value == 'Hour':
                                                ws.merge_cells('A' + str(row) + ':B' + str(row))
                                        except:
                                            pass
                                type_row[type_tags.index('zone')] = row
                                row += 1
                            type_row[type_tags.index('tech')] = row
                            break
                    except:
                        pass
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    try:
                        if ws.cell(row=row, column=col).value[0] != '<':
                            continue
                        if ws.cell(row=row, column=col).value == '<title>':
                            titl = ''
                            for stn in self.stations:
                                if stn.scenario not in titl:
                                    titl += stn.scenario + '; '
                            try:
                                titl = titl[:-2]
                                titl = titl.replace('.xlsx', '')
                                titl = titl.replace('.xls', '')
                                ws.cell(row=row, column=col).value = titl
                            except:
                                ws.cell(row=row, column=col).value = None
                        elif ws.cell(row=row, column=col).value == '<date>':
                            ws.cell(row=row, column=col).value = dte
                        elif ws.cell(row=row, column=col).value == '<period>':
                            per_row[1] = row
                            per_col[1] = col
                            ws.cell(row=row, column=col).value = None
                        elif ws.cell(row=row, column=col).value == '<hour>':
                            per_row[0] = row
                            per_col[0] = col
                            ws.cell(row=row, column=col).value = None
                        elif ws.cell(row=row, column=col).value == '<year>':
                            ws.cell(row=row, column=col).value = str(self.base_year)
                        elif ws.cell(row=row, column=col).value == '<growth>':
                            if self.load_multiplier != 0:
                                ws.cell(row=row, column=col).value = self.load_multiplier
                            else:
                                ws.cell(row=row, column=col).value = None
                        elif ws.cell(row=row, column=col).value.find('_') > 0:
                            bit = ws.cell(row=row, column=col).value[1:-1].split('_')
                            te = tech_tags.index(bit[0])
                            ty = type_tags.index(bit[-1])
                            type_row[ty] = row
                            if not self.plots['save_zone'] or te == 0:
                                tech_col[te] = col
                            ws.cell(row=row, column=col).value = None
                    except:
                        pass
            data_file = 'Powermatch_data_%s.xlsx' % (
                        QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                        'yyyy-MM-dd_hhmm'))
            data_file = QtWidgets.QFileDialog.getSaveFileName(None,
                        'Save Powermatch data file',
                        self.scenarios + data_file, 'Excel Files (*.xlsx)')[0]
            if data_file == '':
                return
            if data_file[-5:] != '.xlsx':
                data_file += '.xlsx'
            if os.path.exists(data_file):
                if os.path.exists(data_file + '~'):
                    os.remove(data_file + '~')
                os.rename(data_file, data_file + '~')
            # stations summaries
            stns = {}
            for i in range(len(self.power_summary)):
                stns[self.power_summary[i].name] = i
            stn_rows(ws)
            try: # in case they have a separate station worksheet
                ws2 = ts.worksheets[1]
                stn_rows(ws2)
            except:
                pass
            datas = {}
            if self.plots['save_zone']:
                for te in range(1, len(tech_col)):
                    tech_col[te] = tech_col[te - 1] + 1
                if tech_col[-1] > ws.max_column:
                    for col in range(ws.max_column, tech_col[-1]):
                        for row in range(1, ws.max_row + 1):
                            cell_style(ws.cell(row=row, column=col),
                                       ws.cell(row=row, column=col + 1),
                                       value=True)
                elif tech_col[-1] < ws.max_column:
                    ws.delete_cols(tech_col[-1] + 1, ws.max_column - tech_col[-1])
                for i in range(len(self.power_summary)):
                    try:
                        te = tech_names.index(self.power_summary[i].technology)
                    except:
                        try:
                            te = tech_names2.index(self.power_summary[i].technology)
                        except:
                            continue
                    key = '%s.%02d' % (self.power_summary[i].zone, te)
                    if key not in datas.keys():
                        datas[key] = [0., 0., 0.]
                    datas[key][0] += self.power_summary[i].capacity
                    datas[key][1] += self.power_summary[i].generation
                    if self.power_summary[i].transmitted is not None:
                        datas[key][2] += self.power_summary[i].transmitted
                for key, value in iter(datas.items()):
                    zt = zone_techs.index(key)
                    bits = key.split('.')
                    ws.cell(row=type_row[type_tags.index('zone')], column=tech_col[zt]).value = bits[0]
                    i = int(bits[1])
                    if tech_names2[i] != '':
                        tech_name = tech_names2[i]
                    else:
                        tech_name = tech_names[i]
                    ws.cell(row=type_row[type_tags.index('tech')], column=tech_col[zt]).value = tech_name
                    ws.cell(row=type_row[type_tags.index('cap')], column=tech_col[zt]).value = value[0]
                    ws.cell(row=type_row[type_tags.index('cap')], column=tech_col[zt]).value = value[0]
                    ws.cell(row=type_row[type_tags.index('gen')], column=tech_col[zt]).value = value[1]
                    if self.plots['grid_losses']:
                        ws.cell(row=type_row[type_tags.index('tmit')], column=tech_col[zt]).value = value[2]
                    if value[1] > 0:
                        ws.cell(row=type_row[type_tags.index('cf')], column=tech_col[zt]).value = \
                          value[1] / (value[0] * 8760)
                # merge zone cells
                row = type_row[type_tags.index('zone')]
                lst_zone = ws.cell(row=row, column=tech_col[1]).value
                lst_col = tech_col[1]
                for col in range(tech_col[1], tech_col[-1] + 1):
                    if ws.cell(row=row, column=col).value != lst_zone:
                        if col - lst_col > 1:
                            ws.merge_cells(start_row=row, start_column=lst_col, end_row=row, end_column=col - 1)
                        lst_col = col
                        lst_zone = ws.cell(row=row, column=col).value
            else:
                for i in range(len(self.power_summary)):
                    key = self.power_summary[i].technology
                    if key in datas.keys():
                        continue
                    datas[key] = [0., 0., 0.]
                for key, value in iter(stns.items()):
                    datas[self.power_summary[value].technology][0] += self.power_summary[value].capacity
                    datas[self.power_summary[value].technology][1] += self.power_summary[value].generation
                    if self.power_summary[value].transmitted is not None:
                        datas[self.power_summary[value].technology][2] += self.power_summary[value].transmitted
                for key, value in iter(datas.items()):
                    try:
                        te = tech_names.index(key)
                    except:
                        try:
                            te = tech_names2.index(key)
                        except:
                            continue
                    ws.cell(row=type_row[type_tags.index('cap')], column=tech_col[te]).value = value[0]
                    ws.cell(row=type_row[type_tags.index('gen')], column=tech_col[te]).value = value[1]
                    if self.plots['grid_losses']:
                        ws.cell(row=type_row[type_tags.index('tmit')], column=tech_col[te]).value = value[2]
                    if value[1] > 0:
                        ws.cell(row=type_row[type_tags.index('cf')], column=tech_col[te]).value = \
                          value[1] / (value[0] * 8760)
            if per_row[0] > 0:
                for i in range(8760):
                    ws.cell(row=per_row[0] + i, column=per_col[0]).value = shortstuff[i].hour
                    cell_format(ws.cell(row=per_row[0], column=per_col[0]), ws.cell(row=per_row[0] + i, column=per_col[0]))
            if per_row[1] > 0:
                for i in range(8760):
                    ws.cell(row=per_row[1] + i, column=per_col[1]).value = shortstuff[i].period
                    cell_format(ws.cell(row=per_row[1], column=per_col[1]), ws.cell(row=per_row[1] + i, column=per_col[1]))
            # load
            if type_row[-1] > 0:
                for i in range(8760):
                    ws.cell(row=type_row[-1] + i, column=tech_col[0]).value = shortstuff[i].load
                    cell_format(ws.cell(row=type_row[-1], column=tech_col[0]), ws.cell(row=type_row[-1] + i,
                                column=tech_col[0]))
            ly_keys = []
            if self.plots['save_zone']:
                ly_keys = []
                for zt in zone_techs:
                    ly_keys.append([])
                if self.plots['by_station']:
                    k = 0
                    for key in self.ly.keys():
                        try:
                            te = tech_names.index(self.stn_tech[k])
                        except:
                            try:
                                te = tech_names2.index(self.stn_tech[k])
                            except:
                                continue
                        zk = '%s.%02d' % (self.stn_zone[k], te)
                        te = zone_techs.index(zk)
                        ly_keys[te].append(key)
                        k += 1
                else:
                    for key in self.ly.keys():
                        bits = key.split('.')
                        try:
                            te = tech_names.index(bits[1])
                        except:
                            try:
                                te = tech_names2.index(bits[1])
                            except:
                                continue
                        zk = '%s.%02d' % (bits[0], te)
                        te = zone_techs.index(zk)
                        ly_keys[te].append(key)
                for te in range(len(zone_techs)):
                    if len(ly_keys[te]) == 0:
                        continue
                    hrly = [0.] * 8760
                    doit = False
                    for key in ly_keys[te]:
                        try:
                            values = self.ly[key]
                            for h in range(len(hrly)):
                                hrly[h] += values[h]
                                if hrly[h] != 0:
                                    doit = True
                        except:
                            pass
                    if doit or not doit:
                        for h in range(len(hrly)):
                            ws.cell(row=type_row[-1] + h, column=tech_col[te]).value = hrly[h]
                            cell_format(ws.cell(row=type_row[-1], column=tech_col[te]), \
                                        ws.cell(row=type_row[-1] + h, column=tech_col[te]))
            else:
                for t in range(len(tech_names)):
                    ly_keys.append([])
                if self.plots['by_station']:
                    k = 0
                    for key in self.ly.keys():
                        try:
                            te = tech_names.index(self.stn_tech[k])
                        except:
                            try:
                                te = tech_names2.index(self.stn_tech[k])
                            except:
                                continue
                        ly_keys[te].append(key)
                        k += 1
                else:
                    for key in self.ly.keys():
                        try:
                            te = tech_names.index(key)
                        except:
                            try:
                                te = tech_names2.index(key)
                            except:
                                continue
                        ly_keys[te].append(key)
                for te in range(len(tech_col)):
                    if tech_col[te] == 0 or len(ly_keys[te]) == 0:
                        continue
                    hrly = [0.] * 8760
                    doit = False
                    for key in ly_keys[te]:
                        try:
                            values = self.ly[key]
                            for h in range(len(hrly)):
                                hrly[h] += values[h]
                                if hrly[h] != 0:
                                    doit = True
                        except:
                            pass
                    if doit or not doit:
                        for h in range(len(hrly)):
                            ws.cell(row=type_row[-1] + h, column=tech_col[te]).value = hrly[h]
                            cell_format(ws.cell(row=type_row[-1], column=tech_col[te]), \
                                        ws.cell(row=type_row[-1] + h, column=tech_col[te]))
            ts.save(data_file)

        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        try:
            mapc = config.get('Map', 'map_choice')
        except:
            mapc = ''
        self.colours = {'cumulative': '#006400', 'gross_load': '#a9a9a9', 'load': '#000000',
                        'shortfall': '#8b0000', 'wind': '#6688bb'}
        try:
            colors = config.items('Colors')
            for item, colour in colors:
                if item in self.technologies or item in self.colours:
                    itm = techClean(item)
                    self.colours[itm] = colour
        except:
            pass
        if mapc != '':
            try:
                colors = config.items('Colors' + mapc)
                for item, colour in colors:
                    if item in self.technologies or item in self.colours:
                        itm = techClean(item)
                        self.colours[itm] = colour
            except:
                pass
        papersizes = {'a0': '33.1,46.8', 'a1': '23.4,33.1', 'a2': '16.5,23.4',
                    'a3': '11.7,16.5', 'a4': '8.3,11.7', 'a5': '5.8,8.3',
                    'a6': '4.1,5.8', 'a7': '2.9,4.1', 'a8': '2,2.9',
                    'a9': '1.5,2', 'a10': '1,1.5', 'b0': '39.4,55.7',
                    'b1': '27.8,39.4', 'b2': '19.7,27.8', 'b3': '13.9,19.7',
                    'b4': '9.8,13.9', 'b5': '6.9,9.8', 'b6': '4.9,6.9',
                    'b7': '3.5,4.9', 'b8': '2.4,3.5', 'b9': '1.7,2.4',
                    'b10': '1.2,1.7', 'foolscap': '8.0,13.0', 'ledger': '8.5,14.0',
                    'legal': '8.5,14.09', 'letter': '8.5,11.0'}
        landscape = False
        papersize = ''
        self.other_width = 2.
        seasons = []
        periods = []
        try:
            items = config.items('Power')
        except:
            seasons = [[], [], [], []]
            seasons[0] = ['Summer', 11, 0, 1]
            seasons[1] = ['Autumn', 2, 3, 4]
            seasons[2] = ['Winter', 5, 6, 7]
            seasons[3] = ['Spring', 8, 9, 10]
            periods = [[], []]
            periods[0] = ['Winter', 4, 5, 6, 7, 8, 9]
            periods[1] = ['Summer', 10, 11, 0, 1, 2, 3]
        for item, values in items:
            if item[:6] == 'season':
                if item == 'season':
                    continue
                i = int(item[6:]) - 1
                if i >= len(seasons):
                    seasons.append([])
                seasons[i] = values.split(',')
                for j in range(1, len(seasons[i])):
                    seasons[i][j] = int(seasons[i][j]) - 1
            elif item[:6] == 'period':
                if item == 'period':
                    continue
                i = int(item[6:]) - 1
                if i >= len(periods):
                    periods.append([])
                periods[i] = values.split(',')
                for j in range(1, len(periods[i])):
                    periods[i][j] = int(periods[i][j]) - 1
            elif item == 'other_width':
                try:
                    self.other_width = float(values)
                except:
                    pass
            elif item == 'save_format':
                plt.rcParams['savefig.format'] = values
            elif item == 'figsize':
                try:
                    papersize = papersizes[values]
                except:
                    papersize = values
            elif item == 'orientation':
                if values.lower()[0] == 'l':
                    landscape = True
            elif item == 'debug_sam':
                if values.lower() in ['true', 'yes', 'on']:
                    self.debug = True
                else:
                    self.debug = False
        if papersize != '':
            if landscape:
                bit = papersize.split(',')
                plt.rcParams['figure.figsize'] = bit[1] + ',' + bit[0]
            else:
                plt.rcParams['figure.figsize'] = papersize
        try:
            self.pm_template = config.get('Power', 'pm_template')
        except:
            try:
                self.pm_template = config.get('Files', 'pm_template')
            except:
                self.pm_template = False
        if self.pm_template:
            try:
                parents = getParents(config.items('Parents'))
                for key, value in parents:
                    self.pm_template = self.pm_template.replace(key, value)
                self.pm_template = self.pm_template.replace('$USER$', getUser())
                if not os.path.exists(self.pm_template):
                    self.pm_template = False
            except:
                pass
        mth_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct',
                      'Nov', 'Dec']
        ssn_labels = []
        for i in range(len(seasons)):
            if len(seasons[i]) == 2:
                ssn_labels.append('%s (%s)' % (seasons[i][0], mth_labels[seasons[i][1]]))
            else:
                ssn_labels.append('%s (%s-%s)' % (seasons[i][0], mth_labels[seasons[i][1]],
                                   mth_labels[seasons[i][-1]]))
        smp_labels = []
        for i in range(len(periods)):
            if len(periods[i]) == 2:
                smp_labels.append('%s (%s)' % (periods[i][0], mth_labels[periods[i][1]]))
            else:
                smp_labels.append('%s (%s-%s)' % (periods[i][0], mth_labels[periods[i][1]],
                                   mth_labels[periods[i][-1]]))
        labels = ['0:00', '4:00', '8:00', '12:00', '16:00', '20:00', '24:00']
        mth_xlabels = ['0:', '4:', '8:', '12:', '16:', '20:', '24:']
        pct_labels = ['0%', '20%', '40%', '60%', '80%', '100%']
        m = 0
        d = 1
        day_labels = []
        while m < len(the_days):
            day_labels.append('%s %s' % (str(d), mth_labels[m]))
            d += 7
            if d > the_days[m]:
                d = d - the_days[m]
                m += 1
        lbl_font = FontProperties()
        lbl_font.set_size('small')
        x24 = []
        l24 = {}
        m24 = {}
        q24 = {}
        s24 = {}
        d365 = {}
        for i in range(24):
            x24.append(i + 1)
        for key in list(ydata.keys()):
            if self.plots['total']:
                l24[key] = []
                for j in range(24):
                    l24[key].append(0.)
            if self.plots['month'] or self.plots['by_month']:
                m24[key] = []
                for m in range(12):
                    m24[key].append([])
                    for j in range(24):
                        m24[key][m].append(0.)
            if self.plots['season'] or self.plots['by_season']:
                q24[key] = []
                for q in range(len(seasons)):
                    q24[key].append([])
                    for j in range(24):
                        q24[key][q].append(0.)
            if self.plots['period'] or self.plots['by_period']:
                s24[key] = []
                for s in range(len(periods)):
                    s24[key].append([])
                    for j in range(24):
                        s24[key][s].append(0.)
            if self.plots['by_day']:
                d365[key] = []
                for j in range(365):
                    d365[key].append([0.])
        the_qtrs = []
        for i in range(len(seasons)):
            d = 0
            for j in range(1, len(seasons[i])):
                d += the_days[seasons[i][j]]
            the_qtrs.append(d)
        the_ssns = []
        for i in range(len(periods)):
            d = 0
            for j in range(1, len(periods[i])):
                d += the_days[periods[i][j]]
            the_ssns.append(d)
        the_hours = [0]
        i = 0
        for m in range(len(the_days)):
            i = i + the_days[m] * 24
            the_hours.append(i)
        d = -1
        for i in range(0, len(x), 24):
            m = 11
            d += 1
            while i < the_hours[m] and m > 0:
                m -= 1
            for k in range(24):
                for key, value in iter(sorted(ydata.items())):
                    if key == 'Generation':
                        continue
                    if self.plots['total']:
                        l24[key][k] += value[i + k]
                    if self.plots['by_day']:
                        d365[key][d][0] += value[i + k]
                    if self.plots['month'] or self.plots['by_month']:
                        m24[key][m][k] = m24[key][m][k] + value[i + k]
                    if self.plots['season'] or self.plots['by_season']:
                        for q in range(len(seasons)):
                            if m in seasons[q]:
                                break
                        q24[key][q][k] = q24[key][q][k] + value[i + k]
                    if self.plots['period'] or self.plots['by_period']:
                        for s in range(len(periods)):
                            if m in periods[s]:
                                break
                        s24[key][s][k] = s24[key][s][k] + value[i + k]
        if self.plots['cumulative']:
            pc = 1
        else:
            pc = 0
        if self.plots['gross_load']:
            pc += 1
        if self.plots['shortfall']:
            pc += 1
        colours = ['r', 'g', 'b', 'c', 'm', 'y', 'orange', 'darkcyan', 'darkmagenta',
                   'darkolivegreen', 'darkorange', 'darkturquoise', 'darkviolet', 'violet']
        colour_index = 0
        linestyles = ['-', '--', '-.', ':']
        line_index = 0
        self.linestyle = {}
        for key in self.colours:
            self.linestyle[key] = '-'
        for key in ydata:
            if key not in self.colours:
                if key[:4] == 'Load':
                    try:
                        self.colours[key] = self.colours['load']
                    except:
                        self.colours[key] = 'black'
                    self.linestyle[key] = '-'
                else:
                    self.colours[key] = colours[colour_index]
                    self.linestyle[key] = linestyles[line_index]
                    colour_index += 1
                    if colour_index >= len(colours):
                        colour_index = 0
                        line_index += 1
                        if line_index >= len(linestyles):
                            line_index = 0
        if self.plots['by_day']:
            stepPlot(self, 'day', d365, day_labels)
        if self.plots['by_month']:
            stepPlot(self, 'month', m24, mth_labels)
        if self.plots['by_season']:
            stepPlot(self, 'season', q24, ssn_labels)
        if self.plots['by_period']:
            stepPlot(self, 'period', s24, smp_labels)
        for key in list(ydata.keys()):
            for k in range(24):
                if self.plots['total']:
                    l24[key][k] = l24[key][k] / 365
                if self.plots['month']:
                    for m in range(12):
                        m24[key][m][k] = m24[key][m][k] / the_days[m]
                if self.plots['season']:
                    for q in range(len(seasons)):
                        q24[key][q][k] = q24[key][q][k] / the_qtrs[q]
                if self.plots['period']:
                    for s in range(len(periods)):
                        s24[key][s][k] = s24[key][s][k] / the_ssns[s]
        if self.plots['hour']:
            if self.plots['save_plot']:
                sp_vals = ['hour']
                sp_data = []
                sp_data.append(x[1:])
                sp_data[-1].append(len(x))
                sp_vals.append('period')
                sp_data.append([])
                for i in range(len(x)):
                    sp_data[-1].append(the_date(self.load_year, i))
            hdr = self.hdrs['hour'].replace('Power - ', '')
            fig = plt.figure(hdr + self.suffix)
            plt.grid(True)
            hx = plt.subplot(111)
            plt.title(self.hdrs['hour'] + self.suffix)
            maxy = 0
            storage = None
            if self.plots['cumulative']:
                cumulative = []
                for i in range(len(x)):
                    cumulative.append(0.)
            if self.plots['gross_load']:
                gross_load = []
                for i in range(len(x)):
                    gross_load.append(0.)
            if self.plots['show_pct']:
                load_sum = 0.
                gen_sum = 0.
            for key, value in iter(sorted(ydata.items())):
                if key == 'Generation':
                    continue
                if self.plots['show_pct']:
                    for i in range(len(x)):
                        if key[:4] == 'Load':
                            load_sum += value[i]
                        elif key == 'Storage':
                            pass
                        else:
                            gen_sum += value[i]
                            if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                load_sum += value[i]
                maxy = max(maxy, max(value))
                lw = self.other_width
                if self.plots['cumulative'] and key[:4] != 'Load' and key != 'Storage':
                    lw = 1.0
                    for i in range(len(x)):
                        cumulative[i] += value[i]
                if self.plots['gross_load'] and (key[:4] == 'Load' or
                      key == 'Existing Rooftop PV'):
                    for i in range(len(x)):
                        gross_load[i] += value[i]
                if self.plots['shortfall'] and key[:4] == 'Load':
                    load = value
                if self.plots['shortfall'] and key == 'Storage':
                    storage = value
                hx.plot(x, value, linewidth=lw, label=shrinkKey(key), color=self.colours[key],
                        linestyle=self.linestyle[key])
                if self.plots['save_plot']:
                    sp_vals.append(shrinkKey(key))
                    sp_data.append(value)
            if self.plots['cumulative']:
                hx.plot(x, cumulative, linewidth=self.other_width, label='Tot. Generation', color=self.colours['cumulative'])
                maxy = max(maxy, max(cumulative))
                if self.plots['save_plot']:
                    sp_vals.append('Tot. Generation')
                    sp_data.append(cumulative)
            if self.plots['gross_load'] and 'Existing Rooftop PV' in list(ydata.keys()):
                hx.plot(x, gross_load, linewidth=1.0, label='Gross Load', color=self.colours['gross_load'])
                maxy = max(maxy, max(gross_load))
                if self.plots['save_plot']:
                    sp_vals.append('Gross Load')
                    sp_data.append(gross_load)
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                rndup = 0
            if self.plots['shortfall'] and self.do_load:
                load2 = []
                if storage is None:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] - load[i])
                else:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] + storage[i] - load[i])
                hx.plot(x, load2, linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                plt.axhline(0, color='black')
                miny = min(load2)
                if rndup != 0 and miny < 0:
                    miny = -ceil(-miny / rndup) * rndup
                if self.plots['save_plot']:
                    sp_vals.append('Shortfall')
                    sp_data.append(load2)
            else:
                miny = 0
            if self.plots['save_plot']:
                titl = 'Hour'
                decpts = [3] * len(sp_vals)
                decpts[0] = decpts[1] = 0
                dialog = displaytable.Table(list(map(list, list(zip(*sp_data)))), title=titl, fields=sp_vals,
                                            save_folder=self.scenarios, decpts=decpts)
                dialog.exec_()
                del dialog, sp_data, sp_vals
            plt.ylim([miny, maxy])
            plt.xlim([0, len(x)])
            if (len(ydata) + pc) > 9:
                 # Shrink current axis by 5%
                box = hx.get_position()
                hx.set_position([box.x0, box.y0, box.width * 0.95, box.height])
                 # Put a legend to the right of the current axis
                hx.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop=lbl_font)
            else:
                hx.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=(len(ydata) + pc),
                          prop=lbl_font)
            plt.xticks(list(range(12, len(x), 168)))
            hx.set_xticklabels(day_labels, rotation='vertical')
            hx.set_xlabel('Month of the year')
            hx.set_ylabel('Power (MW)')
            if self.plots['show_pct']:
                self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                plt.title(self.hdrs['hour'] + self.suffix + self.gen_pct)
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(hx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()
            del zp
        if self.plots['augment'] and self.do_load:
            hdr = self.hdrs['augment'].replace('Power - ', '')
            fig = plt.figure(hdr + self.suffix)
            plt.grid(True)
            hx = plt.subplot(111)
            plt.title(self.hdrs['augment'] + self.suffix)
            maxy = 0
            miny = 0
            storage = None
            cumulative = []
            for i in range(len(x)):
                cumulative.append(0.)
            if self.plots['gross_load']:
                gross_load = []
                for i in range(len(x)):
                    gross_load.append(0.)
            if self.plots['show_pct']:
                load_sum = 0.
                gen_sum = 0.
            for key, value in iter(sorted(ydata.items())):
                if key == 'Generation' or key == 'Excess': # might need to keep excess
                    continue
                if self.plots['show_pct']:
                    for i in range(len(x)):
                        if key[:4] == 'Load':
                            load_sum += value[i]
                        elif key == 'Storage':
                            pass
                        else:
                            gen_sum += value[i]
                            if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                load_sum += value[i]
                maxy = max(maxy, max(value))
                lw = self.other_width
                if key[:4] != 'Load' and key != 'Storage':
                    for i in range(len(x)):
                        cumulative[i] += value[i]
                if self.plots['gross_load'] and (key[:4] == 'Load' or
                      key == 'Existing Rooftop PV'):
                    for i in range(len(x)):
                        gross_load[i] += value[i]
                if key[:4] == 'Load':
                    load = value
                if key == 'Storage':
                    storage = value
            maxy = max(maxy, max(cumulative))
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                rndup = 0
            regen = cumulative[:]
            for r in range(len(regen)):
                if regen[r] > load[r]:
                    regen[r] = load[r]
            hx.fill_between(x, 0, regen, color=self.colours['cumulative']) #'#004949')
            if storage is not None:
                for r in range(len(storage)):
                    storage[r] += cumulative[r]
                for r in range(len(storage)):
                    if storage[r] > load[r]:
                        storage[r] = load[r]
                hx.fill_between(x, regen, storage, color=self.colours['wind']) #'#006DDB')
                hx.fill_between(x, storage, load, color=self.colours['shortfall']) #'#920000')
            else:
                hx.fill_between(x, load, regen, color=self.colours['shortfall']) #'#920000')
            hx.plot(x, cumulative, linewidth=self.other_width, label='RE', linestyle='--', color=self.colours['gross_load'])
            if self.plots['save_plot']:
                sp_vals = ['hour']
                sp_data = []
                sp_tots = ['']
                sp_pts = [0]
                sp_data.append(x[1:])
                sp_data[-1].append(len(x))
                sp_vals.append('period')
                sp_data.append([])
                sp_tots.append('')
                sp_pts.append(0)
                for i in range(len(x)):
                    sp_data[-1].append(the_date(self.load_year, i))
                sp_vals.append('load')
                sp_data.append(load)
                l = len(sp_data) - 1
                sp_tots.append(0.)
                sp_pts.append(4)
                for ld in load:
                    sp_tots[l] += ld
                sp_vals.append('renewable')
                sp_data.append(regen)
                r = len(sp_data) - 1
                sp_tots.append(0.)
                sp_pts.append(4)
                for re in regen:
                    sp_tots[-1] += re
                if storage is not None:
                    sp_vals.append('storage')
                    sp_data.append(storage)
                    s = len(sp_data) - 1
                    sp_tots.append(0.)
                    sp_pts.append(4)
                else:
                    s = 0
                sp_vals.append('re gen.')
                sp_data.append(cumulative)
                e = len(sp_data) - 1
                sp_tots.append(0.)
                sp_pts.append(4)
                titl = 'Augmented'
                dialog = displaytable.Table(list(map(list, list(zip(*sp_data)))), title=titl, fields=sp_vals,
                                            save_folder=self.scenarios, decpts=sp_pts)
                dialog.exec_()
                del dialog
                if s > 0:
                    for i in range(len(sp_data[s])):
                        sp_data[s][i] = sp_data[s][i] - sp_data[r][i]
                        sp_tots[s] += sp_data[s][i]
                sp_vals.append('excess')
                sp_vals[e] = 'augment'
                sp_data.append([])
                sp_tots.append(0.)
                sp_pts.append(4)
                for i in range(len(sp_data[r])):
                    sp_data[-1].append(sp_data[e][i] - sp_data[r][i])
                    sp_tots[-1] += sp_data[-1][i]
                if s > 0:
                    for i in range(len(sp_data[e])):
                        sp_data[e][i] = sp_data[l][i] - sp_data[r][i] - sp_data[s][i]
                        sp_tots[e] += sp_data[e][i]
                        sp_data[-1][i] -= sp_data[s][i]
                        sp_tots[-1] -= sp_data[s][i]
                else:
                    for i in range(len(sp_data[e])):
                        sp_data[e][i] = sp_data[l][i] - sp_data[r][i]
                        sp_tots[e] += sp_data[e][i]
                titl = 'augmented2'
                dialog = displaytable.Table(list(map(list, list(zip(*sp_data)))), title=titl, fields=sp_vals,
                                            save_folder=self.scenarios, decpts=sp_pts)
                dialog.exec_()
                fields = ['row', 'component', 'MWh', 'Load %']
                values = []
                sp_pts = [0, 0, 4, 1]
                for i in range(2, len(sp_vals)):
                    values.append([i - 1, sp_vals[i].title(), sp_tots[i], 0.])
                    values[-1][-1] = (sp_tots[i] * 100.) / sp_tots[l]
                titl = 'augmented3'
                dialog = displaytable.Table(values, fields=fields, title=titl, save_folder=self.scenarios, decpts=sp_pts)
                dialog.exec_()
                del dialog, sp_vals, sp_data, sp_tots
            plt.ylim([miny, maxy])
            plt.xlim([0, len(x)])
            plt.xticks(list(range(12, len(x), 168)))
            hx.set_xticklabels(day_labels, rotation='vertical')
            hx.set_xlabel('Month of the year')
            hx.set_ylabel('Power (MW)')
            zp = ZoomPanX()
            f = zp.zoom_pan(hx, base_scale=1.2) # enable scrollable zoom
            if self.plots['show_pct']:
                self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                plt.title(self.hdrs['hour'] + self.suffix + self.gen_pct)
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(hx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()
            del zp
    #        shortstuff = []
    #        vals = ['load', 'renewable', 'storage', 'cumulative']
    #        for i in range(0, len(load)):
    #            if storage is None:
    #                shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i), [load[i],
    #                                  regen[i], 0., cumulative[i]], values=vals))
    #            else:
    #                shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i), [load[i],
    #                                  regen[i], storage[i], cumulative[i]], values=vals))
    #        vals.insert(0, 'period')
    #        vals.insert(0, 'hour')
    #        dialog = displaytable.Table(shortstuff, title='Augmented',
    #                                        save_folder=self.scenarios, fields=vals)
    #        dialog.exec_()
    #        del dialog
        if self.plots['duration']:
            hdr = self.hdrs['duration'].replace('Power - ', '')
            fig = plt.figure(hdr + self.suffix)
            plt.grid(True)
            dx = plt.subplot(111)
            plt.title(self.hdrs['duration'] + self.suffix)
            maxy = 0
            if self.plots['cumulative']:
                cumulative = []
                for i in range(len(x)):
                    cumulative.append(0.)
            if self.plots['show_pct']:
                load_sum = 0.
                gen_sum = 0.
            for key, value in iter(sorted(ydata.items())):
                if key == 'Generation':
                    continue
                if self.plots['show_pct']:
                    for i in range(len(x)):
                        if key[:4] == 'Load':
                            load_sum += value[i]
                        elif key == 'Storage':
                            pass
                        else:
                            gen_sum += value[i]
                            if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                load_sum += value[i]
                sortydata = sorted(value, reverse=True)
                lw = self.other_width
                if self.plots['cumulative'] and key[:4] != 'Load' and key != 'Storage':
                    lw = 1.0
                    for i in range(len(x)):
                        cumulative[i] += value[i]
                maxy = max(maxy, max(sortydata))
                dx.plot(x, sortydata, linewidth=lw, label=shrinkKey(key), color=self.colours[key],
                        linestyle=self.linestyle[key])
            if self.plots['cumulative']:
                sortydata = sorted(cumulative, reverse=True)
                dx.plot(x, sortydata, linewidth=self.other_width, label='Tot. Generation', color=self.colours['cumulative'])
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                pass
            plt.ylim([0, maxy])
            plt.xlim([0, len(x)])
            if (len(ydata) + pc) > 9:
                 # Shrink current axis by 10%
                box = dx.get_position()
                dx.set_position([box.x0, box.y0, box.width * 0.95, box.height])

                 # Put a legend to the right of the current axis
                dx.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop=lbl_font)
            else:
                dx.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=(len(ydata) + pc),
                          prop=lbl_font)
            tics = int(len(x) / (len(pct_labels) - 1))
            plt.xticks(list(range(0, len(x) + 1, tics)))
            dx.set_xticklabels(pct_labels)
            dx.set_xlabel('Percentage of Year')
            dx.set_ylabel('Power (MW)')
            if self.plots['show_pct']:
                self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                plt.title(self.hdrs['duration'] + self.suffix + self.gen_pct)
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(dx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()
            if self.do_load:
                hdr = self.hdrs['duration'].replace('Power - ', '')
           #      fig = plt.figure(hdr + self.suffix)
                plt.figure(hdr + ' 2')
                plt.grid(True)
                plt.title(self.hdrs['duration'] + ' with renewable contribution')
                lx = plt.subplot(111)
                maxy = 0
                miny = 0
                load = []  # use for this and next graph
                rgen = []  # use for this and next graph
                rgendiff = []
                for i in range(len(self.x)):
                    rgen.append(0.)
                    rgendiff.append(0.)
                if self.plots['show_pct']:
                    load_sum = 0.
                    gen_sum = 0.
                for key, value in ydata.items():
                    if key == 'Generation':
                        continue
                    if self.plots['show_pct']:
                        for i in range(len(value)):
                            if key[:4] == 'Load':
                                load_sum += value[i]
                            elif key == 'Storage':
                                pass
                            else:
                                gen_sum += value[i]
                                if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                    load_sum += value[i]
                    if key[:4] == 'Load':
                        load = value
                    else:
                        for i in range(len(value)):
                            rgen[i] += value[i]
                for i in range(len(load)):
                    rgendiff[i] = load[i] - rgen[i]
                sortly1 = sorted(load, reverse=True)
                maxy = max(maxy, max(load))
                maxy = max(maxy, max(rgendiff))
                miny = min(miny, min(rgendiff))
                try:
                    rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                    maxy = ceil(maxy / rndup) * rndup
                    miny = -ceil(-miny / rndup) * rndup
                except:
                    pass
                if self.load_multiplier != 0:
                    load_key = 'Load ' + self.load_year
                else:
                    load_key = 'Load'
                lx.plot(x, sortly1, linewidth=self.other_width, label=load_key)
                sortly2 = sorted(rgendiff, reverse=True)
                lx.plot(x, sortly2, linewidth=self.other_width, label='Tot. Generation')
                lx.fill_between(x, sortly1, sortly2, facecolor=self.colours['cumulative'])
                plt.ylim([miny, maxy])
                plt.xlim([0, len(x)])
                lx.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=2, prop=lbl_font)
                tics = int(len(x) / (len(pct_labels) - 1))
                plt.xticks(list(range(0, len(x) + 1, tics)))
                lx.set_xticklabels(pct_labels)
                lx.set_xlabel('Percentage of Year')
                lx.set_ylabel('Power (MW)')
                lx.axhline(0, color='black')
                if self.plots['show_pct']:
                    self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                    plt.title(self.hdrs['duration'] + ' with renewable contribution' +
                              self.gen_pct)
                if self.plots['maximise']:
                    mng = plt.get_current_fig_manager()
                    if sys.platform == 'win32' or sys.platform == 'cygwin':
                        if plt.get_backend() == 'TkAgg':
                            mng.window.state('zoomed')
                        elif plt.get_backend() == 'Qt5Agg':
                            mng.window.showMaximized()
                    else:
                        mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(lx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()
        if not self.plots['block']:
            plt.show(block=True)
        if (self.plots['shortfall_detail'] or self.plots['save_match']) and self.do_load:
            load = []
            rgen = []
            shortfall = [[], [], [], []]
            generation = []
            for i in range(len(self.x)):
                rgen.append(0.)
                shortfall[0].append(0.)
            for key, value in ydata.items():
                if key == 'Generation':
                    generation = value
                elif key[:4] == 'Load':
                    load = value
                else:
                    for i in range(len(value)):
                        rgen[i] += value[i]
            shortfall[0][0] = rgen[0] - load[0]
            for i in range(1, len(load)):
                shortfall[0][i] = shortfall[0][i - 1] + rgen[i] - load[i]
            d_short = [[], [0], [0, 0]]
            for i in range(0, len(load), 24):
                d_short[0].append(0.)
                for j in range(i, i + 24):
                    d_short[0][-1] += rgen[i] - load[i]
            if self.iterations > 0:
                for i in range(1, len(d_short[0])):
                    d_short[1].append((d_short[0][i - 1] + d_short[0][i]) / 2)
                for i in range(2, len(d_short[0])):
                    d_short[2].append((d_short[0][i - 2] + d_short[0][i - 1] + d_short[0][i]) / 3)
                d_short[1][0] = d_short[1][1]
                d_short[2][0] = d_short[2][1] = d_short[2][2]
                shortstuff = []
                vals = ['shortfall', 'iteration 1', 'iteration 2']
                for i in range(len(d_short[0])):
                    shortstuff.append(DailyData(i + 1, the_date(self.load_year, i * 24)[:10],
                                      [d_short[0][i], d_short[1][i], d_short[2][i]], values=vals))
                vals.insert(0, 'date')
                vals.insert(0, 'day')
                dialog = displaytable.Table(shortstuff, title='Daily Shortfall',
                         save_folder=self.scenarios, fields=vals)
                dialog.exec_()
                del dialog
                del shortstuff
                xs = []
                for i in range(len(d_short[0])):
                    xs.append(i + 1)
                plt.figure('daily shortfall')
                plt.grid(True)
                plt.title('Daily Shortfall')
                sdfx = plt.subplot(111)
                for i in range(self.iterations):
                    sdfx.step(xs, d_short[i], linewidth=self.other_width,
                              label=str(i + 1) + ' day average',
                              color=colours[i])
                plt.xticks(list(range(0, len(xs), 7)))
                tick_spot = []
                for i in range(0, len(xs), 7):
                    tick_spot.append(i + .5)
                sdfx.set_xticks(tick_spot)
                sdfx.set_xticklabels(day_labels, rotation='vertical')
                sdfx.set_xlabel('Day of the year')
                sdfx.set_ylabel('Power (MW)')
                plt.xlim([0, len(xs)])
                sdfx.legend(loc='best')
                for i in range(len(d_short)):
                    lin = min(d_short[i])
                    sdfx.axhline(lin, linestyle='--', color=colours[i])
                    lin = max(d_short[i])
                    sdfx.axhline(lin, linestyle='--', color=colours[i])
                lin = sum(d_short[0]) / len(d_short[0])
                sdfx.axhline(lin, linestyle='--', color='black')
                if self.plots['maximise']:
                    mng = plt.get_current_fig_manager()
                    if sys.platform == 'win32' or sys.platform == 'cygwin':
                        if plt.get_backend() == 'TkAgg':
                            mng.window.state('zoomed')
                        elif plt.get_backend() == 'Qt5Agg':
                            mng.window.showMaximized()
                    else:
                        mng.resize(*mng.window.maxsize())
                zp = ZoomPanX()
                f = zp.zoom_pan(sdfx, base_scale=1.2) # enable scrollable zoom
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
                h_storage = [-(shortfall[0][-1] / len(shortfall[0]))]  # average shortfall
                for s in range(1, self.iterations + 1):
                    for i in range(len(self.x)):
                        shortfall[s].append(0.)
                    ctr = 0
                    still_short = [0, 0]
                    if rgen[0] - load[0] + h_storage[-1] < 0:
                        still_short[0] += rgen[0] - load[0] + h_storage[-1]
                        ctr += 1
                    else:
                        still_short[1] += rgen[0] - load[0] + h_storage[-1]
                    shortfall[s][0] = rgen[0] - load[0] + h_storage[-1]
                    for i in range(1, len(load)):
                        shortfall[s][i] = shortfall[s][i - 1] + rgen[i] - load[i] + h_storage[-1]
                        if rgen[i] - load[i] + h_storage[-1] < 0:
                            still_short[0] += rgen[i] - load[i] + h_storage[-1]
                            ctr += 1
                        else:
                            still_short[1] += rgen[i] - load[i] + h_storage[-1]
    #                 h_storage.append(h_storage[-1] - still_short[0] / len(self.x))
                    h_storage.append(-(shortfall[0][-1] + still_short[0]) / len(self.x))
                dimen = log10(fabs(shortfall[0][-1]))
                unit = 'MW'
                if dimen > 11:
                    unit = 'PW'
                    div = 9
                elif dimen > 8:
                    unit = 'TW'
                    div = 6
                elif dimen > 5:
                    unit = 'GW'
                    div = 3
                else:
                    div = 0
                if div > 0:
                    for s in range(self.iterations + 1):
                        for i in range(len(shortfall[s])):
                            shortfall[s][i] = shortfall[s][i] / pow(10, div)
                plt.figure('cumulative shortfall')
                plt.grid(True)
                plt.title('Cumulative Shortfall')
                sfx = plt.subplot(111)
                sfx.plot(x, shortfall[0], linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                for s in range(1, self.iterations + 1):
                    if h_storage[s - 1] > 1:
                        amt = '{:0,.0f}'.format(h_storage[s - 1])
                    else:
                        amt = '{:0,.1f}'.format(h_storage[s - 1])
                    lbl = 'iteration %s - add %s MW to generation' % (s, amt )
                    sfx.plot(x, shortfall[s], linewidth=self.other_width, label=lbl, color=colours[s])
                plt.xticks(list(range(0, len(x), 168)))
                tick_spot = []
                for i in range(0, len(x), 168):
                    tick_spot.append(i + .5)
                box = sfx.get_position()
                sfx.set_position([box.x0, box.y0, box.width, box.height])
                sfx.set_xticks(tick_spot)
                sfx.set_xticklabels(day_labels, rotation='vertical')
                plt.xlim([0, len(x)])
                sfx.set_xlabel('Day of the year')
                sfx.set_ylabel('Power (' + unit + ')')
                sfx.legend(loc='best', prop=lbl_font)
                if self.plots['maximise']:
                    mng = plt.get_current_fig_manager()
                    if sys.platform == 'win32' or sys.platform == 'cygwin':
                        if plt.get_backend() == 'TkAgg':
                            mng.window.state('zoomed')
                        elif plt.get_backend() == 'Qt5Agg':
                            mng.window.showMaximized()
                    else:
                        mng.resize(*mng.window.maxsize())
                zp = ZoomPanX()
                f = zp.zoom_pan(sfx, base_scale=1.2) # enable scrollable zoom
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
                for i in range(0, len(load)):
                    shortfall[0][i] = rgen[i] - load[i]
                for s in range(1, self.iterations + 1):
                    for i in range(0, len(load)):
                        shortfall[s][i] = rgen[i] - load[i] + h_storage[s - 1]
                plt.figure('shortfall')
                plt.grid(True)
                plt.title('Shortfall')
                sfx = plt.subplot(111)
                sfx.plot(x, shortfall[0], linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                for s in range(1, self.iterations + 1):
                    if h_storage[s - 1] > 1:
                        amt = '{:0,.0f}'.format(h_storage[s - 1])
                    else:
                        amt = '{:0,.1f}'.format(h_storage[s - 1])
                    lbl = 'iteration %s - add %s MW to generation' % (s, amt )
                    sfx.plot(x, shortfall[s], linewidth=self.other_width, label=lbl, color=colours[s])
                plt.axhline(0, color='black')
                plt.xticks(list(range(0, len(x), 168)))
                tick_spot = []
                for i in range(0, len(x), 168):
                    tick_spot.append(i + .5)
                box = sfx.get_position()
                sfx.set_position([box.x0, box.y0, box.width, box.height])
                sfx.set_xticks(tick_spot)
                sfx.set_xticklabels(day_labels, rotation='vertical')
                sfx.set_xlabel('Day of the year')
                sfx.set_ylabel('Power (MW)')
                plt.xlim([0, len(x)])
                sfx.legend(loc='best', prop=lbl_font)
                if self.plots['maximise']:
                    mng = plt.get_current_fig_manager()
                    if sys.platform == 'win32' or sys.platform == 'cygwin':
                        if plt.get_backend() == 'TkAgg':
                            mng.window.state('zoomed')
                        elif plt.get_backend() == 'Qt5Agg':
                            mng.window.showMaximized()
                    else:
                        mng.resize(*mng.window.maxsize())
                zp = ZoomPanX()
                f = zp.zoom_pan(sfx, base_scale=1.2) # enable scrollable zoom
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                for i in range(0, len(load)):
                    shortfall[0][i] = rgen[i] - load[i]
            shortstuff = []
            if self.plots['grid_losses']:
                vals = ['load', 'generation', 'transmitted', 'shortfall']
                short2 = [shortfall[0][0]]
                for i in range(1, len(self.x)):
                  #   short2.append(shortfall[0][i] - shortfall[0][i - 1])
                    short2.append(shortfall[0][i])
                for i in range(0, len(load)):
                    shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i), [load[i],
                                      generation[i], rgen[i], short2[i]], values=vals))
            else:
                vals = ['load', 'generation', 'shortfall']
                for i in range(0, len(load)):
                    shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i), [load[i],
                                      rgen[i], shortfall[0][i]], values=vals))
            vals.insert(0, 'period')
            vals.insert(0, 'hour')
            if self.plots['shortfall_detail'] and self.plots['save_plot']:
                dialog = displaytable.Table(shortstuff, title='Hourly Shortfall',
                                            save_folder=self.scenarios, fields=vals)
                dialog.exec_()
                del dialog
            if self.plots['save_match']:
                if self.pm_template:
                    saveMatch(self, shortstuff)
                else:
                    saveBalance(self, shortstuff)
            del shortstuff
        if self.plots['total']:
            maxy = 0
            if self.plots['cumulative']:
                cumulative = []
                for i in range(len(x24)):
                    cumulative.append(0.)
            if self.plots['gross_load']:
                gross_load = []
                for i in range(len(x24)):
                    gross_load.append(0.)
            if self.plots['save_plot']:
                sp_data = []
                sp_data.append(x24)
                sp_vals = ['hour']
            hdr = self.hdrs['total'].replace('Power - ', '')
            plt.figure(hdr + self.suffix)
            plt.grid(True)
            plt.title(self.hdrs['total'] + self.suffix)
            tx = plt.subplot(111)
            storage = None
            if self.plots['show_pct']:
                load_sum = 0.
                gen_sum = 0.
            for key in iter(sorted(l24.keys())):
                if key == 'Generation':
                    continue
                if self.plots['show_pct']:
                    for j in range(len(x24)):
                        if key[:4] == 'Load':
                            load_sum += l24[key][j]
                        elif key == 'Storage':
                            pass
                        else:
                            gen_sum += l24[key][j]
                            if self.plots['gross_load'] and key == 'Existing Rooftop PV':
                                load_sum += l24[key][j]
                maxy = max(maxy, max(l24[key]))
                lw = self.other_width
                if self.plots['cumulative'] and key[:4] != 'Load' and key != 'Storage':
                    lw = 1.0
                    for j in range(len(x24)):
                        cumulative[j] += l24[key][j]
                if self.plots['gross_load'] and (key[:4] == 'Load' or
                      key == 'Existing Rooftop PV'):
                    for j in range(len(x24)):
                        gross_load[j] += l24[key][j]
                if self.plots['shortfall'] and key[:4] == 'Load':
                    load = l24[key]
                if self.plots['shortfall'] and key == 'Storage':
                    storage = l24[key]
                tx.plot(x24, l24[key], linewidth=lw, label=shrinkKey(key), color=self.colours[key],
                        linestyle=self.linestyle[key])
                if self.plots['save_plot']:
                    sp_vals.append(key)
                    sp_data.append(l24[key])
            if self.plots['cumulative']:
                tx.plot(x24, cumulative, linewidth=self.other_width, label='Tot. Generation', color=self.colours['cumulative'])
                maxy = max(maxy, max(cumulative))
                if self.plots['save_plot']:
                    sp_vals.append('Tot. Generation')
                    sp_data.append(cumulative)
            if self.plots['gross_load'] and 'Existing Rooftop PV' in list(ydata.keys()):
                tx.plot(x24, gross_load, linewidth=1.0, label='Gross Load', color=self.colours['gross_load'])
                maxy = max(maxy, max(gross_load))
                if self.plots['save_plot']:
                    sp_vals.append('Gross Load')
                    sp_data.append(gross_load)
            try:
                rndup = pow(10, round(log10(maxy * 1.5) - 1)) / 2
                maxy = ceil(maxy / rndup) * rndup
            except:
                pass
            if self.plots['shortfall'] and self.do_load:
                load2 = []
                if storage is None:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] - l24[self.load_key][i])
                else:
                    for i in range(len(cumulative)):
                        load2.append(cumulative[i] + storage[i] - l24[self.load_key][i])
                tx.plot(x24, load2, linewidth=self.other_width, label='Shortfall', color=self.colours['shortfall'])
                plt.axhline(0, color='black')
                miny = min(load2)
                if rndup != 0 and miny < 0:
                    miny = -ceil(-miny / rndup) * rndup
                if self.plots['save_plot']:
                    sp_vals.append('Shortfall')
                    sp_data.append(load2)
            else:
                miny = 0
            if self.plots['save_plot']:
                titl = 'Total'
                decpts = [3] * len(sp_vals)
                decpts[0] = 0
                dialog = displaytable.Table(list(map(list, list(zip(*sp_data)))), title=titl, fields=sp_vals,
                                            save_folder=self.scenarios, decpts=decpts)
                dialog.exec_()
                del dialog, sp_data, sp_vals
            plt.ylim([miny, maxy])
            plt.xlim([1, 25])
            plt.xticks(list(range(0, 25, 4)))
          #   tx.legend(loc='lower left', numpoints = 2, prop=lbl_font)
            tx.set_xticklabels(labels)
            tx.set_xlabel('Hour of the Day')
            tx.set_ylabel('Power (MW)')
            if (len(ydata) + pc) > 9:
                 # Shrink current axis by 5%
                box = tx.get_position()
                tx.set_position([box.x0, box.y0, box.width * 0.95, box.height])
                 # Put a legend to the right of the current axis
                tx.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop=lbl_font)
            else:
                tx.legend(bbox_to_anchor=[0.5, -0.1], loc='center', ncol=(len(ydata) + pc),
                          prop=lbl_font)
            if self.plots['show_pct']:
                self.gen_pct = ' (%s%% of load)' % '{:0,.1f}'.format(gen_sum * 100. / load_sum)
                plt.title(self.hdrs['total'] + self.suffix + self.gen_pct)
            if self.plots['maximise']:
                mng = plt.get_current_fig_manager()
                if sys.platform == 'win32' or sys.platform == 'cygwin':
                    if plt.get_backend() == 'TkAgg':
                        mng.window.state('zoomed')
                    elif plt.get_backend() == 'Qt5Agg':
                        mng.window.showMaximized()
                else:
                    mng.resize(*mng.window.maxsize())
            zp = ZoomPanX()
            f = zp.zoom_pan(tx, base_scale=1.2) # enable scrollable zoom
            if self.plots['block']:
                plt.show(block=True)
                if matplotlib.__version__ > '3.5.1':
                    plt.waitforbuttonpress()
            else:
                plt.draw()
        if self.plots['month']:
            dayPlot(self, 'month', m24, mth_labels, mth_xlabels)
        if self.plots['season']:
            dayPlot(self, 'season', q24, ssn_labels, labels)
        if self.plots['period']:
            dayPlot(self, 'period', s24, smp_labels, labels)
        if not self.plots['block']:
            plt.show(block=True)

    def save_detail(self, data_file, techs, keys=None):
        if self.suffix != '':
            i = data_file.rfind('.')
            if i > 0:
                data_file = data_file[:i] + '_' + self.suffix + data_file[i:]
            else:
                data_file = data_file + '_' + self.suffix
        if data_file[-4:] == '.csv' or data_file[-4:] == '.xls' \
          or data_file[-5:] == '.xlsx':
            pass
        else:
            data_file += '.xlsx'
        data_file = QtWidgets.QFileDialog.getSaveFileName(None, 'Save power data file',
                    self.scenarios + data_file, 'Excel Files (*.xls*);;CSV Files (*.csv)')[0]
        if data_file == '':
            return
        if self.load_multiplier != 0:
            the_year = self.load_year
        else:
            the_year = self.base_year
        if data_file[-4:] == '.csv' or data_file[-4:] == '.xls' \
          or data_file[-5:] == '.xlsx':
            pass
        else:
            data_file += '.xlsx'
        if os.path.exists(data_file):
            if os.path.exists(data_file + '~'):
                os.remove(data_file + '~')
            os.rename(data_file, data_file + '~')
        if keys is None:
            keys = sorted(techs.keys())
        if data_file[-4:] == '.csv':
            tf = open(data_file, 'w')
            line = 'Hour,Period,'
            if self.load_multiplier != 0:
                the_year = self.load_year
            else:
                the_year = self.base_year
            max_outs = 0
            lines = []
            for i in range(8760):
                lines.append(str(i+1) + ',' + str(the_date(the_year, i)) + ',')
            for key in keys:
                if key[:4] == 'Load' and self.load_multiplier != 0:
                    line += 'Load ' + self.load_year + ','
                else:
                    line += key + ','
                for i in range(len(techs[key])):
                    lines[i] += str(round(techs[key][i], 3)) + ','
            tf.write(line + '\n')
            for i in range(len(lines)):
                tf.write(lines[i] + '\n')
            tf.close()
            del lines
        elif data_file[-4:] == '.xls':
            wb = xlwt.Workbook()
            ws = wb.add_sheet('Detail')
            ws.write(0, 0, 'Hour')
            ws.write(0, 1, 'Period')
            for i in range(len(self.x)):
                ws.write(i + 1, 0, i + 1)
                ws.write(i + 1, 1, the_date(the_year, i))
            if 16 * 275 > ws.col(1).width:
                ws.col(1).width = 16 * 275
            c = 2
            for key in keys:
                if key[:4] == 'Load' and self.load_multiplier != 0:
                    ws.write(0, c, 'Load ' + self.load_year)
                else:
                    ws.write(0, c, key)
                if len(key) * 275 > ws.col(c).width:
                    ws.col(c).width = len(key) * 275
                for r in range(len(techs[key])):
                    ws.write(r + 1, c, round(techs[key][r], 3))
                c += 1
            ws.set_panes_frozen(True)  # frozen headings instead of split panes
            ws.set_horz_split_pos(1)  # in general, freeze after last heading row
            ws.set_remove_splits(True)  # if user does unfreeze, dont leave a split there
            wb.save(data_file)
            del wb
        else: # .xlsx
            wb = oxl.Workbook()
            ws = wb.active
            ws.title = 'Detail'
            normal = oxl.styles.Font(name='Arial', size='10')
            ws.cell(row=1, column=1).value = 'Hour'
            ws.cell(row=1, column=1).font = normal
            ws.cell(row=1, column=2).value = 'Period'
            ws.cell(row=1, column=2).font = normal
            for i in range(len(self.x)):
                ws.cell(row=i + 2, column=1).value = i + 1
                ws.cell(row=i + 2, column=1).font = normal
                ws.cell(row=i + 2, column=2).value = the_date(the_year, i)
                ws.cell(row=i + 2, column=2).font = normal
            ws.column_dimensions['B'].width = 16
            c = 3
            for key in keys:
                if key[:4] == 'Load' and self.load_multiplier != 0:
                    ws.cell(row=1, column=c).value = 'Load ' + self.load_year
                else:
                    ws.cell(row=1, column=c).value = key
                ws.cell(row=1, column=c).font = normal
                ws.column_dimensions[ssCol(c)].width = len(key)
                for r in range(len(techs[key])):
                    ws.cell(row=r + 2, column=c).value = round(techs[key][r], 3)
                    ws.cell(row=r + 2, column=c).font = normal
                c += 1
            ws.freeze_panes = 'A2'
            wb.save(data_file)
            wb.close()
            del wb

#       __init__ for PowerModel
    def __init__(self, stations, year=None, status=None, visualise=None, loadonly=False, progress=None):
        self.something = visualise
        self.something.power_signal = self
        self.status = status
        self.stations = stations
        self.progress = progress
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        self.expert = False
        try:
            expert = config.get('Base', 'expert_mode')
            if expert in ['true', 'on', 'yes']:
                self.expert = True
        except:
            pass
        if year is None:
            try:
                self.base_year = config.get('Base', 'year')
            except:
                self.base_year = '2012'
        else:
            self.base_year = year
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', self.base_year)
            i = self.scenarios.rfind('/')
            self.scenarios = self.scenarios[:i + 1]
        except:
            self.scenarios = ''
        try:
            self.load_file = config.get('Files', 'load')
            for key, value in parents:
                self.load_file = self.load_file.replace(key, value)
            self.load_file = self.load_file.replace('$USER$', getUser())
            self.load_file = self.load_file.replace('$YEAR$', self.base_year)
        except:
            self.load_file = ''
        try:
            zone_file = config.get('Files', 'grid_zones')
            zone_file = True
        except:
            zone_file = False
        self.data_file = ''
        try:
            self.data_file = config.get('Files', 'data_file')
        except:
            try:
                self.data_file = config.get('Power', 'data_file')
            except:
                pass
        for key, value in parents:
           self.data_file = self.data_file.replace(key, value)
        self.data_file = self.data_file.replace('$USER$', getUser())
        self.data_file = self.data_file.replace('$YEAR$', self.base_year)
        try:
            helpfile = config.get('Files', 'help')
            for key, value in parents:
                helpfile = helpfile.replace(key, value)
            helpfile = helpfile.replace('$USER$', getUser())
            helpfile = helpfile.replace('$YEAR$', self.base_year)
        except:
            helpfile = ''
        if self.progress is None:
            self.show_progress = None
        else:
            progress_bar = True
            try:
                progress_bar = config.get('View', 'progress_bar')
                if progress_bar in ['false', 'no', 'off']:
                    self.show_progress = None
                else:
                    self.show_progress = True
                    try:
                        self.progress_bar = int(self.progress_bar)
                    except:
                        self.progress_bar = 0
            except:
                self.show_progress = True
                self.progress_bar = 0
#
#       choose what power data to collect (once only)
#
        self.plot_order = ['show_menu', 'actual', 'cumulative', 'by_station', 'adjust',
                           'show_load', 'shortfall', 'grid_losses', 'gross_load', 'save_plot', 'visualise',
                           'show_pct', 'maximise', 'block', 'by_day', 'by_month', 'by_season',
                           'by_period', 'hour', 'total', 'month', 'season', 'period',
                           'duration', 'augment', 'shortfall_detail', 'summary', 'save_zone', 'save_data',
                           'save_detail', 'save_tech', 'save_match', 'financials']
        self.initials = ['actual', 'by_station', 'grid_losses', 'save_zone', 'save_data', 'gross_load',
                         'summary', 'financials'] #, 'show_menu']
        self.hdrs = {'show_menu': 'Check / Uncheck all',
                'actual': 'Generation - use actual generation figures',
                'cumulative': 'Generation - total (cumulative)',
                'by_station': 'Generation - from chosen stations',
                'adjust': 'Generation - adjust generation',
                'show_load': 'Generation - show Load',
                'shortfall': 'Generation - show shortfall from Load',
                'grid_losses': 'Generation - reduce generation by grid losses',
                'gross_load': 'Add Existing Rooftop PV to Load (Gross Load)',
                'save_plot': 'Save chart data',
                'visualise': 'Visualise generation',
                'show_pct': 'Show generation as a percentage of load',
                'maximise': 'Maximise chart windows',
                'block': 'Show charts one at a time',
                'by_day': 'Energy by day',
                'by_month': 'Energy by month',
                'by_season': 'Energy by season',
                'by_period': 'Energy by period',
                'hour': 'Power by hour',
                'total': 'Power - diurnal profile',
                'month': 'Power - diurnal profile by month',
                'season': 'Power - diurnal profile by season',
                'period': 'Power - diurnal profile by period',
                'duration': 'Power - Load duration',
                'augment': 'Power - augmented by hour',
                'shortfall_detail': 'Power - Shortfall analysis',
                'summary': 'Show Summary/Other Tables',
                'save_zone': 'Generation by Zone',
                'save_data': 'Save initial Hourly Data Output',
                'save_detail': 'Save Hourly Data Output by Station',
                'save_tech': 'Save Hourly Data Output by Technology',
                'save_match': 'Save Powermatch Inputs',
                'financials': 'Run Financial Models'}
        self.spacers = {'actual': 'Show in Chart',
                   'save_plot': 'Choose charts (all use a full year of data)',
                   'summary': 'Choose tables'}
        self.plots = {}
        for i in range(len(self.plot_order)):
            self.plots[self.plot_order[i]] = False
        self.load_year = self.base_year
        if loadonly:
            if self.load_file == '' or not os.path.exists(self.load_file):
                self.load_file = None
                return
            plot_order = ['show_menu', 'save_plot', 'maximise', 'block', 'by_day', 'by_month', 'by_season',
                           'by_period', 'hour', 'total', 'month', 'season', 'period']
            spacers = {'maximise': 'Choose plots (all use a full year of data)'}
            self.plots['show_load'] == True
            what_plots = whatPlots(self.plots, plot_order, self.hdrs, spacers, 0., self.base_year, self.load_year,
                                   0, [0, 0], [0, 0], [0, 0], [], initial=False, helpfile=helpfile)
            what_plots.exec_()
            return
        self.technologies = ''
        self.load_growth = 0.
        self.storage = [0., 0.]
        self.recharge = [0., 1.]
        self.discharge = [0., 1.]
        plot_opts = []
        try:
            plot_opts = config.items('Power')
        except:
            pass
        for key, value in plot_opts:
            if key == 'save_balance': # old name for save_match
                key = 'save_match'
            if key in self.plots:
                if value.lower() in ['true', 'yes', 'on']:
                    self.plots[key] = True
            elif key == 'load_growth':
                if value[-1] == '%':
                    self.load_growth = float(value[:-1]) / 100.
                else:
                    self.load_growth = float(value)
            elif key == 'storage':
                if ',' in value:
                    bits = value.split(',')
                    self.storage = [float(bits[0].strip()), float(bits[1].strip())]
                else:
                    self.storage = [float(value), 0.]
            elif key == 'technologies':
                self.technologies = value
            elif key == 'shortfall_iterations':
                self.iterations = int(value)
        try:
            storage = config.get('Storage', 'storage')
            if ',' in storage:
                bits = storage.split(',')
                self.storage = [float(bits[0].strip()), float(bits[1].strip())]
            else:
                self.storage = [float(storage), 0.]
        except:
            pass
        try:
            self.show_menu = self.plots['show_menu']
        except:
            self.show_menu = True
        try:
            self.discharge[0] = float(config.get('Storage', 'discharge_max'))
            self.discharge[1] = float(config.get('Storage', 'discharge_eff'))
            if self.discharge[1] < 0.5:
                self.discharge[1] = 1 - self.discharge[1]
            self.recharge[0] = float(config.get('Storage', 'recharge_max'))
            self.recharge[1] = float(config.get('Storage', 'recharge_eff'))
            if self.recharge[1] < 0.5:
                self.recharge[1] = 1 - self.recharge[1]
        except:
            pass
        if __name__ == '__main__':
            self.show_menu = True
            self.plots['save_data'] = True
        if not self.plots['save_data']:
            self.plot_order.remove('save_data')
        if len(self.stations) == 1:
            self.plot_order.remove('cumulative')
            self.plot_order.remove('by_station')
            self.plot_order.remove('gross_load')
      # check if we can find a load file
        if self.load_file == '' or not os.path.exists(self.load_file):
            self.can_do_load = False
            self.plot_order.remove('augment')
            self.plot_order.remove('duration')
            self.plot_order.remove('show_load')
            self.plot_order.remove('show_pct')
            self.plot_order.remove('shortfall')
            self.plot_order.remove('shortfall_detail')
        else:
            self.can_do_load = True
        if not zone_file:
            self.plot_order.remove('save_zone')
            self.plots['save_zone'] = False
        if self.show_menu:
            if __name__ == '__main__':
                app = QtWidgets.QApplication(sys.argv)
            what_plots = whatPlots(self.plots, self.plot_order, self.hdrs, self.spacers,
                                   self.load_growth, self.base_year, self.load_year,
                                   self.iterations, self.storage, self.discharge,
                                   self.recharge, initial=True, helpfile=helpfile)
            what_plots.exec_()
            self.plots, self.load_growth, self.load_year, self.load_multiplier, self.iterations, \
              self.storage, self.discharge, self.recharge = what_plots.getValues()
            if self.plots is None:
                self.something.power_signal = None
                return
        self.selected = None
        if self.plots['by_station']:
            self.selected = []
            if len(stations) == 1:
                self.selected.append(stations[0].name)
            else:
                selected = whatStations(stations, self.plots['gross_load'],
                                        self.plots['actual'])
                selected.exec_()
                self.selected = selected.getValues()
                if self.selected is None:
                    return
#
#       collect the data (once only)
#
        self.stn_outs = []
        self.model = SuperPower(stations, self.plots, False, year=self.base_year,
                                selected=self.selected, status=status, progress=self.progress)
        self.model.getPower()
        if len(self.model.power_summary) == 0:
            return
        self.power_summary = self.model.power_summary
        self.ly, self.x = self.model.getLy()
        if self.plots['save_data'] or self.plots['financials'] or self.plots['save_detail']:
            self.stn_outs, self.stn_tech, self.stn_size, self.stn_pows, self.stn_grid, \
              self.stn_path = self.model.getStnOuts()
        elif self.plots['save_tech'] or self.plots['save_match']:
            self.stn_outs, self.stn_tech = self.model.getStnTech()
            if self.plots['visualise']:
                _, self.stn_pows = self.model.getStnPows()
        elif self.plots['visualise']:
            self.stn_outs, self.stn_pows = self.model.getStnPows()
        self.suffix = ''
        if len(self.stations) == 1:
            self.suffix = ' - ' + self.stations[0].name
        elif len(self.stn_outs) == 1:
            self.suffix = ' - ' + self.stn_outs[0]
        elif self.plots['by_station']:
            if len(self.ly) == 1:
                self.suffix = ' - ' + list(self.ly.keys())[0]
            else:
                self.suffix = ' - Chosen Stations'
        if self.plots['save_zone']:
            self.stn_zone = self.model.getStnZones()
        if self.plots['save_data']:
            if self.data_file == '':
                data_file = 'Power_Table_%s.xlsx' % \
                            QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                            'yyyy-MM-dd_hhmm')
            else:
                data_file = self.data_file
            stnsh = {}
            # if load
            if self.plots['show_load']:
                stnsh['Load'] = self.load_data[:]
            for i in range(len(self.stn_outs)):
                stnsh[self.stn_outs[i]] = self.stn_pows[i][:]
            self.save_detail(data_file, stnsh)
            del stnsh
        if self.plots['summary']:
            fields = ['name', 'technology', 'capacity', 'cf', 'generation']
            sumfields = ['capacity', 'generation']
            decpts = [0, 0, 1, 1, 2, 0, 1]
            if getattr(self.power_summary[0], 'transmitted') != None:
                fields.append('transmitted')
                sumfields.append('transmitted')
                decpts.append([0, 1])
            if self.plots['save_zone']:
                fields.insert(1, 'zone')
                decpts.insert(1, 0)
            dialog = displaytable.Table(self.power_summary, sumfields=sumfields,
                     units='capacity=MW generation=MWh transmitted=MWh', sumby='technology',
                     decpts=decpts, fields=fields, save_folder=self.scenarios)
            dialog.exec_()
            del dialog
        if self.plots['financials']:
            do_financials = True
        else:
            do_financials = False
        if self.plots['save_data'] or self.plots['summary']:
            show_summ = True
        else:
            show_summ = False
        do_plots = True
#
#       loop around processing plots
#
        if do_plots:
            if matplotlib.__version__ <= '3.5.1':
                if matplotlib.get_backend() != 'TkAgg':
                    plt.switch_backend('TkAgg')
            self.gen_pct = None
            self.load_data = None
            if self.plots['save_detail']:
                pass
            else:
                self.initials.append('save_detail')
                if not self.plots['save_tech']:
                    self.initials.append('save_tech')
                if not self.plots['visualise']:
                    self.initials.append('visualise')
            self.load_key = ''
            self.adjustby = None
            while True:
                if self.plots['visualise'] and self.something is not None:
                    vis2 = Visualise(self.stn_outs, self.stn_pows, self.something, year=self.base_year)
                    vis2.setWindowModality(QtCore.Qt.WindowModal)
                    vis2.setWindowFlags(vis2.windowFlags() |
                                 QtCore.Qt.WindowSystemMenuHint |
                                 QtCore.Qt.WindowMinMaxButtonsHint)
                    vis2.exec_()
                wrkly = {}
                summs = {}
                if self.load_key != '':
                    try:
                        del wrkly[self.load_key]
                    except:
                        pass
                    self.load_key = ''
                if (self.plots['show_load'] or self.plots['save_match'] or self.plots['shortfall'] \
                    or self.plots['shortfall_detail']) and self.can_do_load:
                    other_load_year = False
                    if self.load_year != self.base_year and self.load_growth == 0: # see if other load file
                        load_file = self.load_file.replace(self.base_year, self.load_year)
                        if os.path.exists(load_file):
                            self.load_file = load_file
                            other_load_year = True
                            self.load_data = None
                    if self.load_data is None:
                        tf = open(self.load_file, 'r')
                        lines = tf.readlines()
                        tf.close()
                        self.load_data = []
                        bit = lines[0].rstrip().split(',')
                        if len(bit) > 0: # multiple columns
                            for b in range(len(bit)):
                                if bit[b][:4].lower() == 'load':
                                    if bit[b].lower().find('kwh') > 0: # kWh not MWh
                                        for i in range(1, len(lines)):
                                            bit = lines[i].rstrip().split(',')
                                            self.load_data.append(float(bit[b]) * 0.001)
                                    else:
                                        for i in range(1, len(lines)):
                                            bit = lines[i].rstrip().split(',')
                                            self.load_data.append(float(bit[b]))
                        else:
                            for i in range(1, len(lines)):
                                self.load_data.append(float(lines[i].rstrip()))
                    if self.load_multiplier != 0 or other_load_year:
                        key = 'Load ' + self.load_year
                    else:
                        key = 'Load'  # lines[0].rstrip()
                    self.load_key = key
                    wrkly[key] = []
                    if self.load_multiplier != 0:
                        for i in range(len(self.load_data)):
                            wrkly[key].append(self.load_data[i] * (1 + self.load_multiplier))
                    else:
                        wrkly[key] = self.load_data[:]
                else:
                    self.plots['show_pct'] = False
                if self.plots['adjust']:
                    if self.load_key == '':
                        if self.adjustby is None:
                            adjust = Adjustments(list(self.ly.keys()))
                        else:
                            adjust = Adjustments(self.adjustby)
                    else:
                        if self.adjustby is None:
                            adjust = Adjustments(list(self.ly.keys()), self.load_key, wrkly[self.load_key], self.ly,
                                                 self.load_year)
                        else:
                            adjust = Adjustments(self.adjustby, self.load_key, wrkly[self.load_key], self.ly, self.load_year)
                    adjust.exec_()
                    self.adjustby = adjust.getValues()
                else:
                    self.adjustby = None
                for key in self.ly:
                    if self.adjustby is None:
                        wrkly[key] = self.ly[key][:]
                    else:
                        wrkly[key] = []
                        if key == 'Generation':
                            for i in range(len(self.ly[key])):
                                wrkly[key].append(self.ly[key][i])
                        else:
                            for i in range(len(self.ly[key])):
                                wrkly[key].append(self.ly[key][i] * self.adjustby[key])
                if self.plots['shortfall'] or self.plots['shortfall_detail'] or self.plots['save_match']:
                    self.plots['show_load'] = True
                    self.plots['cumulative'] = True
                try:
                    del wrkly['Storage']
                except:
                    pass
                if self.load_data is None:
                    self.do_load = False
                else:
                    self.do_load = True
                if self.plots['show_load']:
                    total_gen = []
                    for i in range(len(self.x)):
                        total_gen.append(0.)
                    for key, value in wrkly.items():
                        if key == 'Generation':
                            continue
                        if key == 'Storage' or key == 'Excess':
                            continue
                        elif key[:4] == 'Load':
                            pass
                        else:
                            for i in range(len(value)):
                                total_gen[i] += value[i]
                    if self.storage[0] > 0:
                        wrkly['Storage'] = []
                        wrkly['Excess'] = []
                        for i in range(len(self.x)):
                            wrkly['Storage'].append(0.)
                            wrkly['Excess'].append(0.)
                        storage_cap = self.storage[0] * 1000.
                        if self.storage[1] > self.storage[0]:
                            storage_carry = self.storage[0] * 1000.
                        else:
                            storage_carry = self.storage[1] * 1000.
                        storage_bal = []
                        storage_losses = []
                        for i in range(len(self.x)):
                            gap = gape = total_gen[i] - wrkly[self.load_key][i]
                            storage_loss = 0.
                            if gap >= 0:  # excess generation
                                if self.recharge[0] > 0 and gap > self.recharge[0]:
                                    gap = self.recharge[0]
                                if storage_carry >= storage_cap:
                                    wrkly['Excess'][i] = gape
                                else:
                                    if storage_carry + gap > storage_cap:
                                        gap = (storage_cap - storage_carry) * (1 / self.recharge[1])
                                    storage_loss = gap - gap * self.recharge[1]
                                    storage_carry += gap - storage_loss
                                    if gape - gap > 0:
                                        wrkly['Excess'][i] = gape - gap
                                    if storage_carry > storage_cap:
                                        storage_carry = storage_cap
                            else:
                                if self.discharge[0] > 0 and -gap > self.discharge[0]:
                                    gap = -self.discharge[0]
                                if storage_carry > -gap / self.discharge[1]:  # extra storage
                                    wrkly['Storage'][i] = -gap
                                    storage_loss = gap * self.discharge[1] - gap
                                    storage_carry += gap - storage_loss
                                else:  # not enough storage
                                    if self.discharge[0] > 0 and storage_carry > self.discharge[0]:
                                        storage_carry = self.discharge[0]
                                    wrkly['Storage'][i] = storage_carry * (1 / (2 - self.discharge[1]))
                                    storage_loss = storage_carry - wrkly['Storage'][i]
                                    storage_carry = 0 # ???? bug ???
                            storage_bal.append(storage_carry)
                            storage_losses.append(storage_loss)
                        if self.plots['shortfall_detail']:
                            shortstuff = []
                            for i in range(len(self.x)):
                                shortfall = total_gen[i] + wrkly['Storage'][i] - wrkly[self.load_key][i]
                                if shortfall > 0:
                                    shortfall = 0
                                shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i),
                                                  [wrkly[self.load_key][i], total_gen[i],
                                                  wrkly['Storage'][i], storage_losses[i],
                                                  storage_bal[i], shortfall, wrkly['Excess'][i]],
                                                  values=['load', 'generation', 'storage_used',
                                                          'storage_loss', 'storage_balance',
                                                          'shortfall', 'excess']))
                            dialog = displaytable.Table(shortstuff, title='Storage',
                                                        save_folder=self.scenarios,
                                                        fields=['hour', 'period', 'load', 'generation',
                                                                'storage_used', 'storage_loss',
                                                                'storage_balance', 'shortfall', 'excess'])
                            dialog.exec_()
                            del dialog
                            del shortstuff
                        if show_summ:
                            summs['Shortfall'] = ['', '', 0., 0]
                            summs['Excess'] = ['', '', 0., 0]
                            for i in range(len(self.x)):
                                if total_gen[i] > wrkly[self.load_key][i]:
                                    summs['Excess'][2] += total_gen[i] - wrkly[self.load_key][i]
                                else:
                                    summs['Shortfall'][2] += total_gen[i]  - wrkly[self.load_key][i]
                            summs['Shortfall'][2] = round(summs['Shortfall'][2], 1)
                            summs['Excess'][2] = round(summs['Excess'][2], 1)
                    elif show_summ or self.plots['shortfall_detail']:
                        if self.plots['shortfall_detail']:
                            shortstuff = []
                            for i in range(len(self.x)):
                                if total_gen[i] > wrkly[self.load_key][i]:
                                    excess = total_gen[i] - wrkly[self.load_key][i]
                                    shortfall = 0
                                else:
                                    shortfall = total_gen[i]  - wrkly[self.load_key][i]
                                    excess = 0
                                shortstuff.append(ColumnData(i + 1, the_date(self.load_year, i),
                                                  [wrkly[self.load_key][i], total_gen[i],
                                                   shortfall, excess],
                                                  values=['load', 'generation',
                                                          'shortfall', 'excess']))
                            dialog = displaytable.Table(shortstuff, title='Hourly Shortfall',
                                                        save_folder=self.scenarios,
                                                        fields=['hour', 'period', 'load', 'generation',
                                                                'shortfall', 'excess'])
                            dialog.exec_()
                            del dialog
                            del shortstuff
                        else:
                            summs['Shortfall'] = ['', '', 0., 0]
                            summs['Excess'] = ['', '', 0., 0]
                            for i in range(len(self.x)):
                                if total_gen[i] > wrkly[self.load_key][i]:
                                    summs['Excess'][2] += total_gen[i] - wrkly[self.load_key][i]
                                else:
                                    summs['Shortfall'][2] += total_gen[i] - wrkly[self.load_key][i]
                            summs['Shortfall'][2] = round(summs['Shortfall'][2], 1)
                            summs['Excess'][2] = round(summs['Excess'][2], 1)
                if show_summ and self.adjustby is not None:
                    keys = []
                    for key in wrkly:
                        keys.append(key)
                        gen = 0.
                        for i in range(len(wrkly[key])):
                            gen += wrkly[key][i]
                        if key[:4] == 'Load':
                            incr = 1 + self.load_multiplier
                        else:
                            try:
                                incr = self.adjustby[key]
                            except:
                                incr = ''
                        try:
                            summs[key] = [0., round(incr, 2), round(gen, 1), 0]
                            if key[:4] == 'Load':
                                summs[key][0] = ''
                        except:
                            summs[key] = ['', '', round(gen, 1), 0]
                    keys.sort()
                    xtra = ['Generation', 'Load', 'Gen. - Load', 'Storage Capacity', 'Storage', 'Shortfall', 'Excess']
                    o = 0
                    gen = 0.
                    if self.storage[0] > 0:
                        summs['Storage Capacity'] = [self.storage[0] * 1000., '', '', 0]
                    for i in range(len(keys)):
                        if keys[i][:4] == 'Load':
                            xtra[1] = keys[i]
                        elif keys[i] in xtra:
                            continue
                        else:
                            o += 1
                            summs[keys[i]][3] = o
                            gen += summs[keys[i]][2]
                    if xtra[0] not in list(summs.keys()):
                        summs[xtra[0]] = ['', '', gen, 0]
                    if xtra[1] in list(summs.keys()):
                        summs[xtra[2]] = ['', '', round(gen - summs[xtra[1]][2], 1), 0]
                    for i in range(len(xtra)):
                        if xtra[i] in list(summs.keys()):
                            o += 1
                            summs[xtra[i]][3] = o
                    try:
                        summs['Storage Used'] = summs.pop('Storage')
                    except:
                        pass
                    try:
                        summs['Excess Gen.'] = summs.pop('Excess')
                    except:
                        pass
                    for it in self.power_summary:
                        if self.plots['by_station']:
                            try:
                                summs[it.name][0] = it.capacity
                            except:
                                pass
                        else:
                            try:
                                summs[it.technology][0] += it.capacity
                            except:
                                pass
                    for key, value in summs.items():
                        try:
                            value[0] = round(value[0], 2)
                        except:
                            pass
                    dialog = displaytable.Table(summs, title='Generation Summary',
                                                save_folder=self.scenarios,
                                                fields=['component', 'capacity', 'multiplier', 'generation', 'row'],
                                                units='generation=MWh', sortby='row')
                    dialog.exec_()
                    del dialog
                if self.plots['save_detail'] or self.plots['save_tech']:
                    dos = []
                    if self.plots['save_detail']:
                        dos.append('')
                    if self.plots['save_tech']:
                        dos.append('Technology_')
                    for l in range(len(dos)):
                        if self.data_file == '':
                            if year is None:
                                data_file = 'Power_Detail_%s%s.xlsx' % ( dos[l] ,
                                QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                                'yyyy-MM-dd_hhmm'))
                            else:
                                data_file = 'Power_Detail_%s%s_%s.xlsx' % ( dos[l] , str(year),
                                QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                                'yyyy-MM-dd_hhmm'))
                        else:
                             data_file = self.data_file
                        keys = []
                        keys2 = []
                        if dos[l] != '':
                            techs = {}
                            for key, value in iter(wrkly.items()):
                                try:
                                    i = self.stn_outs.index(key)
                                    if self.stn_tech[i] in list(techs.keys()):
                                        for j in range(len(value)):
                                            techs[self.stn_tech[i]][j] += value[j]
                                    else:
                                        techs[self.stn_tech[i]] = value[:]
                                        keys.append(self.stn_tech[i])
                                except:
                                    techs[key] = value[:]
                                    keys2.append(key)
                            keys.sort()
                            keys2.extend(keys) # put Load first
                            self.save_detail(data_file, techs, keys=keys2)
                            del techs
                        else:
                            for key in list(wrkly.keys()):
                                try:
                                    i = self.stn_outs.index(key)
                                    keys.append(self.stn_outs[i])
                                except:
                                    keys2.append(key)
                            keys.sort()
                            keys2.extend(keys) # put Load first
                            self.save_detail(data_file, wrkly, keys=keys2)
                self.showGraphs(wrkly, self.x)
                if __name__ == '__main__':
                    self.show_menu = True
                    self.plots['save_data'] = True
                if self.show_menu:
                    what_plots = whatPlots(self.plots, self.plot_order, self.hdrs, self.spacers,
                                 self.load_growth, self.base_year, self.load_year, self.iterations,
                                 self.storage, self.discharge, self.recharge, self.initials, helpfile=helpfile)
                    what_plots.exec_()
                    self.plots, self.load_growth, self.load_year, self.load_multiplier, \
                        self.iterations, self.storage, self.discharge, self.recharge = what_plots.getValues()
                    if self.plots is None:
                        break
                else:
                    break
#
#       loop around doing financials
#
         # run the financials model
        if do_financials:
            self.financial_parms = None
            while True:
                self.financials = FinancialModel(self.stn_outs, self.stn_tech, self.stn_size,
                                  self.stn_pows, self.stn_grid, self.stn_path, year=self.base_year,
                                  parms=self.financial_parms, status=self.status)
                if self.financials.stations is None:
                    break
                self.financial_parms = self.financials.getParms()
                fin_fields = ['name', 'technology', 'capacity', 'generation', 'cf',
                              'capital_cost', 'lcoe_real', 'lcoe_nominal', 'npv']
                fin_sumfields = ['capacity', 'generation', 'capital_cost', 'npv']
                fin_units = 'capacity=MW generation=MWh capital_cost=$ lcoe_real=c/kWh' + \
                            ' lcoe_nominal=c/kWh npv=$'
                tot_capital = 0.
                tot_capacity = 0.
                tot_generation = 0.
                tot_lcoe_real = [0., 0.]
                tot_lcoe_nom = [0., 0.]
                for stn in self.financials.stations:
                    tot_capital += stn.capital_cost
                    tot_capacity += stn.capacity
                    tot_generation += stn.generation
                for stn in self.financials.stations:
                    tot_lcoe_real[0] += stn.lcoe_real * (stn.generation / tot_generation)
                    tot_lcoe_nom[0] += stn.lcoe_nominal * (stn.generation / tot_generation)
                    tot_lcoe_real[1] += stn.lcoe_real * (stn.capacity / tot_capacity)
                    tot_lcoe_nom[1] += stn.lcoe_nominal * (stn.capacity / tot_capacity)
                    if stn.grid_cost > 0:
                        i = fin_fields.index('capital_cost')
                        fin_fields.insert(i + 1, 'grid_cost')
                        fin_sumfields.append('grid_cost')
                        fin_units += ' grid_cost=$'
                        break
                tot_fields = [['cf', tot_generation / tot_capacity / 8760],
                              ['lcoe_real', tot_lcoe_real[0]],
                              ['lcoe_nominal', tot_lcoe_nom[0]]]
                dialog = displaytable.Table(self.financials.stations, fields=fin_fields,
                         sumfields=fin_sumfields, units=fin_units, sumby='technology',
                         save_folder=self.scenarios, title='Financials', totfields=tot_fields)
                dialog.exec_()
                del dialog
        self.something.power_signal = None

    def getValues(self):
        try:
            return self.power_summary
        except:
            return None

    def getPct(self):
        return self.gen_pct

    @QtCore.pyqtSlot()
    def exit(self):
        self.something.power_signal = None
        return #exit()
