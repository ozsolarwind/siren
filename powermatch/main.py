#!/usr/bin/python3
#
#  Copyright (C) 2018-2024 Sustainable Energy Now Inc., Angus King
#
#  powermatch.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#
# Note: Batch process is all rather messy.

from credits import fileVersion
import configparser  # decode .ini file
from copy import copy
from core import (
    AdjustmentsBase, Constraint, Facility, Optimisation, ProgressHandler, ProgressInfo,
    PowerMatchBase, setTransitionBase,compute_lcoe, generate_summary
    )
from datetime import datetime
from displaytable import Table
import displayobject
from editini import EdtDialog, SaveIni
from floaters import ProgressBar, FloatStatus
from getmodels import getModelFile, commonprefix
import matplotlib
if matplotlib.__version__ > '3.5.1':
    matplotlib.use('Qt5Agg')
else:
    matplotlib.use('TkAgg') # so PyQT5 and Matplotlib windows don't interfere
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
# This import registers the 3D projection, but is otherwise unused.
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.ticker import LinearLocator, FormatStrFormatter
import numpy as np
import openpyxl as oxl
import os
from PyQt5 import QtCore, QtGui, QtWidgets
import random
import sys
import shutil
import subprocess
from senutils import ClickableQLabel, getUser, ListWidget, ssCol, WorkBook
import time
from typing import Optional
from zoompan import ZoomPanX
try:
    from opt_debug import optimiseDebug
except:
    pass

from powermatch_lit import *

class MyQDialog(QtWidgets.QDialog):
    ignoreEnter = True

    def keyPressEvent(self, qKeyEvent):
        if qKeyEvent.key() == QtCore.Qt.Key_Return:
            self.ignoreEnter = True
        else:
            self.ignoreEnter = False

    def closeEvent(self, event):
        if self.ignoreEnter:
            self.ignoreEnter = False
            event.ignore()
        else:
            event.accept()

 #   def resizeEvent(self, event):
  #      return super(Window, self).resizeEvent(event)

class Adjustments(AdjustmentsBase, MyQDialog):
    @staticmethod
    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 65 + ctr * 32
        width = window.frameSize().width()
        screen = QtWidgets.QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, data, adjustin, adjust_cap, prefix, show_multipliers=False, save_folder=None,
                 batch_file=None):
        AdjustmentsBase.__init__(self, parent, data, adjustin, adjust_cap, prefix, 
                                show_multipliers=show_multipliers, 
                                save_folder=save_folder, 
                                batch_file=batch_file)
        MyQDialog.__init__(self, parent)  # Initialize the QDialog properly
        self.init_ui(data, adjustin, adjust_cap, prefix, save_folder)

    def init_ui(self, data, adjustin, adjust_cap, prefix, save_folder):
        self.grid = QtWidgets.QGridLayout()
        ctr = 0
        if prefix is not None:
            self.grid.addWidget(QtWidgets.QLabel('Results Prefix:'), ctr, 0)
            self.pfx_fld = QtWidgets.QLineEdit()
            self.pfx_fld.setText(prefix)
            self.grid.addWidget(self.pfx_fld, ctr, 1, 1, 2)
            ctr += 1
        # Note: relies on Load being first entry
        for key, typ, capacity in data:
            self._adjust_typ[key] = typ
            if key != 'Load' and capacity is None:
                continue
       #     if key not in adjustin.keys():
       #         continue
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, typ, adjustin[key])
            except:
                mw = 0
                mwtxt = 'MW'
                mwcty = 0
                div = 0
            self._data[key] = [capacity / pow(10, div), div]
            self._adjust_cty[key] = QtWidgets.QDoubleSpinBox()
            self._adjust_cty[key].setRange(0, capacity / pow(10, div) * adjust_cap)
            self._adjust_cty[key].setDecimals(self._decpts)
            if self.show_multipliers:
                self._adjust_rnd[key] = QtWidgets.QDoubleSpinBox()
                self._adjust_rnd[key].setRange(0, adjust_cap)
                self._adjust_rnd[key].setDecimals(4)
            if key in adjustin.keys():
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    try:
                        self._adjust_mul[key] = adjustin[key] / capacity
                        self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
                    except:
                        self._adjust_mul[key] = 1.
                        self._adjust_rnd[key].setValue(1.)
            else:
                self._adjust_cty[key].setValue(0)
                if self.show_multipliers:
                    self._adjust_mul[key] = 0.
                    self._adjust_rnd[key].setValue(0.)
            self._adjust_cty[key].setObjectName(key)
            self.grid.addWidget(QtWidgets.QLabel(key), ctr, 0)
            self.grid.addWidget(self._adjust_cty[key], ctr, 1)
            self._adjust_txt[key] = QtWidgets.QLabel('')
            self._adjust_txt[key].setObjectName(key + 'label')
            self._adjust_txt[key].setText(mwtxt)
            self.grid.addWidget(self._adjust_txt[key], ctr, 2)
            if self.show_multipliers:
                self._adjust_cty[key].valueChanged.connect(self.adjustCap)
                self._adjust_rnd[key].setSingleStep(.1)
                self._adjust_rnd[key].setObjectName(key)
                self.grid.addWidget(self._adjust_rnd[key], ctr, 3)
                self._adjust_rnd[key].valueChanged.connect(self.adjustMult)
            ctr += 1
            if key == 'Load' and len(data) > 1:
                self.grid.addWidget(QtWidgets.QLabel('Facility'), ctr, 0)
                self.grid.addWidget(QtWidgets.QLabel('Capacity'), ctr, 1)
                if self.show_multipliers:
                    self.grid.addWidget(QtWidgets.QLabel('Multiplier'), ctr, 3)
                ctr += 1
        quit = QtWidgets.QPushButton('Quit', self)
        self.grid.addWidget(quit, ctr, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        self.grid.addWidget(show, ctr, 1)
        show.clicked.connect(self.showClicked)
        if prefix is not None:
            reset = QtWidgets.QPushButton('Reset', self)
            self.grid.addWidget(reset, ctr, 2)
            reset.clicked.connect(self.resetClicked)
        resetload = QtWidgets.QPushButton('Reset Load', self)
        self.grid.addWidget(resetload, ctr, 3)
        resetload.clicked.connect(self.resetloadClicked)
        if save_folder is not None:
            ctr += 1
            save = QtWidgets.QPushButton('Save', self)
            self.grid.addWidget(save, ctr, 0)
            save.clicked.connect(self.saveClicked)
            restore = QtWidgets.QPushButton('Restore', self)
            self.grid.addWidget(restore, ctr, 1)
            restore.clicked.connect(self.restoreClicked)
            listi = QtWidgets.QPushButton('List', self)
            self.grid.addWidget(listi, ctr, 2)
            listi.clicked.connect(self.listClicked)
            if self._batch_file is not None:
                batch = QtWidgets.QPushButton('Add to Batch', self)
                self.grid.addWidget(batch, ctr, 3)
                batch.clicked.connect(self.addtoBatch)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        # self.niceSize(ctr)
        self.setWindowTitle('SIREN - Powermatch - Adjust generators')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()
        

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def restoreClicked(self):
        ini_file = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Adjustments file',
                   self._save_folder, 'Preferences Files (*.ini)')[0]
        if ini_file != '':
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(ini_file)
            try:
                prefix = ini_file[ini_file.rfind('/') + 1: - 4]
            except:
                prefix = ''
            self.getIt(config, prefix)

    def listClicked(self):
        if os.path.exists(self._save_folder):
            names = {}
            techs = []
            ini_files = os.listdir(self._save_folder)
            ini_files.sort()
            for ini_file in ini_files:
                if ini_file[-4:] == '.ini':
                    config = configparser.RawConfigParser()
                    try:
                        config.read(self._save_folder + ini_file)
                    except:
                        continue
                    try:
                        adjustto = config.get('Powermatch', 'adjusted_capacities')
                    except:
                        continue
                    names[ini_file[:-4]] = [0] * len(techs)
                    bits = adjustto.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        key = bi[0]
                        try:
                            mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                                                          float(bi[1]))
                        except:
                            mwcty = 0
                        if mwcty == 0:
                            continue
                        if key not in techs:
                            techs.append(key)
                            names[ini_file[:-4]].append(0)
                        ndx = techs.index(key)
                        names[ini_file[:-4]][ndx] = mwcty
            techs.insert(0, 'Preference File')
            decpts = [1] * len(techs)
            dialog = Table(
                objects=names,  # The data to display
                decpts=decpts,  # The number of decimal points to display
                fields=techs,      # The fields (columns) to display
                save_folder=self._save_folder  # Optional folder to save the table to
            )
            dialog.setWindowTitle(self.sender().text())  # Set the title of the dialog
            dialog.exec_()  # Execute the dialog
            chosen = dialog.getItem(0)
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(self._save_folder + chosen + '.ini')
            self.getIt(config, chosen)
            del dialog  # Clean up the dialog

    def saveClicked(self):
        line = ''
        for key, value in self._adjust_cty.items():
            if self._decpts == 2:
                line += '{}={:.2f},'.format(key, value.value() * pow(10, self._data[key][1]))
            else:
                line += '{}={:.1f},'.format(key, value.value() * pow(10, self._data[key][1]))
        if line != '':
            line = 'adjusted_capacities=' + line[:-1]
            updates = {'Powermatch': ['adjustments=', line]}
            save_file = self._save_folder
            if self.pfx_fld.text() != '':
                save_file += '/' + self.pfx_fld.text()
            inifile = QtWidgets.QFileDialog.getSaveFileName(None, 'Save Adjustments to file',
                      save_file, 'Preferences Files (*.ini)')[0]
            if inifile != '':
                if inifile[-4:] != '.ini':
                    inifile = inifile + '.ini'
                SaveIni(updates, ini_file=inifile)

    def addtoBatch(self):
        check_list = list(self._adjust_cty.keys())[1:]
        wb = oxl.load_workbook(self._batch_file)
        batch_input_sheet = wb.worksheets[0]
        batch_input_sheet.protection.sheet = False
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        col = batch_input_sheet.max_column + 1
        tot_row = -1
        fst_row = -1
        if col == 4: # possibly only chart stuff in columns 2 and 3
            get_out = False
            for col in range(3, 1, -1):
                for row in range(1, batch_input_sheet.max_row + 1):
                    if batch_input_sheet.cell(row=row, column=col).value is not None:
                        col += 1
                        get_out = True
                        break
                    if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                        break
                if get_out:
                    break
        for row in range(1, batch_input_sheet.max_row + 1):
            if batch_input_sheet.cell(row=row, column=1).value is None:
                continue
            if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                new_cell = batch_input_sheet.cell(row=row, column=col)
                new_cell.value = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'MM-dd hh:mm')
                add_msg = new_cell.value
            if batch_input_sheet.cell(row=row, column=1).value == 'Capacity (MW)':
                fst_row = row + 1
                cell = batch_input_sheet.cell(row=row, column=col - 1)
                new_cell = batch_input_sheet.cell(row=row, column=col)
                new_cell.value = 'MW'
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                continue
            for key in self._adjust_cty.keys():
                if key == batch_input_sheet.cell(row=row, column=1).value:
                    cell = batch_input_sheet.cell(row=fst_row, column=col - 1)
                    new_cell = batch_input_sheet.cell(row=row, column=col)
                    new_cell.value = self._adjust_cty[key].value()
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                        if col == 2:
                            new_cell.font = normal
                            new_cell.number_format = '#0.00'
                        else:
                            new_cell.number_format = copy(cell.number_format)
                    elif col == 2:
                        new_cell.font = normal
                        new_cell.number_format = '#0.00'
                    try:
                        i = check_list.index(key)
                        del check_list[i]
                    except:
                        pass
            if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                tot_row = row
           #     if len(check_list) > 0:
           #         tot_row = row
        if len(check_list) > 0:
            check_list.reverse()
            cell = batch_input_sheet.cell(row=fst_row, column=col)
            for key in check_list:
                if self._adjust_cty[key].value() == 0:
                    continue
                batch_input_sheet.insert_rows(tot_row)
                new_cell = batch_input_sheet.cell(row=tot_row, column=1)
                new_cell.value = key
                new_cell = batch_input_sheet.cell(row=tot_row, column=col)
                new_cell.value = self._adjust_cty[key].value()
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                tot_row += 1
        if fst_row > 0 and tot_row > 0:
            new_cell = batch_input_sheet.cell(row=tot_row, column=col)
            new_cell.value = '=SUM(' + ssCol(col) + str(fst_row) + ':' + ssCol(col) + str(tot_row - 1) + ')'
            if col > 2:
                cell = batch_input_sheet.cell(row=tot_row, column=2)
            else:
                cell = batch_input_sheet.cell(row=tot_row, column=col)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
        wb.save(self._batch_file)
        QtWidgets.QMessageBox.about(self, 'SIREN - Add to Batch', "Added to batch as '" + add_msg + "' (column " + ssCol(col) + ')')

class setTransition(setTransitionBase, MyQDialog):
    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 70 + ctr * 32
        width = window.frameSize().width()
        screen = QtWidgets.QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, label, generators, sheet, year):
        setTransitionBase.__init__(self, parent, label, generators, sheet, year)
        MyQDialog.__init__(self, parent)  # Initialize the QDialog properly
        self._results = None
        i = generators.rfind('/')
        generator_file = generators[i + 1:]
        self.grid = QtWidgets.QGridLayout()
        r = 0
        self.grid.addWidget(QtWidgets.QLabel(label + ' File:'), r, 0)
        file_name = QtWidgets.QLabel(generator_file)
        file_name.setStyleSheet("border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.grid.addWidget(file_name, r, 1, 1, 5)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel(label + ' Sheet:'), r, 0)
        self.sheet = QtWidgets.QLineEdit()
        if sheet[-4:].isnumeric():
            sheet = sheet[:-4] + '$YEAR$'
        else:
            sheet = sheet.replace(year, '$YEAR$')
        self.sheet.setText(sheet)
        self.grid.addWidget(self.sheet, r, 1, 1, 2)
        r += 1
        quit = QtWidgets.QPushButton('Quit', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        self.grid.addWidget(show, r, 1)
        show.clicked.connect(self.showClicked)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.niceSize(r)
        self.setWindowTitle('SIREN - Powermatch - Transition files')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def showClicked(self):
        self.ignoreEnter = False
        self._results = self.sheet.text()
        self.close()

class PowerMatch(QtWidgets.QWidget, PowerMatchBase):
    log = QtCore.pyqtSignal()
    progress = QtCore.pyqtSignal()

    def __init__(self, TableClass=Table, *args, **kwargs):
        QtWidgets.QWidget.__init__(self)
        PowerMatchBase.__init__(self)
        self.displaytable = TableClass  # Store the TableClass as an instance variable
        self.help = help
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        self.init_settings(config)
        self.init_ui()

    def init_ui(self):
        # Setup the GUI layout
        self.opt_progressbar = None
        self.floatstatus = None # status window
   #     self.tabs = QtGui.QTabWidget()    # Create tabs
   #     tab1 = QtGui.QWidget()
   #     tab2 = QtGui.QWidget()
   #     tab3 = QtGui.QWidget()
   #     tab4 = QtGui.QWidget()
   #     tab5 = QtGui.QWidget()
        self.grid = QtWidgets.QGridLayout()
        self.updated = False
        edit = [None] * D
        r = 0
        for i in range(len(self.file_labels)):
            if i == R:
                self.grid.addWidget(QtWidgets.QLabel('Results Prefix:'), r, 0)
                self.results_pfx_fld = QtWidgets.QLineEdit()
                self.results_pfx_fld.setText(self.results_prefix)
                self.results_pfx_fld.textChanged.connect(self.pfxChanged)
                self.grid.addWidget(self.results_pfx_fld, r, 1, 1, 2)
                r += 1
            self.labels[i] = QtWidgets.QLabel(self.file_labels[i] + ' File:')
            self.grid.addWidget(self.labels[i], r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
            self.files[i].setText(self.ifiles[i])
            self.files[i].clicked.connect(self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 5)
            button = QtWidgets.QPushButton(f'Open {self.file_labels[i]} file', self)
            self.grid.addWidget(button, r, 6)
            button.clicked.connect(self.openClicked)
            if i < D:
                r += 1
                self.grid.addWidget(QtWidgets.QLabel(self.file_labels[i] + ' Sheet:'), r, 0)
                self.sheets[i] = QtWidgets.QComboBox()
                try:
                    curfile = self.get_filename(self.ifiles[i])
                    ts = WorkBook()
                    ts.open_workbook(curfile)
                    ndx = 0
                    j = -1
                    for sht in ts.sheet_names():
                        j += 1
                        self.sheets[i].addItem(sht)
                        if sht == self.isheets[i]:
                            ndx = j
                    self.sheets[i].setCurrentIndex(ndx)
                    self.ws = ts.sheet_by_index(ndx)
                    if i == C:
                        self.constraints = self.getConstraints(self.ws)
                    elif i == G:
                        self.generators = self.getGenerators(self.ws)
                    elif i == O:
                        self.optimisation = self.getOptimisation(self.ws)
                    ts.close()
                    del ts
                except Exception as error:
                    print('Get constraints, generators and optimisation:', error)
                self.sheets[i].addItem(self.isheets[i])
                self.grid.addWidget(self.sheets[i], r, 1, 1, 3)
                self.sheets[i].currentIndexChanged.connect(self.sheetChanged)
                edit[i] = QtWidgets.QPushButton(self.file_labels[i], self)
                self.grid.addWidget(edit[i], r, 4, 1, 2)
                edit[i].clicked.connect(self.editClicked)
            elif i == D and self.load_files != '':
                r += 1
                self.grid.addWidget(QtWidgets.QLabel('Load Folder:'), r, 0)
                self.load_dir = ClickableQLabel()
                try:
                    self.load_dir.setText(self.load_files[:self.load_files.rfind('/')])
                except:
                    self.load_dir.setText('')
                self.load_dir.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
                self.load_dir.clicked.connect(self.loaddirChanged)
                self.grid.addWidget(self.load_dir, r, 1, 1, 5)
                r += 1
                self.grid.addWidget(QtWidgets.QLabel('Load Year:'), r, 0)
                self.load_years = self.get_load_years()
                self.loadCombo = QtWidgets.QComboBox()
                for choice in self.load_years:
                    self.loadCombo.addItem(choice)
                    if choice == self.load_year:
                        self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
                self.loadCombo.currentIndexChanged.connect(self.changes)
                self.grid.addWidget(self.loadCombo, r, 1)
                self.grid.addWidget(QtWidgets.QLabel("(To to use a different load year to the data file. Otherwise choose 'n/a')"), r, 2, 1, 4)
            r += 1
      #  wdth = edit[1].fontMetrics().boundingRect(edit[1].text()).width() + 9
        self.grid.addWidget(QtWidgets.QLabel('Replace Last:'), r, 0)
        if self.batch_new_file:
            msg = '(check to replace an existing Results workbook)'
        else:
            msg = '(check to replace last Results worksheet in Batch spreadsheet)'
        self.replace_last = QtWidgets.QCheckBox(msg, self)
        self.replace_last.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.replace_last, r, 1, 1, 3)
        self.grid.addWidget(QtWidgets.QLabel('Prefix facility names in Batch report:'), r, 4)
        self.batch_prefix_check = QtWidgets.QCheckBox('', self)
        if self.batch_prefix:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Checked)
        else:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.batch_prefix_check, r, 5)
        self.batch_prefix_check.stateChanged.connect(self.bpcchanged)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Discount Rate:'), r, 0)
        self.discount = QtWidgets.QDoubleSpinBox()
        self.discount.setRange(0, 100)
        self.discount.setDecimals(2)
        self.discount.setSingleStep(.5)
        try:
            self.discount.setValue(self.discount_rate * 100.)
        except:
            self.discount.setValue(0.)
        self.grid.addWidget(self.discount, r, 1)
        self.discount.valueChanged.connect(self.drchanged)
        self.grid.addWidget(QtWidgets.QLabel('(%. Only required if using input costs rather than reference LCOE)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Carbon Price:'), r, 0)
        self.carbon = QtWidgets.QDoubleSpinBox()
        self.carbon.setRange(0, self.carbon_price_max)
        self.carbon.setDecimals(2)
        try:
            self.carbon.setValue(self.carbon_price)
        except:
            self.carbon.setValue(0.)
        self.grid.addWidget(self.carbon, r, 1)
        self.carbon.valueChanged.connect(self.cpchanged)
        self.grid.addWidget(QtWidgets.QLabel('($/tCO2e. Use only if LCOE excludes carbon price)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Adjust Generators:'), r, 0)
        self.adjust = QtWidgets.QCheckBox('(check to adjust generators capacity data)', self)
        if self.adjust_gen:
            self.adjust.setCheckState(QtCore.Qt.Checked)
        self.grid.addWidget(self.adjust, r, 1, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Dispatch Order:\n(move to right\nto exclude)'), r, 0)
        self.order = ListWidget(self) #QtWidgets.QListWidget()
      #  self.order.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.order, r, 1, 1, 3)
        self.ignore = ListWidget(self) # QtWidgets.QListWidget()
      #  self.ignore.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.ignore, r, 4, 1, 3)
        r += 1
        self.log = QtWidgets.QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, r, 1, 1, 6)
        r += 1
        self.progressbar = QtWidgets.QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(20) # was 10 set to 20 to get 5% steps
        self.progressbar.setValue(0)
        self.progressbar.setStyleSheet('QProgressBar {border: 1px solid grey; border-radius: 2px; text-align: center;}' \
                                       + 'QProgressBar::chunk { background-color: #06A9D6;}')
        self.grid.addWidget(self.progressbar, r, 1, 1, 6)
        self.progressbar.setHidden(True)
        r += 1
        r += 1
        quit = QtWidgets.QPushButton('Done', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.progress_handler = QtProgressHandler(self.progressbar)
        pms = QtWidgets.QPushButton('Summary', self)
        self.grid.addWidget(pms, r, 1)
        # pms.clicked.connect(self.pmClicked)
        pms.clicked.connect(lambda checked, handler=self.progress_handler: self.pmClicked('Summary', handler))
        pm = QtWidgets.QPushButton('Detail', self)
     #   pm.setMaximumWidth(wdth)
        self.grid.addWidget(pm, r, 2)
        # pm.clicked.connect(self.pmClicked)
        pm.clicked.connect(lambda checked, handler=self.progress_handler: self.pmClicked('Detail', handler))
        btch = QtWidgets.QPushButton('Batch', self)
        self.grid.addWidget(btch, r, 3)
        # btch.clicked.connect(self.pmClicked)
        btch.clicked.connect(lambda checked, handler=self.progress_handler, adjuster=self.create_adjustments: self.pmClicked('Batch', handler, adjuster))
        trns = QtWidgets.QPushButton('Transition', self)
        self.grid.addWidget(trns, r, 4)
        # trns.clicked.connect(self.pmClicked)
        trns.clicked.connect(lambda checked, handler=self.progress_handler, set_transition=self.create_setTransitions: self.pmClicked('Transition', handler, None, set_transition))
        opt = QtWidgets.QPushButton('Optimise', self)
        self.grid.addWidget(opt, r, 5)
        # opt.clicked.connect(self.pmClicked)
        opt.clicked.connect(lambda checked, handler=self.progress_handler, optimiser=self.optClicked: self.pmClicked('Optimise', handler, optimiser))
        help = QtWidgets.QPushButton('Help', self)
     #   help.setMaximumWidth(wdth)
      #  quit.setMaximumWidth(wdth)
        self.grid.addWidget(help, r, 6)
        help.clicked.connect(self.helpClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
     #   self.grid.setColumnStretch(0, 2)
        r += 1
        editini = QtWidgets.QPushButton('Preferences', self)
     #   editini.setMaximumWidth(wdth)
        self.grid.addWidget(editini, r, 0)
        editini.clicked.connect(self.editIniFile)
        do_tml = False
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                do_tml = True
        elif os.path.exists('pmtmldetail.py'):
            do_tml = True
        if do_tml:
            tmld = QtWidgets.QPushButton('TML Detail', self)
            self.grid.addWidget(tmld, r, 1)
            tmld.clicked.connect(self.tmlClicked)
        self.setOrder()
        if len(self.iorder) > 0:
            self.order.clear()
            self.ignore.clear()
            for gen in self.iorder:
                self.order.addItem(gen)
            try:
                for gen in self.generators.keys():
                    if (gen in tech_names and gen not in self.dispatchable) or gen in self.iorder:
                        continue
                    try:
                        chk = gen[gen.find('.') + 1:]
                        if chk in tech_names and chk not in self.dispatchable:
                            continue
                    except:
                        pass
                    self.ignore.addItem(gen)
            except:
                pass
        if self.adjust_gen and self.adjustto is None:
           self.adjustto = {}
           self.adjustto['Load'] = 0
           for gen in tech_names:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
           for gen in self.iorder:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
    #    tab1.setLayout(self.layout)
    #    self.tabs.addTab(tab1,'Parms')
    #    self.tabs.addTab(tab2,'Constraints')
    #    self.tabs.addTab(tab3,'Generators')
    #    self.tabs.addTab(tab4,'Summary')
    #    self.tabs.addTab(tab5,'Details')
        self.setWindowTitle('SIREN - powermatch (' + fileVersion() + ') - Powermatch')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        if self.restorewindows:
            try:
                rw = config.get('Windows', 'powermatch_size').split(',')
                self.resize(int(rw[0]), int(rw[1]))
                mp = config.get('Windows', 'powermatch_pos').split(',')
                self.move(int(mp[0]), int(mp[1]))
            except:
                pass
        else:
            self.center()
            self.resize(int(self.sizeHint().width() * 1.2), int(self.sizeHint().height() * 1.2))

    def create_adjustments(self, pmss_data, load_col):
        if self.adjust_gen:
            generated = sum(pmss_data[load_col])
            datain = [['Load', 'L', generated]]
            adjustto = {'Load': generated}
            adjust = Adjustments(self, datain, adjustto, self.adjust_cap, None,
                                        show_multipliers=self.show_multipliers)
            adjust.exec_()
            return adjust
        return None
    
    def create_setTransitions(self, label, generators, sheet, year):
        files = setTransition(self, label, generators,
                                      sheet, year)
        files.exec_()
        return files

    def optClicked(self, in_year, sender_text, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, data_file, opt_progress_handler: Optional[ProgressHandler] = None, ):

        def create_starting_population(individuals, chromosome_length):
            # Set up an initial array of all zeros
            population = np.zeros((individuals, chromosome_length))
            # Loop through each row (individual)
            for i in range(individuals):
                # Choose a random number of ones to create but at least one 1
                ones = random.randint(1, chromosome_length)
                # Change the required number of zeros to ones
                population[i, 0:ones] = 1
                # Sfuffle row
                np.random.shuffle(population[i])
            return population

        def select_individual_by_tournament(population, *argv):
            # Get population size
            population_size = len(population)

            # Pick individuals for tournament
            fighter = [0, 0]
            fighter_fitness = [0, 0]
            fighter[0] = random.randint(0, population_size - 1)
            fighter[1] = random.randint(0, population_size - 1)

            # Get fitness score for each
            if len(argv) == 1:
                fighter_fitness[0] = argv[0][fighter[0]]
                fighter_fitness[1] = argv[0][fighter[1]]
            else:
                for arg in argv:
                    min1 = min(arg)
                    max1 = max(arg)
                    for f in range(len(fighter)):
                        try:
                            fighter_fitness[f] += (arg[f] - min1) / (max1 - min1)
                        except:
                            pass
            # Identify individual with lowest score
            # Fighter 1 will win if score are equal
            if fighter_fitness[0] <= fighter_fitness[1]:
                winner = fighter[0]
            else:
                winner = fighter[1]

            # Return the chromsome of the winner
            return population[winner, :]

        def breed_by_crossover(parent_1, parent_2, points=2):
            # Get length of chromosome
            chromosome_length = len(parent_1)

            # Pick crossover point, avoiding ends of chromsome
            if points == 1:
                crossover_point = random.randint(1,chromosome_length-1)
            # Create children. np.hstack joins two arrays
                child_1 = np.hstack((parent_1[0:crossover_point],
                                     parent_2[crossover_point:]))
                child_2 = np.hstack((parent_2[0:crossover_point],
                                     parent_1[crossover_point:]))
            else: # only do 2 at this    stage
                crossover_point_1 = random.randint(1, chromosome_length - 2)
                crossover_point_2 = random.randint(crossover_point_1 + 1, chromosome_length - 1)
                child_1 = np.hstack((parent_1[0:crossover_point_1],
                                     parent_2[crossover_point_1:crossover_point_2],
                                     parent_1[crossover_point_2:]))
                child_2 = np.hstack((parent_2[0:crossover_point_1],
                                     parent_1[crossover_point_1:crossover_point_2],
                                     parent_2[crossover_point_2:]))
            # Return children
            return child_1, child_2

        def randomly_mutate_population(population, mutation_probability):
            # Apply random mutation
            random_mutation_array = np.random.random(size=(population.shape))
            random_mutation_boolean = random_mutation_array <= mutation_probability
        #    random_mutation_boolean[0][:] = False # keep the best multi and lcoe
       #     random_mutation_boolean[1][:] = False
            population[random_mutation_boolean] = np.logical_not(population[random_mutation_boolean])
            # Return mutation population
            return population

        def calculate_fitness(population):
            lcoe_fitness_scores = [] # scores = LCOE values
            multi_fitness_scores = [] # scores = multi-variable weight
            multi_values = [] # values for each of the six variables
            if len(population) == 1:
                option = O1
            else:
                option = O
            if self.debug:
                self.popn += 1
                self.chrom = 0
            for chromosome in population:
                # now get random amount of generation per technology (both RE and non-RE)
                for fac, value in opt_order.items():
                    capacity = value[2]
                    for c in range(value[0], value[1]):
                        if chromosome[c]:
                            capacity = capacity + capacities[c]
                    try:
                        pmss_details[fac].multiplier = capacity / pmss_details[fac].capacity
                    except:
                        print('PME2:', gen, capacity, pmss_details[fac].capacity)
                multi_value, op_data, extra = self.doDispatch(year, sender_text, option, pmss_details, pmss_data, re_order,
                                              dispatch_order, pm_data_file, data_file, progress_handler=opt_progress_handler)
                if multi_value['load_pct'] < self.targets['load_pct'][3]:
                    if multi_value['load_pct'] == 0:
                        print('PME3:', multi_value['lcoe'], self.targets['load_pct'][3], multi_value['load_pct'])
                        lcoe_fitness_scores.append(1)
                    else:
                        try:
                            lcoe_fitness_scores.append(pow(multi_value['lcoe'],
                                self.targets['load_pct'][3] / multi_value['load_pct']))
                        except OverflowError as err:
                            self.setStatus(f"Overflow error: {err}; POW({multi_value['lcoe']:,}, " \
                                         + f"{self.targets['load_pct'][3] / multi_value['load_pct']:,}) " \
                                         + f"({self.targets['load_pct'][3]:,} / {multi_value['load_pct']:,} )")
                        except:
                            pass
                else:
                    lcoe_fitness_scores.append(multi_value['lcoe'])
                multi_values.append(multi_value)
                multi_fitness_scores.append(calc_weight(multi_value))
                if self.debug:
                    self.chrom += 1
                    line = str(self.popn) + ',' + str(self.chrom) + ','
                    for fac, value in opt_order.items():
                        try:
                            line += str(pmss_details[fac].capacity * pmss_details[fac].multiplier) + ','
                        except:
                            line += ','
                    for key in self.targets.keys():
                        try:
                            line += '{:.3f},'.format(multi_value[key])
                        except:
                            line += multi_value[key] + ','
                    line += '{:.5f},'.format(multi_fitness_scores[-1])
                    self.db_file.write(line + '\n')
            # alternative approach to calculating fitness
            multi_fitness_scores1 = []
            maxs = {}
            mins = {}
            tgts = {}
            for key in multi_value.keys():
                if key[-4:] == '_pct':
                    tgts[key] = abs(self.targets[key][2] - self.targets[key][3])
                else:
                    maxs[key] = 0
                    mins[key] = -1
                    for popn in multi_values:
                        try:
                            tgt = abs(self.targets[key][2] - popn[key])
                        except:
                            continue
                        if tgt > maxs[key]:
                            maxs[key] = tgt
                        if mins[key] < 0 or tgt < mins[key]:
                            mins[key] = tgt
            for popn in multi_values:
                weight = 0
                for key, value in multi_value.items():
                    if self.targets[key][1] <= 0:
                        continue
                    try:
                        tgt = abs(self.targets[key][2] - popn[key])
                    except:
                        continue
                    if key[-4:] == '_pct':
                        if tgts[key] != 0:
                            if tgt > tgts[key]:
                                weight += 1 * self.targets[key][1]
                            else:
                                try:
                                    weight += 1 - ((tgt / tgts[key]) * self.targets[key][1])
                                except:
                                    pass
                    else:
                        try:
                            weight += 1 - (((maxs[key] - tgt) / (maxs[key] - mins[key])) \
                                      * self.targets[key][1])
                        except:
                            pass
                multi_fitness_scores1.append(weight)
            if len(population) == 1: # return the table for best chromosome
                return op_data, multi_values
            else:
                return lcoe_fitness_scores, multi_fitness_scores, multi_values

        def optQuitClicked(event):
            self.optExit = True
            optDialog.close()

        def chooseClicked(event):
            self.opt_choice = sender_text
            chooseDialog.close()

        def calc_weight(multi_value, calc=0):
            weight = [0., 0.]
            if calc == 0:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    elif value[2] > value[3]: # wants higher target
                        if multi_value[key] > value[2]: # high no weight
                            w = 0.
                        elif multi_value[key] < value[3]: # low maximum weight
                            w = 2.
                        else:
                            w = 1 - (multi_value[key] - value[3]) / (value[2] - value[3])
                    else: # lower target
                        if multi_value[key] == -1 or multi_value[key] > value[3]: # high maximum weight
                            w = 2.
                        elif multi_value[key] < value[2]: # low no weight
                            w = 0.
                        else:
                            w = multi_value[key] / (value[3] - value[2])
                    weight[0] += w * value[1]
            elif calc == 1:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if multi_value[key] < 0:
                        w = 1
                    elif value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    else: # target range
                        w = min(abs(value[2] - multi_value[key]) / abs(value[2] - value[3]), 1)
                    weight[1] += w * value[1]
            return weight[calc]

        def show_multitable(best_score_progress, multi_best, multi_order, title):
            def pwr_chr(amt):
                pwr_chrs = ' KMBTPEZY'
                pchr = ''
                divisor = 1.
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if amt > pow(10, pwr * 3):
                        pchr = pwr_chrs[pwr]
                        divisor = 1. * pow(10, pwr * 3)
                        break
                return amt / divisor, pchr

            def opt_fmat(amt, fmat):
                tail = ' '
                p = fmat.find('pwr_chr')
                if p > 0:
                    amt, tail = pwr_chr(amt)
                    fmat = fmat.replace('pwr_chr', '')
              #  d = fmat.find('d')
             #   if d > 0:
              #      amt = amt * 100.
                fmat = fmat.replace('.1f', '.2f')
                i = fmat.find('%')
                fmt = fmat[:i] + '{:> 8,' + fmat[i:].replace('%', '').replace('d', '.2%') + '}' + tail
                return fmt.format(amt)

            best_table = []
            best_fmate = []
            for b in range(len(multi_best)):
                bl = list(multi_best[b].values())
                bl.insert(0, best_score_progress[b])
                bl.insert(0, b + 1)
                best_table.append(bl)
                best_fmate.append([b + 1])
                best_fmate[-1].insert(1, best_score_progress[b])
                for f in range(2, len(best_table[-1])):
                    if target_keys[f - 2] in ['load_pct', 'surplus_pct', 're_pct']:
                        best_fmate[-1].append(opt_fmat(bl[f], '%d%%'))
                    else:
                        best_fmate[-1].append(opt_fmat(bl[f], target_fmats[f - 2]))
            fields = target_names[:]
            fields.insert(0, 'weight')
            fields.insert(0, 'iteration')
            dialog = self.displaytable.Table(best_fmate, fields=fields, txt_align='R', decpts=[0, 4],
                     title=title, sortby='weight')
            dialog.exec_()
            b = int(dialog.getItem(0)) - 1
            del dialog
            pick = [b]
            for fld in multi_order[:3]:
                i = target_keys.index(fld)
                pick.append(best_table[b][i + 2])
            return [pick]

#       optClicked mainline starts here
        year = in_year
        option = in_option
        pmss_details = dict(in_pmss_details)
        pmss_data = in_pmss_data[:]
        re_order = in_re_order[:]
        dispatch_order = in_dispatch_order[:]
        if self.optimise_debug:
            self.debug = True
        else:
            self.debug = False
        missing = []
        for fac in re_order:
            if fac == 'Load':
                continue
            if fac not in self.optimisation.keys():
                missing.append(fac)
        for gen in dispatch_order:
            if gen not in self.optimisation.keys():
                missing.append(gen)
        if len(missing) > 0:
            bad = False
            if self.optimise_default is not None:
                defaults = self.optimise_default.split(',')
                if len(defaults) < 1 or len(defaults) > 3:
                    bad = True
                else:
                    try:
                        for miss in missing:
                            if len(defaults) == 2:
                                minn = 0
                            else:
                                if defaults[0][-1] == 'd':
                                    minn = pmss_details[miss].capacity * float(defaults[0][:-1])
                                elif defaults[0][-1] == 'c':
                                    minn = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[0][:-1])
                                else:
                                    minn = float(defaults[0])
                            if defaults[-2][-1] == 'd':
                                maxx = pmss_details[miss].capacity * float(defaults[-2][:-1])
                            elif defaults[-2][-1] == 'c':
                                maxx = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[-2][:-1])
                            else:
                                maxx = float(defaults[-2])
                            if len(defaults) == 3:
                                if defaults[-1][-1].lower() == 'd':
                                    step = capacity * float(defaults[-1][:-1])
                                elif defaults[-1][-1].lower() == 'c':
                                    step = capacity * multiplier * float(defaults[-1][:-1])
                                else:
                                    step = float(defaults[-1])
                            else:
                                step = (maxx - minn) / float(defaults[-1])
                            self.optimisation[miss] =  Optimisation(miss, 'None', None)
                            self.optimisation[miss].approach = 'Range'
                            self.optimisation[miss].capacity_min = minn
                            self.optimisation[miss].capacity_max = maxx
                            self.optimisation[miss].capacity_step = step
                    except:
                        bad = True
                check = ''
                for miss in missing:
                    check += miss + ', '
                check = check[:-2]
                if bad:
                    self.setStatus('Key Error: Missing Optimisation entries for: ' + check)
                    return
                self.setStatus('Missing Optimisation entries added for: ' + check)
        self.optExit = False
        self.setStatus('Optimise processing started')
        err_msg = ''
        
        optDialog = QtWidgets.QDialog()
        grid = QtWidgets.QGridLayout()
        grid.addWidget(QtWidgets.QLabel('Adjust load'), 0, 0)
        self.optLoad = QtWidgets.QDoubleSpinBox()
        self.optLoad.setRange(-1, self.adjust_cap)
        self.optLoad.setDecimals(4)
        self.optLoad.setSingleStep(.1)
        rw = 0
        grid.addWidget(self.optLoad, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Multiplier for input Load'), rw, 2, 1, 3)
        self.optLoad.setValue(pmss_details['Load'].multiplier)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Population size'), rw, 0)
        optPopn = QtWidgets.QSpinBox()
        optPopn.setRange(10, 500)
        optPopn.setSingleStep(10)
        optPopn.setValue(self.optimise_population)
        optPopn.valueChanged.connect(self.changes)
        grid.addWidget(optPopn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Size of population'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('No. of iterations'), rw, 0, 1, 3)
        optGenn = QtWidgets.QSpinBox()
        optGenn.setRange(10, 500)
        optGenn.setSingleStep(10)
        optGenn.setValue(self.optimise_generations)
        optGenn.valueChanged.connect(self.changes)
        grid.addWidget(optGenn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Number of iterations (generations)'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Mutation probability'), rw, 0)
        optMutn = QtWidgets.QDoubleSpinBox()
        optMutn.setRange(0, 1)
        optMutn.setDecimals(4)
        optMutn.setSingleStep(0.001)
        optMutn.setValue(self.optimise_mutation)
        optMutn.valueChanged.connect(self.changes)
        grid.addWidget(optMutn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Add in mutation'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Exit if stable'), rw, 0)
        optStop = QtWidgets.QSpinBox()
        optStop.setRange(0, 50)
        optStop.setSingleStep(10)
        optStop.setValue(self.optimise_stop)
        optStop.valueChanged.connect(self.changes)
        grid.addWidget(optStop, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Exit if LCOE/weight remains the same after this many iterations'),
                       rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Optimisation choice'), rw, 0)
        optCombo = QtWidgets.QComboBox()
        choices = ['LCOE', 'Multi', 'Both']
        for choice in choices:
            optCombo.addItem(choice)
            if choice == self.optimise_choice:
                optCombo.setCurrentIndex(optCombo.count() - 1)
        grid.addWidget(optCombo, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Choose type of optimisation'),
                       rw, 2, 1, 3)
        rw += 1
        # for each variable name
        grid.addWidget(QtWidgets.QLabel('Variable'), rw, 0)
        grid.addWidget(QtWidgets.QLabel('Weight'), rw, 1)
        grid.addWidget(QtWidgets.QLabel('Better'), rw, 2)
        grid.addWidget(QtWidgets.QLabel('Worse'), rw, 3)
        rw += 1
        ndx = grid.count()
        for key in self.targets.keys():
            self.targets[key][4] = ndx
            ndx += 4
        for key, value in self.targets.items():
            if value[2] == value[3]:
                ud = '(=)'
            elif value[2] < 0:
                ud = '(<html>&uarr;</html>)'
            elif value[3] < 0 or value[3] > value[2]:
                ud = '(<html>&darr;</html>)'
            else:
                ud = '(<html>&uarr;</html>)'
            grid.addWidget(QtWidgets.QLabel(value[0] + ': ' + ud), rw, 0)
            weight = QtWidgets.QDoubleSpinBox()
            weight.setRange(0, 1)
            weight.setDecimals(2)
            weight.setSingleStep(0.05)
            weight.setValue(value[1])
            grid.addWidget(weight, rw, 1)
            if key[-4:] == '_pct':
                minim = QtWidgets.QDoubleSpinBox()
                minim.setRange(-.1, 1.)
                minim.setDecimals(2)
                minim.setSingleStep(0.1)
                minim.setValue(value[2])
                grid.addWidget(minim, rw, 2)
                maxim = QtWidgets.QDoubleSpinBox()
                maxim.setRange(-.1, 1.)
                maxim.setDecimals(2)
                maxim.setSingleStep(0.1)
                maxim.setValue(value[3])
                grid.addWidget(maxim, rw, 3)
            else:
                minim = QtWidgets.QLineEdit()
                minim.setValidator(QtGui.QDoubleValidator())
                minim.validator().setDecimals(2)
                minim.setText(str(value[2]))
                grid.addWidget(minim, rw, 2)
                maxim = QtWidgets.QLineEdit()
                maxim.setValidator(QtGui.QDoubleValidator())
                maxim.validator().setDecimals(2)
                maxim.setText(str(value[3]))
                grid.addWidget(maxim, rw, 3)
            rw += 1
        quit = QtWidgets.QPushButton('Quit', self)
        grid.addWidget(quit, rw, 0)
        quit.clicked.connect(self._optQuitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        grid.addWidget(show, rw, 1)
        show.clicked.connect(optDialog.close)
        optDialog.setLayout(grid)
        optDialog.setWindowTitle('Choose Optimisation Parameters')
        optDialog.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        optDialog.exec_()
        if self.optExit: # a fudge to exit
            self.setStatus('Execution aborted.')
            return
        # check we have optimisation entries for generators and storage
        # update any changes to targets
        self.optimise_choice = optCombo.currentText()
        for key in self.targets.keys():
            weight = grid.itemAt(self.targets[key][4] + 1).widget()
            self.targets[key][1] = weight.value()
            minim = grid.itemAt(self.targets[key][4] + 2).widget()
            try:
                self.targets[key][2] = minim.value()
            except:
                self.targets[key][2] = float(minim.text())
            maxim = grid.itemAt(self.targets[key][4] + 3).widget()
            try:
                self.targets[key][3] = maxim.value()
            except:
                self.targets[key][3] = float(maxim.text())
        # might want to save load value if changed
        pmss_details['Load'].multiplier = self.optLoad.value()
    
        updates = {}
        lines = []
        lines.append('optimise_choice=' + self.optimise_choice)
        lines.append('optimise_generations=' + str(self.optimise_generations))
        lines.append('optimise_mutation=' + str(self.optimise_mutation))
        lines.append('optimise_population=' + str(self.optimise_population))
        lines.append('optimise_stop=' + str(self.optimise_stop))
        if self.optimise_choice == 'LCOE':
            do_lcoe = True
            do_multi = False
        elif self.optimise_choice == 'Multi':
            do_lcoe = False
            do_multi = True
        else:
            do_lcoe = True
            do_multi = True
        multi_order = []
        for key, value in self.targets.items():
            line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
            lines.append(line)
            multi_order.append('{:.2f}{}'.format(value[1], key))
        updates['Powermatch'] = lines
        SaveIni(updates)
        multi_order.sort(reverse=True)
    #    multi_order = multi_order[:3] # get top three weighted variables - but I want them all
        multi_order = [o[4:] for o in multi_order]
        self.adjust_gen = True
        orig_load = []
        load_col = -1
        orig_tech = {}
        orig_capacity = {}
        opt_order = {} # rely on it being processed in added order
        # each entry = [first entry in chrom, last entry, minimum capacity]
        # first get original renewables generation from data sheet
        for fac in re_order:
            if fac == 'Load':
                continue
            opt_order[fac] = [0, 0, 0]
        # now add scheduled generation
        for gen in dispatch_order:
            opt_order[gen] = [0, 0, 0]
        capacities = []
        for gen in opt_order.keys():
            opt_order[gen][0] = len(capacities) # first entry
            try:
                if self.optimisation[gen].approach == 'Discrete':
                    capacities.extend(self.optimisation[gen].capacities)
                    opt_order[gen][1] = len(capacities) # last entry
                elif self.optimisation[gen].approach == 'Range':
                    if self.optimisation[gen].capacity_max == self.optimisation[gen].capacity_min:
                        capacities.extend([0])
                        opt_order[gen][1] = len(capacities)
                        opt_order[gen][2] = self.optimisation[gen].capacity_min
                        continue
                    ctr = int((self.optimisation[gen].capacity_max - self.optimisation[gen].capacity_min) / \
                              self.optimisation[gen].capacity_step)
                    if ctr < 1:
                        self.setStatus("Error with Optimisation table entry for '" + gen + "'")
                        return
                    capacities.extend([self.optimisation[gen].capacity_step] * ctr)
                    tot = self.optimisation[gen].capacity_step * ctr + self.optimisation[gen].capacity_min
                    if tot < self.optimisation[gen].capacity_max:
                        capacities.append(self.optimisation[gen].capacity_max - tot)
                    opt_order[gen][1] = len(capacities)
                    opt_order[gen][2] = self.optimisation[gen].capacity_min
                else:
                    opt_order[gen][1] = len(capacities)
            except KeyError as err:
                self.setStatus('Key Error: No Optimisation entry for ' + str(err))
                opt_order[gen] = [len(capacities), len(capacities) + 5, 0]
                capacities.extend([pmss_details[gen].capacity / 5.] * 5)
            except ZeroDivisionError as err:
                self.setStatus('Zero capacity: ' + gen + ' ignored')
            except:
                err = str(sys.exc_info()[0]) + ',' + str(sys.exc_info()[1]) + ',' + gen + ',' \
                      + str(opt_order[gen])
                self.setStatus('Error: ' + str(err))
                return
        # chromosome = [1] * int(len(capacities) / 2) + [0] * (len(capacities) - int(len(capacities) / 2))
        # we have the original data - from here down we can do our multiple optimisations
        # Set general parameters
        self.setStatus('Optimisation choice is ' + self.optimise_choice)
        chromosome_length = len(capacities)
        self.setStatus(f'Chromosome length: {chromosome_length}; {pow(2, chromosome_length):,} permutations')
        self.optimise_population = optPopn.value()
        population_size = self.optimise_population
        self.optimise_generations = optGenn.value()
        maximum_generation = self.optimise_generations
        self.optimise_mutation = optMutn.value()
        self.optimise_stop = optStop.value()
        lcoe_scores = []
        multi_scores = []
        multi_values = []
     #   if do_lcoe:
      #      lcoe_target = 0. # aim for this LCOE
        if do_multi:
            multi_best = [] # list of six variables for best weight
            multi_best_popn = [] # list of chromosomes for best weight
        self.show_ProgressBar(maximum=optGenn.value(), msg='Process iterations', title='SIREN - Powermatch Progress')
        self.opt_progressbar.setVisible(True)
        start_time = time.time()
        # Create starting population
        self.opt_progressbar.barProgress(1, 'Processing iteration 1')
        QtCore.QCoreApplication.processEvents()
        population = create_starting_population(population_size, chromosome_length)
        # calculate best score(s) in starting population
        # if do_lcoe best_score = lowest non-zero lcoe
        # if do_multi best_multi = lowest weight and if not do_lcoe best_score also = best_weight
        if self.debug:
            filename = self.scenarios + 'opt_debug_' + \
                datetime.now().strftime('_%Y-%M-%d_%H%M') + '.csv'
            self.db_file = open(filename, 'w')
            line0 = 'Popn,Chrom,'
            line1 = 'Weights,,'
            line2 = 'Targets,,'
            line3 = 'Range,' + str(population_size) + ','
            for gen, value in opt_order.items():
                 line0 += gen + ','
                 line1 += ','
                 line2 += ','
                 line3 += ','
            for key in self.targets.keys():
                 line0 += key + ','
                 line1 += str(self.targets[key][1]) + ','
                 line2 += str(self.targets[key][2]) + ','
                 if key[-4:] == '_pct':
                     line3 += str(abs(self.targets[key][2] - self.targets[key][3])) + ','
                 else:
                     line3 += ','
            line0 += 'Score'
            self.db_file.write(line0 + '\n' + line1 + '\n' + line2 + '\n' + line3 + '\n')
            self.popn = 0
            self.chrom = 0
        lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
        if do_lcoe:
            try:
                best_score = np.min(lcoe_scores)
            except:
                print('PME4:', lcoe_scores)
            best_ndx = lcoe_scores.index(best_score)
            lowest_chrom = population[best_ndx]
            self.setStatus('Starting LCOE: $%.2f' % best_score)
        if do_multi:
            if self.more_details: # display starting population ?
                pick = self._plot_multi(multi_scores, multi_values, multi_order, 'starting population')
            # want maximum from first round to set base upper limit
            for key in self.targets.keys():
                if self.targets[key][2] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][2] = setit
                if self.targets[key][3] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][3] = setit
            # now we can find the best weighted result - lowest is best
            best_multi = np.min(multi_scores)
            best_mndx = multi_scores.index(best_multi)
            multi_lowest_chrom = population[best_mndx]
            multi_best_popn.append(multi_lowest_chrom)
            multi_best.append(multi_values[best_mndx])
            self.setStatus('Starting Weight: %.4f' % best_multi)
            multi_best_weight = best_multi
            best_multi_progress = [best_multi]
            if not do_lcoe:
                best_score = best_multi
            last_multi_score = best_multi
            lowest_multi_score = best_multi
            mud = '='
        # Add starting best score to progress tracker
        best_score_progress = [best_score]
        best_ctr = 1
        last_score = best_score
        lowest_score = best_score
        lud = '='
        # Now we'll go through the generations of genetic algorithm
        for generation in range(1, maximum_generation):
            lcoe_status = ''
            multi_status = ''
            if do_lcoe:
                lcoe_status = ' %s $%.2f ;' % (lud, best_score)
            if do_multi:
                multi_status = ' %s %.4f ;' % (mud, best_multi)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = ' (%s%s %.1f secs)' % (lcoe_status, multi_status, tim)
            else:
                tim = ' (%s%s %.2f mins)' % (lcoe_status, multi_status, tim / 60.)
            self.opt_progressbar.barProgress(generation + 1,
                'Processing iteration ' + str(generation + 1) + tim)
            QtWidgets.QApplication.processEvents()
            if not self.opt_progressbar.be_open:
                break
        # Create an empty list for new population
            new_population = []
        # Using elitism approach include best individual
            if do_lcoe:
                new_population.append(lowest_chrom)
            if do_multi:
                new_population.append(multi_lowest_chrom)
            # Create new population generating two children at a time
            if do_lcoe:
                if do_multi:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
                else:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
            else:
                for i in range(int(population_size/2)):
                    parent_1 = select_individual_by_tournament(population, multi_scores)
                    parent_2 = select_individual_by_tournament(population, multi_scores)
                    child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                    new_population.append(child_1)
                    new_population.append(child_2)
            # get back to original size (after elitism adds)
            if do_lcoe:
                new_population.pop()
            if do_multi:
                new_population.pop()
            # Replace the old population with the new one
            population = np.array(new_population)
            if self.optimise_mutation > 0:
                population = randomly_mutate_population(population, self.optimise_mutation)
            # Score best solution, and add to tracker
            lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
            if do_lcoe:
                best_lcoe = np.min(lcoe_scores)
                best_ndx = lcoe_scores.index(best_lcoe)
                best_score = best_lcoe
            # now we can find the best weighted result - lowest is best
            if do_multi:
                best_multi = np.min(multi_scores)
                best_mndx = multi_scores.index(best_multi)
                multi_lowest_chrom = population[best_mndx]
                multi_best_popn.append(multi_lowest_chrom)
                multi_best.append(multi_values[best_mndx])
           #     if multi_best_weight > best_multi:
                multi_best_weight = best_multi
                if not do_lcoe:
                    best_score = best_multi
                best_multi_progress.append(best_multi)
            best_score_progress.append(best_score)
            if best_score < lowest_score:
                lowest_score = best_score
                if do_lcoe:
                    lowest_chrom = population[best_ndx]
                else: #(do_multi only)
                    multi_lowest_chrom = population[best_mndx]
            if self.optimise_stop > 0:
                if best_score == last_score:
                    best_ctr += 1
                    if best_ctr >= self.optimise_stop:
                        break
                else:
                    last_score = best_score
                    best_ctr = 1
            last_score = best_score
            if do_lcoe:
                if best_score == best_score_progress[-2]:
                    lud = '='
                elif best_score < best_score_progress[-2]:
                    lud = '<html>&darr;</html>'
                else:
                    lud = '<html>&uarr;</html>'
            if do_multi:
                if best_multi == last_multi_score:
                    mud = '='
                elif best_multi < last_multi_score:
                    mud = '<html>&darr;</html>'
                else:
                    mud = '<html>&uarr;</html>'
                last_multi_score = best_multi
        if self.debug:
            try:
                self.db_file.close()
                optimiseDebug(self.db_file.name)
                os.remove(self.db_file.name)
            except:
                pass
            self.debug = False
        self.opt_progressbar.setVisible(False)
        self.opt_progressbar.close()
        tim = (time.time() - start_time)
        if tim < 60:
            tim = '%.1f secs)' % tim
        else:
            tim = '%.2f mins)' % (tim / 60.)
        msg = 'Optimise completed (%0d iterations; %s' % (generation + 1, tim)
        if best_score > lowest_score:
            msg += ' Try more iterations.'
        # we'll keep two or three to save re-calculating_fitness
        op_data = [[], [], [], [], []]
        score_data = [None, None, None, None, None]
        if do_lcoe:
            op_data[0], score_data[0] = calculate_fitness([lowest_chrom])
        if do_multi:
            op_data[1], score_data[1] = calculate_fitness([multi_lowest_chrom])
        self.setStatus(msg)
        QtWidgets.QApplication.processEvents()
        self.progressbar.setHidden(True)
        self.progressbar.setValue(0)
        # GA has completed required generation
        if do_lcoe:
            self.setStatus('Final LCOE: $%.2f' % best_score)
            fig = 'optimise_lcoe'
            titl = 'Optimise LCOE using Genetic Algorithm'
            ylbl = 'Best LCOE ($/MWh)'
        else:
            fig = 'optimise_multi'
            titl = 'Optimise Multi using Genetic Algorithm'
            ylbl = 'Best Weight'
        if do_multi:
            self.setStatus('Final Weight: %.4f' % multi_best_weight)
        # Plot progress
        x = list(range(1, len(best_score_progress)+ 1))
        matplotlib.rcParams['savefig.directory'] = self.scenarios
        plt.figure(fig + datetime.now().strftime('_%Y-%M-%d_%H%M'))
        lx = plt.subplot(111)
        plt.title(titl)
        lx.plot(x, best_score_progress)
        lx.set_xlabel('Optimise Cycle (' + str(len(best_score_progress)) + ' iterations)')
        lx.set_ylabel(ylbl)
        zp = ZoomPanX()
        f = zp.zoom_pan(lx, base_scale=1.2, annotate=True)
        plt.show()
        pick = None
        pickf = None
        if do_multi:
            if self.optimise_multiplot:
                pick = self._plot_multi(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                if self.more_details:
                    pickf = self._plot_multi(multi_scores, multi_values, multi_order, 'final iteration')
            if self.optimise_multitable:
                pick2 = show_multitable(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                try:
                    pick = pick + pick2
                except:
                    pick = pick2
                if self.more_details:
                    pick2 = show_multitable(multi_scores, multi_values, multi_order, 'final iteration')
                    try:
                        pickf = pickf + pick2
                    except:
                        pickf = pick2
        op_pts = [0] * len(headers)
        for p in [st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
            op_pts[p] = 2
        op_pts[st_cap] = 3
        if self.more_details:
            if do_lcoe:
                list(map(list, list(zip(*op_data[0]))))
                dialog = self.displaytable.Table(op_data[0], title=sender_text, fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
            if do_multi:
                list(map(list, list(zip(*op_data[1]))))
                dialog = self.displaytable.Table(op_data[1], title='Multi_' + sender_text, fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
        # now I'll display the resulting capacities for LCOE, lowest weight, picked
        # now get random amount of generation per technology (both RE and non-RE)
        its = {}
        for fac, value in opt_order.items():
            its[fac] = []
        chrom_hdrs = []
        chroms = []
        ndxes = []
        if do_lcoe:
            chrom_hdrs = ['Lowest LCOE']
            chroms = [lowest_chrom]
            ndxes = [0]
        if do_multi:
            chrom_hdrs.append('Lowest Weight')
            chroms.append(multi_lowest_chrom)
            ndxes.append(1)
        if pickf is not None:
            for p in range(len(pickf)):
                if pick is None:
                    pick = [pickf[f]]
                else:
                    pick.append(pickf[p][:])
                pick[-1][0] = len(multi_best_popn)
                multi_best_popn.append(population[pickf[p][0]])
        if pick is not None:
            # at present I'll calculate the best weight for the chosen picks. Could actually present all for user choice
            if len(pick) <= 3:
                multi_lowest_chrom = multi_best_popn[pick[0][0]]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = self.displaytable.Table(op_data[2], title='Pick_' + sender_text, fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
                if len(pick) >= 2:
                    multi_lowest_chrom = multi_best_popn[pick[1][0]]
                    op_data[3], score_data[3] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[3]))))
                        dialog = self.displaytable.Table(op_data[3], title='Pick_' + sender_text, fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 2nd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(3)
                if len(pick) == 3:
                    multi_lowest_chrom = multi_best_popn[pick[2][0]]
                    op_data[4], score_data[4] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[4]))))
                        dialog = self.displaytable.Table(op_data[4], title='Pick_' + sender_text, fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 3rd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(4)
            else:
                picks = []
                for pck in pick:
                    picks.append(multi_best_popn[pck[0]])
                a, b, c = calculate_fitness(picks)
                best_multi = np.min(b)
                best_mndx = b.index(best_multi)
                multi_lowest_chrom = picks[best_mndx]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = self.displaytable.Table(op_data[2], title='Pick_' + sender_text, fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
        for chromosome in chroms:
            for fac, value in opt_order.items():
                capacity = opt_order[fac][2]
                for c in range(value[0], value[1]):
                    if chromosome[c]:
                        capacity = capacity + capacities[c]
                its[fac].append(capacity / pmss_details[fac].capacity)
                
        chooseDialog = QtWidgets.QDialog()
        hbox = QtWidgets.QHBoxLayout()
        grid = [QtWidgets.QGridLayout()]
        label = QtWidgets.QLabel('<b>Facility</b>')
        label.setAlignment(QtCore.Qt.AlignCenter)
        grid[0].addWidget(label, 0, 0)
        for h in range(len(chrom_hdrs)):
            grid.append(QtWidgets.QGridLayout())
            label = QtWidgets.QLabel('<b>' + chrom_hdrs[h] + '</b>')
            label.setAlignment(QtCore.Qt.AlignCenter)
            grid[-1].addWidget(label, 0, 0, 1, 3)
        rw = 1
        for key, value in its.items():
            grid[0].addWidget(QtWidgets.QLabel(key), rw, 0)
            if pmss_details[key].fac_type == 'S':
                typ = ' MWh'
            else:
                typ = ' MW'
            if self.show_multipliers:
                for h in range(len(chrom_hdrs)):
                    label = QtWidgets.QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0)
                    label = QtWidgets.QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 1)
                    label = QtWidgets.QLabel('({:.2f})'.format(value[h]))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 2)
            else:
                for h in range(len(chrom_hdrs)):
                    label = QtWidgets.QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0, 1, 2)
                    label = QtWidgets.QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 2)
            rw += 1

        max_amt = [0., 0.]
        if do_lcoe:
            max_amt[0] = score_data[0][0]['cost']
            max_amt[1] = score_data[0][0]['co2']
        if do_multi:
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
        pwr_chr = ['', '']
        divisor = [1., 1.]
        pwr_chrs = ' KMBTPEZY'
        for m in range(2):
            for pwr in range(len(pwr_chrs) - 1, -1, -1):
                if max_amt[m] > pow(10, pwr * 3):
                    pwr_chr[m] = pwr_chrs[pwr]
                    divisor[m] = 1. * pow(10, pwr * 3)
                    break
        self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
        self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
        for key in multi_order:
            lbl = QtWidgets.QLabel('<i>' + self.targets[key][0] + '</i>')
            lbl.setAlignment(QtCore.Qt.AlignCenter)
            grid[0].addWidget(lbl, rw, 0)
            for h in range(len(chrom_hdrs)):
                if key == 'cost':
                    amt = score_data[ndxes[h]][0][key] / divisor[0] # cost
                elif key == 'co2':
                    amt = score_data[ndxes[h]][0][key] / divisor[1] # co2
                elif key[-4:] == '_pct': # percentage
                    amt = score_data[ndxes[h]][0][key] * 100.
                else:
                    amt = score_data[ndxes[h]][0][key]
                txt = '<i>' + self.targets[key][5] + '</i>'
                try:
                    label = QtWidgets.QLabel(txt % amt)
                except:
                    label = QtWidgets.QLabel('?')
                    print('PME5:', key, txt, amt)
                label.setAlignment(QtCore.Qt.AlignCenter)
                grid[h + 1].addWidget(label, rw, 0, 1, 3)
            rw += 1
        cshow = QtWidgets.QPushButton('Quit', self)
        grid[0].addWidget(cshow)
        cshow.clicked.connect(chooseDialog.close)
        for h in range(len(chrom_hdrs)):
            button = QtWidgets.QPushButton(chrom_hdrs[h], self)
            grid[h + 1].addWidget(button, rw, 0, 1, 3)
            button.clicked.connect(chooseClicked) #(chrom_hdrs[h]))
        for gri in grid:
            frame = QtWidgets.QFrame()
            frame.setFrameStyle(QtWidgets.QFrame.Box)
            frame.setLineWidth(1)
            frame.setLayout(gri)
            hbox.addWidget(frame)
   #     grid.addWidget(show, rw, 1)
    #    show.clicked.connect(optDialog.close)
        chooseDialog.setLayout(hbox)
        chooseDialog.setWindowTitle('Choose Optimal Generator Mix')
        chooseDialog.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        
     #  this is a big of a kluge but I couldn't get it to behave
        self.opt_choice = ''
        chooseDialog.exec_()
        del chooseDialog
        try:
            h = chrom_hdrs.index(self.opt_choice)
        except:
            return
        op_data[h], score_data[h] = calculate_fitness([chroms[h]]) # make it current
        msg = chrom_hdrs[h] + ': '
        for key in multi_order[:3]:
            msg += self.targets[key][0] + ': '
            if key == 'cost':
                amt = score_data[h][0][key] / divisor[0] # cost
            elif key == 'co2':
                amt = score_data[h][0][key] / divisor[1] # co2
            elif key[-4:] == '_pct': # percentage
                amt = score_data[h][0][key] * 100.
            else:
                amt = score_data[h][0][key]
            txt = self.targets[key][5]
            txt = txt % amt
            msg += txt + '; '
        self.setStatus(msg)
        list(map(list, list(zip(*op_data[h]))))
        op_data[h].append(' ')
        op_data[h].append('Optimisation Parameters')
        op_op_prm = len(op_data[h])
        op_data[h].append(['Population size', str(self.optimise_population)])
        op_data[h].append(['No. of iterations', str(self.optimise_generations)])
        op_data[h].append(['Mutation probability', '%0.4f' % self.optimise_mutation])
        op_data[h].append(['Exit if stable', str(self.optimise_stop)])
        op_data[h].append(['Optimisation choice', self.optimise_choice])
        op_data[h].append(['Variable', 'Weight', 'Better', 'Worse'])
        for i in range(len(target_keys)):
            op_data[h].append([])
            for j in range(4):
                if j == 0:
                    op_data[h][-1].append(self.targets[target_keys[i]][j])
                else:
                    op_data[h][-1].append('{:.2f}'.format(self.targets[target_keys[i]][j]))
        op_max_row = len(op_data[h])
        for key in self.optimisation.keys():
            op_data[h].append(['Max. ' + key, self.optimisation[key].capacity_max])
        
        dialog = Table(
            objects=op_data[h],  # The data to display
            fields=headers,      # The fields (columns) to display
            parent=None          # Optional parent for the dialog
        )
        dialog.setWindowTitle('Chosen_' + sender_text)  # Set the title of the dialog
        dialog.exec_()  # Execute the dialog
        del dialog  # Clean up the dialog

        if self.optimise_to_batch:
            msgbox = QtWidgets.QMessageBox()
            msgbox.setWindowTitle('SIREN - Add to Batch')
            msgbox.setText("Press 'Yes' to add to Batch file")
            msgbox.setIcon(QtWidgets.QMessageBox.Question)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            reply = msgbox.exec_()
            if reply == QtWidgets.QMessageBox.Yes:
                check_list = []
                tot_row = 0
                save_opt_rows = False
                for o_r in range(len(op_data[h])):
                    if op_data[h][o_r][0] == 'Total':
                        break
                    if op_data[h][o_r][0] != 'RE Contribution To Load':
                        check_list.append(o_r)
                if self.files[B].text() != '':
                    newfile = self.get_filename(self.files[B].text())
                else:
                    curfile = self.scenarios[:-1]
                    newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Create and save ' + self.file_labels[B] + ' file',
                              curfile, 'Excel Files (*.xlsx)')[0]
                    if newfile == '':
                        return
                if os.path.exists(newfile):
                    wb = oxl.load_workbook(newfile)
                elif self.batch_template == '':
                    return
                else:
                    wb = oxl.load_workbook(self.batch_template)   #copy batch
                    if newfile[: len(self.scenarios)] == self.scenarios:
                        self.files[B].setText(newfile[len(self.scenarios):])
                    else:
                        if newfile.rfind('/') > 0:
                            that_len = len(commonprefix([self.scenarios, newfile]))
                            if that_len > 0:
                                bits = self.scenarios[that_len:].split('/')
                                pfx = ('..' + '/') * (len(bits) - 1)
                                newfile = pfx + newfile[that_len + 1:]
                        if newfile[-5:] != '.xlsx':
                            newfile += '.xlsx'
                        self.files[B].setText(newfile)
                if wb.worksheets[0].max_column > 1024:
                    self.clean_batch_sheet()
                    ds = oxl.load_workbook(self.get_filename(self.files[B].text()))
                batch_input_sheet = wb.worksheets[0]
                batch_input_sheet.protection.sheet = False
                normal = oxl.styles.Font(name='Arial')
                bold = oxl.styles.Font(name='Arial', bold=True)
                col = batch_input_sheet.max_column + 1
                fst_row = -1
                if col == 4: # possibly only chart stuff in columns 2 and 3
                    get_out = False
                    for col in range(3, 1, -1):
                        for row in range(1, batch_input_sheet.max_row + 1):
                            if batch_input_sheet.cell(row=row, column=col).value is not None:
                                col += 1
                                get_out = True
                                break
                            if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                                break
                        if get_out:
                            break
                for row in range(1, batch_input_sheet.max_row + 1):
                    if batch_input_sheet.cell(row=row, column=1).value is None:
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        new_cell.value = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'MM-dd hh:mm')
                        msg += " Added to batch as '" + new_cell.value + "' (column " + ssCol(col) + ')'
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value == 'Capacity (MW)':
                        fst_row = row + 1
                        cell = batch_input_sheet.cell(row=row, column=col - 1)
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        new_cell.value = 'MW'
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                        continue
                    if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                        save_opt_rows = True
                        break
                    for o_r in range(len(op_data[h])):
                        if op_data[h][o_r][0] == batch_input_sheet.cell(row=row, column=1).value:
                            if op_data[h][o_r][0] == 'Total' and col > 2:
                                cell = batch_input_sheet.cell(row=row, column=2)
                            else:
                                cell = batch_input_sheet.cell(row=row, column=col - 1)
                            new_cell = batch_input_sheet.cell(row=row, column=col)
                            try:
                                new_cell.value = float(op_data[h][o_r][1])
                            except:
                                try:
                                    new_cell.value = op_data[h][o_r][1]
                                except:
                                    pass
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                                if col == 2:
                                    new_cell.font = normal
                                    new_cell.number_format = '#0.00'
                                else:
                                    new_cell.number_format = copy(cell.number_format)
                            elif col == 2:
                                new_cell.font = normal
                                new_cell.number_format = '#0.00'
                            try:
                                i = check_list.index(o_r)
                                del check_list[i]
                            except:
                                pass
                    if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                        tot_row = row
                if save_opt_rows: # want optimisation?
                    for o_r in range(op_op_prm, len(op_data[h])):
                        row += 1
                        new_cell = batch_input_sheet.cell(row=row, column=1)
                        new_cell.value = op_data[h][o_r][0]
                        new_cell = batch_input_sheet.cell(row=row, column=col)
                        try:
                            new_cell.value = float(op_data[h][o_r][1])
                        except:
                            new_cell.value = op_data[h][o_r][1]
                if len(check_list) > 0:
                    check_list.reverse()
                    if col > 2:
                        cell = batch_input_sheet.cell(row=fst_row, column=2)
                    else:
                        cell = batch_input_sheet.cell(row=fst_row, column=col)
                    for o_r in check_list:
                        batch_input_sheet.insert_rows(tot_row)
                        new_cell = batch_input_sheet.cell(row=tot_row, column=1)
                        new_cell.value = op_data[h][o_r][0]
                        new_cell = batch_input_sheet.cell(row=tot_row, column=col)
                        try:
                            new_cell.value = float(op_data[h][o_r][1])
                        except:
                            new_cell.value = op_data[h][o_r][1]
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                wb.save(self.get_filename(self.files[B].text()))
                self.setStatus(msg)
        
        if self.adjust_gen:
            self.adjustto = {}
            for fac, value in sorted(pmss_details.items()):
                self.adjustto[fac] = value.capacity * value.multiplier
        return

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        centerPoint = QtWidgets.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        self.setStatus('')
        for i in range(len(self.file_labels)):
            if self.files[i].hasFocus():
                break
        if self.files[i].text() == '':
            curfile = self.scenarios[:-1]
        else:
            curfile = self.get_filename(self.files[i].text())
        if i == R:
            if self.files[i].text() == '':
                curfile = self.get_filename(self.files[D].text())
                curfile = curfile.replace('data', 'results')
                curfile = curfile.replace('Data', 'Results')
                if curfile == self.scenarios + self.files[D].text():
                    j = curfile.find(' ')
                    if j > 0:
                        jnr = ' '
                    else:
                        jnr = '_'
                    j = curfile.rfind('.')
                    curfile = curfile[:j] + jnr + 'Results' + curfile[j:]
            else:
                curfile = self.get_filename(self.files[R].text())
            newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)')[0]
        elif i == B and not self.batch_new_file:
            options = QtWidgets.QFileDialog.Options()
            # options |= QFileDialog.DontUseNativeDialog
            newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Open/Create and save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)', options=options)[0]
        else:
            newfile = QtWidgets.QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[i] + ' file',
                      curfile)[0]
        if newfile != '':
            if i < D:
                if i == C:
                    self.constraints = None
                elif i == G:
                    self.generators = None
                elif i == O:
                    self.optimisation = None
                ts = WorkBook()
                ts.open_workbook(newfile)
                ndx = 0
                self.sheets[i].clear()
                j = -1
                for sht in ts.sheet_names():
                    j += 1
                    self.sheets[i].addItem(sht)
                    if len(sht) >= len(self.file_labels[i]):
                        if sht[:len(self.file_labels[i])].lower() == self.file_labels[i].lower():
                            ndx = j
                self.sheets[i].setCurrentIndex(ndx)
                if i == G:
                    ws = ts.sheet_by_index(ndx)
                    self.getGenerators(ws)
                    self.setOrder()
                ts.close()
                del ts
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                if newfile.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newfile]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newfile = pfx + newfile[that_len + 1:]
                self.files[i].setText(newfile)
            if i == D and self.change_res:
                newfile = self.files[D].text()
                newfile = newfile.replace('data', 'results')
                newfile = newfile.replace('Data', 'Results')
                if newfile != self.files[D].text():
                    self.files[R].setText(newfile)
            self.updated = True

    def pfxChanged(self):
        self.results_prefix = self.results_pfx_fld.text()
     #   self.setStatus('Results filename will be ' + self.results_pfx_fld.text() + '_' + self.files[R].text())
        self.updated = True

    def sheetChanged(self, i):
        try:
            for i in range(3):
                if self.sheets[i].hasFocus():
                    break
            else:
                return
        except:
            return # probably file changed
        self.setStatus('')
        newfile = self.get_filename(self.files[i].text())
        ts = WorkBook()
        ts.open_workbook(newfile)
        ws = ts.sheet_by_name(self.sheets[i].currentText())
        self.setStatus('Sheet ' + self.sheets[i].currentText() + ' loaded')
        if i == C:
            self.getConstraints(ws)
        elif i == G:
            self.getGenerators(ws)
            self.setOrder()
        elif i == O:
            self.getOptimisation(ws)
        ts.close()
        del ts

    def loaddirChanged(self):
        curdir = self.load_dir.text()
        newdir = QtWidgets.QFileDialog.getExistingDirectory(self, 'Choose Load File Folder',
                 curdir, QtWidgets.QFileDialog.ShowDirsOnly)
        if newdir != '':
            try:
                self.load_files = newdir + self.load_files[self.load_files.rfind('/'):]
            except:
                self.load_files = newdir + self.load_files
            if newdir[: len(self.scenarios)] == self.scenarios:
                self.load_dir.setText(newdir[len(self.scenarios):])
            else:
                if newdir.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newdir]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newdir = pfx + newdir[that_len + 1:]
                self.load_dir.setText(newdir)
            self.load_years = self.get_load_years()
            self.loadCombo.clear()
            for choice in self.load_years:
                self.loadCombo.addItem(choice)
                if choice == self.load_year:
                    self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
            self.updated = True

    def helpClicked(self):
        dialog = displayobject.AnObject(QtWidgets.QDialog(), self.help,
                 title='Help for powermatch (' + fileVersion() + ')', section='powermatch')
        dialog.exec_()

    def drchanged(self):
        self.updated = True
        self.discount_rate = self.discount.value() / 100.

    def cpchanged(self):
        self.updated = True
        self.carbon_price = self.carbon.value()

    def changes(self):
        self.updated = True

    def bpcchanged(self):
        if self.batch_prefix_check.isChecked():
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        self.updated = True

    def openClicked(self):
        bit = self.sender().text().split()
        fnr = self.file_labels.index(bit[1])
        curfile = self.get_filename(self.files[fnr].text())
        if not os.path.exists(curfile):
            if fnr == R and self.results_pfx_fld.text() != '':
                i = curfile.rfind('/')
                curfile = curfile[:i + 1] + self.results_pfx_fld.text() + '_' + curfile[i+1:]
                print(curfile)
                if not os.path.exists(curfile):
                    self.setStatus(self.file_labels[fnr] + ' not found.')
                    return
            else:
                self.setStatus(self.file_labels[fnr] + ' not found.')
                return
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            os.startfile(curfile)
        elif sys.platform == 'darwin':
            try:
                subprocess.call('open', curfile)
            except:
                try:
                    subprocess.call('open', '-a', 'Microsoft Excel', curfile)
                except:
                    self.setStatus("Can't 'launch' '" + self.file_labels[fnr] + "' file")
                    return
        elif sys.platform == 'linux2' or sys.platform == 'linux':
            subprocess.call(('xdg-open', curfile))
        self.setStatus(self.file_labels[fnr] + ' file "launched".')

    def quitClicked(self):
        if self.updated or self.order.updated or self.ignore.updated:
            updates = {}
            lines = []
            lines.append('adjust_generators=' + str(self.adjust.isChecked()))
            lines.append('adjustments=') # clean up the old way
            if self.adjustto is not None:
                line = ''
                for key, value in self.adjustto.items():
                    line += '{}={:.1f},'.format(key, value)
                if line != '':
                    lines.append('adjusted_capacities=' + line[:-1])
            line = 'batch_prefix='
            if self.batch_prefix:
                line += 'True'
            lines.append(line)
            lines.append('carbon_price=' + str(self.carbon_price))
            lines.append('discount_rate=' + str(self.discount_rate))
            line = ''
            for itm in range(self.order.count()):
                line += self.order.item(itm).text() + ','
            lines.append('dispatch_order=' + line[:-1])
            for i in range(len(self.file_labels)):
                lines.append(self.file_labels[i].lower() + '_file=' + self.files[i].text().replace(getUser(), '$USER$'))
            for i in range(D):
                lines.append(self.file_labels[i].lower() + '_sheet=' + self.sheets[i].currentText())
            line = 'load='
            if self.load_dir.text() != self._load_folder:
                line += self.load_files.replace(getUser(), '$USER$')
            lines.append(line)
            line = 'load_year='
            if self.loadCombo.currentText() != 'n/a':
                line += self.loadCombo.currentText()
            lines.append(line)
            lines.append('optimise_choice=' + self.optimise_choice)
            lines.append('optimise_generations=' + str(self.optimise_generations))
            lines.append('optimise_mutation=' + str(self.optimise_mutation))
            lines.append('optimise_population=' + str(self.optimise_population))
            lines.append('optimise_stop=' + str(self.optimise_stop))
            for key, value in self.targets.items():
                line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
                lines.append(line)
            lines.append('results_prefix=' + self.results_prefix)
            updates['Powermatch'] = lines
            SaveIni(updates)
        self.close()

    def closeEvent(self, event):
        if self.floatstatus is not None:
            self.floatstatus.exit()
        if self.restorewindows:
            updates = {}
            lines = []
            add = int((self.frameSize().width() - self.size().width()) / 2)   # need to account for border
            lines.append('powermatch_pos=%s,%s' % (str(self.pos().x() + add), str(self.pos().y() + add)))
            lines.append('powermatch_size=%s,%s' % (str(self.width()), str(self.height())))
            updates['Windows'] = lines
            SaveIni(updates)
        event.accept()

    def tmlClicked(self):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                pid = subprocess.Popen(['pmtmldetail.exe', config_file]).pid
            elif os.path.exists('pmtmldetail.py'):
                pid = subprocess.Popen(['pmtmldetail.py', config_file], shell=True).pid
        else:
            pid = subprocess.Popen(['python3', 'pmtmldetail.py', config_file]).pid # python3
        return

    def editIniFile(self):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        before = os.stat(config_file).st_mtime
        dialr = EdtDialog(config_file, section='[Powermatch]')
        dialr.exec_()
        after = os.stat(config_file).st_mtime
        if after == before:
            return
     #   self.get_config()   # refresh config values
        config = configparser.RawConfigParser()
        config.read(config_file)
        try:
            st = config.get('Powermatch', 'save_tables')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.save_tables = True
        else:
            self.save_tables = False
        try:
            st = config.get('Powermatch', 'more_details')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.more_details = True
        else:
            self.more_details = False
        try:
            st = config.get('Powermatch', 'optimise_to_batch')
        except:
            st = 'True'
        if st.lower() in ['true', 'yes', 'on']:
            self.optimise_to_batch = True
        else:
            self.optimise_to_batch = False
        try:
            st = config.get('Powermatch', 'show_multipliers')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.show_multipliers = True
        else:
            self.show_multipliers = False
        try:
            st = config.get('Powermatch', 'batch_new_file')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_new_file = True
            msg = '(check to replace an existing Results workbook)'
        else:
            self.batch_new_file = False
            msg = '(check to replace last Results worksheet in Batch spreadsheet)'
        self.replace_last = QtWidgets.QCheckBox(msg, self)
        try:
            st = config.get('Powermatch', 'batch_prefix')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        QtWidgets.QApplication.processEvents()
        self.setStatus(config_file + ' edited. Reload may be required.')

    def editClicked(self):
        def update_dictionary(it, source):
            new_keys = list(source.keys())
            # first we delete and add keys to match updated dictionary
            if it == C:
                old_keys = list(self.constraints.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.constraints[key]
                for key in new_keys:
                    self.constraints[key] = Constraint(key, '<category>',
                                            0., 1., 0, 1., 0, 0., 1., 1., 0., 1., 0., 0, 0)
                target = self.constraints
            elif it == G:
                old_keys = list(self.generators.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.generators[key]
                for key in new_keys:
                    self.generators[key] = Facility(name=key, constraint='<constraint>')
                target = self.generators
            elif it == O:
                old_keys = list(self.optimisation.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.optimisation[key]
                for key in new_keys:
                    self.optimisation[key] = Optimisation(key, 'None', None)
                target = self.optimisation
            # now update the data
            for key in list(target.keys()):
                for prop in dir(target[key]):
                    if prop[:2] != '__' and prop[-2:] != '__':
                        try:
                            if prop == 'lifetime' and source[key][prop] == 0:
                                setattr(target[key], prop, 20)
                            else:
                                setattr(target[key], prop, source[key][prop])
                        except:
                            pass

        self.setStatus('')
        msg = ''
        ts = None
        it = self.file_labels.index(self.sender().text())
        if it == C and self.constraints is not None:
            pass
        elif it == G and self.generators is not None:
            pass
        elif it == O and self.optimisation is not None:
            pass
        else:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[it].text()))
                try:
                    sht = self.sheets[it].currentText()
                except:
                    self.setStatus(self.sheets[it].currentText() + ' not found in ' \
                                     + self.file_labels[it] + ' spreadsheet.')
                    return
                self.ws = ts.sheet_by_name(sht)
            except:
                ws = None
        if it == C: # self.constraints
            if self.constraints is None:
                try:
                    self.getConstraints()
                except:
                    return
            sp_pts = [2] * 15
            sp_pts[4] = 3 # discharge loss
            sp_pts[8] = 3 # parasitic loss
            sp_pts[11] = 3 # recharge loss
            dialog = Table(
                objects=self.constraints,  # The data to display
                decpts=sp_pts,  # The number of decimal points to display
                save_folder=self.scenarios,  # Optional folder to save the table to
                edit=True,  # Allow the user to edit the table
                abbr=False  # Use abbreviations for the column names
            )
            dialog.setWindowTitle(self.sender().text())  # Set the title of the dialog
            dialog.exec_()  # Execute the dialog
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                msg = ' table updated'
                
        elif it == G: # generators
            if self.generators is None:
                try:
                    self.getGenerators(ws)
                except:
                    return
            sp_pts = []
            for key in self.generators.keys():
                break
            for prop in dir(self.generators[key]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop == 'name':
                        sp_pts.insert(0, 0)
                    elif prop in ['capex', 'constraint', 'fixed_om', 'order']:
                        sp_pts.append(0)
                    elif prop == 'disc_rate' or prop == 'emissions':
                        sp_pts.append(3)
                    elif prop == 'area':
                        sp_pts.append(5)
                    else:
                        sp_pts.append(2)
            
            dialog = Table(
                objects=self.generators,  # The data to display
                decpts=sp_pts,  # The number of decimal points to display
                save_folder=self.scenarios,  # Optional folder to save the table to
                edit=True,  # Allow the user to edit the table
                abbr=False  # Use abbreviations for the column names
            )
            dialog.setWindowTitle(self.sender().text())  # Set the title of the dialog
            dialog.exec_()  # Execute the dialog
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                self.setOrder()
                msg = ' table updated'
            
        elif it == O: # self.optimisation
            if self.optimisation is None:
                try:
                    self.getOptimisation(ws)
                except:
                    return

            dialog = Table(
                objects=self.optimisation,  # The data to display
                decpts=sp_pts,  # The number of decimal points to display
                save_folder=self.scenarios,  # Optional folder to save the table to
                edit=True  # Allow the user to edit the table
            )
            dialog.setWindowTitle(self.sender().text())  # Set the title of the dialog
            dialog.exec_()  # Execute the dialog
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                for key in self.optimisation.keys():
                    if self.optimisation[key].approach == 'Discrete':
                        caps = self.optimisation[key].capacities.split()
                        self.optimisation[key].capacities = []
                        cap_max = 0.
                        for cap in caps:
                            try:
                                self.optimisation[key].capacities.append(float(cap))
                                cap_max += float(cap)
                            except:
                                pass
                        self.optimisation[key].capacity_min = 0
                        self.optimisation[key].capacity_max = round(cap_max, 3)
                        self.optimisation[key].capacity_step = None
                msg = ' table updated'

        if ts is not None:
            ts.close()
            del ts
        newfile = dialog.savedfile
        if newfile is not None:
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[it].setText(newfile[len(self.scenarios):])
            else:
                self.files[it].setText(newfile)
            if msg == '':
                msg = ' table exported'
            else:
                msg += ' and exported'
        if msg != '':
            self.setStatus(self.file_labels[it] + msg)

    def get_constraint_args(self):
        # Return the specific arguments needed for this implementation
        return (self.ws,)

    def fetchConstraint_data(self, ws):
        if ws is None:
            return None

        # Dictionary to map spreadsheet columns to Constraint parameter names
        column_mapping = {
            'Name': 'name',
            'Category': 'category',
            'Capacity Min': 'capacity_min',
            'Capacity Max': 'capacity_max',
            'Discharge Loss': 'discharge_loss',
            'Discharge Max': 'discharge_max',
            'Discharge Start': 'discharge_start',
            'Min Run Time': 'min_run_time',
            'Parasitic Loss': 'parasitic_loss',
            'Rampdown Max': 'rampdown_max',
            'Rampup Max': 'rampup_max',
            'Recharge Loss': 'recharge_loss',
            'Recharge Max': 'recharge_max',
            'Recharge Start': 'recharge_start',
            'Warmup Time': 'warm_time'
        }
        
        col_names = list(column_mapping.keys())
        col_no = [-1] * len(col_names)
        
         # Initialize column mapping
        if ws.cell_value(1, 0) == 'Name' and ws.cell_value(1, 1) == 'Category':
            col = 0
            while col < ws.ncols:
                if ws.cell_value(0, col) == 'Capacity':
                    for c2 in range(col, col + 2):
                       try:
                           ndx = col_names.index('Capacity ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except ValueError:
                            pass
                    col += 2
                elif ws.cell_value(0, col) == 'Discharge':
                    for c2 in range(col, col + 3):
                       try:
                           ndx = col_names.index('Discharge ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except ValueError:
                            pass
                    col += 3
                elif ws.cell_value(0, col) == 'Recharge':
                    for c2 in range(col, col + 3):
                       try:
                           ndx = col_names.index('Recharge ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except ValueError:
                            pass
                    col += 3
                else:
                    try:
                        ndx = col_names.index(ws.cell_value(1, col))
                        col_no[ndx] = col
                    except ValueError:
                        pass
                    col += 1
            strt_row = 2
        elif ws.cell_value(0, 0) == 'Name': # saved file
            for col in range(ws.ncols):
                try:
                    ndx = col_names.index(ws.cell_value(0, col))
                    col_no[ndx] = col
                except ValueError:
                    pass
            strt_row = 1
        else:
            self.setStatus('Not a ' + self.file_labels[C] + ' worksheet.')
            return None

        if col_no[0] < 0:
            self.setStatus('Not a ' + self.file_labels[C] + ' worksheet.')
            return None

        constraint_data = []
        for row in range(strt_row, ws.nrows):
            # Create a dictionary for each row
            record = {}
            for ndx, col_name in enumerate(col_names):
                if col_no[ndx] >= 0:
                    # Map the spreadsheet column name to the Constraint parameter name
                    constraint_param_name = column_mapping[col_name]
                    record[constraint_param_name] = ws.cell_value(row, col_no[ndx])
            constraint_data.append(record)

        return constraint_data

    def fetchGenerator_data(self, ws):
        if ws is None:
            self.generators = {}
            args = {'name': '<name>', 'constraint': '<constraint>'}
            self.generators['<name>'] = Facility(**args)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not a ' + self.file_labels[G] + ' worksheet.')
            return
        args = ['name', 'order', 'constraint', 'capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial',
                'capex', 'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']
        possibles = {'name': 0}
        for col in range(ws.ncols):
            try:
                arg = ws.cell_value(0, col).lower()
            except:
                continue
            if arg in args:
                possibles[arg] = col
            elif ws.cell_value(0, col)[:9] == 'Capital':
                possibles['capex'] = col
            elif ws.cell_value(0, col)[:8] == 'Discount':
                possibles['disc_rate'] = col
            elif ws.cell_value(0, col)[:8] == 'Dispatch':
                possibles['order'] = col
            elif ws.cell_value(0, col)[:9] == 'Emissions':
                possibles['emissions'] = col
            elif ws.cell_value(0, col) == 'FOM':
                possibles['fixed_om'] = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                possibles['lcoe_cf'] = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                possibles['lcoe'] = col
            elif ws.cell_value(0, col) == 'VOM':
                possibles['variable_om'] = col
        generator_data = []
        for row in range(1, ws.nrows):
            if ws.cell_value(row, 0) is None:
                continue
            in_args = {}
            for key, value in possibles.items():
                in_args[key] = ws.cell_value(row, value)
            generator_data.append(in_args)
        return generator_data

    def fetchOptimisation_data(self, ws):
        """
        Fetches optimisation data from an Excel worksheet.

        Args:
            ws: The worksheet object.

        Returns:
            A list of dictionaries containing optimisation data.
        """
        if ws is None:
            self.setStatus('Worksheet not supplied.')
            return None
        # Validate the worksheet's header
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return None
        # Define expected columns
        cols = ['Name', 'Approach', 'Values', 'Capacity Max', 'Capacity Min',
                'Capacity Step', 'Capacities']
        coln = {col: -1 for col in cols}
        
        # Map column headers to indices
        for col in range(ws.ncols):
            header = ws.cell_value(0, col)
            if header in cols:
                coln[header] = col
                
         # Check if the required "Name" column exists
        if coln['Name'] < 0:
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return None
        
        # Parse data row
        optimisation_data = []
        for row in range(1, ws.nrows):
            tech = ws.cell_value(row, coln['Name'])
            if not tech: # Skip rows with no technology name
                continue
            data = {'name': tech}
            
            if coln['Values'] >= 0: # values format
                data['approach'] = ws.cell_value(row, coln['Approach'])
                data['values'] = ws.cell_value(row, coln['Values'])
            else:
                # Fudge "Values" format based on "Approach"
                approach = ws.cell_value(row, coln['Approach'])
                data['approach'] = approach
                if approach == 'Discrete':
                    data['values'] = ws.cell_value(row, coln['Capacities'])
                else:
                    # Handle individual attributes
                    for key, col_index in coln.items():
                        if col_index > 0:
                            attr = key.lower().replace(' ', '_')
                            data[attr] = ws.cell_value(row, col_index)
                
            optimisation_data.append(data)
        return optimisation_data

    def clean_batch_sheet(self):
        msgbox = QtWidgets.QMessageBox()
        msgbox.setWindowTitle('SIREN - Powermatch Batch')
        msgbox.setText("Batch worksheet has more that 1,024 columns.\nSome may be invalid/empty. Would you like these to be removed")
        msgbox.setIcon(QtWidgets.QMessageBox.Warning)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        reply = msgbox.exec_()
        if reply != QtWidgets.QMessageBox.Yes:
            return
        batch_report_file = self.get_filename(self.files[B].text())
        if os.path.exists(batch_report_file + '~'):
            os.remove(batch_report_file + '~')
        shutil.copy2(batch_report_file, batch_report_file + '~')
        wb = oxl.load_workbook(batch_report_file)
        ws = wb.worksheets[0]
        for row in range(1, 4):
            try:
                if ws.cell(row=row, column=1).value.lower() in ['model', 'model label', 'technology']:
                    break
            except:
                pass
        else:
            return # bail out
        for col in range(ws.max_column, 1, -1):
            if ws.cell(row=row, column=col).value is None:
               ws.delete_cols(col, 1)
        wb.save(batch_report_file)

    def show_ProgressBar(self, maximum, msg, title):
        if self.opt_progressbar is None:
            self.opt_progressbar = ProgressBar(maximum=maximum, msg=msg, title=title)
            self.opt_progressbar.setWindowModality(QtCore.Qt.WindowModal)
            self.opt_progressbar.show()
            self.opt_progressbar.setVisible(False)
            self.activateWindow()
        else:
            self.opt_progressbar.barRange(0, maximum, msg=msg)

    def show_FloatStatus(self):
        if not self.log_status:
            return
        if self.floatstatus is None:
            self.floatstatus = FloatStatus(self, self.scenarios, None, program='Powermatch')
            self.floatstatus.setWindowModality(QtCore.Qt.WindowModal)
            self.floatstatus.setWindowFlags(self.floatstatus.windowFlags() |
                         QtCore.Qt.WindowSystemMenuHint |
                         QtCore.Qt.WindowMinMaxButtonsHint)
            self.floatstatus.procStart.connect(self.getStatus)
            self.floatstatus.show()
            self.activateWindow()

    def setStatus(self, text):
        if self.log.text() == text:
            return
        self.log.setText(text)
        if text == '':
            return
        if self.floatstatus and self.log_status:
            self.floatstatus.log(text)
            QtWidgets.QApplication.processEvents()

    def _makeAdjustments(self, pmss_data, load_col, ws, typ_row, icap_row, zone_row, tech_names, zone_techs, do_zone):
        if self.adjustto is None:
            self.adjustto = {}
            self.adjustto['Load'] = 0
            if do_zone:
                tns = zone_techs[:]
            else:
                tns = tech_names[:]
            for gen in tns:
                try:
                    if self.generators[gen].capacity > 0:
                        self.adjustto[gen] = self.generators[gen].capacity
                except:
                    pass
            for i in range(self.order.count()):
                gen = self.order.item(i).text()
                try:
                    if self.generators[gen].capacity > 0:
                        self.adjustto[gen] = self.generators[gen].capacity
                except:
                    pass
        generated = sum(pmss_data[load_col])
        datain = []
        datain.append(['Load', 'L', generated])
        if self.adjustto['Load'] == 0:
            self.adjustto['Load'] = generated
        for col in range(4, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                continue
            key = tech_names[i]
            if do_zone:
                cell = ws.cell(row=zone_row, column=col)
                if type(cell).__name__ == 'MergedCell':
                    pass
                else:
                    zone = ws.cell(row=zone_row, column=col).value
                if zone is None or zone == '':
                    zone_tech = valu
                else:
                    zone_tech = zone + '.' + valu
                key = zone_tech
            try:
                typ = self.constraints[tech_names[i]].category[0]
                if typ == '':
                    typ = 'R'
                datain.append([key, typ, float(ws.cell(row=icap_row, column=col).value)])
            except:
                try:
                    datain.append([key, 'R', float(ws.cell(row=icap_row, column=col).value)])
                except:
                    pass
        for i in range(self.order.count()):
            try:
                if self.generators[self.order.item(i).text()].capacity > 0:
                    gen = self.order.item(i).text()
                    try:
                        if self.generators[gen].constraint in self.constraints and \
                            self.constraints[self.generators[gen].constraint].category == 'Generator':
                            typ = 'G'
                        else:
                            typ = 'S'
                    except:
                        continue
                    datain.append([gen, typ, self.generators[gen].capacity])
            except:
                pass
        adjust = Adjustments(self, datain, self.adjustto, self.adjust_cap, self.results_prefix,
                                show_multipliers=self.show_multipliers, save_folder=self.scenarios,
                                batch_file=self.get_filename(self.files[B].text()))
        adjust.exec_()
        if adjust.getValues() is None:
            self.setStatus('Execution aborted.')
            if progress_handler:
                progress_handler.hide()
            return
        self.adjustto = adjust.getValues()
        results_prefix = adjust.getPrefix()
        if results_prefix != self.results_prefix:
            self.results_prefix = results_prefix
            self.results_pfx_fld.setText(self.results_prefix)
        self.updated = True
        do_adjust = True

    def _plot_multi(multi_scores, multi_best, multi_order, title):
        data = [[], [], []]
        max_amt = [0., 0.]
        for multi in multi_best:
            max_amt[0] = max(max_amt[0], multi['cost'])
            max_amt[1] = max(max_amt[1], multi['co2'])
        pwr_chr = ['', '']
        divisor = [1., 1.]
        pwr_chrs = ' KMBTPEZY'
        for m in range(2):
            for pwr in range(len(pwr_chrs) - 1, -1, -1):
                if max_amt[m] > pow(10, pwr * 3):
                    pwr_chr[m] = pwr_chrs[pwr]
                    divisor[m] = 1. * pow(10, pwr * 3)
                    break
        self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
        self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
        for multi in multi_best:
            for axis in range(3): # only three axes in plot
                if multi_order[axis] == 'cost':
                    data[axis].append(multi[multi_order[axis]] / divisor[0]) # cost
                elif multi_order[axis] == 'co2':
                    data[axis].append(multi[multi_order[axis]] / divisor[1]) # co2
                elif multi_order[axis][-4:] == '_pct': # percentage
                    data[axis].append(multi[multi_order[axis]] * 100.)
                else:
                    data[axis].append(multi[multi_order[axis]])
        # create colour map
        colours = multi_scores[:]
        cmax = max(colours)
        cmin = min(colours)
        if cmin == cmax:
            return
        for c in range(len(colours)):
            colours[c] = (colours[c] - cmin) / (cmax - cmin)
        scolours = sorted(colours)
        cvals  = [-1., 0, 1]
        colors = ['green' ,'orange', 'red']
        norm = plt.Normalize(min(cvals), max(cvals))
        tuples = list(zip(map(norm,cvals), colors))
        cmap = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples)
        fig = plt.figure(title + QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), '_yyyy-MM-dd_hhmm'))
        mx = plt.axes(projection='3d')
        plt.title('\n' + title.title() + '\n')
        try:
            for i in range(len(data[0])):
                mx.scatter3D(data[2][i], data[1][i], data[0][i], picker=True, color=cmap(colours[i]), cmap=cmap)
                if title[:5] == 'start':
                    mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(i+1))
                else:
                    j = scolours.index(colours[i])
                    if j < 10:
                        mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(j+1))
        except:
            return
        if self.optimise_multisurf:
            cvals_r  = [-1., 0, 1]
            colors_r = ['red' ,'orange', 'green']
            norm_r = plt.Normalize(min(cvals_r), max(cvals_r))
            tuples_r = list(zip(map(norm_r, cvals_r), colors_r))
            cmap_r = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples_r)
            # https://www.fabrizioguerrieri.com/blog/surface-graphs-with-irregular-dataset/
            triang = mtri.Triangulation(data[2], data[1])
            mx.plot_trisurf(triang, data[0], cmap=cmap_r)
        mx.xaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[2]][5]))
        mx.yaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[1]][5]))
        mx.zaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[0]][5]))
        mx.set_xlabel(self.targets[multi_order[2]][6])
        mx.set_ylabel(self.targets[multi_order[1]][6])
        mx.set_zlabel(self.targets[multi_order[0]][6])
        zp = ZoomPanX()
        f = zp.zoom_pan(mx, base_scale=1.2, annotate=True)
        plt.show()
        if zp.datapoint is not None: # user picked a point
            if zp.datapoint[0][0] < 0: # handle problem in matplotlib sometime after version 3.0.3
                best = [0, 999]
                for p in range(len(multi_best)):
                    diff = 0
                    for v in range(3):
                        key = multi_order[v]
                        valu = multi_best[p][key]
                        if key[-4:] == '_pct':
                            valu = valu * 100.
                        diff += abs((valu - zp.datapoint[0][v + 1]) / valu)
                    if diff < best[1]:
                        best = [p, diff]
                zp.datapoint = [[best[0]]]
                for v in range(3):
                    key = multi_order[v]
                    zp.datapoint[0].append(multi_best[p][key])
            if self.more_details:
                for p in zp.datapoint:
                    msg = 'iteration ' + str(p[0]) + ': '
                    mult = []
                    for i in range(3):
                        if self.targets[multi_order[i]][6].find('%') > -1:
                            mult.append(100.)
                        else:
                            mult.append(1.)
                        msg += self.targets[multi_order[i]][6].replace('%', '%%') + ': ' + \
                                self.targets[multi_order[i]][5] + '; '
                    msg = msg % (p[1] * mult[0], p[2] * mult[1], p[3] * mult[2])
                    self.setStatus(msg)
        return zp.datapoint

    def _optQuitClicked(self):
        self.optExit = True
        self.optDialog.close()

    def _display_table(self, objects, fields=None, title=None, save_folder='', sortby=None, decpts=None):
        dialog = Table(objects, title=title, fields=fields,
                 save_folder=save_folder, sortby=sortby, decpts=decpts)
        dialog.exec_()
        del dialog

    @QtCore.pyqtSlot(str)
    def getStatus(self, text):
        if text == 'goodbye':
            self.floatstatus = None

    def exit(self):
        self.updated = False
        self.order.updated = False
        self.ignore.updated = False
        if self.floatstatus is not None:
            self.floatstatus.exit()
        self.close()

class QtProgressHandler(ProgressHandler):
    """Qt-specific progress handler"""
    def __init__(self, progressbar):
        self.progressbar = progressbar
        
    def update(self, progress: ProgressInfo) -> None:
        self.progressbar.setValue(int(progress.current))
        if progress.message:
            self.progressbar.setFormat(f"%p% - {progress.message}")
            
    def show(self):
        self.progressbar.setHidden(False)
        QtWidgets.QApplication.processEvents()
    
    def hide(self):
        self.progressbar.setHidden(True)
        QtWidgets.QApplication.processEvents()
        
    def finish(self) -> None:
        self.progressbar.setValue(100)

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ex = PowerMatch()
    ex.show_FloatStatus() # status window
    ex.show()
    app.exec_()
    app.deleteLater()
    sys.exit()
