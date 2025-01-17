# core.py
"""
Core logic for the powermatch application, refactored to separate processing logic
from input/output and user interaction.
"""
from abc import ABC, abstractmethod
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from editini import SaveIni
import glob
from math import log10
import numpy as np
import openpyxl as oxl
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
import os
from powermatch_lit import *
from senutils import getParents, getUser, ssCol, techClean, WorkBook
import sys
from typing import List, Any, Optional
import time
def get_value(ws, row, col):
    def get_range(text, alphabet=None, base=1):
        if len(text) < 1:
            return None
        if alphabet is None:
            alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if alphabet[0] == ' ':
            alphabet = alphabet[1:]
        alphabet = alphabet.upper()
        bits = ['', '']
        b = 0
        in_char = True
        for char in text:
            if char.isdigit():
                if in_char:
                    in_char = False
                    b += 1
            else:
                if alphabet.find(char.upper()) < 0:
                    continue
                if not in_char:
                    in_char = True
                    b += 1
            if b >= len(bits):
                break
            bits[b] += char.upper()
        try:
            bits[1] = int(bits[1]) - (1 - base)
        except:
            pass
        row = 0
        ndx = 1
        for c in range(len(bits[0]) -1, -1, -1):
            ndx1 = alphabet.index(bits[0][c]) + 1
            row = row + ndx1 * ndx
            ndx = ndx * len(alphabet)
        bits[0] = row - (1 - base)
        for c in bits:
            if c == '':
                return None
    try:
        while ws.cell(row=row, column=col).value[0] == '=':
            row, col = get_range(ws.cell(row=row, column=col).value)
    except:
        return ''
    return ws.cell(row=row, column=col).value

class setTransitionBase():
    def niceSize(): # works for Adjustments window (probably because less that 640*480)
        pass

    def __init__(self, parent, label, generators, sheet, year):
        pass

    def quitClicked(self):
        pass

    def showClicked(self):
        pass

    def getValues(self):
        return self._results

class Constraint:
    def __init__(self, name, category, capacity_min, capacity_max, discharge_loss, 
                 discharge_max, discharge_start, min_run_time, parasitic_loss,
                 rampdown_max, rampup_max, recharge_loss, recharge_max, recharge_start,
                 warm_time):
        """
        after name and category all variables are passed in alphabetical order
        Constraint(key, '<category>', 0., 1., 0, 1., 0, 0., 1., 1., 0., 1., 0., 0, 0)
        """
        self.name = name.strip()
        self.category = category
        self.capacity_min = float(capacity_min) if capacity_min else 0.0
        self.capacity_max = float(capacity_max) if capacity_max else 1.0
        self.discharge_loss = float(discharge_loss) if discharge_loss else 0.0
        self.discharge_max = float(discharge_max) if discharge_max else 1.0
        try:
            if isinstance(discharge_start, datetime.time):
                self.discharge_start = discharge_start.hour
            else:
                self.discharge_start = float(discharge_start)
                if self.discharge_start >= 1:
                    self.discharge_start = int(self.discharge_start)
                else:
                    self.discharge_start = int((self.discharge_start + 1/3600) * 24)
        except:
            self.discharge_start = 0
        self.min_run_time = int(min_run_time) if min_run_time else 0
        self.parasitic_loss = float(parasitic_loss) if parasitic_loss else 0.0
        self.rampdown_max = float(rampdown_max) if rampdown_max else 1.0
        self.rampup_max = float(rampup_max) if rampup_max else 1.0
        self.recharge_loss = float(recharge_loss) if recharge_loss else 0.0
        self.recharge_max = float(recharge_max) if recharge_max else 1.0
        try:
            if isinstance(recharge_start, datetime.time):
                self.recharge_start = recharge_start.hour
            else:
                self.recharge_start = float(recharge_start)
                if self.recharge_start >= 1:
                    self.recharge_start = int(self.recharge_start)
                else:
                    self.recharge_start = int((self.recharge_start + 1/3600) * 24)
        except:
            self.recharge_start = 0
        try:
            self.warm_time = float(warm_time)
            if self.warm_time >= 1:
                self.warm_time = self.warm_time / 60
                if self.warm_time > 1:
                    self.warm_time = 1
            elif self.warm_time > 0:
                if self.warm_time <= 1 / 24.:
                    self.warm_time = self.warm_time * 24
        except:
            self.warm_time = 0

class Facility:
    def __init__(self, **kwargs):
        kwargs = {**kwargs}
      #  return
        self.name = ''
        self.constraint = ''
        self.order = 0
        self.lifetime = 20
        self.area = None
        for attr in ['capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial', 'capex',
                     'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']:
            setattr(self, attr, 0.)
        for key, value in kwargs.items():
            if value != '' and value is not None:
                if key == 'lifetime' and value == 0:
                    setattr(self, key, 20)
                else:
                    setattr(self, key, value)

class PM_Facility:
    def __init__(self, name, generator, capacity, fac_type, col, multiplier):
        self.name = name
        if name.find('.') > 0:
            self.zone = name[:name.find('.')]
        else:
            self.zone = ''
        self.generator = generator
        self.capacity = capacity
        self.fac_type = fac_type
        self.col = col
        self.multiplier = multiplier
        
class Optimisation:
    def __init__(self, name, approach, values):
        """
        Initialize an optimisation object.
        """
        self.name = name.strip()
        self.approach = approach
        self._parse_values(values)

    def _parse_values(self, values):
        """Parse values depending on approach type."""
        caps = values.split()
        if self.approach == 'Discrete':
            self.capacities = [float(cap) for cap in caps]
            self.capacity_min = 0
            self.capacity_max = sum(self.capacities)
        elif self.approach == 'Range':
            self.capacity_min, self.capacity_max, self.capacity_step = map(float, caps)
        else:
            self.capacity_min = self.capacity_max = self.capacity_step = 0

# Utility functions
def compute_lcoe(cost, capacity, emissions):
    """
    Example core computation for LCOE (Levelized Cost of Energy).
    """
    if capacity == 0:
        return float('inf')
    return (cost + emissions * 10) / capacity

def generate_summary(data):
    """
    Generate a summary report based on the input data.
    """
    summary = {}
    for item in data:
        summary[item['name']] = compute_lcoe(
            item['cost'], item['capacity'], item.get('emissions', 0)
        )
    return summary

@dataclass
class ProgressInfo:
    """Data class to hold progress information"""
    current: int
    total: Optional[int] = None
    percentage: Optional[float] = None
    message: Optional[str] = None

class ProgressHandler(ABC):
    """Abstract base class for progress handlers"""
    @abstractmethod
    def update(self, progress: ProgressInfo) -> None:
        """Update progress"""
        pass
    
    @abstractmethod
    def show(self):
        pass

    @abstractmethod
    def hide(self):
        pass

    @abstractmethod
    def finish(self) -> None:
        """Called when processing is complete"""
        pass

class AdjustmentsBase():
    def setAdjValueUnits(self, key, typ, capacity):
        if key != 'Load':
            mw = capacity
            if typ == 'S':
                unit = 'MWh'
            else:
                unit = 'MW'
            dp = self._decpts
            div = 0
        else:
            dimen = log10(capacity)
            unit = 'MWh'
            if dimen > 11:
                unit = 'PWh'
                div = 9
            elif dimen > 8:
                unit = 'TWh'
                div = 6
            elif dimen > 5:
                unit = 'GWh'
                div = 3
            else:
                div = 0
            mw = capacity / pow(10, div)
            dp = None
        mwtxt = unit
        mwcty = round(mw, dp)
        return mw, mwtxt, mwcty, div

    def __init__(self, parent, data, adjustin, adjust_cap, prefix, show_multipliers=False, save_folder=None,
                 batch_file=None):
        self.ignoreEnter = False
        self._adjust_typ = {} # facility type = G, S or L
        self._adjust_cty = {} # (actual) adjust capacity
        self.show_multipliers = show_multipliers
        if self.show_multipliers:
            self._adjust_mul = {} # (actual) adjust multiplier
            self._adjust_rnd = {} # multiplier widget (rounded to 4 digits)
        self._adjust_txt = {} # string with capacity units
        self._save_folder = save_folder
        self._batch_file = None
        if batch_file is not None:
            if os.path.isfile(batch_file):
                self._batch_file = batch_file
        self._ignore = False
        self._results = None
        self._data = {}

        self._decpts = 1
        for key, typ, capacity in data:
            if key == 'Load' or capacity is None:
                continue
            dimen = log10(capacity)
            if dimen < 2.:
                if dimen < 1.:
                    self._decpts = 2
                elif self._decpts != 2:
                    self._decpts = 1
        self.pfx_fld = ''

    def getValues(self):
        return self._results

    def adjustMult(self):
        key = self.sender().objectName()
        if not self._ignore:
            self._adjust_mul[key] = self._adjust_rnd[key].value()
            self._adjust_cty[key].setValue(self._data[key][0] * self._adjust_rnd[key].value())
        mw, mwtxt, mwstr, div = self.setAdjValueUnits(key, self._adjust_typ[key], self._data[key][0])
        self._adjust_txt[key].setText(mwtxt)
     #   if not self._ignore:
      #      self._adjust_val[key].setText(mwstr)
        self._ignore = False

    def adjustCap(self):
        if self._ignore:
            return
        key = self.sender().objectName()
        if key != 'Load':
            adj = self._adjust_cty[key].value() / self._data[key][0]
         #   self._adjust_rnd[key].setValue(adj)
        else:
            dimen = log10(self._data[key][0])
            if dimen > 11:
                mul = 9
            elif dimen > 8:
                mul = 6
            elif dimen > 5:
                mul = 3
            else:
                mul = 0
            adj = (self._adjust_cty[key].value() * pow(10, mul)) / self._data[key][0]
        self._adjust_mul[key] = adj
      #  self._adjust_cty[key] = self._data[key] * adj
        self._ignore = True
        self._adjust_rnd[key].setValue(round(adj, 4))
        self._ignore = False

    def resetClicked(self, to):
        if to is None:
            to = 0.
        else:
            to = 1.
        if self.show_multipliers:
            for key in self._adjust_rnd.keys():
                self._adjust_rnd[key].setValue(to)
        else:
            if to == 0:
                for key in self._adjust_cty.keys():
                    self._adjust_cty[key].setValue(0.)
            else:
                for key in self._adjust_cty.keys():
                    self._adjust_cty[key].setValue(self._data[key][0])
        self.pfx_fld.setText('')

    def resetloadClicked(self, to):
        if isinstance(to, bool):
            to = 1.
        if self.show_multipliers:
            self._adjust_rnd['Load'].setValue(to)
        else:
            self._adjust_cty['Load'].setValue(self._data['Load'][0])

    def getIt(self, config, prefix=''):
        try:
            adjustto = config.get('Powermatch', 'adjusted_capacities')
        except:
            return
        self.resetClicked(to=None)
        bits = adjustto.split(',')
        for bit in bits:
            bi = bit.split('=')
            key = bi[0]
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                        float(bi[1]))
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    self._adjust_mul[key] = float(bi[1]) / (self._data[key][0] * pow(10, self._data[key][1]))
                    self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
            except:
                pass
        self._ignore = False
        self.pfx_fld.setText(prefix)
        
    def showClicked(self):
        self.ignoreEnter = False
        self._results = {}
        for key in list(self._adjust_cty.keys()):
            self._results[key] = self._adjust_cty[key].value() * pow(10, self._data[key][1])
        self.close()

    def getPrefix(self):
        return self.pfx_fld.text()
    
class PowerMatchBase:
    
    def __init__(self, config=None):
        self.config = config
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results', 'Batch']
        self.ifiles = [''] * len(self.file_labels)
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.sheets = self.file_labels[:]
        del self.sheets[-2:]
        self.batch_new_file = False
        self.batch_prefix = False
        self.more_details = False
        self.constraints = None
        self.generators = None
        self.optimisation = None
        self.adjustto = None # adjust capacity to this
        self.adjust_cap = 25
        self.adjust_gen = False
        self.change_res = True
        self.adjusted_lcoe = True
        self.carbon_price = 0.
        self.carbon_price_max = 200.
        self.discount_rate = 0.
        self.load_folder = ''
        self.load_year = 'n/a'
        self.optimise_choice = 'LCOE'
        self.optimise_generations = 20
        self.optimise_mutation = 0.005
        self.optimise_population = 50
        self.optimise_stop = 0
        self.optimise_debug = False
        self.optimise_default = None
        self.optimise_multiplot = True
        self.optimise_multisurf = False
        self.optimise_multitable = False
        self.optimise_to_batch = True
        self.remove_cost = True
        self.results_prefix = ''
        self.dispatchable = ['Biomass', 'Geothermal', 'Pumped Hydro', 'Solar Thermal', 'CST'] # RE dispatchable
        self.save_tables = False
        self.show_multipliers = False
        self.show_correlation = False
        self.summary_sources = True
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        self.underlying = ['Rooftop PV'] # technologies contributing to underlying (but not operational) load
        self.operational = []
        self.iorder = []
        self.targets = {}
        self.order = []
        self.loadCombo = 'n/a'
        for t in range(len(target_keys)):
            if target_keys[t] in ['re_pct', 'surplus_pct']:
                self.targets[target_keys[t]] = [target_names[t], 0., -1, 0., 0, target_fmats[t],
                                                 target_titles[t]]
            else:
                self.targets[target_keys[t]] = [target_names[t], 0., 0., -1, 0, target_fmats[t],
                                                 target_titles[t]]
        try:
            dts = config.get('Grid', 'dispatchable').split(' ')
            dispatchable = []
            for dt in dts:
                dispatchable.append(techClean(dt.replace('_', ' ').title()))
            self.dispatchable = dispatchable
        except:
            pass

    def init_settings(self, config):
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            base_year = config.get('Base', 'year')
        except:
            base_year = '2012'
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.batch_template = config.get('Files', 'pmb_template')
            for key, value in parents:
                self.batch_template = self.batch_template.replace(key, value)
            self.batch_template = self.batch_template.replace('$USER$', getUser())
            if not os.path.exists(self.batch_template):
                self.batch_template = ''
        except:
            self.batch_template = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', base_year)
            self.scenarios = self.scenarios[: self.scenarios.rfind('/') + 1]
            if self.scenarios[:3] == '../':
                ups = self.scenarios.split('../')
                me = os.getcwd().split(os.sep)
                me = me[: -(len(ups) - 1)]
                me.append(ups[-1])
                self.scenarios = '/'.join(me)
        except:
            self.scenarios = ''
        try:
            self.load_files = config.get('Files', 'load')
            for key, value in parents:
                self.load_files = self.load_files.replace(key, value)
            self.load_files = self.load_files.replace('$USER$', getUser())
        except:
            self.load_files = ''
        try:
            self._load_folder = self.load_files[:self.load_files.rfind('/')]
        except:
            self._load_folder = ''
        self.log_status = True
        try:
            rw = config.get('Windows', 'log_status')
            if rw.lower() in ['false', 'no', 'off']:
                self.log_status = False
        except:
            pass
        try:
            adjust_cap = config.get('Power', 'adjust_cap')
            try:
                self.adjust_cap = float(adjust_cap)
            except:
                try:
                    self.adjust_cap = eval(adjust_cap)
                except:
                    pass
            if self.adjust_cap < 0:
                self.adjust_cap = pow(10, 12)
        except:
            pass
        try:
            items = config.items('Powermatch')
            for key, value in items:
                if key == 'batch_new_file':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_new_file = True
                elif key == 'batch_prefix':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_prefix = True
                elif key[:4] == 'tml_':
                    continue
                elif key[-5:] == '_file':
                    ndx = self.file_labels.index(key[:-5].title())
                    self.ifiles[ndx] = value.replace('$USER$', getUser())
                elif key[-6:] == '_sheet':
                    ndx = self.file_labels.index(key[:-6].title())
                    self.isheets[ndx] = value
                elif key == 'adjust_generators':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.adjust_gen = True
                elif key == 'adjusted_capacities':
                    self.adjustto = {}
                    bits = value.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        self.adjustto[bi[0]] = float(bi[1])
                elif key == 'carbon_price':
                    try:
                        self.carbon_price = float(value)
                    except:
                        pass
                elif key == 'carbon_price_max':
                    try:
                        self.carbon_price_max = float(value)
                    except:
                        pass
                elif key == 'change_results':
                    if value.lower() in ['false', 'off', 'no']:
                        self.change_res = False
                elif key == 'adjusted_lcoe' or key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.adjusted_lcoe = False
                elif key == 'discount_rate':
                    try:
                        self.discount_rate = float(value)
                    except:
                        pass
                elif key == 'dispatch_order':
                    self.iorder = value.split(',')
                elif key == 'load':
                    try:
                        self.load_files = value
                        for ky, valu in parents:
                            self.load_files = self.load_files.replace(ky, valu)
                        self.load_files = self.load_files.replace('$USER$', getUser())
                    except:
                        pass
                elif key == 'load_year':
                    self.load_year = value
                elif key == 'log_status':
                    if value.lower() in ['false', 'no', 'off']:
                        self.log_status = False
                elif key == 'more_details':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.more_details = True
                elif key == 'optimise_debug':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_debug = True
                elif key == 'optimise_default':
                    self.optimise_default = value
                elif key == 'optimise_choice':
                    self.optimise_choice = value
                elif key == 'optimise_generations':
                    try:
                        self.optimise_generations = int(value)
                    except:
                        pass
                elif key == 'optimise_multiplot':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_multiplot = False
                    elif value.lower() in ['surf', 'tri-surf', 'trisurf']:
                        self.optimise_multisurf = True
                elif key == 'optimise_multitable':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_multitable = True
                elif key == 'optimise_mutation':
                    try:
                        self.optimise_mutation = float(value)
                    except:
                        pass
                elif key == 'optimise_population':
                    try:
                        self.optimise_population = int(value)
                    except:
                        pass
                elif key == 'optimise_stop':
                    try:
                        self.optimise_stop = int(value)
                    except:
                        pass
                elif key == 'optimise_to_batch':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_to_batch = False
                elif key[:9] == 'optimise_':
                    try:
                        bits = value.split(',')
                        t = target_keys.index(key[9:])
                        # name, weight, minimum, maximum, widget index
                        self.targets[key[9:]] = [target_names[t], float(bits[0]), float(bits[1]),
                                                float(bits[2]), 0, target_fmats[t],
                                                 target_titles[t]]
                    except:
                        pass
                elif key == 'remove_cost':
                    if value.lower() in ['false', 'off', 'no']:
                        self.remove_cost = False
                elif key == 'results_prefix':
                    self.results_prefix = value
                elif key == 'save_tables':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.save_tables = True
                elif key == 'show_correlation':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_correlation = True
                elif key == 'show_multipliers':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_multipliers = True
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
                elif key == 'summary_sources':
                    if value.lower() in ['false', 'off', 'no']:
                        self.summary_sources = False
                elif key == 'underlying':
                    self.underlying = value.split(',')
                elif key == 'operational':
                    self.operational = value.split(',')
        except:
            print('PME1: Error with', key)
            pass
        self.restorewindows = False
        try:
            rw = config.get('Windows', 'restorewindows')
            if rw.lower() in ['true', 'yes', 'on']:
                self.restorewindows = True
        except:
            pass

    def get_filename(self, filename):
        if filename.find('/') == 0: # full directory in non-Windows
            return filename
        elif (sys.platform == 'win32' or sys.platform == 'cygwin') \
          and filename[1:3] == ':/': # full directory for Windows
            return filename
        elif filename[:3] == '../': # directory upwards of scenarios
            ups = filename.split('../')
            scens = self.scenarios.split('/')
            scens = scens[: -(len(ups) - 1)]
            scens.append(ups[-1])
            return '/'.join(scens)
        else: # subdirectory of scenarios
            return self.scenarios + filename

    def clean_batch_sheet(self):
        pass

    def get_load_years(self):
        load_years = ['n/a']
        i = self.load_files.find('$YEAR$')
        if i < 0:
            return load_years
        j = len(self.load_files) - i - 6
        files = glob.glob(self.load_files[:i] + '*' + self.load_files[i + 6:])
        for fil in files:
            load_years.append(fil[i:len(fil) - j])
        return sorted(load_years, reverse=True)

    def getConstraints(self, ws):
        """
        Template method to load constraint data. Calls the `fetchConstraint_data` method, 
        which can be overridden by subclasses to fetch data from different sources.
        """
        raw_data = self.fetchConstraint_data(*self.get_constraint_args(), **self.get_constraint_kwargs())
        if raw_data:
            constraints = {}
            for record in raw_data:
                constraint = self.create_constraint(record)
                constraints[constraint.name] = constraint
        else:
            name = '<name>'
            category = '<category>'
            capacity_min = 0.
            capacity_max = 1.
            discharge_loss = 0.
            discharge_max = 1.
            discharge_start = 0
            min_run_time = 0
            parasitic_loss = 0.
            rampdown_max = 1.
            rampup_max = 1.
            recharge_loss = 0.
            recharge_max = 1.
            recharge_start = 0
            warm_time = 0
            constraints = {}
            constraints['<name>'] = Constraint(
                name, category, capacity_min,
                capacity_max, discharge_loss, discharge_max,
                discharge_start, min_run_time, parasitic_loss,
                rampdown_max, rampup_max, recharge_loss,
                recharge_max, recharge_start, warm_time
                )
        return constraints
    
    def get_constraint_args(self):
        # Override this in derived class if needed
        return ()  # Return empty tuple by default
        
    def get_constraint_kwargs(self):
        # Override this in derived class if needed
        return {}  # Return empty dict by default
    
    def fetchConstraint_data(self, *args, **kwargs):
        """
        Placeholder method to be overridden by subclasses.
        Should return raw constraint data as a list of dictionaries.
        """
        raise NotImplementedError("Subclasses must override fetchConstraint_data.")

    def getOptimisation(self, ws):
        """
        Template method to load generator data. Calls the `fetchOptimisation_data` method, 
        which can be overridden by subclasses to fetch data from different sources.
        """
        raw_data = self.fetchOptimisation_data(*self.get_constraint_args(), **self.get_constraint_kwargs())
        optimisations = {}
        if raw_data:
            for record in raw_data:
                optimisation = self.create_optimisation(record)
                optimisations[optimisation.name] = optimisation
        else:
            optimisations['<name>'] = Optimisation('<name>', 'None', None)
        return optimisations

    def get_optimisation_args(self):
        # Override this in derived class if needed
        return ()  # Return empty tuple by default
        
    def get_optimisation_kwargs(self):
        # Override this in derived class if needed
        return {}  # Return empty dict by default

    def fetchOptimisation_data(self):
        """
        Placeholder method to be overridden by subclasses.
        Should return raw constraint data as a list of dictionaries.
        """
        raise NotImplementedError("Subclasses must override fetchOptimisation_data.")
    
    def getGenerators(self, ws):
        """
        Template method to load generator data. Calls the `fetchGenerator_data` method, 
        which can be overridden by subclasses to fetch data from different sources.
        """
        raw_data = self.fetchGenerator_data(*self.get_constraint_args(), **self.get_constraint_kwargs())
        generators = {}
        for record in raw_data:
            facility = self.create_facility(record)
            generators[facility.name] = facility
        return generators
    
    def get_generator_args(self):
        # Override this in derived class if needed
        return ()  # Return empty tuple by default
        
    def get_generator_kwargs(self):
        # Override this in derived class if needed
        return {}  # Return empty dict by default
    
    def fetchGenerator_data(self):
        """
        Placeholder method to be overridden by subclasses.
        Should return raw generator data as a list of dictionaries.
        """
        raise NotImplementedError("Subclasses must override fetch_Generator_data.")

    def setOrder(self):
        self.order.clear()
        self.ignore.clear()
        self.re_capacity = {}
        if self.generators is None:
            order = ['Storage', 'Biomass', 'PHES', 'Gas', 'CCG1', 'Other', 'Coal']
            for stn in order:
                self.order.addItem(stn)
        else:
            order = []
            zero = []
            for key, value in self.generators.items():
            #    if value.capacity == 0:
            #        continue
                if key in tech_names and key not in self.dispatchable:
                    self.re_capacity[key] = value.capacity
                    continue
                try:
                    gen = key[key.find('.') + 1:]
                    if gen in tech_names and gen not in self.dispatchable:
                        self.re_capacity[key] = value.capacity
                        continue
                except:
                    pass
                try:
                    o = int(value.order)
                    if o > 0:
                        while len(order) <= o:
                            order.append([])
                        order[o - 1].append(key)
                    elif o == 0:
                        zero.append(key)
                except:
                    pass
            order.append(zero)
            for cat in order:
                for stn in cat:
                    self.order.addItem(stn)

    def create_constraint(self, record):
        """
        Converts raw data into a Constraint object.
        """
        return Constraint(**record)
    
    def create_facility(self, record):
        """
        Converts raw data into a Facility object.
        """
        return Facility(**record)
    
    def create_optimisation(self, record):
        """
        Converts raw data into a Optimisation object.
        """
        return Optimisation(**record)

    def data_sources(self, sheet, sheet_row, pm_data_file, option):
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        sheet.cell(row=sheet_row, column=1).value = 'Data sources'
        sheet.cell(row=sheet_row, column=1).font = bold
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Scenarios folder'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = self.scenarios
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Powermatch data file'
        sheet.cell(row=sheet_row, column=1).font = normal
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        sheet.cell(row=sheet_row, column=2).value = pm_data_file
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        try:
            if self.loadCombo.currentText() != 'n/a':
                sheet.cell(row=sheet_row, column=1).value = 'Load file'
                sheet.cell(row=sheet_row, column=1).font = normal
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sheet.cell(row=sheet_row, column=2).value = load_file
                sheet.cell(row=sheet_row, column=2).font = normal
                sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
                sheet_row += 1
        except:
            pass
        sheet.cell(row=sheet_row, column=1).value = 'Constraints worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = str(self.files[C].text()) \
               + '.' + str(self.sheets[C].currentText())
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Generators worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        if option == T:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text()
        else:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text() \
                   + '.' + self.sheets[G].currentText()
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        return sheet_row

    def getBatch(self, ws, option):
        global columns, rows, values
        def recurse(lvl):
            if lvl >= len(rows) - 1:
                return
            for i in range(len(values[lvl])):
                columns[lvl] = columns[lvl] + [values[lvl][i]] * cols[lvl+1]
                recurse(lvl + 1)

        def step_split(steps):
            bits = steps.split(',')
            if len(bits) == 1:
                bits = steps.split(';')
            try:
                strt = int(bits[0])
            except:
                return 0, 0, 0, -1
            try:
                stop = int(bits[1])
                step = int(bits[2])
                try:
                    frst = int(bits[3])
                except:
                    frst = -1
            except:
                return strt, strt, strt, frst
            return strt, stop, step, frst

        if ws is None:
            self.setStatus(self.file_labels[B] + ' worksheet missing.')
            return False
        istrt = 0
        year_row = -1
        for row in range(3):
            if ws.cell_value(row, 0) in ['Model', 'Model Label', 'Technology']:
                istrt = row + 1
                break
        else:
            self.setStatus('Not a ' + self.file_labels[B] + ' worksheet.')
            return False
        self.batch_models = [{}] # cater for a range of capacities
        self.batch_report = [['Capacity (MW/MWh)', 1]]
        self.batch_tech = []
        istop = ws.nrows
        inrows = False
        for row in range(istrt, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is not None and tech != '':
                if year_row < 0 and tech[:4].lower() == 'year':
                    year_row = row
                    continue
                inrows = True
                if tech[:8].lower() != 'capacity':
                    if tech.find('.') > 0:
                        tech = tech[tech.find('.') + 1:]
                    if tech != 'Total' and tech not in self.generators.keys():
                        self.setStatus('Unknown technology - ' + tech + ' - in batch file.')
                        return False
                    self.batch_tech.append(ws.cell_value(row, 0))
                else:
                    self.batch_report[0][1] = row + 1
            elif inrows:
                istop = row
                break
            if tech[:5] == 'Total':
                istop = row + 1
                break
        if len(self.batch_tech) == 0:
            self.setStatus('No input technologies found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).')
            return False
        carbon_row = -1
        discount_row = -1
        for row in range(istop, ws.nrows):
            if ws.cell_value(row, 0) is not None and ws.cell_value(row, 0) != '':
                if ws.cell_value(row, 0).lower() in ['chart', 'graph', 'plot']:
                    self.batch_report.append(['Chart', row + 1])
                    break
                if ws.cell_value(row, 0).lower() in ['carbon price', 'carbon price ($/tco2e)']:
                    carbon_row = row
                if ws.cell_value(row, 0).lower() == 'discount rate' or ws.cell_value(row, 0).lower() == 'wacc':
                    discount_row = row
                self.batch_report.append([techClean(ws.cell_value(row, 0), full=True), row + 1])
        range_rows = {}
        for col in range(1, ws.ncols):
            model = ws.cell_value(istrt - 1, col)
            if model is None:
                break
            self.batch_models[0][col] = {'name': model}
            if option == T and year_row < 0:
                self.batch_models[0][col]['year'] = str(model)
            for row in range(istrt, istop):
                if row == year_row:
                    if ws.cell_value(row, col) is not None and ws.cell_value(row, col) != '':
                        self.batch_models[0][col]['year'] = str(ws.cell_value(row, col))
                    continue
                tech = ws.cell_value(row, 0)
                try:
                    if ws.cell_value(row, col) > 0:
                        self.batch_models[0][col][tech] = ws.cell_value(row, col)
                except:
                    if ws.cell_value(row, col) is None:
                        pass
                    elif ws.cell_value(row, col).find(',') >= 0 or ws.cell_value(row, col).find(';') >= 0:
                        try:
                            range_rows[col].append(row)
                        except:
                            range_rows[col] = [row]
                        try:
                            strt, stop, step, frst = step_split(ws.cell_value(row, col))
                            self.batch_models[0][col][tech] = strt
                            if frst >= 0 and len(range_rows[col]) > 1:
                                del range_rows[col][-1]
                                range_rows[col].insert(0, row)
                        except:
                            pass
                    pass
            if carbon_row >= 0:
                if isinstance(ws.cell_value(carbon_row, col), float):
                    self.batch_models[0][col]['Carbon Price'] = ws.cell_value(carbon_row, col)
                elif isinstance(ws.cell_value(carbon_row, col), int):
                    self.batch_models[0][col]['Carbon Price'] = float(ws.cell_value(carbon_row, col))
            if discount_row >= 0:
                if isinstance(ws.cell_value(discount_row, col), float):
                    self.batch_models[0][col]['Discount Rate'] = ws.cell_value(discount_row, col)
                elif isinstance(ws.cell_value(discount_row, col), int):
                    self.batch_models[0][col]['Discount Rate'] = float(ws.cell_value(discount_row, col))
        if len(self.batch_models[0]) == 0:
            self.setStatus('No models found in ' + self.file_labels[B] + ' worksheet (try opening and re-saving the workbook).')
            return False
        if len(range_rows) == 0:
            return True
        # cater for ranges - so multiple batch_models lists
        for rcol, ranges in range_rows.items():
            rows = {}
            for rw in ranges:
                rows[rw] = ws.cell_value(rw, rcol)
            if len(ranges) > 1: # create sheet for each range else one sheet
                values = []
                cols = [1]
                for i in range(len(ranges) -1, 0, -1):
                    strt, stop, step, frst = step_split(rows[ranges[i]])
                    values.insert(0, [])
                    for stp in range(strt, stop + step, step):
                        values[0].append(stp)
                    cols.insert(0, cols[0] * len(values[0]))
                columns = [[]] * len(rows)
                recurse(0)
                my_tech = ws.cell_value(ranges[0], 0)
                tech_2 = ws.cell_value(ranges[1], 0)
              # produce new batch_models entry for first range tech
                techs = {}
                for c in range(1, len(ranges)):
                    techs[ws.cell_value(ranges[c], 0)] = c - 1
                bits = my_tech.split('.')
                strt, stop, step, frst = step_split(rows[ranges[0]])
                for sht in range(strt, stop + step, step):
                    self.batch_models.append({})
                    for c2 in range(len(columns[0])):
                        self.batch_models[-1][c2] = {}
                        for key, value in self.batch_models[0][rcol].items():
                            self.batch_models[-1][c2][key] = value
                        self.batch_models[-1][c2][my_tech] = sht
                        for key, value in techs.items():
                            self.batch_models[-1][c2][key] = columns[value][c2]
                        self.batch_models[-1][c2]['name'] = f'{bits[-1]}_{sht}_{tech_2}'
            else:
                my_tech = ws.cell_value(ranges[0], 0)
                self.batch_models.append({})
                strt, stop, step, frst = step_split(rows[ranges[0]])
                c2 = -1
                for ctr in range(strt, stop + step, step):
                    c2 += 1
                    self.batch_models[-1][c2] = {}
                    if c2 == 0:
                        self.batch_models[-1][c2]['hdr'] = ws.cell_value(ranges[0], 0) # fudge to get header name
                    for key, value in self.batch_models[0][rcol].items():
                        self.batch_models[-1][c2][key] = value
                    self.batch_models[-1][c2][my_tech] = ctr
                #    for key, value in techs.items():
                 #       self.batch_models[-1][c2][key] = columns[value][c2]
                    self.batch_models[-1][c2]['name'] = f'Model {c2 + 1}'
        return True
 
    def pmClicked(
        self, sender_text,  progress_handler: Optional[ProgressHandler] = None, 
        adjuster = None, 
        set_transition = None,
        optimiser = None):
        def get_load_data(load_file):
            try:
                tf = open(load_file, 'r')
                lines = tf.readlines()
                tf.close()
            except:
                return None
            load_data = []
            bit = lines[0].rstrip().split(',')
            if len(bit) > 0: # multiple columns
                for b in range(len(bit)):
                    if bit[b][:4].lower() == 'load':
                        if bit[b].lower().find('kwh') > 0: # kWh not MWh
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]) * 0.001)
                        else:
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]))
            else:
                for i in range(1, len(lines)):
                    load_data.append(float(lines[i].rstrip()))
            return load_data

        def get_batch_prefix(report_group):
            if report_group == 'Lifetime Emissions':
                return 'LES_'
            if report_group in ['Correlation To Load', 'Static Variables']:
                return ''
            bits = report_group.split(' ')
            for i in range(len(bits) -1, -1, -1):
                if bits[i][0] == '(' and bits[i][-1] == ')':
                    del bits[i]
            if len(bits) == 1:
                abr = bits[0][0] + bits[0][-1]
            else:
                abr = ''
                for bit in bits:
                    abr += bit[0]
            return abr.upper() + '_'

        col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.setStatus(sender_text + ' processing started')
        if sender_text == 'Detail': # detailed spreadsheet?
            option = D
        elif sender_text == 'Optimise': # do optimisation?
            option = O
            self.optExit = False #??
        elif sender_text == 'Batch': # do batch processsing
            option = B
        elif sender_text == 'Transition': # do transition processsing
            option = T
        else:
            option = S
        if option != O:
            if progress_handler:
                progress = ProgressInfo(
                    current=0,
                    total=20,
                    percentage=0
                )
                progress_handler.update(progress)
                progress_handler.show()
        err_msg = ''
        if self.constraints is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[C].text()))
                ws = ts.sheet_by_name(self.sheets[C].currentText())
                self.getConstraints(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                err_msg = 'Constraints file not found - ' + self.files[C].text()
                self.getConstraints(None)
            except Exception as e:
                err_msg = 'Error accessing Constraints file ' + str(e)
                self.getConstraints(None)
        if self.generators is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[G].text()))
                ws = ts.sheet_by_name(self.sheets[G].currentText())
                self.getGenerators(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Generators - ' + self.files[G].text()
                else:
                    err_msg = 'Generators file not found - ' + self.files[G].text()
                self.getGenerators(None)
            except Exception as e:
                if err_msg != '':
                    err_msg += ' and Generators'
                else:
                    err_msg = 'Error accessing Generators'  + str(e)
                self.getGenerators(None)
        if option == B or option == T: # has to be xlsx workbook
            try:
                ts = WorkBook()
                bwbopen_start = time.time()
                ts.open_workbook(self.get_filename(self.files[B].text()))
                ws = ts.sheet_by_index(0)
                if ws.ncols > 1024:
                    ts.close()
                    self.clean_batch_sheet()
                    ts = WorkBook()
                    ts.open_workbook(self.get_filename(self.files[B].text()))
                    ws = ts.sheet_by_index(0)
                tim = time.time() - bwbopen_start
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'{self.file_labels[B]} workbook opened ({tim})')
                ok = self.getBatch(ws, option)
                ts.close()
                del ts
                if not ok:
                    return
            except FileNotFoundError:
                err_msg = 'Batch file not found - ' + self.files[B].text()
            except Exception as e:
                err_msg = 'Error accessing Batch file ' + str(e)
        if option == O and self.optimisation is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[O].text()))
                ws = ts.sheet_by_name(self.sheets[O].currentText())
                self.getOptimisation(ws)
                ts.close()
                del ts
                if self.optimisation is None:
                    if err_msg != '':
                        err_msg += ' not an Optimisation worksheet'
                    else:
                        err_msg = 'Not an optimisation worksheet'
            except FileNotFoundError:
                if err_msg != '':
                    err_msg += ' nor Optimisation - ' + self.files[O].text()
                else:
                    err_msg = 'Optimisation file not found - ' + self.files[O].text()
            except:
                if err_msg != '':
                    err_msg += ' and Optimisation'
                else:
                    err_msg = 'Error accessing Optimisation'
            if self.optimisation is None:
                self.getOptimisation(None)
        if err_msg != '':
            self.setStatus(err_msg)
            return
        pm_data_file = self.get_filename(self.files[D].text())
        if pm_data_file[-5:] != '.xlsx': #xlsx format only
            self.setStatus('Not a Powermatch data spreadsheet (1)')
            if progress_handler:
                progress_handler.hide()
            return
        try:
            ts = oxl.load_workbook(pm_data_file)
        except FileNotFoundError:
            self.setStatus('Data file not found - ' + self.files[D].text())
            return
        except:
            self.setStatus('Error accessing Data file - ' + self.files[D].text())
            return
        ws = ts.worksheets[0]
        top_row = ws.max_row - 8760
        if top_row < 1 or (ws.cell(row=top_row, column=1).value != 'Hour' \
                           or ws.cell(row=top_row, column=2).value != 'Period'):
            self.setStatus(f'Not a Powermatch data spreadsheet (2; {top_row})')
            if progress_handler:
                progress_handler.hide()
            return
        typ_row = top_row - 1
        gen_row = typ_row
        while typ_row > 0:
            if ws.cell(row=typ_row, column=1).value[:9] == 'Generated':
                gen_row = typ_row
            if ws.cell(row=typ_row, column=3).value in tech_names:
                break
            typ_row -= 1
        else:
            self.setStatus('no suitable data')
            return
        do_zone = False
        zone_row = typ_row - 1
        try:
            if ws.cell(row=zone_row, column=1).value.lower() == 'zone':
                do_zone = True
                zone_techs = []
        except:
            pass
        icap_row = typ_row + 1
        while icap_row < top_row:
            if ws.cell(row=icap_row, column=1).value[:8] == 'Capacity':
                break
            icap_row += 1
        else:
            self.setStatus('no capacity data')
            return
        year = ws.cell(row=top_row + 1, column=2).value[:4]
        pmss_details = {} # contains name, generator, capacity, fac_type, col, multiplier
        pmss_data = []
        re_order = [] # order for re technology
        dispatch_order = [] # order for dispatchable technology
        load_columns = {}
        load_col = -1
        strt_col = 3
        try:
            if self.load_year != 'n/a':
                year = self.load_year
                strt_col = 4
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
                pmss_details['Load'] = PM_Facility('Load', 'Load', 0, 'L', len(pmss_data), 1)
                load_columns[year] = len(pmss_data)
                pmss_data.append([])
                load_file = self.load_files.replace('$YEAR$', year)
                pmss_data[-1] = get_load_data(load_file)
                re_order.append('Load')
        except:
            pass
        zone = ''
        for col in range(strt_col, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                continue
            key = tech_names[i]
            if key == 'Load':
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
            else:
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                    zone_techs.append(key)
                else: # temp
                    if len(self.re_capacity) > 0 and tech_names[i] not in self.re_capacity.keys():
                        continue
                try:
                    capacity = float(ws.cell(row=icap_row, column=col).value)
                except:
                    continue
                if capacity <= 0:
                    continue
                typ = 'R'
                if do_zone:
                    fctr = 1
                elif tech_names[i] in self.re_capacity and capacity > 0:
                    fctr = self.re_capacity[tech_names[i]] / capacity
                else:
                    fctr = 1
            pmss_details[key] = PM_Facility(key, tech_names[i], capacity, typ, len(pmss_data), fctr)
            if key == 'Load':
                load_columns[year] = len(pmss_data)
            pmss_data.append([])
            re_order.append(key)
            for row in range(top_row + 1, ws.max_row + 1):
                pmss_data[-1].append(ws.cell(row=row, column=col).value)
        pmss_details['Load'].capacity = sum(pmss_data[load_col])
        do_adjust = False
        if option == O:
            for itm in range(self.order.count()):
                gen = self.order.item(itm).text()
                try:
                    if self.generators[gen].capacity <= 0:
                        continue
                except KeyError as err:
                    self.setStatus('Key Error: No Generator entry for ' + str(err))
                    continue
                try:
                    if self.generators[gen].constraint in self.constraints and \
                      self.constraints[self.generators[gen].constraint].category == 'Generator':
                        typ = 'G'
                    else:
                        typ = 'S'
                except:
                    continue
                dispatch_order.append(gen)
                pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
            if self.adjust_gen:
                 pmss_details['Load'].multiplier = self.adjustto['Load'] / pmss_details['Load'].capacity
            self.optClicked(year, sender_text, option, pmss_details, pmss_data, re_order, dispatch_order,
                            None, None, progress_handler)
            return
        if self.adjust_gen and option != B and option != T:
            self._makeAdjustments(pmss_data,load_col, ws, typ_row, icap_row, zone_row, tech_names, zone_techs, do_zone)
        ts.close()
        if progress_handler:
            progress = ProgressInfo(
                    current=0,
                )
            progress_handler.update(progress)

        if self.files[R].text() == '':
            i = pm_data_file.rfind('/')
            if i >= 0:
                data_file = pm_data_file[i + 1:]
            else:
                data_file = pm_data_file
            data_file = data_file.replace('data', 'results')
            data_file = data_file.replace('Data', 'Results')
            if data_file == pm_data_file[i + 1:]:
                j = data_file.find(' ')
                if j > 0:
                    jnr = ' '
                else:
                    jnr = '_'
                j = data_file.rfind('.')
                data_file = data_file[:j] + jnr + 'Results' + data_file[j:]
            self.files[R].setText(data_file)
        else:
            data_file = self.get_filename(self.files[R].text())
        if self.results_prefix != '':
            j = data_file.rfind('/')
            if data_file[j + 1:j + 1 + len(self.results_prefix)] != self.results_prefix:
                data_file = data_file[: j + 1] + self.results_prefix + '_' + data_file[j + 1:]
        for itm in range(self.order.count()):
            gen = self.order.item(itm).text()
            try:
                if self.generators[gen].capacity <= 0:
                    continue
            except KeyError as err:
                self.setStatus('Key Error: No Generator entry for ' + str(err))
                continue
            except:
                continue
            if do_adjust:
                try:
                    if self.adjustto[gen] <= 0:
                        continue
                except:
                    pass
            try:
                if self.generators[gen].constraint in self.constraints and \
                  self.constraints[self.generators[gen].constraint].category == 'Generator':
                    typ = 'G'
                else:
                    typ = 'S'
            except:
                continue
            dispatch_order.append(gen)
            pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
        if option == B or option == T:
            if option == T:
                # files = setTransition(self, self.file_labels[G], self.get_filename(self.files[G].text()),
                #                       self.sheets[G].currentText(), self.loadCombo.currentText())
                # files.exec_()
                files = set_transition(
                    self.file_labels[G], 
                    self.get_filename(self.files[G].text()), 
                    self.sheets[G].currentText(), self.loadCombo
                    )
                if files.getValues() is None:
                    self.setStatus('Execution aborted.')
                    if progress_handler:
                        progress_handler.hide()
                    return
                gen_sheet = files.getValues()
                trn_year = ''
                newfile = self.get_filename(self.files[G].text())
                gen_book = WorkBook()
                gen_book.open_workbook(newfile)
                pmss_details['Load'].multiplier = 1
            elif self.adjust_gen:
                generated = sum(pmss_data[load_col])
                datain = [['Load', 'L', generated]]
                adjustto = {'Load': generated}
                adjust = adjuster(pmss_data, load_col)
                # adjust = Adjustments(self, datain, adjustto, self.adjust_cap, None,
                #                      show_multipliers=self.show_multipliers)
                # adjust.exec_()
                if adjust.getValues() is None:
                    self.setStatus('Execution aborted.')
                    if progress_handler:
                        progress_handler.hide()
                    return
                adjustto = adjust.getValues()
                pmss_details['Load'].multiplier = adjustto['Load'] / pmss_details['Load'].capacity
       #     start_time = time.time() # just for fun
            batch_details = {'Capacity (MW/MWh)': [st_cap, '#,##0.00'],
                             'To Meet Load (MWh)': [st_tml, '#,##0'],
                             'Generation (MWh)': [st_sub, '#,##0'],
                             'Capacity Factor': [st_cfa, '#,##0.0%'],
                             'Cost ($/Yr)': [st_cst, '#,##0'],
                             'LCOG ($/MWh)': [st_lcg, '#,##0.00'],
                             'LCOE ($/MWh)': [st_lco, '#,##0.00'],
                             'Emissions (tCO2e)': [st_emi, '#,##0'],
                             'Emissions Cost': [st_emc, '#,##0'],
                             'LCOE With CO2 ($/MWh)': [st_lcc, '#,##0.00'],
                             'Max MWh': [st_max, '#,##0'],
                             'Capital Cost': [st_cac, '#,##0'],
                             'Lifetime Cost': [st_lic, '#,##0'],
                             'Lifetime Emissions': [st_lie, '#,##0'],
                             'Lifetime Emissions Cost': [st_lec, '#,##0'],
                             'Area': [st_are, '#,###0.00']}
            batch_extra = {'RE': ['#,##0.00', ['RE %age', st_cap], ['Storage %age', st_cap], ['RE %age of Total Load', st_cap]],
                           'Load Analysis': ['#,##0', ['Load met', st_tml], ['Load met %age', st_cap], ['Shortfall', st_tml], ['Total Load', st_tml],
                           ['Largest Shortfall', st_cap], ['Storage losses', st_sub], ['Surplus', st_sub], ['Surplus %age', st_cap]],
                           'Carbon': ['#,##0.00', ['Carbon Price', st_cap], ['Carbon Cost', st_emc], ['LCOE incl. Carbon Cost', st_lcc],
                           ['Lifetime Emissions Cost', st_lec]],
                           'Correlation To Load': ['0.0000', ['RE Contribution', st_cap], ['RE plus Storage', st_cap],
                           ['To Meet Load', st_cap]],
                           'Static Variables': ['#,##0.00', ['Carbon Price', st_cap], ['Lifetime', st_cap],
                           ['Discount Rate', st_cap]],
                           'Optimisation Parameters': ['#,##0.00', ['Population size', 1], ['No. of iterations', 1],
                           ['Mutation probability', 1], ['Exit if stable', 1], ['Optimisation choice', 1],
                           ['Variable', 1], ['LCOE', 1], ['Load%', 1], ['Surplus%', 1], ['RE%', 1],
                           ['Cost', 1], ['CO2', 1]]}
                           # LCOE (incl. CO2)
         #   batch_extra['Optimisation Parameters'] = []
            batch_extra['LCOE ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                if tech == 'Total':
                    batch_extra['LCOE ($/MWh)'].append([tech + ' LCOE ($/MWh)'])
                else:
                    batch_extra['LCOE ($/MWh)'].append([tech])
            batch_extra['LCOE ($/MWh)'].append(['LCOE', st_lco])
            batch_extra['LCOE With CO2 ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                batch_extra['LCOE With CO2 ($/MWh)'].append([tech])
            batch_extra['LCOE With CO2 ($/MWh)'].append(['LCOE incl. Carbon Cost', st_lcc])
         #   batch_extra['To Meet Load (MWh)'] = ['#,##0.00', ['Total', st_tml]]
            wbopen_start = time.time()
            wb = oxl.load_workbook(self.get_filename(self.files[B].text()))
            tim = time.time() - wbopen_start
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.file_labels[B]} workbook re-opened for update ({tim})')
            batch_input_sheet = wb.worksheets[0]
            rpt_time = datetime.now().strftime('_%Y-%M-%d_%H%M')
            if self.batch_new_file:
                wb.close()
                i = self.files[B].text().rfind('.')
                suffix = '_report_' + rpt_time
                batch_report_file = self.get_filename(self.files[B].text()[:i] + suffix + self.files[B].text()[i:])
                if batch_report_file == '':
                    self.setStatus(sender_text + ' aborted')
                    return
                if batch_report_file[-5:] != '.xlsx':
                    batch_report_file += '.xlsx'
                if os.path.exists(batch_report_file) and not self.replace_last.isChecked():
                    wb = oxl.load_workbook(batch_report_file)
                    bs = wb.create_sheet('Results_' + rpt_time)
                else:
                    wb = oxl.Workbook()
                    bs = wb.active
                    bs.title = 'Results_' + rpt_time
            else:
                batch_report_file = self.get_filename(self.files[B].text())
                if self.replace_last.isChecked():
                    del_sht = ''
                    for sht in wb.sheetnames:
                        if sht[:8] == 'Results_' and sht > del_sht:
                            del_sht = sht
                    if del_sht != '':
                        del wb[del_sht]
                        del_sht = del_sht.replace('Results', 'Charts')
                        if del_sht in wb.sheetnames:
                            del wb[del_sht]
                bs = wb.create_sheet('Results_' + rpt_time)
            start_time = time.time() # just for fun
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            grey = oxl.styles.colors.Color(rgb='00f2f2f2')
            grey_fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor=grey)
            total_models = 0
            for sht in range(len(self.batch_models)):
                total_models = total_models + len(self.batch_models[sht])
            try:
                incr = 20 / total_models
            except:
                incr = .05
            prgv = incr
            prgv_int = 0
            model_row = False
            model_row_no = 0
            sht_nam_len = max(len(str(len(self.batch_models))), 2)
            for sht in range(len(self.batch_models)):
                sheet_start = time.time()
                if sht == 0: # normal case
                   # copy header rows to new worksheet
                   merged_cells = []
                   merge_cells = None
                   model_row = False
                   model_cols = len(self.batch_models[sht])
                   for row in range(1, self.batch_report[0][1] + 2):
                       if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology']:
                           model_row = True
                           model_row_no = row
                       else:
                           model_row = False
                       for col in range(1, model_cols + 2):
                           cell = batch_input_sheet.cell(row=row, column=col)
                           if type(cell).__name__ == 'MergedCell':
                               if merge_cells is None:
                                   merge_cells = [row, col - 1, col]
                               else:
                                   merge_cells[2] = col
                               continue
                           if model_row and col > 1:
                               new_cell = bs.cell(row=row, column=col, value=self.batch_models[sht][col - 1]['name'])
                           else:
                               new_cell = bs.cell(row=row, column=col, value=cell.value)
                           if cell.has_style:
                               new_cell.font = copy(cell.font)
                               new_cell.border = copy(cell.border)
                               new_cell.fill = copy(cell.fill)
                               new_cell.number_format = copy(cell.number_format)
                               new_cell.protection = copy(cell.protection)
                               new_cell.alignment = copy(cell.alignment)
                           if merge_cells is not None:
                               bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                               merged_cells.append(merge_cells)
                               merge_cells = None
                       if merge_cells is not None:
                           bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                           merged_cells.append(merge_cells)
                           merge_cells = None
                   try:
                       normal = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz)
                       bold = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz, bold=True)
                   except:
                       pass
                else:
                    sheet_name = f'{sht:0{sht_nam_len}}'
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
                        if 'Charts_' + sheet_name in wb.sheetnames:
                            del wb['Charts_' + sheet_name]
                    bs = wb.create_sheet(sheet_name)
                    if model_row_no > 1:
                        title = self.batch_models[sht][0]['name']
                        tech_2 = title.split('_')
                        if len(tech_2) > 1:
                            tech_2 = tech_2[-1]
                            bits_2 = tech_2.split('.')[-1]
                            title = title.replace(tech_2, bits_2)
                            cap_2 = self.batch_models[sht][0][tech_2]
                            fst_col = 2
                            bs.cell(row=1, column=2).value = f'{title}_{cap_2}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            g = 1
                            for i in range(1, len(self.batch_models[sht])):
                                if self.batch_models[sht][i][tech_2] != cap_2:
                                    bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 1)
                                    fst_col = i + 2
                                    cap_2 = self.batch_models[sht][i][tech_2]
                                    bs.cell(row=1, column=fst_col).value = f'{title}_{cap_2}'
                                    if g == 0:
                                        g = 1
                                    else:
                                        bs.cell(row=1, column=fst_col).fill = grey_fill
                                        g = 0
                                    bs.cell(row=1, column=fst_col).font = normal
                                    bs.cell(row=1, column=fst_col).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 2)
                        else:
                            try:
                                title = self.batch_models[sht][0]['hdr'].split('.')[-1]
                                del self.batch_models[sht][0]['hdr']
                            except:
                                pass
                            bs.cell(row=1, column=2).value = f'{title}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(self.batch_models[sht]) + 1)
                column = 1
                gndx = self.batch_report[0][1] # Capacity group starting row
                do_opt_parms = [False, 0, 0, 0]
                total_load_row = 0
                if self.discount_rate > 0:
                    batch_disc_row = 0
                else:
                    batch_disc_row = -1
                if self.carbon_price > 0:
                    batch_carbon_row = 0
                else:
                    batch_carbon_row = -1
                batch_lifetime = False
                batch_data_sources_row = 0
                re_tml_row = 0
                max_load_row = -1
                report_keys = []
                for g in range(len(self.batch_report)):
                    report_keys.append(self.batch_report[g][0])
                if 'Lifetime Cost' in report_keys:
                    batch_lifetime = True
                for g in range(len(self.batch_report)):
                    if self.batch_report[g][0] == 'Chart':
                        continue
                    elif self.batch_report[g][0] == 'Carbon Price':
                        batch_carbon_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0] == 'Discount Rate' or self.batch_report[g][0].lower() == 'wacc':
                        batch_disc_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0].lower() == 'data sources':
                        batch_data_sources_row = gndx
                        gndx += 6
                        try:
                            if self.loadCombo.currentText() != 'n/a':
                                gndx += 1
                        except:
                            pass
                        continue
                    if self.batch_report[g][0] not in batch_details.keys() and self.batch_report[g][0] not in batch_extra.keys():
                        continue
                    self.batch_report[g][1] = gndx
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix(self.batch_report[g][0])
                    else:
                        batch_pfx = ''
                    bs.cell(row=gndx, column=1).value = self.batch_report[g][0]
                    bs.cell(row=gndx, column=1).font = bold
                    if self.batch_report[g][0] in batch_extra.keys():
                        key = self.batch_report[g][0]
                        if self.batch_report[g][0] == 'Optimisation Parameters':
                            for row in range(1, batch_input_sheet.max_row + 1):
                                if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                                    do_opt_parms[0] = True
                                    do_opt_parms[1] = gndx
                                    do_opt_parms[2] = row
                                    break
                            for row in range(row, batch_input_sheet.max_row + 1):
                                gndx += 1
                                if batch_input_sheet.cell(row=row, column=1).value == '':
                                    break
                            do_opt_parms[3] = row
                            continue
                        for sp in range(1, len(batch_extra[key])):
                            if batch_extra[key][sp][0] == 'Total Load':
                                total_load_row = gndx + sp
                            elif batch_extra[key][sp][0] == 'Carbon Price':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' ($/tCO2e)'
                            elif batch_extra[key][sp][0] == 'Lifetime':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' (years)'
                            elif batch_extra[key][sp][0] == 'Total incl. Carbon Cost':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + 'LCOE incl. Carbon Cost'
                            else:
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0]
                            if batch_extra[key][sp][0] in ['RE %age of Total Load', 'Total incl. Carbon Cost'] or \
                              batch_extra[key][sp][0].find('LCOE') >= 0 and batch_extra[key][sp][0].find('Total LCOE') < 0:
                                bs.cell(row=gndx + sp, column=1).font = bold
                            else:
                                bs.cell(row=gndx + sp, column=1).font = normal
                        gndx += len(batch_extra[key]) + 1
                        if key == 'Carbon':
                            if not batch_lifetime:
                                gndx -= 1
                                tot_carb_row = gndx - 3
                            else:
                                tot_carb_row = gndx - 4
                        elif key == 'LCOE ($/MWh)':
                            tot_lco_row = gndx - 2
                        elif key == 'LCOE With CO2 ($/MWh)':
                            tot_lcc_row = gndx - 2
                    else:
                        if self.batch_report[g][0] not in batch_details.keys():
                            continue
                        if self.batch_prefix:
                            batch_pfx = get_batch_prefix(self.batch_report[g][0])
                        else:
                            batch_pfx = ''
                        for sp in range(len(self.batch_tech)):
                        #    if self.batch_report[g][0] == 'To Meet Load (MWh)' and sp == 0:
                         #       bs.cell(row=gndx + sp + 1, column=1).value = 'RE Contribution To Load'
                            if self.batch_report[g][0] != 'Capacity Factor' or self.batch_tech[sp] != 'Total':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp]
                            if self.batch_report[g][0] == 'Max MWh' and self.batch_tech[sp] == 'Total':
                                max_load_row = gndx + sp + 1
                                bs.cell(row=max_load_row, column=1).value = batch_pfx + 'Max Load'
                            elif self.batch_tech[sp] == 'Total' and self.batch_report[g][0] != 'Capacity Factor':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp] + ' ' + self.batch_report[g][0]
                            bs.cell(row=gndx + sp + 1, column=1).font = normal
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            batch_disc_row = gndx + sp + 2
                            bs.cell(row=batch_disc_row, column=1).value = batch_pfx + 'Discount Rate'
                            bs.cell(row=batch_disc_row, column=1).font = normal
                        if self.batch_report[g][0] == 'Capacity Factor' and self.batch_tech[-1] == 'Total':
                            gndx += len(self.batch_tech) + 1
                        else:
                            gndx += len(self.batch_tech) + 2
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            gndx += 1
                        if self.batch_report[g][0] == 'To Meet Load (MWh)':
                            re_tml_row = gndx - 1
                            bs.cell(row=re_tml_row, column=1).value = batch_pfx + 'RE Contribution To Load'
                            bs.cell(row=re_tml_row, column=1).font = normal
                            bs.cell(row=re_tml_row + 1, column=1).value = batch_pfx + 'Storage Contribution To Load'
                            bs.cell(row=re_tml_row + 1, column=1).font = normal
                            gndx += 2
                merge_col = 1
                last_name = ''
                # find first varying capacity to create model name
                model_key = ''
                model_nme = ''
                if sht > 0:
                    for key in self.batch_models[sht][0].keys():
                        if key == 'name':
                            continue
                        try:
                            if self.batch_models[sht][0][key] != self.batch_models[sht][1][key]:
                                model_key = key
                                bits = key.split('.')[-1].split(' ')
                                for bit in bits:
                                    model_nme += bit.strip('()')[0]
                                model_nme += '-'
                                break
                        except:
                            pass
                if option == T:
                    capex_table = {}
                    for fac in pmss_details.keys():
                        capex_table[fac] = {'cum': 0}
                for model, capacities in self.batch_models[sht].items():
                    if option == T:
                        if capacities['year'] != trn_year:
                            # get generators and load for new year
                            trn_year = capacities['year']
                            year = str(trn_year)
                            ws = gen_book.sheet_by_name(gen_sheet.replace('$YEAR$', year))
                            self.getGenerators(ws)
                            if year not in load_columns.keys():
                                load_columns[year] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', year)
                                pmss_data[-1] = get_load_data(load_file)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            pmss_details['Load'].capacity = sum(pmss_data[load_columns[year]])
                            pmss_details['Load'].col = load_columns[year]
                            continue
                        pmss_details[fac].multiplier = 0
                    if int(prgv) > prgv_int:
                        prgv_int = int(prgv)
                        if progress_handler:
                            progress = ProgressInfo(
                                current=int(prgv)
                            )
                            progress_handler.update(progress)
                    prgv += incr
                    column += 1
                    dispatch_order = []
                    for key, capacity in capacities.items(): # cater for zones
                        if key in ['Carbon Price', 'Discount Rate', 'Total']:
                            continue
                        if key == 'name' and model_row_no > 0:
                            if model_key != '':
                                bs.cell(row=model_row_no, column=column).value = f'{model_nme}{capacities[model_key]}'
                            elif option == T:
                                bs.cell(row=model_row_no, column=column).value = f'{capacity}'
                            else:
                                bs.cell(row=model_row_no, column=column).value = f'Model {model + 1}'
                            bs.cell(row=model_row_no, column=column).font = normal
                            bs.cell(row=model_row_no, column=column).alignment = oxl.styles.Alignment(wrap_text=True,
                                    vertical='bottom', horizontal='center')
                            continue
                        if key == 'year':
                            if capacity in load_columns.keys():
                                pmss_details['Load'].col = load_columns[capacity]
                            else:
                                load_columns[capacity] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', capacity)
                                pmss_data[-1] = get_load_data(load_file)
                                pmss_details['Load'].col = load_columns[capacity]
                            pmss_details['Load'].capacity = sum(pmss_data[pmss_details['Load'].col])
                            continue
                        if key not in re_order:
                            dispatch_order.append(key)
                        if key not in pmss_details.keys():
                            gen = key[key.find('.') + 1:]
                            if gen in re_order:
                                typ = 'R'
                            elif self.generators[gen].constraint in self.constraints and \
                              self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                            pmss_details[key] = PM_Facility(key, gen, capacity, typ, -1, 1)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            continue
                        gen = pmss_details[fac].generator
                        try:
                            pmss_details[fac].multiplier = capacities[fac] * 1.0 / pmss_details[fac].capacity
                        except:
                            pass
                        if option == T:
                            if fac not in capex_table.keys():
                                capex_table[fac] = {'cum': 0}
                            if year not in capex_table[fac].keys():
                                try:
                                    capex_table[fac][year] = [self.generators[fac].capex, 0]
                                except:
                                    capex_table[fac][year] = [self.generators[fac[fac.find('.') + 1:]].capex, 0]
                            capx = pmss_details[fac].multiplier * pmss_details[fac].capacity
                            capex_table[fac][year][1] = capx - capex_table[fac]['cum']
                            capex_table[fac]['cum'] = capx
                    if option == T:
                        for fac in capex_table.keys():
                            if capex_table[fac]['cum'] == 0:
                                continue
                            capx = 0
                            for key, detail in capex_table[fac].items():
                                if key == 'cum':
                                    continue
                                capx = capx + detail[0] * detail[1]
                            capx = capx / capex_table[fac]['cum']
                            try:
                                self.generators[fac].capex = round(capx)
                            except:
                                self.generators[fac[fac.find('.') + 1:]].capex = round(capx)
                    save_carbon_price = None
                    if 'Carbon Price' in capacities.keys():
                        save_carbon_price = self.carbon_price
                        self.carbon_price = capacities['Carbon Price']
                    if 'Discount Rate' in capacities.keys():
                        save_discount_rate = self.discount_rate
                        self.discount_rate = capacities['Discount Rate']
                    sp_data = self.doDispatch(year, sender_text, option, pmss_details, pmss_data, re_order, dispatch_order,
                              pm_data_file, data_file, progress_handler=progress_handler, title=capacities['name'])
                    if 'Carbon Price' in capacities.keys():
                        self.carbon_price = save_carbon_price
                    # first the Facility/technology table at the top of sp_data
                    for sp in range(len(self.batch_tech) + 1):
                        if sp_data[sp][st_fac] in self.batch_tech:
                            tndx = self.batch_tech.index(sp_data[sp][st_fac]) + 1
                            for group in self.batch_report:
                                if group[0] in batch_details.keys():
                                    gndx = group[1]
                                    col = batch_details[group[0]][0]
                                    if group[0] == 'Capacity Factor' and sp_data[sp][0] == 'Total':
                                        continue
                                    if group[0] == 'Capacity Factor' and isinstance(sp_data[sp][col], str):
                                        bs.cell(row=gndx + tndx, column=column).value = float(sp_data[sp][col].strip('%')) / 100.
                                    else:
                                        bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = batch_details[group[0]][1]
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                        if sp_data[sp][st_fac] == 'Total':
                            break
                    if batch_disc_row > 1:
                         bs.cell(row=batch_disc_row, column=column).value = self.discount_rate
                         bs.cell(row=batch_disc_row, column=column).number_format = '#0.00%'
                         bs.cell(row=batch_disc_row, column=column).font = normal
                    # save details from Total row
                    for group in self.batch_report:
                        if group[0] == 'LCOE ($/MWh)':
                            try:
                                col = batch_details['LCOE ($/MWh)'][0]
                                bs.cell(row=tot_lco_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lco_row, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                bs.cell(row=tot_lco_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'LCOE With CO2 ($/MWh)':
                            try:
                                col = batch_details['LCOE With CO2 ($/MWh)'][0]
                                bs.cell(row=tot_lcc_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lcc_row, column=column).number_format = batch_details['LCOE With CO2 ($/MWh)'][1]
                                bs.cell(row=tot_lcc_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'Carbon':
                            try:
                                bs.cell(row=tot_carb_row, column=column).value = sp_data[sp][st_emc]
                                bs.cell(row=tot_carb_row, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row, column=column).font = normal
                                bs.cell(row=tot_carb_row + 1, column=column).value = sp_data[sp][st_lcc]
                                bs.cell(row=tot_carb_row + 1, column=column).number_format = '#,##0.00'
                                bs.cell(row=tot_carb_row + 1, column=column).font = bold
                                bs.cell(row=tot_carb_row + 2, column=column).value = sp_data[sp][st_lec]
                                bs.cell(row=tot_carb_row + 2, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row + 2, column=column).font = normal
                            except:
                                pass
                    if 'Discount Rate' in capacities.keys():
                        self.discount_rate = save_discount_rate
                    # now the other stuff in sp_data
                    for sp in range(sp + 1, len(sp_data)):
                        if sp_data[sp][st_fac] == '':
                            continue
                        i = sp_data[sp][st_fac].find(' (')
                        if i >= 0:
                            tgt = sp_data[sp][st_fac][: i]
                        else:
                            tgt = sp_data[sp][st_fac]
                        if tgt == 'RE %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Storage %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'LCOE':
                            for group in self.batch_report:
                                if group[0] == 'LCOE ($/MWh)':
                                    try:
                                        col = batch_details['LCOE ($/MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Carbon Price':
                            for group in batch_extra['Carbon'][1:]:
                                if group[0] == 'Carbon Price':
                                    try:
                                        col = group[1]
                                        bs.cell(row=tot_carb_row - 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=tot_carb_row - 1, column=column).number_format = batch_extra['Carbon'][0]
                                        bs.cell(row=tot_carb_row - 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt[:10] == 'Total Load':
                            for group in self.batch_report:
                                if group[0] == 'Max MWh':
                                    try:
                                        col = batch_details['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=max_load_row, column=column).number_format = batch_extra['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        for key, details in batch_extra.items():
                            try:
                                x = [x for x in details if tgt in x][0]
                                for group in self.batch_report:
                                    if group[0] == key:
                                        gndx = group[1]
                                        break
                                else:
                                    continue
                                tndx = details.index(x)
                                col = x[1]
                                bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                if key == 'RE' or (key == 'Static Variables' and x[0] == 'Discount Rate'):
                                    pct = float(sp_data[sp][col].strip('%')) / 100.
                                    bs.cell(row=gndx + tndx, column=column).value = pct
                                    bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                else:
                                    bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = details[0]
                                bs.cell(row=gndx + tndx, column=column).font = normal
                                if sp_data[sp][st_fac] == 'RE %age of Total Load' or \
                                  sp_data[sp][st_fac].find('LCOE') >= 0 or \
                                  sp_data[sp][st_fac].find('incl.') >= 0:
                                    bs.cell(row=gndx + tndx, column=column).font = bold
                                else:
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                                if key == 'Load Analysis':
                                    if x[0] in ['Load met', 'Surplus']:
                                        tndx += 1
                                        col = batch_extra['Load Analysis'][tndx][1]
                                        pct = float(sp_data[sp][col].strip('%')) / 100.
                                        bs.cell(row=gndx + tndx, column=column).value = pct
                                        bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                        bs.cell(row=gndx + tndx, column=column).font = normal
                            except:
                                pass
                tim = (time.time() - sheet_start)
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                timt = (time.time() - start_time)
                if timt < 60:
                    timt = '%.1f secs' % timt
                else:
                    hhmm = timt / 60.
                    timt = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'Processed sheet {sht + 1} of {len(self.batch_models)}; ({len(self.batch_models[sht])} models; {tim}. Total {timt})')
                
                if total_load_row > 0:
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix('Load Analysis')
                    if option == T:
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load'
                    else:
                        load_mult = ''
                        try:
                            mult = round(pmss_details['Load'].multiplier, 3)
                            if mult != 1:
                                load_mult = ' x ' + str(mult)
                        except:
                            pass
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load - ' + year + load_mult
                if do_opt_parms[0]:
                    t_row = do_opt_parms[1]
                    for row in range(do_opt_parms[2], do_opt_parms[3] + 1):
                        for col in range(1, batch_input_sheet.max_column + 1):
                            cell = batch_input_sheet.cell(row=row, column=col)
                            new_cell = bs.cell(row=t_row, column=col, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                        t_row += 1
                del_rows = []
                for group in self.batch_report:
                    if group[0] in ['Generation (MWh)']:
                        # remove storage or RE
                        gndx = group[1]
                        if group[0] == 'Generation (MWh)':
                            tst = 'S'
                        else:
                            tst = 'R' # probably redundant
                        for row in range(gndx, gndx + len(self.batch_tech)):
                            try:
                                if pmss_details[bs.cell(row=row, column=1).value].fac_type == tst:
                                    del_rows.append(row)
                            except:
                                pass
                for row in sorted(del_rows, reverse=True):
                    bs.delete_rows(row, 1)
                for column_cells in bs.columns:
                    length = 0
                    for cell in column_cells:
                        if cell.row < self.batch_report[0][1] - 1:
                            continue
                        try:
                            value = str(round(cell.value, 2))
                        except:
                            value = cell.value
                        if value is None:
                            continue
                        if len(value) > length:
                            length = len(value)
                    if isinstance(cell.column, int):
                        cel = ssCol(cell.column)
                    else:
                        cel = cell.column
                    bs.column_dimensions[cel].width = max(length * 1.05, 10)
                if batch_data_sources_row > 0:
                    i = self.data_sources(bs, batch_data_sources_row - len(del_rows), pm_data_file, option)
                bs.freeze_panes = 'B' + str(self.batch_report[0][1])
                bs.activeCell = 'B' + str(self.batch_report[0][1])
                for sheet in wb:
                    wb[sheet.title].views.sheetView[0].tabSelected = False
                wb.active = bs
                # check if any charts/graphs
                if self.batch_report[-1][0] == 'Chart':
                    bold = oxl.styles.Font(name='Arial', bold=True)
                    min_col = 2
                    max_col = len(self.batch_models[sht]) + 1
                    chs = None
                    in_chart = False
                    cht_cells = ['N', 'B']
                    cht_row = -27
                    tndx_rows = max(9, len(self.batch_tech) + 4)
                    cats = None
                    chart_group = ''
                    for row in range(self.batch_report[-1][1], batch_input_sheet.max_row + 1):
                        if batch_input_sheet.cell(row=row, column=1).value is None:
                            continue
                        if batch_input_sheet.cell(row=row, column=1).value.lower() in ['chart', 'graph', 'plot']:
                            if in_chart:
                                charts[-1].width = 20
                                charts[-1].height = 12
                                for s in range(len(charts[-1].series)):
                                    ser = charts[-1].series[s]
                                    ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                              #      ser.graphicalProperties.line.solidFill = "00AAAA"
                                if charts2[-1] is not None:
                                    for s in range(len(charts2[-1].series)):
                                        ser = charts2[-1].series[s]
                                        ser.marker.symbol = 'triangle'
                               #         ser.graphicalProperties.line.solidFill = "00AAAA"
                                    charts2[-1].y_axis.crosses = 'max'
                                    charts[-1] += charts2[-1]
                                if cats is not None:
                                    charts[-1].set_categories(cats)
                                if len(charts) % 2:
                                    cht_row += 30
                                if chart_group != '':
                                    cht_col = col_letters.index(cht_cells[len(charts) % 2])
                                    chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                                    chs.cell(row=cht_row - 1, column=cht_col).font = bold
                                chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
                            in_chart = True
                            if chs is None:
                                if bs.title.find('Results') >= 0:
                                    txt = bs.title.replace('Results', 'Charts')
                                else:
                                    txt = 'Charts_' + bs.title
                                chs = wb.create_sheet(txt)
                                charts = []
                                charts2 = []
                            charts.append(LineChart())
                            charts2.append(None)
                            if batch_input_sheet.cell(row=row, column=2).value is None or len(merged_cells) == 0:
                                min_col = 2
                                max_col = len(self.batch_models[sht]) + 1
                                chart_group = ''
                            else:
                                merge_group = get_value(batch_input_sheet, row, 2)
                                for i in range(len(merged_cells) -1, -1, -1):
                                    merge_value = get_value(batch_input_sheet, merged_cells[i][0], merged_cells[i][1])
                                    if merge_value == merge_group:
                                        min_col = merged_cells[i][1]
                                        max_col = merged_cells[i][2]
                                        chart_group = merge_group
                                        break
                        elif not in_chart:
                            continue
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'title':
                            charts[-1].title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'x-title':
                            charts[-1].x_axis.title = get_value(batch_input_sheet, row, 2)
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title':
                            charts[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title2':
                            if charts2[-1] is None:
                                charts2[-1] = LineChart()
                                charts2[-1].x_axis.title = None
                            charts2[-1].y_axis.axId = 200
                            charts2[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() in ['categories', 'y-labels', 'data', 'data2']:
                            dgrp = get_value(batch_input_sheet, row, 2)
                            if batch_input_sheet.cell(row=row, column=1).value.lower() == 'categories' \
                              and dgrp.lower() in ['model', 'model label', 'technology']: # models as categories
                                rw = self.batch_report[0][1] - 1
                                cats = Reference(bs, min_col=min_col, min_row=rw, max_col=max_col, max_row=rw)
                                continue
                            if dgrp.lower() in ['capacity (mw)', 'capacity (mw/mwh)']:
                                gndx = self.batch_report[0][1]
                            else:
                                for group in self.batch_report:
                                    if group[0].lower() == dgrp.lower():
                                        gndx = group[1]
                                        break
                                else:
                                     continue
                                # backup a bit in case rows deleted
                                for r in range(len(del_rows)):
                                    try:
                                        if bs.cell(row=gndx, column=1).value.lower() == group[0].lower():
                                            break
                                    except:
                                        pass
                                    gndx -= 1
                            ditm = get_value(batch_input_sheet, row, 3)
                            for tndx in range(tndx_rows):
                                if bs.cell(row=gndx + tndx, column=1).value is None:
                                    break
                                if bs.cell(row=gndx + tndx, column=1).value.lower() == ditm.lower():
                                    if batch_input_sheet.cell(row=row, column=1).value.lower() == 'data':
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts[-1].append(series)
                                    elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'data2':
                                        if charts2[-1] is None:
                                            charts2[-1] = LineChart()
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        charts2[-1].append(series)
                                    else:
                                        cats = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                    break
                    if in_chart:
                        charts[-1].width = 20
                        charts[-1].height = 12
                        for s in range(len(charts[-1].series)):
                            ser = charts[-1].series[s]
                            ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                        if charts2[-1] is not None:
                            for s in range(len(charts2[-1].series)):
                                ser = charts2[-1].series[s]
                                ser.marker.symbol = 'triangle'
                            charts2[-1].y_axis.crosses = 'max'
                            charts[-1] += charts2[-1]
                        if cats is not None:
                            charts[-1].set_categories(cats)
                        if len(charts) % 2:
                            cht_row += 30
                        if chart_group != '':
                            cht_col = col_letters.index(cht_cells[len(charts) % 2])
                            chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                            chs.cell(row=cht_row - 1, column=cht_col).font = bold
                        chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
            if len(self.batch_models) > 1 and len(self.batch_models[0]) == 1:
                try:
                    del wb['Results_' + rpt_time]
                    del wb['Charts_' + rpt_time]
                except:
                    pass
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'Saving {sender_text} report ({total_models:,} models; {tim})')
           #     self.setStatus('Saving %s report' % (sender_text))
            if progress_handler:
                progress = ProgressInfo(
                        current=20,
                )
                progress_handler.update(progress)
            wb.save(batch_report_file)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{sender_text} completed ({len(self.batch_models)} sheets, {total_models:,} models; {tim}). You may need to open and save the workbook to reprocess it.')
            return
        if do_adjust:
            if self.adjustto is not None:
                for fac, value in self.adjustto.items():
                    try:
                        pmss_details[fac].multiplier = value / pmss_details[fac].capacity
                    except:
                        pass
        self.doDispatch(year, sender_text, option, pmss_details, pmss_data, re_order, dispatch_order,
                        pm_data_file, data_file, progress_handler=progress_handler)

    def doDispatch(self, year, sender_text, option, pmss_details, pmss_data, re_order, dispatch_order,
                   pm_data_file, data_file, title=None,  progress_handler=None):
        def calcLCOE(annual_output, capital_cost, annual_operating_cost, discount_rate, lifetime):
            # Compute levelised cost of electricity
            if discount_rate > 0:
                annual_cost_capital = capital_cost * discount_rate * pow(1 + discount_rate, lifetime) / \
                                      (pow(1 + discount_rate, lifetime) - 1)
            else:
                annual_cost_capital = capital_cost / lifetime
            total_annual_cost = annual_cost_capital + annual_operating_cost
            try:
                return total_annual_cost / annual_output
            except:
                return total_annual_cost

        def format_period(per):
            hr = per % 24
            day = int((per - hr) / 24)
            mth = 0
            while day > the_days[mth] - 1:
                day -= the_days[mth]
                mth += 1
            return '{}-{:02d}-{:02d} {:02d}:00'.format(year, mth+1, day+1, hr)

        def summary_totals(title=''):
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = title + 'Total'
            sp_d[st_cap] = cap_sum
            sp_d[st_tml] = tml_sum
            sp_d[st_sub] = gen_sum
            sp_d[st_cst] = cost_sum
            sp_d[st_lcg] = gs
            sp_d[st_lco] = gsw
            sp_d[st_emi] = co2_sum
            sp_d[st_emc] = co2_cost_sum
            sp_d[st_lcc] = gswc
            sp_d[st_cac] = capex_sum
            sp_d[st_lic] = lifetime_sum
            sp_d[st_lie] = lifetime_co2_sum
            sp_d[st_lec] = lifetime_co2_cost
            sp_d[st_are] = total_area
            sp_data.append(sp_d)
            if (self.carbon_price > 0 or option == B or option == T):
                sp_d = [' '] * len(headers)
                cc = co2_sum * self.carbon_price
                cl = cc * max_lifetime
                if self.adjusted_lcoe and tml_sum > 0:
                    cs = (cost_sum + cc) / tml_sum
                else:
                    if gen_sum > 0:
                        cs = (cost_sum + cc) / gen_sum
                    else:
                        cs = ''
                sp_d[st_fac] = title + 'Total incl. Carbon Cost'
                sp_d[st_cst] = cost_sum + cc
                sp_d[st_lic] = lifetime_sum + cl
                sp_data.append(sp_d)
            if tml_sum > 0:
                sp_d = [' '] * len(headers)
             #   sp_d[st_fac] = 'RE Direct Contribution to ' + title + 'Load'
                sp_d[st_fac] = 'RE %age'
                re_pct = (tml_sum - sto_sum - ff_sum) / tml_sum
                sp_d[st_cap] = '{:.1f}%'.format(re_pct * 100.)
                sp_d[st_tml] = tml_sum - ff_sum - sto_sum
                sp_data.append(sp_d)
                if sto_sum > 0:
                    sp_d = [' '] * len(headers)
                 #   sp_d[st_fac] = 'RE Contribution to ' + title + 'Load via Storage'
                    sp_d[st_fac] = 'Storage %age'
                    sp_d[st_cap] = '{:.1f}%'.format(sto_sum * 100. / tml_sum)
                    sp_d[st_tml] = sto_sum
                    sp_data.append(sp_d)
            sp_data.append([' '])
            sp_data.append([title + 'Load Analysis'])
            if sp_load != 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Load met'
                load_pct = (sp_load - sf_sums[0]) / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(load_pct * 100)
                sp_d[st_tml] = sp_load - sf_sums[0]
                sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Shortfall'
                sp_d[st_cap] = '{:.1f}%'.format(sf_sums[0] * 100 / sp_load)
                sp_d[st_tml] = sf_sums[0]
                sp_data.append(sp_d)
                if option == B or option == T:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = title + 'Total Load'
                    sp_d[st_tml] = sp_load
                    if title == '':
                        sp_d[st_max] = load_max
                    sp_data.append(sp_d)
                else:
                    load_mult = ''
                    try:
                        mult = round(pmss_details['Load'].multiplier, 3)
                        if mult != 1:
                            load_mult = ' x ' + str(mult)
                    except:
                        pass
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Total ' + title + 'Load - ' + year + load_mult
                    sp_d[st_tml] = sp_load
                    if title == '' or option == S:
                        sp_d[st_max] = load_max
                        sp_d[st_bal] = ' (' + format_period(load_hr)[5:] + ')'
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'RE %age of Total ' + title + 'Load'
                sp_d[st_cap] = '{:.1f}%'.format((sp_load - sf_sums[0] - ff_sum) * 100. / sp_load)
                sp_data.append(sp_d)
                sp_data.append([' '])
                if tot_sto_loss != 0:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Storage losses'
                    sp_d[st_sub] = tot_sto_loss
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Surplus'
                surp_pct = -sf_sums[1] / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(surp_pct * 100)
                sp_d[st_sub] = -sf_sums[1]
                sp_data.append(sp_d)
            else:
                load_pct = 0
                surp_pct = 0
                re_pct = 0
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Largest Shortfall'
                sp_d[st_sub] = round(max_short[1], 2)
                sp_d[st_cfa] = ' (' + format_period(max_short[0])[5:] + ')'
                sp_data.append(sp_d)
            if option == O or option == O1:
                try:
                    re_pct
                except NameError:
                    print('No RE %age for: ',sp_d)
                    re_pct = 0
                return load_pct, surp_pct, re_pct

        def do_detail(fac, col, ss_row):
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            col += 1
            sp_cols.append(fac)
            sp_cap.append(pmss_details[fac].capacity * pmss_details[fac].multiplier)
            if do_zone and pmss_details[fac].zone != '':
                ns.cell(row=zone_row, column=col).value = pmss_details[fac].zone
                ns.cell(row=zone_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            try:
                ns.cell(row=what_row, column=col).value = fac[fac.find('.') + 1:]
            except:
                ns.cell(row=what_row, column=col).value = fac # gen
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=cap_row, column=col).value = sp_cap[-1]
            ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
            # capacity
            ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) \
                    + str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=col).number_format = '#,##0'
            # To meet load MWh
            ns.cell(row=tml_row, column=col).value = fac_tml[fac]
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            # subtotal MWh
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) +'/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return col
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + \
                        str(cost_row) + '/' + ssCol(col) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe > 0:
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ssCol(col) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            di = pmss_details[fac].col
            if pmss_details[fac].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows] * \
                                                         pmss_details[fac].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            return col

        def do_detail_summary(fac, col, ss_row, dd_tml_sum, dd_re_sum):
            if do_zone and pmss_details[fac].zone != '':
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(zone_row) + \
                                                      '&"."&Detail!' + ssCol(col) + str(what_row)
            else:
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(what_row)
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col) + str(cap_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # To meet load MWh
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col) + str(tml_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            dd_tml_sum += ssCol(st_tml+1) + str(ss_row) + '+'
            # subtotal MWh
            ss.cell(row=ss_row, column=st_sub+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(sum_row) + ',"")'
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            dd_re_sum += ssCol(st_sub+1) + str(ss_row) + '+'
            # CF
            ss.cell(row=ss_row, column=st_cfa+1).value = '=IF(Detail!' + ssCol(col) + str(cf_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(cf_row) + ',"")'
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return dd_tml_sum, dd_re_sum
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=IF(Detail!' + ssCol(col) + str(lcoe_row) \
                                                      + '>0,Detail!' + ssCol(col) + str(lcoe_row) + ',"")'
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ssCol(col) + str(cap_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cap_row) + '*'  \
                                                        + str(self.generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            # lifetime cost
            ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                    + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                   + '>0,Detail!' + ssCol(col) + str(max_row) + ',"")'
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            if self.generators[gen].emissions > 0:
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
            if self.generators[gen].area > 0:
                ss.cell(row=ss_row, column=st_are+1).value = '=Detail!' + ssCol(col) + str(cap_row) +\
                                                             '*' + str(self.generators[gen].area)
                ss.cell(row=ss_row, column=st_are+1).number_format = '#,##0.00'
            return dd_tml_sum, dd_re_sum

        def detail_summary_total(ss_row, title='', base_row='', back_row=''):
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Total'
            for col in range(1, len(headers) + 1):
                ss.cell(row=3, column=col).font = bold
                ss.cell(row=ss_row, column=col).font = bold
            for col in [st_cap, st_tml, st_sub, st_cst, st_emi, st_emc, st_cac, st_lic, st_lie, st_lec, st_are]:
                if back_row != '':
                    strt = ssCol(col, base=0) + back_row + '+'
                else:
                    strt = ''
                ss.cell(row=ss_row, column=col+1).value = '=' + strt + 'SUM(' + ssCol(col, base=0) + \
                        base_row + ':' + ssCol(col, base=0) + str(ss_row - 1) + ')'
                if col in [st_cap, st_are]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0.00'
                elif col in [st_tml, st_sub, st_emi, st_lie]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0'
                else:
                    ss.cell(row=ss_row, column=col+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lcg+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_sub+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
            ss.cell(row=ss_row, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_tml+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
            if self.carbon_price > 0:
                ss.cell(row=ss_row, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(ss_row) + \
                    '+' + ssCol(st_emc+1) + str(ss_row) + ')/' + ssCol(st_tml+1) + str(ss_row)
                ss.cell(row=ss_row, column=st_lcc+1).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=st_lcc+1).font = bold
            last_col = ssCol(ns.max_column)
            r = 1
            if self.carbon_price > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Total incl. Carbon Cost'
                ss.cell(row=ss_row, column=st_cst+1).value = '=' + ssCol(st_cst+1) + str(ss_row - 1) + \
                        '+' + ssCol(st_emc+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ss.cell(row=ss_row, column=st_lic+1).value = '=' + ssCol(st_lic+1) + str(ss_row - r) + \
                                                             '+' + ssCol(st_lec+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                r += 1
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'RE %age'
            ss.cell(row=ss_row, column=st_tml+1).value = ns_tml_sum[:-1] + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' +\
                                                         ssCol(st_tml+1) + str(ss_row - r)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_re_row = ss_row
            ss_sto_row = -1
            # if storage
            if ns_sto_sum != '':
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Storage %age'
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + ns_sto_sum[1:]
                ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ns_sto_sum[1:] + ')/' + ssCol(st_tml+1) + \
                                                             str(ss_row - r - 1)
                ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
                ss_sto_row = ss_row
            # now do the LCOE and LCOE with CO2 stuff
            if base_row == '4':
                base_col = 'C'
                if ss_sto_row >= 0:
                    for rw in range(ss_re_fst_row, ss_re_lst_row + 1):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(AND(' + ssCol(st_lcg+1) + str(rw) + '<>"",' + \
                                ssCol(st_lcg+1) + str(rw) + '>0),' + \
                                ssCol(st_cst+1) + str(rw) + '/(' + ssCol(st_tml+1) + str(rw) + '+(' + \
                                ssCol(st_tml+1) + '$' + str(ss_sto_row) + '*' + ssCol(st_tml+1) + str(rw) + \
                                ')/' + ssCol(st_tml+1) + '$' + str(ss_re_row) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                    ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/(' + \
                                    ssCol(st_tml+1) + str(rw) + '+(' + ssCol(st_tml+1) + '$' + str(ss_sto_row) + \
                                    '*' + ssCol(st_tml+1) + str(rw) + ')/' + ssCol(st_tml+1) + '$' + \
                                    str(ss_re_row) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                else:
                    for rw in range(ss_re_fst_row, ss_re_lst_row):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(' + ssCol(st_lcg+1) + str(rw) + '>0,' + \
                                ssCol(st_cst+1) + str(rw) + '/' + ssCol(st_tml+1) + str(rw) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                    ssCol(st_cst+1) + str(rw) + ssCol(st_emc+1) + str(rw) + ')/' + \
                                    ssCol(st_tml+1) + str(rw) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                for rw in range(ss_re_lst_row + 1, ss_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=IF(AND(' + ssCol(st_tml+1) + str(rw) + '<>"",' + \
                                    ssCol(st_tml+1) + str(rw) + '>0),' + ssCol(st_cst+1) + str(rw) + \
                                                             '/' + ssCol(st_tml+1) + str(rw) + ',"")'
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/' + \
                                ssCol(st_tml+1) + str(rw) + ',"")'
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            else:
                base_col = ssCol(next_col)
                for rw in range(ul_fst_row, ul_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(rw) + \
                                                             '/' + ssCol(st_tml+1) + str(rw)
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(rw) + \
                            '+' + ssCol(st_emc+1) + str(rw) + ')/' + ssCol(st_tml+1) + str(rw)
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = title + 'Load Analysis'
            ss.cell(row=ss_row, column=1).font = bold
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Load met'
      ##      lm_row = ss_row
      #      if self.surplus_sign < 0:
      #          addsub = ')+' + base_col
      #      else:
      #          addsub = ')-' + base_col
      #      ss.cell(row=ss_row, column=st_tml+1).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[0] + '=0",Detail!C' + str(hrows) \
      #          + ':Detail!C' + str(hrows + 8759) + ')+SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
      #          + last_col + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!C' + str(hrows) + ':Detail!C' \
      #          + str(hrows + 8759) + addsub + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + base_col + str(sum_row) + '-' + base_col + str(ss_row + 1)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ss_row + 2)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Shortfall'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[0] + '0",Detail!' + last_col \
                      + str(hrows) + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign > 0:
                ss.cell(row=ss_row, column=st_tml+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ss_row + 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ld_row = ss_row
            load_mult = ''
            try:
                mult = round(pmss_details['Load'].multiplier, 3)
                if mult != 1:
                    load_mult = ' x ' + str(mult)
            except:
                pass
            ss.cell(row=ss_row, column=1).value = 'Total ' + title + 'Load - ' + year + load_mult
            ss.cell(row=ss_row, column=1).font = bold
            ss.cell(row=ss_row, column=st_tml+1).value = '=SUM(' + ssCol(st_tml+1) + str(ss_row - 2) + ':' + \
                                                         ssCol(st_tml+1) + str(ss_row - 1) + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_tml+1).font = bold
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + base_col + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=st_max+1).font = bold
            ss.cell(row=ss_row, column=st_bal+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + ',MATCH(Detail!' + \
                    base_col + str(max_row) + ',Detail!' + base_col + str(hrows) + ':Detail!' + base_col + \
                    str(hrows + 8759) + ',0),0)&")"'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'RE %age of Total ' + title + 'Load'
            ss.cell(row=ss_row, column=1).font = bold
            if ns_sto_sum == '':
                ss.cell(row=ss_row, column=st_cap+1).value = ssCol(st_tml+1) + str(ss_re_row - 1) + \
                                                             '/' + ssCol(st_tml+1) + str(ss_row - 1)
            else:
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ssCol(st_tml+1) + str(ss_re_row) + '+' + \
                                                             ssCol(st_tml+1) + str(ss_sto_row) + ')/' + \
                                                             ssCol(st_tml+1) + str(ss_row - 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss.cell(row=ss_row, column=st_cap+1).font = bold
            ss_row += 2
            if ns_loss_sum != '':
                ss.cell(row=ss_row, column=1).value = title + 'Storage Losses'
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + ns_loss_sum[1:]
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
                ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Surplus'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!' + last_col + str(hrows) \
                      + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign < 0:
                ss.cell(row=ss_row, column=st_sub+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_sub+1) + str(ss_row) + '/' + ssCol(st_tml+1) + str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = 'Largest ' + title + 'Shortfall:'
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + last_col + str(hrows + max_short[0])
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=st_cfa+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + \
                        ',MATCH(' + ssCol(st_sub+1) + str(ss_row) + ',Detail!' + last_col + str(hrows) + \
                        ':Detail!' + last_col + str(hrows + 8759) + ',0),0)&")"'
            return ss_row, ss_re_row

    # The "guts" of Powermatch processing. Have a single calculation algorithm
    # for Summary, Powermatch (detail), and Optimise. The detail makes it messy
    # Note: For Batch pmss_data is reused so don't update it in doDispatch
        the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if self.surplus_sign < 0:
            sf_test = ['>', '<']
            sf_sign = ['+', '-']
        else:
            sf_test = ['<', '>']
            sf_sign = ['-', '+']
        sp_cols = []
        sp_cap = []
        shortfall = [0.] * 8760
        re_tml_sum = 0. # keep tabs on how much RE is used
        start_time = time.time()
        do_zone = False # could pass as a parameter
        max_lifetime = 0
        # find max. lifetime years for all technologies selected
        for key in pmss_details.keys():
            if key == 'Load'or key == 'Total':
                continue
            if pmss_details[key].capacity * pmss_details[key].multiplier > 0:
             #   gen = key.split('.')[-1]
                gen = pmss_details[key].generator
                max_lifetime = max(max_lifetime, self.generators[gen].lifetime)
        for key in pmss_details.keys():
            if key.find('.') > 0:
                do_zone = True
                break
        underlying_facs = []
        undercol = [] * len(self.underlying)
        operational_facs = []
        fac_tml = {}
        for fac in re_order:
            if fac == 'Load':
                continue
            fac_tml[fac] = 0.
            if fac in self.operational:
              #  operational_facs.append(fac)
                continue
            if fac.find('.') > 0:
                if fac[fac.find('.') + 1:] in self.underlying:
                    underlying_facs.append(fac)
                    continue
            elif fac in self.underlying:
                underlying_facs.append(fac)
                continue
        load_col = pmss_details['Load'].col
        for h in range(len(pmss_data[load_col])):
            load_h = pmss_data[load_col][h] * pmss_details['Load'].multiplier
            shortfall[h] = load_h
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    continue
                shortfall[h] -= pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
            if shortfall[h] >= 0:
                alloc = 1.
            else:
                alloc = load_h / (load_h - shortfall[h])
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                else:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier * alloc
            line = ''
        fac_tml_sum = 0
        for fac in fac_tml.keys():
            fac_tml_sum += fac_tml[fac]
        if self.show_correlation:
            col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                df1 = pmss_data[col]
            else:
                tgt = []
                for h in range(len(pmss_data[col])):
                    tgt.append(pmss_data[col][h] * pmss_details['Load'].multiplier)
                df1 = tgt
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data = [['Correlation To Load']]
            corr_data.append(['RE Contribution', corr])
        else:
            corr_data = None
        if option == D:
            wb = oxl.Workbook()
            ns = wb.active
            ns.title = 'Detail'
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            ss = wb.create_sheet('Summary', 0)
            ns_re_sum = '=('
            ns_tml_sum = '=('
            ns_sto_sum = ''
            ns_loss_sum = ''
            ns_not_sum = ''
            cap_row = 1
            ns.cell(row=cap_row, column=2).value = 'Capacity (MW/MWh)' #headers[1].replace('\n', ' ')
            ss.row_dimensions[3].height = 40
            ss.cell(row=3, column=st_fac+1).value = headers[st_fac] # facility
            ss.cell(row=3, column=st_cap+1).value = headers[st_cap] # capacity
            ini_row = 2
            ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
            tml_row = 3
            ns.cell(row=tml_row, column=2).value = headers[st_tml].replace('\n', ' ')
            ss.cell(row=3, column=st_tml+1).value = headers[st_tml] # to meet load
            sum_row = 4
            ns.cell(row=sum_row, column=2).value = headers[st_sub].replace('\n', ' ')
            ss.cell(row=3, column=st_sub+1).value = headers[st_sub] # subtotal MWh
            cf_row = 5
            ns.cell(row=cf_row, column=2).value = headers[st_cfa].replace('\n', ' ')
            ss.cell(row=3, column=st_cfa+1).value = headers[st_cfa] # CF
            cost_row = 6
            ns.cell(row=cost_row, column=2).value = headers[st_cst].replace('\n', ' ')
            ss.cell(row=3, column=st_cst+1).value = headers[st_cst] # Cost / yr
            lcoe_row = 7
            ns.cell(row=lcoe_row, column=2).value = headers[st_lcg].replace('\n', ' ')
            ss.cell(row=3, column=st_lcg+1).value = headers[st_lcg] # LCOG
            ss.cell(row=3, column=st_lco+1).value = headers[st_lco] # LCOE
            emi_row = 8
            ns.cell(row=emi_row, column=2).value = headers[st_emi].replace('\n', ' ')
            ss.cell(row=3, column=st_emi+1).value = headers[st_emi] # emissions
            ss.cell(row=3, column=st_emc+1).value = headers[st_emc] # emissions cost
            ss.cell(row=3, column=st_lcc+1).value = headers[st_lcc] # LCOE with CO2
            ss.cell(row=3, column=st_max+1).value = headers[st_max] # max. MWh
            ss.cell(row=3, column=st_bal+1).value = headers[st_bal] # max. balance
            ss.cell(row=3, column=st_cac+1).value = headers[st_cac] # capital cost
            ss.cell(row=3, column=st_lic+1).value = headers[st_lic] # lifetime cost
            ss.cell(row=3, column=st_lie+1).value = headers[st_lie] # lifetime emissions
            ss.cell(row=3, column=st_lec+1).value = headers[st_lec] # lifetime emissions cost
            ss.cell(row=3, column=st_are+1).value = headers[st_are] # area
            ss.cell(row=3, column=st_rlc+1).value = headers[st_rlc] # reference lcoe
            ss.cell(row=3, column=st_rcf+1).value = headers[st_rcf] # reference cf
            ss_row = 3
            ss_re_fst_row = 4
            fall_row = 9
            ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
            max_row = 10
            ns.cell(row=max_row, column=2).value = 'Maximum (MW/MWh)'
            hrs_row = 11
            ns.cell(row=hrs_row, column=2).value = 'Hours of usage'
            if do_zone:
                zone_row = 12
                what_row = 13
                hrows = 14
                ns.cell(row=zone_row, column=1).value = 'Zone'
            else:
                what_row = 12
                hrows = 13
            ns.cell(row=what_row, column=1).value = 'Hour'
            ns.cell(row=what_row, column=2).value = 'Period'
            ns.cell(row=what_row, column=3).value = 'Load'
            ns.cell(row=sum_row, column=3).value = '=SUM(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=3).number_format = '#,##0'
            ns.cell(row=max_row, column=3).value = '=MAX(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=3).number_format = '#,##0.00'
            o = 4
            col = 3
            # hour, period
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=1).value = row - hrows + 1
                ns.cell(row=row, column=2).value = format_period(row - hrows)
            # and load
            load_col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows] * \
                            pmss_details['Load'].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            # here we're processing renewables (so no storage)
            for fac in re_order:
                if fac == 'Load':
                    continue
                if fac in underlying_facs:
                    continue
                if pmss_details[fac].col <= 0:
                    continue
                ss_row += 1
                col = do_detail(fac, col, ss_row)
                ns_tml_sum, ns_re_sum = do_detail_summary(fac, col, ss_row, ns_tml_sum, ns_re_sum)
            ss_re_lst_row = ss_row
            col += 1
            shrt_col = col
            ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ssCol(shrt_col) \
                            + str(hrows) + ':' + ssCol(shrt_col) + str(hrows + 8759) + \
                            ',"' + sf_test[0] + '0")'
            ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
            ns.cell(row=what_row, column=shrt_col).value = 'Shortfall (' + sf_sign[0] \
                    + ') /\nSurplus (' + sf_sign[1] + ')'
            ns.cell(row=max_row, column=shrt_col).value = '=MAX(' + ssCol(shrt_col) + str(hrows) + \
                                           ':' + ssCol(shrt_col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=shrt_col).number_format = '#,##0.00'
            for col in range(3, shrt_col + 1):
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                for col in range(3, shrt_col + 1):
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                ns.cell(row=row, column=col).number_format = '#,##0.00'
            col = shrt_col + 1
            ns.cell(row=tml_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                   ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            ns.cell(row=what_row, column=col).value = 'RE Contrib.\nto Load'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            for row in range(hrows, 8760 + hrows):
                if shortfall[row - hrows] < 0:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * pmss_details['Load'].multiplier
                else:
                    if pmss_details['Load'].multiplier == 1:
                        rec = pmss_data[load_col][row - hrows] - shortfall[row - hrows]
                    else:
                        rec = pmss_data[load_col][row - hrows] * pmss_details['Load'].multiplier - \
                              shortfall[row - hrows]
                ns.cell(row=row, column=col).value = rec
               # the following formula will do the same computation
               # ns.cell(row=row, column=col).value = '=IF(' + ssCol(shrt_col) + str(row) + '>0,' + \
               #                            ssCol(3) + str(row) + ',' + ssCol(3) + str(row) + \
               #                            '+' + ssCol(shrt_col) + str(row) + ')'
                ns.cell(row=row, column=col).number_format = '#,##0.00'
          #  shrt_col += 1
           # col = shrt_col + 1
            ul_re_sum = ns_re_sum
            ul_tml_sum = ns_tml_sum
            nsul_sums = ['C']
            nsul_sum_cols = [3]
            for fac in underlying_facs:
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                col = do_detail(fac, col, -1)
                nsul_sums.append(ssCol(col))
                nsul_sum_cols.append(col)
            if col > shrt_col + 1: # underlying
                col += 1
                ns.cell(row=what_row, column=col).value = 'Underlying\nLoad'
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                for row in range(hrows, 8760 + hrows):
                    txt = '='
                    for c in nsul_sums:
                        txt += c + str(row) + '+'
                    ns.cell(row=row, column=col).value = txt[:-1]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            next_col = col
            col += 1
        else: # O, O1, B, T
            sp_data = []
            sp_load = 0. # load from load curve
            hrows = 10
            load_max = 0
            load_hr = 0
            tml = 0.
            try:
                load_col = pmss_details['Load'].col
            except:
                load_col = 0
            if (option == B or option == T) and len(underlying_facs) > 0:
                load_facs = underlying_facs[:]
                load_facs.insert(0, 'Load')
                for h in range(len(pmss_data[load_col])):
                    amt = 0
                    for fac in load_facs:
                        amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
                    sp_load += amt
                underlying_facs = []
            else:
                fac = 'Load'
                sp_load = sum(pmss_data[load_col]) * pmss_details[fac].multiplier
                for h in range(len(pmss_data[load_col])):
                    amt = pmss_data[load_col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
            for fac in re_order:
                if fac == 'Load' or fac in underlying_facs:
                    continue
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = fac
                sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                try:
                    sp_d[st_tml] = fac_tml[fac]
                except:
                    pass
                sp_d[st_sub] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_data.append(sp_d)
       #     for h in range(len(shortfall)):
        #        if shortfall[h] < 0:
         #           tml += pmss_data[load_col][h] * pmss_details['Load'].multiplier
          #      else:
           #         tml += pmss_data[load_col][h] * pmss_details['Load'].multiplier - shortfall[h]
        if option not in [O, O1, B, T]:
            if progress_handler:
                progress = ProgressInfo(
                        current=6,
                )
                progress_handler.update(progress)
        storage_names = []
        # find any minimum generation for generators
        short_taken = {}
        short_taken_tot = 0
        for gen in dispatch_order:
            if pmss_details[gen].fac_type == 'G': # generators
                try:
                    const = self.generators[gen].constraint
                except:
                    try:
                        g2 = gen[gen.find('.') + 1:]
                        const = self.generators[g2].constraint
                    except:
                        continue
                if self.constraints[const].capacity_min != 0:
                    try:
                        short_taken[gen] = pmss_details[gen].capacity * pmss_details[gen].multiplier * \
                            self.constraints[const].capacity_min
                    except:
                        short_taken[gen] = pmss_details[gen].capacity * \
                            self.constraints[const].capacity_min
                    short_taken_tot += short_taken[gen]
                    for row in range(8760):
                        shortfall[row] = shortfall[row] - short_taken[gen]
        tot_sto_loss = 0.
        for gen in dispatch_order:
         #   min_after = [0, 0, -1, 0, 0, 0] # initial, low balance, period, final, low after, period
         #  Min_after is there to see if storage is as full at the end as at the beginning
            try:
                capacity = pmss_details[gen].capacity * pmss_details[gen].multiplier
            except:
                try:
                    capacity = pmss_details[gen].capacity
                except:
                    continue
            if gen not in self.generators.keys():
                continue
            if self.generators[gen].constraint in self.constraints and \
              self.constraints[self.generators[gen].constraint].category == 'Storage': # storage
                storage_names.append(gen)
                storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                storage[0] = capacity
                if option == D:
                    ns.cell(row=cap_row, column=col + 2).value = capacity
                    ns.cell(row=cap_row, column=col + 2).number_format = '#,##0.00'
                try:
                    storage[1] = self.generators[gen].initial * pmss_details[gen].multiplier
                except:
                    storage[1] = self.generators[gen].initial
                if self.constraints[self.generators[gen].constraint].capacity_min > 0:
                    storage[2] = capacity * self.constraints[self.generators[gen].constraint].capacity_min
                if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                    storage[3] = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                else:
                    storage[3] = capacity
                recharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].recharge_max > 0:
                    recharge[0] = capacity * self.constraints[self.generators[gen].constraint].recharge_max
                else:
                    recharge[0] = capacity
                if self.constraints[self.generators[gen].constraint].recharge_loss > 0:
                    recharge[1] = self.constraints[self.generators[gen].constraint].recharge_loss
                discharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].discharge_max > 0:
                    discharge[0] = capacity * self.constraints[self.generators[gen].constraint].discharge_max
                if self.constraints[self.generators[gen].constraint].discharge_loss > 0:
                    discharge[1] = self.constraints[self.generators[gen].constraint].discharge_loss
                if self.constraints[self.generators[gen].constraint].parasitic_loss > 0:
                    parasite = self.constraints[self.generators[gen].constraint].parasitic_loss / 24.
                else:
                    parasite = 0.
                in_run = [False, False]
                min_run_time = self.constraints[self.generators[gen].constraint].min_run_time
                in_run[0] = True # start off in_run
                if min_run_time > 0 and self.generators[gen].initial == 0:
                    in_run[0] = False
                warm_time = self.constraints[self.generators[gen].constraint].warm_time
                storage_carry = storage[1] # self.generators[gen].initial
                if option == D:
                    ns.cell(row=ini_row, column=col + 2).value = storage_carry
                    ns.cell(row=ini_row, column=col + 2).number_format = '#,##0.00'
                storage_bal = []
                storage_can = 0.
                use_max = [0, None]
                sto_max = storage_carry
                for row in range(8760):
                    storage_loss = 0.
                    storage_losses = 0.
                    if storage_carry > 0:
                        loss = storage_carry * parasite
                        # for later: record parasitic loss
                        storage_carry = storage_carry - loss
                        storage_losses -= loss
                    if shortfall[row] < 0:  # excess generation
                        if row % 24 >= self.constraints[self.generators[gen].constraint].recharge_start:
                            if min_run_time > 0:
                                in_run[0] = False
                            if warm_time > 0:
                                in_run[1] = False
                            can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                            if can_use < 0: # can use some
                                if shortfall[row] > can_use:
                                    can_use = shortfall[row]
                                if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                    can_use = - recharge[0]
                            else:
                                can_use = 0.
                            # for later: record recharge loss
                            storage_losses += can_use * recharge[1]
                            storage_carry -= (can_use * (1 - recharge[1]))
                            shortfall[row] -= can_use
                            if corr_data is not None:
                               corr_src[row] += can_use
                        else:
                            can_use = 0
                    else: # shortfall
                        # This is code to support delaying battery usage until a certain time
                        # to implement fully need an additional facility variable to indicate start time
                        # Ref 2024 WEM ESOO 2.5 (ESROI)
                        if row % 24 >= self.constraints[self.generators[gen].constraint].discharge_start:
                            if min_run_time > 0 and shortfall[row] > 0:
                                if not in_run[0]:
                                    if row + min_run_time <= 8759:
                                        for i in range(row + 1, row + min_run_time + 1):
                                            if shortfall[i] <= 0:
                                                break
                                        else:
                                            in_run[0] = True
                            if in_run[0]:
                                can_use = shortfall[row] * (1 / (1 - discharge[1]))
                                can_use = min(can_use, discharge[0])
                                if can_use > storage_carry - storage[2]:
                                    can_use = storage_carry - storage[2]
                                if warm_time > 0 and not in_run[1]:
                                    in_run[1] = True
                                    can_use = can_use * (1 - warm_time)
                            else:
                                can_use = 0
                            if can_use > 0:
                                storage_loss = can_use * discharge[1]
                                storage_losses -= storage_loss
                                storage_carry -= can_use
                                can_use = can_use - storage_loss
                                shortfall[row] -= can_use
                                if corr_data is not None:
                                    corr_src[row] += can_use
                                if storage_carry < 0:
                                    storage_carry = 0
                            else:
                                can_use = 0.
                        else:
                            can_use = 0.
                    if can_use < 0:
                        if use_max[1] is None or can_use < use_max[1]:
                            use_max[1] = can_use
                    elif can_use > use_max[0]:
                        use_max[0] = can_use
                    storage_bal.append(storage_carry)
                    if storage_bal[-1] > sto_max:
                        sto_max = storage_bal[-1]
                    if option == D:
                        if can_use > 0:
                            ns.cell(row=row + hrows, column=col).value = 0
                            ns.cell(row=row + hrows, column=col + 2).value = can_use * self.surplus_sign
                        else:
                            ns.cell(row=row + hrows, column=col).value = can_use * -self.surplus_sign
                            ns.cell(row=row + hrows, column=col + 2).value = 0
                        ns.cell(row=row + hrows, column=col + 1).value = storage_losses
                        ns.cell(row=row + hrows, column=col + 3).value = storage_carry
                        ns.cell(row=row + hrows, column=col + 4).value = (shortfall[row] + short_taken_tot) * -self.surplus_sign
                        for ac in range(5):
                            ns.cell(row=row + hrows, column=col + ac).number_format = '#,##0.00'
                            ns.cell(row=max_row, column=col + ac).value = '=MAX(' + ssCol(col + ac) + \
                                    str(hrows) + ':' + ssCol(col + ac) + str(hrows + 8759) + ')'
                            ns.cell(row=max_row, column=col + ac).number_format = '#,##0.00'
                    else:
                        tot_sto_loss += storage_losses
                        if can_use > 0:
                            storage_can += can_use
                if option == D:
                    ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 1).value = '=SUMIF(' + ssCol(col + 1) + \
                            str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + ',"<0")'
                    ns.cell(row=sum_row, column=col + 1).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 2).value = '=SUMIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col + 2).value = '=IF(' + ssCol(col + 2) + str(cap_row) + '>0,' + \
                            ssCol(col + 2) + str(sum_row) + '/' + ssCol(col + 2) + '1/8760,"")'
                    ns.cell(row=cf_row, column=col + 2).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col + 2).value = '=COUNTIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 3).value = '=' + ssCol(col + 2) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 3).number_format = '#,##0.0%'
                    col += 5
                else:
                    if storage[0] == 0:
                        continue
               #     tml_tot += storage_can
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = storage[0]
                    sp_d[st_tml] = storage_can
                    sp_d[st_max] = use_max[0]
                    sp_d[st_bal] = sto_max
                    sp_data.append(sp_d)
            else: # generator
                try:
                    if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                        cap_capacity = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                    else:
                        cap_capacity = capacity
                except:
                    cap_capacity = capacity
                if gen in short_taken.keys():
                    for row in range(8760):
                        shortfall[row] = shortfall[row] + short_taken[gen]
                    short_taken_tot -= short_taken[gen]
                    min_gen = short_taken[gen]
                else:
                    min_gen = 0
                if option == D:
                    ns.cell(row=cap_row, column=col).value = capacity
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                ns.cell(row=row + hrows, column=col).value = cap_capacity
                            elif shortfall[row] < min_gen:
                                ns.cell(row=row + hrows, column=col).value = min_gen
                                shortfall[row] -= min_gen
                            else:
                                ns.cell(row=row + hrows, column=col).value = shortfall[row]
                                shortfall[row] = 0
                        else:
                            shortfall[row] -= min_gen
                            ns.cell(row=row + hrows, column=col).value = min_gen
                        ns.cell(row=row + hrows, column=col + 1).value = (shortfall[row] + short_taken_tot) * -self.surplus_sign
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                            ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                            ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                                str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 1).value = '=' + ssCol(col) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 1).number_format = '#,##0.0%'
                    col += 2
                else:
                    gen_can = 0.
                    gen_max = 0
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                gen_can += cap_capacity
                                if cap_capacity > gen_max:
                                    gen_max = cap_capacity
                            elif shortfall[row] < min_gen:
                                gen_can += min_gen
                                if min_gen > gen_max:
                                    gen_max = min_gen
                                shortfall[row] -= min_gen
                            else:
                                gen_can += shortfall[row]
                                if shortfall[row] > gen_max:
                                    gen_max = shortfall[row]
                                shortfall[row] = 0
                        else:
                            if min_gen > gen_max:
                                gen_max = min_gen
                            gen_can += min_gen
                            shortfall[row] -= min_gen # ??
                    if capacity == 0:
                        continue
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = capacity
                    sp_d[st_tml] = gen_can
                    sp_d[st_sub] = gen_can
                    sp_d[st_max] = gen_max
                    sp_data.append(sp_d)
#        if option == D: # Currently calculated elsewhere
#            if self.surplus_sign > 0:
#                maxmin = 'MIN'
#            else:
#                maxmin = 'MAX'
#            ns.cell(row=max_row, column=col-1).value = '=' + maxmin + '(' + \
#                    ssCol(col-1) + str(hrows) + ':' + ssCol(col - 1) + str(hrows + 8759) + ')'
#            ns.cell(row=max_row, column=col-1).number_format = '#,##0.00'
        if option not in [O, O1, B, T]:
            if progress_handler:
                progress = ProgressInfo(
                        current=8,
                )
                progress_handler.update(progress)
        if corr_data is not None:
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['RE plus Storage', corr])
            col = pmss_details['Load'].col
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['To Meet Load', corr])
            for c in range(1, len(corr_data)):
                if abs(corr_data[c][1]) < 0.1:
                    corr_data[c].append('None')
                elif abs(corr_data[c][1]) < 0.3:
                    corr_data[c].append('Little if any')
                elif abs(corr_data[c][1]) < 0.5:
                    corr_data[c].append('Low')
                elif abs(corr_data[c][1]) < 0.7:
                    corr_data[c].append('Moderate')
                elif abs(corr_data[c][1]) < 0.9:
                    corr_data[c].append('High')
                else:
                    corr_data[c].append('Very high')
        if option != D:
            load_col = pmss_details['Load'].col
            cap_sum = 0.
            gen_sum = 0.
            re_sum = 0.
            tml_sum = 0.
            ff_sum = 0.
            sto_sum = 0.
            cost_sum = 0.
            co2_sum = 0.
            co2_cost_sum = 0.
            capex_sum = 0.
            lifetime_sum = 0.
            lifetime_co2_sum = 0.
            lifetime_co2_cost = 0.
            total_area = 0.
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    sto_sum += sp_data[sp][2]
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen in tech_names or gen2 in tech_names:
                        re_sum += sp_data[sp][st_sub]
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    ndx = 2
                else:
                    if gen in self.generators.keys():
                        pass
                    else:
                        try:
                            gen = gen[gen.find('.') + 1:]
                        except:
                            pass
                    ndx = 3
                try:
                    if sp_data[sp][st_cap] > 0:
                        cap_sum += sp_data[sp][st_cap]
                        if self.generators[gen].lcoe > 0:
                            sp_data[sp][st_cfa] = sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 # need number for now
                        else:
                            sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 * 100)
                    gen_sum += sp_data[sp][st_sub]
                except:
                    pass
                try:
                    tml_sum += sp_data[sp][st_tml]
                except:
                    pass
                if gen not in self.generators.keys():
                    continue
                ndx = 3
                if gen in storage_names:
                    ndx = 2
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen not in tech_names and gen2 not in tech_names:
                        ff_sum += sp_data[sp][ndx]
                if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                  or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    capex = sp_data[sp][st_cap] * self.generators[gen].capex
                    capex_sum += capex
                    opex = sp_data[sp][st_cap] * self.generators[gen].fixed_om \
                           + sp_data[sp][ndx] * self.generators[gen].variable_om \
                           + sp_data[sp][ndx] * self.generators[gen].fuel
                    disc_rate = self.generators[gen].disc_rate
                    if disc_rate == 0:
                        disc_rate = self.discount_rate
                    lifetime = self.generators[gen].lifetime
                    sp_data[sp][st_lcg] = calcLCOE(sp_data[sp][ndx], capex, opex, disc_rate, lifetime)
                    sp_data[sp][st_cst] = sp_data[sp][ndx] * sp_data[sp][st_lcg]
                    if gen in tech_names or gen2 in tech_names:
                        sp_data[sp][st_lco] = sp_data[sp][st_cst] / (sp_data[sp][st_tml] + (sto_sum * sp_data[sp][st_tml] / fac_tml_sum))
                    else:
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_cac] = capex
                elif self.generators[gen].lcoe > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    if self.generators[gen].lcoe_cf > 0:
                        lcoe_cf = self.generators[gen].lcoe_cf
                    else:
                        lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_data[sp][st_cap]
                    if sp_data[sp][st_cfa] > 0:
                        sp_data[sp][st_lcg] = sp_data[sp][st_cst] / sp_data[sp][ndx]
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][st_cfa] * 100.)
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_rlc] = self.generators[gen].lcoe
                    sp_data[sp][st_rcf] = '{:.1f}%'.format(lcoe_cf * 100.)
                elif self.generators[gen].lcoe_cf == 0: # no cost facility
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = 0
                    cost_sum += sp_data[sp][st_cst]
                sp_data[sp][st_lic] = sp_data[sp][st_cst] * max_lifetime
                lifetime_sum += sp_data[sp][st_lic]
                if self.generators[gen].emissions > 0 and sp_data[sp][st_tml]> 0:
                    sp_data[sp][st_emi] = sp_data[sp][ndx] * self.generators[gen].emissions
                    co2_sum += sp_data[sp][st_emi]
                    sp_data[sp][st_emc] = sp_data[sp][st_emi] * self.carbon_price
                    if sp_data[sp][st_cst] == 0:
                        sp_data[sp][st_lcc] = sp_data[sp][st_emc] / sp_data[sp][st_tml]
                    else:
                        sp_data[sp][st_lcc] = sp_data[sp][st_lco] * ((sp_data[sp][st_cst] + sp_data[sp][st_emc]) / sp_data[sp][st_cst])
                    co2_cost_sum += sp_data[sp][st_emc]
                    sp_data[sp][st_lie] = sp_data[sp][st_emi] * max_lifetime
                    lifetime_co2_sum += sp_data[sp][st_lie]
                    sp_data[sp][st_lec] = sp_data[sp][st_lie] * self.carbon_price
                    lifetime_co2_cost += sp_data[sp][st_lec]
                else:
                    sp_data[sp][st_lcc] = sp_data[sp][st_lco]
                if self.generators[gen].area > 0:
                    sp_data[sp][st_are] = sp_data[sp][st_cap] * self.generators[gen].area
                    total_area += sp_data[sp][st_are]
            sf_sums = [0., 0., 0.]
            for sf in range(len(shortfall)):
                if shortfall[sf] > 0:
                    sf_sums[0] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
                else:
                    sf_sums[1] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
            if gen_sum > 0:
                gs = cost_sum / gen_sum
            else:
                gs = ''
            if tml_sum > 0:
                gsw = cost_sum / tml_sum # LCOE
                gswc = (cost_sum + co2_cost_sum) / tml_sum
            else:
                gsw = ''
                gswc = ''
            if option == O or option == O1:
                load_pct, surp_pct, re_pct = summary_totals()
            else:
                summary_totals()
            do_underlying = False
            if len(underlying_facs) > 0:
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier > 0:
                        do_underlying = True
                        break
            if do_underlying:
                sp_data.append([' '])
                sp_data.append(['Additional Underlying Load'])
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                        continue
                    if fac in self.generators.keys():
                        gen = fac
                    else:
                        gen = pmss_details[fac].generator
                    col = pmss_details[fac].col
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = fac
                    sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                    cap_sum += sp_d[st_cap]
                    sp_d[st_tml] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    tml_sum += sp_d[st_tml]
                    sp_d[st_sub] = sp_d[st_tml]
                    gen_sum += sp_d[st_tml]
                    sp_load += sp_d[st_tml]
                    sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_sub] / sp_d[st_cap] / 8760 * 100.)
                    sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                      or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                        capex = sp_d[st_cap] * self.generators[gen].capex
                        capex_sum += capex
                        opex = sp_d[st_cap] * self.generators[gen].fixed_om \
                               + sp_d[st_tml] * self.generators[gen].variable_om \
                               + sp_d[st_tml] * self.generators[gen].fuel
                        disc_rate = self.generators[gen].disc_rate
                        if disc_rate == 0:
                            disc_rate = self.discount_rate
                        lifetime = self.generators[gen].lifetime
                        sp_d[st_lcg] = calcLCOE(sp_d[st_tml], capex, opex, disc_rate, lifetime)
                        sp_d[st_cst] = sp_d[st_tml] * sp_d[st_lcg]
                        cost_sum += sp_d[st_cst]
                        sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cac] = capex
                    elif self.generators[gen].lcoe > 0:
                        if self.generators[gen].lcoe_cf > 0:
                            lcoe_cf = self.generators[gen].lcoe_cf
                        else:
                            lcoe_cf = sp_d[st_cfa]
                        sp_d[st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_d[st_cap]
                        cost_sum += sp_d[st_cst]
                        if sp_d[st_cfa] > 0:
                            sp_d[st_lcg] = sp_d[st_cst] / sp_d[st_tml]
                            sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_cfa] * 100.)
                        sp_d[st_rlc] = self.generators[gen].lcoe
                        sp_d[st_rcf] = '{:.1f}%'.format(lcoe_cf * 100.)
                    elif self.generators[gen].lcoe_cf == 0: # no cost facility
                        sp_d[st_cst] = 0
                        sp_d[st_lcg] = 0
                        sp_d[st_lco] = 0
                        sp_d[st_rlc] = self.generators[gen].lcoe
                    sp_d[st_lic] = sp_d[st_cst] * max_lifetime
                    lifetime_sum += sp_d[st_lic]
                    if self.generators[gen].emissions > 0:
                        sp_d[st_emi] = sp_d[st_tml] * self.generators[gen].emissions
                        co2_sum += sp_d[st_emi]
                        sp_d[st_emc] = sp_d[st_emi] * self.carbon_price
                        if sp_d[st_cst] > 0:
                            sp_d[st_lcc] = sp_d[st_lco] * ((sp_d[st_cst] + sp_d[st_emc]) / sp_d[st_cst])
                        else:
                            sp_d[st_lcc] = sp_d[st_emc] / sp_d[st_tml]
                        co2_cost_sum += sp_d[st_emc]
                        sp_d[st_lie] = sp_d[st_emi] * max_lifetime
                        lifetime_co2_sum += sp_d[st_lie]
                        sp_d[st_lec] = sp_d[st_lie] * self.carbon_price
                        lifetime_co2_cost += sp_d[st_lec]
                    else:
                        sp_d[st_lcc] = sp_d[st_lco]
                    if self.generators[gen].area > 0:
                        sp_d[st_are] = sp_d[st_cap] * self.generators[gen].area
                    sp_data.append(sp_d)
                if gen_sum > 0:
                    gs = cost_sum / gen_sum
                else:
                    gs = ''
                if tml_sum > 0:
                    gsw = cost_sum / tml_sum # LCOE
                    gswc = (cost_sum + co2_cost_sum) / tml_sum
                else:
                    gsw = ''
                    gswc = ''
                # find maximum underlying load
                if option == S:
                    load_max = 0
                    load_hr = 0
                    load_col = pmss_details['Load'].col
                    for h in range(len(pmss_data[load_col])):
                        amt = pmss_data[load_col][h] * pmss_details['Load'].multiplier
                        for fac in underlying_facs:
                            amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                        if amt > load_max:
                            load_max = amt
                            load_hr = h
                summary_totals('Underlying ')
            if corr_data is not None:
                sp_data.append([' '])
                sp_data = sp_data + corr_data
            sp_data.append([' '])
            sp_data.append(['Static Variables'])
            if self.carbon_price > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Carbon Price ($/tCO2e)'
                sp_d[st_cap] = self.carbon_price
                sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Lifetime (years)'
            sp_d[st_cap] = max_lifetime
            sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Discount Rate'
            sp_d[st_cap] = '{:.2%}'.format(self.discount_rate)
            sp_data.append(sp_d)
            if option == B or option == T:
                if self.optimise_debug:
                    sp_pts = [0] * len(headers)
                    for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                        sp_pts[p] = 2
                    if corr_data is not None:
                        sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
                    dialog = self.displaytable.Table(sp_data, title='Debug', fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=sp_pts)
                    dialog.exec_()
                return sp_data
            if option == O or option == O1:
                op_load_tot = pmss_details['Load'].capacity * pmss_details['Load'].multiplier
                if gswc != '':
                    lcoe = gswc
                elif self.adjusted_lcoe:
                    lcoe = gsw # target is lcoe
                else:
                    lcoe = gs
                if gen_sum == 0:
                    re_pct = 0
                    load_pct = 0
                    re_pct = 0
                multi_value = {'lcoe': lcoe, #lcoe. lower better
                    'load_pct': load_pct, #load met. 100% better
                    'surplus_pct': surp_pct, #surplus. lower better
                    're_pct': re_pct, # RE pct. higher better
                    'cost': cost_sum, # cost. lower better
                    'co2': co2_sum} # CO2. lower better
                if option == O:
                    if multi_value['lcoe'] == '':
                        multi_value['lcoe'] = 0
                    return multi_value, sp_data, None
                else:
                    extra = [gsw, op_load_tot, sto_sum, re_sum, re_pct, sf_sums]
                    return multi_value, sp_data, extra
        #    list(map(list, list(zip(*sp_data))))
            span = None
            if self.summary_sources: # want data sources
                sp_data.append([' '])
                sp_data.append(['Data sources'])
                span = 'Data sources'
                sp_data.append(['Scenarios folder', self.scenarios])
                if pm_data_file[: len(self.scenarios)] == self.scenarios:
                    pm_data_file = pm_data_file[len(self.scenarios):]
                sp_data.append(['Powermatch data file', pm_data_file])
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sp_data.append(['Load file', load_file])
                sp_data.append(['Constraints worksheet', str(self.files[C].text()) \
                                + '.' + str(self.sheets[C].currentText())])
                sp_data.append(['Generators worksheet', str(self.files[G].text()) \
                                + '.' + str(self.sheets[G].currentText())])
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc, st_are]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            self.setStatus(sender_text + ' completed')
            if title is not None:
                atitle = title
            elif self.results_prefix != '':
                atitle = self.results_prefix + '_' + sender_text
            else:
                atitle = sender_text
            dialog = self.displaytable(sp_data, title=atitle, fields=headers,
                     save_folder=self.scenarios, sortby='', decpts=sp_pts,
                     span=span)
            dialog.exec_()
            if progress_handler:
                progress = ProgressInfo(
                        current=0,
                )
                progress_handler.update(progress)
                progress_handler.hide()
            return # finish if not detailed spreadsheet
        col = next_col + 1
        is_storage = False
        ss_sto_rows = []
        ss_st_row = -1
        for gen in dispatch_order:
            ss_row += 1
            try:
                if self.constraints[self.generators[gen].constraint].category == 'Storage':
                    ss_sto_rows.append(ss_row)
                    nc = 2
                    ns.cell(row=what_row, column=col).value = 'Charge\n' + gen
                    ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    ns.cell(row=what_row, column=col + 1).value = gen + '\nLosses'
                    ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    is_storage = True
                    ns_sto_sum += '+' + ssCol(st_tml+1) + str(ss_row)
                    ns_loss_sum += '+Detail!' + ssCol(col + 1) + str(sum_row)
                else:
                    nc = 0
                    is_storage = False
                    ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            except KeyError as err:
                msg = 'Key Error: No Constraint for ' + gen
                if title is not None:
                    msg += ' (model ' + title + ')'
                self.setStatus(msg)
                nc = 0
                is_storage = False
                ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            ns.cell(row=what_row, column=col + nc).value = gen
            ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col + nc) + str(what_row)
            # facility
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col + nc) + str(cap_row)
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # tml
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            # subtotal
            try:
                if self.constraints[self.generators[gen].constraint].category != 'Storage':
                    ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                    ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            except KeyError as err:
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            # cf
            ss.cell(row=ss_row, column=st_cfa+1).value = '=Detail!' + ssCol(col + nc) + str(cf_row)
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col + nc) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col + nc) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + \
                        '),0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + \
                        str(cost_row) + '/' + ssCol(col + nc) + str(sum_row) + ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ssCol(col + nc) + str(cap_row) \
                                                            + '>0,Detail!' + ssCol(col + nc) + str(cap_row) + '*'  \
                                                            + str(self.generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcoe_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1).number_format = '#,##0.0%'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col + nc).value = '=' + ssCol(col + nc) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col + nc).number_format = '#,##0'
                # emissions
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col + nc) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + ssCol(col + nc) + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            # max balance
            if nc > 0: # storage
                ss.cell(row=ss_row, column=st_bal+1).value = '=Detail!' + ssCol(col + nc + 1) + str(max_row)
                ss.cell(row=ss_row, column=st_bal+1).number_format = '#,##0.00'
            ns.cell(row=what_row, column=col + nc).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + nc + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if is_storage:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col + 2) + str(sum_row) \
                                                        + '>0,Detail!' + ssCol(col + 2) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                # ns.cell(row=what_row, column=col + 1).value = gen
                ns.cell(row=what_row, column=col + 3).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 3).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=what_row, column=col + 4).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 4).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 4).value = '=COUNTIF(' + ssCol(col + 4) \
                        + str(hrows) + ':' + ssCol(col + 4) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 4).number_format = '#,##0'
                col += 5
            else:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ssCol(col + 1) \
                        + str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
        if is_storage:
            ns.cell(row=emi_row, column=col - 2).value = '=MIN(' + ssCol(col - 2) + str(hrows) + \
                    ':' + ssCol(col - 2) + str(hrows + 8759) + ')'
            ns.cell(row=emi_row, column=col - 2).number_format = '#,##0.00'
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            sum_value = 0
            do_sum = False
            do_cost = False
            for cell in column_cells:
                if cell.row >= hrows:
                    if do_sum:
                        try:
                            sum_value += abs(cell.value)
                        except:
                            pass
                    else:
                        try:
                            value = str(round(cell.value, 2))
                            if len(value) > length:
                                length = len(value) + 2
                        except:
                            pass
                elif cell.row > 0:
                    if str(cell.value)[0] != '=':
                        values = str(cell.value).split('\n')
                        for value in values:
                            if cell.row == cost_row:
                                valf = value.split('.')
                                alen = int(len(valf[0]) * 1.6)
                            else:
                                alen = len(value) + 2
                            if alen > length:
                                length = alen
                    else:
                        if cell.row == cost_row:
                            do_cost = True
                        if cell.value[1:4] == 'SUM':
                            do_sum = True
            if sum_value > 0:
                alen = len(str(int(sum_value))) * 1.5
                if do_cost:
                    alen = int(alen * 1.5)
                if alen > length:
                    length = alen
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            ns.column_dimensions[cel].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        ns.column_dimensions['B'].width = 21
        st_row = hrows + 8760
        st_col = col
        for row in range(1, st_row):
            for col in range(1, st_col):
                try:
                    ns.cell(row=row, column=col).font = normal
                except:
                    pass
        if progress_handler:
            progress = ProgressInfo(
                current=12,
            )
            progress_handler.update(progress)
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        if self.results_prefix != '':
            ss.cell(row=1, column=1).value = 'Powermatch - ' + self.results_prefix + ' Summary'
        else:
            ss.cell(row=1, column=1).value = 'Powermatch - Summary'
        ss.cell(row=1, column=1).font = bold
        ss_lst_row = ss_row + 1
        ss_row, ss_re_row = detail_summary_total(ss_row, base_row='4')
        if len(nsul_sum_cols) > 1: # if we have underlying there'll be more than one column
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = 'Additional Underlying Load'
            ss.cell(row=ss_row, column=1).font = bold
            base_row = str(ss_row + 1)
            for col in nsul_sum_cols[1:]:
                ss_row += 1
                ul_tml_sum, ul_re_sum = do_detail_summary(fac, col, ss_row, ul_tml_sum, ul_re_sum)
            ul_fst_row = int(base_row)
            ul_lst_row = ss_row
            ns_re_sum = ul_re_sum
            ns_tml_sum = ul_tml_sum
            ss_row, ss_re_row = detail_summary_total(ss_row, title='Underlying ', base_row=base_row,
                                          back_row=str(ss_lst_row))
        wider = [ssCol(st_cac + 1), ssCol(st_lic + 1)]
        for column_cells in ss.columns:
            length = 0
            value = ''
            for cell in column_cells:
                if str(cell.value)[0] != '=':
                    values = str(cell.value).split('\n')
                    for value in values:
                        if len(value) + 1 > length:
                            length = len(value) + 1
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            if cel in wider:
                ss.column_dimensions[cel].width = max(length, 10) * 1.5
            else:
                ss.column_dimensions[cel].width = max(length, 10) * 1.2

        if corr_data is not None:
            ss_row += 2
            for corr in corr_data:
                ss.cell(row=ss_row, column=1).value = corr[0]
                if len(corr) > 1:
                    ss.cell(row=ss_row, column=2).value = corr[1]
                    ss.cell(row=ss_row, column=2).number_format = '#0.0000'
                    ss.cell(row=ss_row, column=3).value = corr[2]
                ss_row += 1
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Static Variables'
        ss.cell(row=ss_row, column=1).font = bold
        if self.carbon_price > 0:
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'Carbon Price ($/tCO2e)'
            ss.cell(row=ss_row, column=st_cap+1).value = self.carbon_price
            ss.cell(row=ss_row, column=st_cap+1).number_format = '$#,##0.00'
            attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
            carbon_cell = oxl.workbook.defined_name.DefinedName('carbon_price', attr_text=attr_text)
            try: # openpyxl 3.1.x
                wb.defined_names['carbon_price'] = carbon_cell
            except: # openpyxl 3.1.x
                wb.defined_names.append(carbon_cell)
        ss_row += 1
        attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
        lifetime_cell = oxl.workbook.defined_name.DefinedName('lifetime', attr_text=attr_text)
        try:
            wb.defined_names['lifetime'] = lifetime_cell
        except:
            wb.defined_names.append(lifetime_cell)
        ss.cell(row=ss_row, column=1).value = 'Lifetime (years)'
        ss.cell(row=ss_row, column=st_cap+1).value = max_lifetime
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Discount Rate'
        ss.cell(row=ss_row, column=st_cap+1).value = self.discount_rate
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00%'
        ss_row += 2
        ss_row = self.data_sources(ss, ss_row, pm_data_file, option)
        if progress_handler:
            progress = ProgressInfo(
                current=14,
            )
            progress_handler.update(progress)
        for row in range(1, ss_row + 1):
            for col in range(1, len(headers) + 1):
                try:
                    if ss.cell(row=row, column=col).font.name != 'Arial':
                        ss.cell(row=row, column=col).font = normal
                except:
                    pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        if self.save_tables:
            gens = []
            cons = []
            for fac in re_order:
                if fac == 'Load':
                    continue
                if pmss_details[fac].multiplier <= 0:
                    continue
                if fac.find('.') > 0:
                    gens.append(fac[fac.find('.') + 1:])
                else:
                    gens.append(fac)
                cons.append(self.generators[pmss_details[fac].generator].constraint)
            for gen in dispatch_order:
                gens.append(gen)
                cons.append(self.generators[gen].constraint)
            gs = wb.create_sheet(self.sheets[G].currentText())
            fields = []
            col = 1
            row = 1
            if hasattr(self.generators[list(self.generators.keys())[0]], 'name'):
                fields.append('name')
                gs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.generators[list(self.generators.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        txt = prop.replace('_', ' ').title()
                        txt = txt.replace('Cf', 'CF')
                        txt = txt.replace('Lcoe', 'LCOE')
                        txt = txt.replace('Om', 'OM')
                        gs.cell(row=row, column=col).value = txt
                        if prop == 'capex':
                            txt = txt + txt
                        gs.column_dimensions[ssCol(col)].width = max(len(txt) * 1.4, 10)
                        col += 1
            nme_width = 4
            con_width = 4
            for key, value in self.generators.items():
                if key in gens:
                    row += 1
                    col = 1
                    for field in fields:
                        gs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'constraint']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > con_width:
                                    con_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = con_width * 1.4
                        elif field in ['capex', 'fixed_om']:
                            gs.cell(row=row, column=col).number_format = '$#,##0'
                        elif field in ['lcoe', 'variable_om', 'fuel']:
                            gs.cell(row=row, column=col).number_format = '$#,##0.00'
                        elif field in ['disc_rate']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00%'
                        elif field in ['capacity', 'lcoe_cf', 'initial']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00'
                        elif field in ['emissions']:
                            gs.cell(row=row, column=col).number_format = '#,##0.000'
                        elif field in ['lifetime', 'order']:
                            gs.cell(row=row, column=col).number_format = '#,##0'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    gs.cell(row=row, column=col).font = normal
            gs.freeze_panes = 'B2'
            gs.activeCell = 'B2'
            fields = []
            col = 1
            row = 1
            cs = wb.create_sheet(self.sheets[C].currentText())
            if hasattr(self.constraints[list(self.constraints.keys())[0]], 'name'):
                fields.append('name')
                cs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.constraints[list(self.constraints.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        if prop == 'warm_time':
                            cs.cell(row=row, column=col).value = 'Warmup Time'
                        else:
                            cs.cell(row=row, column=col).value = prop.replace('_', ' ').title()
                        cs.column_dimensions[ssCol(col)].width = max(len(prop) * 1.4, 10)
                        col += 1
            nme_width = 4
            cat_width = 4
            for key, value in self.constraints.items():
                if key in cons:
                    row += 1
                    col = 1
                    for field in fields:
                        cs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'category']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > cat_width:
                                    cat_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = cat_width * 1.4
                        elif field == 'warm_time':
                            cs.cell(row=row, column=col).number_format = '#0.00'
                        elif field != 'category':
                            cs.cell(row=row, column=col).number_format = '#,##0%'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    try:
                        cs.cell(row=row, column=col).font = normal
                    except:
                        pass
            cs.freeze_panes = 'B2'
            cs.activeCell = 'B2'
        wb.save(data_file)
        if progress_handler:
            progress = ProgressInfo(
                current=20,
            )
            progress_handler.update(progress)
        j = data_file.rfind('/')
        data_file = data_file[j + 1:]
        msg = '%s created (%.2f seconds)' % (data_file, time.time() - start_time)
        msg = '%s created.' % data_file
        self.setStatus(msg)
        if progress_handler:
            progress_handler.hide()
            progress = ProgressInfo(
                current=0,
            )
            progress_handler.update(progress)

    def setStatus(self, text):
        pass

    def _makeAdjustments(self):
        pass

    def _plot_multi(multi_scores, multi_best, multi_order, title):
        pass
    
    def _display_table(self, objects, fields=None, title=None, save_folder='', sortby=None, decpts=None):
        """
        Placeholder method to be overridden by subclasses.
        Should return raw constraint data as a list of dictionaries.
        """
        raise NotImplementedError("Subclasses must override _display_table.")
        
    def _display_table(self):
        pass