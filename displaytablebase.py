import numpy as np
import os
import pandas as pd
import openpyxl as oxl
from senutils import techClean
from typing import List, Dict, Any, Optional, Union
import xlwt

class TableBase:
    """Base class for table data manipulation and formatting using pandas DataFrame"""
    
    def __init__(self, 
                 objects: Union[List, Dict], 
                 fields: Optional[List[str]] = None,
                 sumby: Optional[Union[str, List[str]]] = None,
                 sumfields: Optional[Union[str, List[str]]] = None,
                 units: str = '',
                 decpts: Optional[List[int]] = None,
                 abbr: bool = True
                 ):
        """
        Initialize TableBase with objects and formatting options
        
        Args:
            objects: Input data as list or dict
            fields: Column fields to display
            sumby: Fields to group by for summation
            sumfields: Fields to sum
            units: Units string for columns
            decpts: Decimal points for each column
            abbr: Use abbreviated numbers
        """
        self.objects = self._convert_to_dataframe(objects, fields)
        self.fields = fields if fields else list(self.objects.columns)
        self.sumby = [sumby] if isinstance(sumby, str) else sumby
        self.sumfields = [sumfields] if isinstance(sumfields, str) else sumfields
        self.units = units
        self.decpts = decpts
        self.abbr = abbr
        self.labels = self._get_column_types()
        self.lens = self._get_column_lengths()
        self.sums = self._calculate_sums() if sumfields else None

    def _convert_to_dataframe(self, objects: Union[List, Dict], fields: Optional[List[str]]) -> pd.DataFrame:
        """Convert input objects to pandas DataFrame"""
        # Pad or trim rows to match the number of fields
        max_length = len(fields)
        normalized_objects = [
            row[:max_length] if len(row) > max_length else row + [''] * (max_length - len(row))
            for row in objects
        ]
        if isinstance(objects, list):
            if len(objects) == 0:
                return pd.DataFrame()
            if isinstance(objects[0], list):
                return pd.DataFrame(normalized_objects, columns=fields)
            return pd.DataFrame([self._object_to_dict(obj) for obj in normalized_objects])
        elif isinstance(objects, dict):
            if fields is None:  # Assume class objects
                first_obj = objects[list(objects.keys())[0]]
                fields = ['name'] if hasattr(first_obj, 'name') else []
                fields.extend([prop for prop in dir(first_obj) 
                             if not prop.startswith('__') and prop != 'name'])
                return pd.DataFrame([self._object_to_dict(normalized_objects, fields) for obj in normalized_objects.values()])
            return pd.DataFrame([[k, v] for k, v in objects.items()], columns=fields)
        
        return pd.DataFrame(normalized_objects)

    def _object_to_dict(self, obj: Any, fields: Optional[List[str]] = None) -> Dict:
        """Convert object attributes to dictionary"""
        if fields is None:
            fields = [attr for attr in dir(obj) if not attr.startswith('__')]
        
        result = {}
        for field in fields:
            value = getattr(obj, field, None)
            if isinstance(value, list):
                value = ' '.join(str(x) for x in value)
            result[field] = value
        return result

    def _get_column_types(self) -> Dict[str, str]:
        """Determine column types (int, float, str)"""
        labels = {}
        for col in self.objects.columns:
            if self.objects[col].dtype in (np.int32, np.int64):
                labels[col] = 'int'
            elif self.objects[col].dtype in (np.float32, np.float64):
                labels[col] = 'float'
            else:
                labels[col] = 'str'
        return labels

    def _get_column_lengths(self) -> Dict[str, List[int]]:
        """Calculate display lengths for columns"""
        lens = {}
        for col in self.objects.columns:
            if self.labels[col] in ('int', 'float'):
                max_val = self.objects[col].max()
                if pd.isna(max_val):
                    lens[col] = [0, 0]
                    continue
                    
                str_val = str(max_val)
                if '.' in str_val:
                    int_part, dec_part = str_val.split('.')
                    lens[col] = [len(int_part), len(dec_part)]
                else:
                    lens[col] = [len(str_val), 0]
            else:
                max_len = self.objects[col].astype(str).str.len().max()
                lens[col] = [max_len if not pd.isna(max_len) else 0, 0]
        return lens

    def _calculate_sums(self) -> Dict[str, List[float]]:
        """Calculate sums and percentages for specified columns"""
        sums = {}
        
        if not self.sumfields:
            return sums
            
        # Calculate total sums
        total_sums = self.objects[self.sumfields].sum().tolist()
        sums['~~'] = total_sums
        
        # Calculate group sums if sumby specified
        if self.sumby:
            groups = self.objects.groupby(self.sumby[0])[self.sumfields].sum()
            for group_name, group_sums in groups.iterrows():
                sums[group_name] = group_sums.tolist()
                
        return sums

    def nice(self, string):
        try:
            self.hdrs
        except:
            self.hdrs = {}
        if string == '':
            strout = string
        else:
            strout = techClean(string, full=True)
            if string != '' and string in self.units:
                i = self.units.find(string)
                j = self.units.find(' ', i)
                if j < 0:
                    j = len(self.units)
                strout = strout + ' (' + self.units[i + 1 + len(string):j] + ')'
        self.hdrs[strout] = string
        return strout
    
    def format_value(self, value: Any, col: str) -> str:
        """Format value for display based on column type and settings"""
        if pd.isna(value):
            return ''
            
        if self.labels[col] in ('int', 'float'):
            if isinstance(value, str):
                try:
                    value = float(value.replace(',', ''))
                    if self.labels[col] == 'int':
                        value = int(value)
                except ValueError:
                    return value
                    
            if value == 0:
                return ''
                
            if self.abbr:
                if value >= 1000000:
                    return f"{value/1000000:.1f}M"
                elif value >= 1000:
                    return f"{value/1000:.1f}K"
                    
            dec_places = self.decpts[self.fields.index(col)] if self.decpts else self.lens[col][1]
            fmt = f"{{:,.{dec_places}f}}"
            return fmt.format(value)
            
        return str(value)

    def sort_by_column(self, col_idx: int, ascending: bool = True) -> None:
        """Sort objects by specified column"""
        if col_idx >= 0:
            sort_col = self.fields[col_idx]
            self.objects.sort_values(by=sort_col, ascending=ascending, inplace=True)

    def get_formatted_data(self) -> List[Dict[str, str]]:
        """Get formatted data for display"""
        formatted = []
        for _, row in self.objects.iterrows():
            formatted_row = {}
            for col in self.fields:
                formatted_row[col] = self.format_value(row[col], col)
            formatted.append(formatted_row)
        return formatted

    def export_to_excel(self, filename: str) -> None:
        """Export data to Excel file"""
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        
        # Format the data
        formatted_data = self.get_formatted_data()
        df = pd.DataFrame(formatted_data)
        
        # Write to Excel with formatting
        df.to_excel(writer, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[writer.sheets.keys()[0]]
        
        # Apply formatting (column widths, number formats, etc.)
        for idx, col in enumerate(df.columns):
            max_length = max(df[col].astype(str).apply(len).max(), len(col))
            worksheet.column_dimensions[oxl.utils.get_column_letter(idx + 1)].width = max_length + 2
            
        writer.save()

    def save_as(self, data_file, table) -> None:
        self.table = table
        if data_file[-4:] == '.csv' or data_file[-4:] == '.xls' or data_file[-5:] == '.xlsx':
            pass
        else:
            data_file += '.xlsx'
        if os.path.exists(data_file):
            if os.path.exists(data_file + '~'):
                os.remove(data_file + '~')
            os.rename(data_file, data_file + '~')
        if data_file[-4:] == '.csv':
            tf = open(data_file, 'w')
            hdr_types = []
            line = ''
            for cl in range(self.table.columnCount()):
                if cl > 0:
                    line += ','
                hdr = self.table.horizontalHeaderItem(cl).text()
                if hdr[0] != '%':
                    txt = hdr
                    if ',' in txt:
                        line += '"' + txt + '"'
                    else:
                        line += txt
                txt = self.hdrs[hdr]
                try:
                    hdr_types.append(self.labels[txt.lower()])
                except:
                    hdr_types.append(self.labels[txt])
            tf.write(line + '\n')
            for rw in range(self.table.rowCount()):
                line = ''
                for cl in range(self.table.columnCount()):
                    if cl > 0:
                        line += ','
                    if self.table.item(rw, cl) is not None:
                        txt = self.table.item(rw, cl).text()
                        if hdr_types[cl] == 'int':
                            try:
                                txt = self.table.item(rw, cl).text().strip()
                            except:
                                pass
                        elif hdr_types[cl] == 'float':
                            try:
                                txt = self.table.item(rw, cl).text().strip()
                                txt = txt.replace(',', '')
                            except:
                                pass
                        if ',' in txt:
                            line += '"' + txt + '"'
                        else:
                            line += txt
                tf.write(line + '\n')
            tf.close()
        elif data_file[-4:] == '.xls':
            wb = xlwt.Workbook()
            for ch in ['\\' , '/' , '*' , '?' , ':' , '[' , ']']:
                if ch in iam:
                    iam = iam.replace(ch, '_')
            if len(iam) > 31:
                iam = iam[:31]
            ws = wb.add_sheet(iam)
            hdr_types = []
            dec_fmts = []
            xl_lens = []
            hdr_rows = 0
            hdr_style = xlwt.XFStyle()
            hdr_style.alignment.wrap = 1
            for cl in range(self.table.columnCount()):
                hdr = self.table.horizontalHeaderItem(cl).text()
                if hdr[0] != '%':
                    ws.write(0, cl, hdr, hdr_style)
                txt = self.hdrs[hdr]
                try:
                    hdr_types.append(self.labels[txt.lower()])
                    txt = txt.lower()
                except:
                    try:
                        hdr_types.append(self.labels[txt])
                    except:
                        hdr_types.append('str')
                style = xlwt.XFStyle()
                try:
                    if self.lens[txt][1] > 0:
                        style.num_format_str = '#,##0.' + '0' * self.lens[txt][1]
                    elif self.labels[txt] == 'int' or self.labels[txt] == 'float':
                        style.num_format_str = '#,##0'
                except:
                    pass
                dec_fmts.append(style)
                bits = hdr.split('\n')
                hdr_rows = max(hdr_rows, len(bits))
                hl = 0
                for bit in bits:
                    hl = max(hl, len(bit) + 1)
                xl_lens.append(hl)
            if hdr_rows > 1:
                ws.row(0).height = 250 * hdr_rows
            in_span = False
            for rw in range(self.table.rowCount()):
                for cl in range(self.table.columnCount()):
                    if self.table.item(rw, cl) is not None:
                        valu = self.table.item(rw, cl).text().strip()
                        if len(valu) < 1:
                            continue
                        if self.span is not None and valu == self.span:
                            in_span = True
                        style = dec_fmts[cl]
                        if valu[-1] == '%':
                            is_pct = True
                            i = valu.rfind('.')
                            if i >= 0:
                                dec_pts = (len(valu) - i - 2)
                                style = xlwt.XFStyle()
                                try:
                                    style.num_format_str = '#,##0.' + '0' * dec_pts + '%'
                                except:
                                    pass
                            else:
                                 dec_pts = 0
                                 style.num_format_str = '#,##0%'
                        else:
                            is_pct = False
                        if hdr_types[cl] == 'int':
                            try:
                                val1 = valu
                                if is_pct:
                                    val1 = val1.strip('%')
                                val1 = val1.replace(',', '')
                                if is_pct:
                                    valu = round(int(val1) / 100., dec_pts + 2)
                                else:
                                    valu = int(val1)
                            except:
                                pass
                        elif hdr_types[cl] == 'float':
                            try:
                                val1 = valu
                                if is_pct:
                                    val1 = val1.strip('%')
                                val1 = val1.replace(',', '')
                                if is_pct:
                                    valu = round(float(val1) / 100., dec_pts + 2)
                                else:
                                    valu = float(val1)
                            except:
                                pass
                        else:
                            if is_pct:
                                try:
                                    val1 = valu.strip('%')
                                    val1 = val1.replace(',', '')
                                    valu = round(float(val1) / 100., dec_pts + 2)
                                except:
                                    pass
                        if not in_span:
                            xl_lens[cl] = max(xl_lens[cl], len(str(valu)))
                        ws.write(rw + 1, cl, valu, style)
            for cl in range(self.table.columnCount()):
                if xl_lens[cl] * 275 > ws.col(cl).width:
                    ws.col(cl).width = xl_lens[cl] * 275
            ws.set_panes_frozen(True)   # frozen headings instead of split panes
            ws.set_horz_split_pos(1)   # in general, freeze after last heading row
            ws.set_remove_splits(True)   # if user does unfreeze, don't leave a split there
            wb.save(data_file)
        else: # .xlsx
            wb = oxl.Workbook()
            ws = wb.active
            for ch in ['\\' , '/' , '*' , '?' , ':' , '[' , ']']:
                if ch in iam:
                    iam = iam.replace(ch, '_')
            if len(iam) > 31:
                iam = iam[:31]
            ws.title = iam
            normal = oxl.styles.Font(name='Arial', size='10')
        #    bold = oxl.styles.Font(name='Arial', bold=True)
            hdr_types = []
            dec_fmts = []
            xl_lens = []
            hdr_rows = 0
        #    hdr_style = xlwt.XFStyle()
        #    hdr_style.alignment.wrap = 1
            for cl in range(self.table.columnCount()):
                hdr = self.table.horizontalHeaderItem(cl).text()
                if hdr[0] != '%':
                    ws.cell(row=1, column=cl + 1).value = hdr
                    ws.cell(row=1, column=cl + 1).font = normal
                    ws.cell(row=1, column=cl + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                txt = self.hdrs[hdr]
                try:
                    hdr_types.append(self.labels[txt.lower()])
                    txt = txt.lower()
                except:
                    try:
                        hdr_types.append(self.labels[txt])
                    except:
                        hdr_types.append('str')
                style = ''
                try:
                    if self.lens[txt][1] > 0:
                        style = '#,##0.' + '0' * self.lens[txt][1]
                    elif self.labels[txt] == 'int' or self.labels[txt] == 'float':
                        style = '#,##0'
                except:
                    pass
                dec_fmts.append(style)
                bits = hdr.split('\n')
                hdr_rows = max(hdr_rows, len(bits))
                hl = 0
                for bit in bits:
                    hl = max(hl, len(bit) + 1)
                xl_lens.append(hl)
            if hdr_rows > 1:
                ws.row_dimensions[1].height = 12 * hdr_rows
            in_span = False
            for rw in range(self.table.rowCount()):
                for cl in range(self.table.columnCount()):
                    if self.table.item(rw, cl) is not None:
                        valu = self.table.item(rw, cl).text().strip()
                        if len(valu) < 1:
                            continue
                        if self.span is not None and valu == self.span:
                            in_span = True
                        style = dec_fmts[cl]
                        if valu[-1] == '%':
                            is_pct = True
                            i = valu.rfind('.')
                            if i >= 0:
                                dec_pts = (len(valu) - i - 2)
                                style = '#,##0.' + '0' * dec_pts + '%'
                            else:
                                dec_pts = 0
                                style = '#,##0%'
                        else:
                            is_pct = False
                        if hdr_types[cl] == 'int':
                            try:
                                val1 = valu
                                if is_pct:
                                    val1 = val1.strip('%')
                                val1 = val1.replace(',', '')
                                if is_pct:
                                    valu = round(int(val1) / 100., dec_pts + 2)
                                else:
                                    valu = int(val1)
                            except:
                                pass
                        elif hdr_types[cl] == 'float':
                            try:
                                val1 = valu
                                if is_pct:
                                    val1 = val1.strip('%')
                                val1 = val1.replace(',', '')
                                if is_pct:
                                    valu = round(float(val1) / 100., dec_pts + 2)
                                else:
                                    valu = float(val1)
                            except:
                                pass
                        else:
                            if is_pct:
                                try:
                                    val1 = valu.strip('%')
                                    val1 = val1.replace(',', '')
                                    valu = round(float(val1) / 100., dec_pts + 2)
                                except:
                                    pass
                        if not in_span:
                            if is_pct:
                                plus = 3
                            else:
                                plus = 0
                            xl_lens[cl] = max(xl_lens[cl], len(str(valu)) + plus)
                        ws.cell(row=rw + 2, column=cl + 1).value = valu
                        ws.cell(row=rw + 2, column=cl + 1).font = normal
                        ws.cell(row=rw + 2, column=cl + 1).number_format = style
            for cl in range(self.table.columnCount()):
                ws.column_dimensions[ssCol(cl + 1)].width = xl_lens[cl]
            ws.freeze_panes = 'A2'
            wb.save(data_file)
            wb.close()
        self.savedfile = data_file
        if not self.edit_table:
            self.close()