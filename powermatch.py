#!/usr/bin/python3
#
#  Copyright (C) 2018-2025 Sustainable Energy Now Inc., Angus King
#
#  powermatch.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#
# Note: Batch process is all rather messy.
from copy import copy
import os
import sys
import datetime
import time
from PyQt5 import QtCore, QtGui, QtWidgets
import displayobject
import displaytable
from credits import fileVersion
import glob
from math import log10
import matplotlib
if matplotlib.__version__ > '3.5.1':
    matplotlib.use('Qt5Agg')
else:
    matplotlib.use('TkAgg') # so PyQT5 and Matplotlib windows don't interfere
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
# This import registers the 3D projection, but is otherwise unused.
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.ticker import LinearLocator, FormatStrFormatter
import numpy as np
import openpyxl as oxl
from openpyxl.chart import (
    BarChart,
    LineChart,
    Reference,
    Series
)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
import random
import shutil
import subprocess
from colours import PlotPalette
from senutils import ClickableQLabel, getParents, getUser, ListWidget, setFontSize, ssCol, techClean, WorkBook
from editini import EdtDialog, SaveIni
from floaters import ProgressBar, FloatStatus
from getmodels import getModelFile, commonprefix
import configparser  # decode .ini file
from zoompan import ZoomPanX
try:
    from opt_debug import optimiseDebug
except:
    pass

tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
# initialise tech_names from .ini file
#            add dispatchable for re from [Grid] dispatchable?
# load data file. If not in data file then include in order and flag as RE
# tracking_pv is a synonym form dual_axis_pv
# phes is a synonym for pumped_hydro
# other1 is a synonym for other - or the other way around
# [Grid]
# dispatchable=pumped_hydro geothermal biomass solar_thermal cst
# consider: hydrogen bess
# [Power]
# technologies=backtrack_pv bess biomass cst fixed_pv geothermal offshore_wind rooftop_pv single_axis_pv solar_thermal tracking_pv wave wind other other_wave
#              add pumped_hydro hydrogen
#              maybe drop bess?
# fossil_technologies=fossil_ccgt fossil_coal fossil_cogen fossil_distillate fossil_gas fossil_mixed fossil_ocgt
target_keys = ['lcoe', 'load_pct', 'surplus_pct', 're_pct', 'cost', 'co2']
target_names = ['LCOE', 'Load%', 'Surplus%', 'RE%', 'Cost', 'CO2']
target_fmats = ['$%.2f', '%.1f%%', '%.1f%%', '%.1f%%', '$%.1fpwr_chr', '%.1fpwr_chr']
target_titles = ['LCOE ($)', 'Load met %', 'Surplus %', 'RE %', 'Total Cost ($)', 'tCO2e']
headers = ['Facility', 'Capacity\n(Gen, MW;\nStor, MWh)', 'To meet\nLoad (MWh)',
           'Subtotal\n(MWh)', 'CF', 'Cost ($/yr)', 'LCOG\nCost\n($/MWh)', 'LCOE\nCost\n($/MWh)',
           'Emissions\n(tCO2e)', 'Emissions\nCost', 'LCOE With\nCO2 Cost\n($/MWh)', 'Max.\nMWH',
           'Max.\nBalance', 'Capital\nCost', 'Lifetime\nCost', 'Lifetime\nEmissions',
           'Lifetime\nEmissions\nCost', 'Area (km^2)', 'Job\nYears', 'Reference\nLCOE', 'Reference\nCF']
# set up columns for summary table. Hopefully to make it easier to add / alter columns
st_fac = 0 # Facility
st_cap = 1 # Capacity\n(Gen, MW;\nStor, MWh)
st_tml = 2 # To meet\nLoad (MWh)
st_sub = 3 # Subtotal\n(MWh)
st_cfa = 4 # CF
st_cst = 5 # Cost ($/yr)
st_lcg = 6 # LCOG\nCost\n($/MWh)
st_lco = 7 # LCOE\nCost\n($/MWh)
st_emi = 8 # Emissions\n(tCO2e)
st_emc = 9 # Emissions\nCost
st_lcc = 10 # LCOE With\nCO2 Cost\n($/MWh)
st_max = 11 # Max.\nMWH
st_bal = 12 # Max.\nBalance'
st_cac = 13 # Capital\nCost'
st_lic = 14 # Lifetime\nCost'
st_lie = 15 # Lifetime\nEmissions
st_lec = 16 # Lifetime\nEmissions\nCost
st_are = 17 # Areas (km^2)
st_job = 18 # Job (years)
st_rlc = 19 # Reference\nLCOE
st_rcf = 20 # Reference\nCF

# same order as self.file_labels
C = 0 # Constraints - xls or xlsx
G = 1 # Generators - xls or xlsx
O = 2 # Optimisation - xls or xlsx
J = 3 # Jobs - xls or xlsx
D = 4 # Data - xlsx
R = 5 # Results - xlsx
B = 6 # Batch input - xlsx
T = 7 # Transition input - xlsx
S = 'S' # Summary
O1 = 'O1'
def get_value(ws, row, col):
    def get_range(text, alphabet=None, base=1):
        if len(text) < 1:
            return None
        if alphabet is None:
            alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if alphabet[0] == ' ':
            alphabet = alphabet[1:]
        alphabet = alphabet.upper()
        bits = ['', '']
        b = 0
        in_char = True
        for char in text:
            if char.isdigit():
                if in_char:
                    in_char = False
                    b += 1
            else:
                if alphabet.find(char.upper()) < 0:
                    continue
                if not in_char:
                    in_char = True
                    b += 1
            if b >= len(bits):
                break
            bits[b] += char.upper()
        try:
            bits[1] = int(bits[1]) - (1 - base)
        except:
            pass
        row = 0
        ndx = 1
        for c in range(len(bits[0]) -1, -1, -1):
            ndx1 = alphabet.index(bits[0][c]) + 1
            row = row + ndx1 * ndx
            ndx = ndx * len(alphabet)
        bits[0] = row - (1 - base)
        for c in bits:
            if c == '':
                return None
    try:
        while ws.cell(row=row, column=col).value[0] == '=':
            row, col = get_range(ws.cell(row=row, column=col).value)
    except:
        return ''
    return ws.cell(row=row, column=col).value

def add_to_batch(wb, batch_file, adj_list):
    def copy_format(cell, new_cell):
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)

    check_list = list(adj_list.keys())
    # remove not of interest
    strt = len(check_list) - 1
    try:
        strt = check_list.index('Optimisation Parameters') - 1
    except:
        pass
    ignore = True
    for i in range(strt, -1, -1):
        if check_list[i] == 'Discount Rate':
            continue
        elif check_list[i] == 'Load':
            if adj_list['Load'] > 10000:
                adj_list['Load'] = round(adj_list['Load'], -4)
            continue
    #    elif check_list[i][:10] == 'Total Load':
     #       adj_list['Load'] = adj_list[check_list[i]]
      #      del adj_list[check_list[i]]
       #     check_list[i] = 'Load'
        #    continue
        elif check_list[i] == 'Carbon Price':
            continue
        elif check_list[i] == 'Carbon Price ($/tCO2e)':
            adj_list['Carbon Price'] = adj_list[check_list[i]]
            del adj_list['Carbon Price ($/tCO2e)']
            check_list[i] = 'Carbon Price'
            continue
        elif check_list[i] == 'Underlying Total':
            del check_list[i]
            ignore = False
            continue
        elif check_list[i] == 'Additional Underlying Load':
            del check_list[i]
            ignore = True
            continue
        elif check_list[i] == 'Total':
            ignore = False
            continue
        if ignore:
            del check_list[i]
    batch_input_sheet = wb.worksheets[0]
    batch_input_sheet.protection.sheet = False
    normal = oxl.styles.Font(name='Arial')
    bold = oxl.styles.Font(name='Arial', bold=True)
    col = batch_input_sheet.max_column + 1
    fst_row = -1
    the_rows = [-1, -1, -1, -1, -1, -1]
    tot_row = 0
    disc_row = 1
    cp_row = 2
    load_row = 3
    opt_row = 4
    max_row = 5
    save_opt_rows = False
    if col == 4: # possibly only chart stuff in columns 2 and 3
        get_out = False
        for col in range(3, 1, -1):
            for row in range(1, batch_input_sheet.max_row + 1):
                if batch_input_sheet.cell(row=row, column=col).value is not None:
                    col += 1
                    get_out = True
                    break
                if batch_input_sheet.cell(row=row, column=1).value == 'Total':
                    break
            if get_out:
                break
    for row in range(1, batch_input_sheet.max_row + 1):
        if batch_input_sheet.cell(row=row, column=1).value is None:
            continue
        if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology', 'Year']:
            new_cell = batch_input_sheet.cell(row=row, column=col)
            new_cell.value = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'MM-dd hh:mm')
            add_msg = f"Added to batch as '{new_cell.value}' (column {ssCol(col)})"
            continue
        if batch_input_sheet.cell(row=row, column=1).value == 'Capacity (MW)':
            fst_row = row + 1
            cell = batch_input_sheet.cell(row=row, column=2)
            new_cell = batch_input_sheet.cell(row=row, column=col)
            new_cell.value = 'MW'
            if cell.has_style:
                copy_format(cell, new_cell)
            continue
        if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
            try:
                i = check_list.index('Optimisation Parameters')
                if i >= 0:
                    the_rows[opt_row] = row
                    save_opt_rows = True
                    opt_list = check_list[i:]
            except:
                pass
            break
        for key, value in adj_list.items():
            if key == batch_input_sheet.cell(row=row, column=1).value:
                cell = batch_input_sheet.cell(row=row, column=2)
                new_cell = batch_input_sheet.cell(row=row, column=col)
                new_cell.value = value
                if cell.has_style:
                    copy_format(cell, new_cell)
                    if col == 2:
                        new_cell.font = normal
                        new_cell.number_format = '#0.00'
                    else:
                        new_cell.number_format = copy(cell.number_format)
                elif col == 2:
                    new_cell.font = normal
                    new_cell.number_format = '#0.00'
                try:
                    i = check_list.index(key)
                    del check_list[i]
                except:
                    pass
        if batch_input_sheet.cell(row=row, column=1).value == 'Total':
            the_rows[tot_row] = row
        elif batch_input_sheet.cell(row=row, column=1).value == 'Discount Rate':
            the_rows[disc_row] = row
        elif batch_input_sheet.cell(row=row, column=1).value == 'Carbon Price':
            the_rows[cp_row] = row
        elif batch_input_sheet.cell(row=row, column=1).value == 'Load':
            the_rows[load_row] = row
    for r in range(len(the_rows)):
        if r == max_row:
            continue
        the_rows[max_row] = max(the_rows[max_row], the_rows[r])
    the_rows[max_row] += 1
    if the_rows[load_row] < 0:
        try:
            i = check_list.index('Load')
            del check_list[i]
        except:
            pass
    try:
        i = check_list.index('Optimisation Parameters')
        if i >= 0:
            del check_list[i:]
    except:
        pass
    if len(check_list) > 0:
        if col > 2:
            cell = batch_input_sheet.cell(row=fst_row, column=2)
        else:
            cell = batch_input_sheet.cell(row=fst_row, column=col)
   #     cell = batch_input_sheet.cell(row=fst_row, column=col)
        for key in check_list:
            if adj_list[key] == 0:
                continue
            if key in ['Discount Rate', 'Carbon Price', 'Load']:
                new_row = the_rows[max_row]
                r1 = 1
            else:
                new_row = the_rows[tot_row] - 1 # to keep SUM formula intact
                r1 = 0
            batch_input_sheet.insert_rows(new_row)
            new_cell = batch_input_sheet.cell(row=new_row, column=1)
            new_cell.value = key
            new_cell = batch_input_sheet.cell(row=new_row, column=col)
            new_cell.value = adj_list[key]
            if cell.has_style:
                copy_format(cell, new_cell)
            for r in range(r1, len(the_rows)):
                if the_rows[r] >= 0:
                    the_rows[r] += 1
    if fst_row > 0 and the_rows[tot_row] > 0:
        new_cell = batch_input_sheet.cell(row=the_rows[tot_row], column=col)
        new_cell.value = '=SUM(' + ssCol(col) + str(fst_row) + ':' + ssCol(col) + str(the_rows[tot_row] - 1) + ')'
        if col > 2:
            cell = batch_input_sheet.cell(row=the_rows[tot_row], column=2)
        else:
            cell = batch_input_sheet.cell(row=the_rows[tot_row], column=col)
        if cell.has_style:
            copy_format(cell, new_cell)
    if save_opt_rows: # want optimisation?
        for row in range(the_rows[opt_row], batch_input_sheet.max_row + 1):
            for key in opt_list:
                if key == batch_input_sheet.cell(row=row, column=1).value:
                    cell = batch_input_sheet.cell(row=fst_row, column=col - 1)
                    new_cell = batch_input_sheet.cell(row=row, column=col)
                    new_cell.value = adj_list[key]
                    if cell.has_style:
                        copy_format(cell, new_cell)
                    try:
                        i = opt_list.index(key)
                        del opt_list[i]
                    except:
                        pass
        if len(opt_list) > 0:
            for key in opt_list:
                row += 1
                new_cell = batch_input_sheet.cell(row=row, column=1)
                new_cell.value = key
                new_cell = batch_input_sheet.cell(row=row, column=col)
                try:
                    new_cell.value = adj_list[key]
                except:
                    pass
    wb.save(batch_file)
    return add_msg


class MyQDialog(QtWidgets.QDialog):
    ignoreEnter = True

    def keyPressEvent(self, qKeyEvent):
        if qKeyEvent.key() == QtCore.Qt.Key_Return:
            self.ignoreEnter = True
        else:
            self.ignoreEnter = False

    def closeEvent(self, event):
        if self.ignoreEnter:
            self.ignoreEnter = False
            event.ignore()
        else:
            event.accept()

 #   def resizeEvent(self, event):
  #      return super(Window, self).resizeEvent(event)


class Constraint:
    def __init__(self, name, category,
                 capacity_min, capacity_max,
                 discharge_loss, discharge_max, discharge_start,
                 min_run_time, parasitic_loss,
                 rampdown_max, rampup_max,
                 recharge_loss, recharge_max, recharge_start,
                 warm_time):
        #after name and category all variables are passed in alphabetical order
        # Constraint(key, '<category>', 0., 1., 0, 1., 0, 0., 1., 1., 0., 1., 0., 0, 0)
        self.name = name.strip()
        self.category = category
        try:
            self.capacity_min = float(capacity_min) # minimum run_rate for generator; don't drain below for storage
        except:
            self.capacity_min = 0.
        try:
            self.capacity_max = float(capacity_max) # maximum run_rate for generator; don't drain more than this for storage
        except:
            self.capacity_max = 1.
        try:
            self.discharge_loss = float(discharge_loss)
        except:
            self.discharge_loss = 0.
        try:
            self.discharge_max = float(discharge_max) # can't discharge more than this per hour
        except:
            self.discharge_max = 1.
        try:
            if isinstance(discharge_start, datetime.time):
                self.discharge_start = discharge_start.hour
            else:
                self.discharge_start = float(discharge_start)
                if self.discharge_start >= 1:
                    self.discharge_start = int(self.discharge_start)
                else:
                    self.discharge_start = int((self.discharge_start + 1/3600) * 24)
        except:
            self.discharge_start = 0
        try:
            self.min_run_time = int(min_run_time)
        except:
            self.min_run_time = 0
        try:
            self.parasitic_loss = float(parasitic_loss) # daily parasitic loss / hourly ?
        except:
            self.parasitic_loss = 0.
        try:
            self.rampdown_max = float(rampdown_max)
        except:
            self.rampdown_max = 1.
        try:
            self.rampup_max = float(rampup_max)
        except:
            self.rampup_max = 1.
        try:
            self.recharge_loss = float(recharge_loss)
        except:
            self.recharge_loss = 0.
        try:
            self.recharge_max = float(recharge_max) # can't charge more than this per hour
        except:
            self.recharge_max = 1.
        try:
            if isinstance(recharge_start, datetime.time):
                self.recharge_start = recharge_start.hour
            else:
                self.recharge_start = float(recharge_start)
                if self.recharge_start >= 1:
                    self.recharge_start = int(self.recharge_start)
                else:
                    self.recharge_start = int((self.recharge_start + 1/3600) * 24)
        except:
            self.recharge_start = 0
        try:
            self.warm_time = float(warm_time)
            if self.warm_time >= 1:
                self.warm_time = self.warm_time / 60
                if self.warm_time > 1:
                    self.warm_time = 1
            elif self.warm_time > 0:
                if self.warm_time <= 1 / 24.:
                    self.warm_time = self.warm_time * 24
        except:
            self.warm_time = 0


class Facility:
    def __init__(self, **kwargs):
        kwargs = {**kwargs}
      #  return
        self.name = ''
        self.constraint = ''
        self.order = 0
        self.lifetime = 20
        self.area = None
        for attr in ['capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial', 'capex',
                     'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']:
            setattr(self, attr, 0.)
        for key, value in kwargs.items():
            if value != '' and value is not None and value != '#N/A':
                if key == 'lifetime' and value == 0:
                    setattr(self, key, 20)
                else:
                    setattr(self, key, value)

class PM_Facility:
    def __init__(self, name, generator, capacity, fac_type, col, multiplier):
        self.name = name
        if name.find('.') > 0:
            self.zone = name[:name.find('.')]
        else:
            self.zone = ''
        self.generator = generator
        self.capacity = capacity
        self.fac_type = fac_type
        self.col = col
        self.multiplier = multiplier


class Optimisation:
    def __init__(self, name, approach, values): #capacity=None, cap_min=None, cap_max=None, cap_step=None, caps=None):
        self.name = name.strip()
        self.approach = approach
        if approach == 'Discrete':
            caps = values.split()
            self.capacities = []
            cap_max = 0.
            for cap in caps:
                try:
                    self.capacities.append(float(cap))
                    cap_max += float(cap)
                except:
                    pass
            self.capacity_min = 0
            self.capacity_max = round(cap_max, 3)
            self.capacity_step = None
        elif approach == 'Range':
            caps = values.split()
            try:
                self.capacity_min = float(caps[0])
            except:
                self.capacity_min = 0.
            try:
                self.capacity_max = float(caps[1])
            except:
                self.capacity_max = 0.
            try:
                self.capacity_step = float(caps[2])
            except:
                self.capacity_step = 0.
            self.capacities = None
        else:
            self.capacity_min = 0.
            self.capacity_max = 0.
            self.capacity_step = 0.
            self.capacities = None
        self.capacity = 0.


class JobFactors():
    def __init__(self, **kwargs):
        kwargs = {**kwargs}
      #  return
        self.name = ''
        for attr in ['years', 'local_pct', 'manufacture', 'install', 'operate', 'dismantle']:
            setattr(self, attr, 0.)
        for key, value in kwargs.items():
            if value != '' and value is not None:
                setattr(self, key, value)


class Adjustments(MyQDialog):
    def setAdjValueUnits(self, key, typ, capacity):
        if key != 'Load':
            mw = capacity
            if typ == 'S':
                unit = 'MWh'
            else:
                unit = 'MW'
            dp = self._decpts
            div = 0
        else:
            dimen = log10(capacity)
            unit = 'MWh'
            if dimen > 11:
                unit = 'PWh'
                div = 9
            elif dimen > 8:
                unit = 'TWh'
                div = 6
            elif dimen > 5:
                unit = 'GWh'
                div = 3
            else:
                div = 0
            mw = capacity / pow(10, div)
            dp = None
        mwtxt = unit
        mwcty = round(mw, dp)
        return mw, mwtxt, mwcty, div

    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 65 + ctr * 32
        width = window.frameSize().width()
        screen = QtWidgets.QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, data, adjustin, adjust_cap, prefix, show_multipliers=False, save_folder=None,
                 batch_file=None):
        super(Adjustments, self).__init__()
        self.ignoreEnter = False
        self._adjust_typ = {} # facility type = G, S or L
        self._adjust_cty = {} # (actual) adjust capacity
        self.show_multipliers = show_multipliers
        if self.show_multipliers:
            self._adjust_mul = {} # (actual) adjust multiplier
            self._adjust_rnd = {} # multiplier widget (rounded to 4 digits)
        self._adjust_txt = {} # string with capacity units
        self._save_folder = save_folder
        self._batch_file = None
        if batch_file is not None:
            if os.path.isfile(batch_file):
                self._batch_file = batch_file
        self._ignore = False
        self._results = None
        self._reset_last = False
        self.grid = QtWidgets.QGridLayout()
        self._data = {}
        ctr = 0
        self._decpts = 1
        for key, typ, capacity in data:
            if key == 'Load' or capacity is None:
                continue
            dimen = log10(capacity)
            if dimen < 2.:
                if dimen < 1.:
                    self._decpts = 2
                elif self._decpts != 2:
                    self._decpts = 1
        if prefix is not None:
            self.grid.addWidget(QtWidgets.QLabel('Results Prefix:'), ctr, 0)
            self.pfx_fld = QtWidgets.QLineEdit()
            self.pfx_fld.setText(prefix)
            self.grid.addWidget(self.pfx_fld, ctr, 1, 1, 2)
            ctr += 1
        # Note: relies on Load being first entry
        for key, typ, capacity in data:
            self._adjust_typ[key] = typ
            if key != 'Load' and capacity is None:
                continue
       #     if key not in adjustin.keys():
       #         continue
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, typ, adjustin[key])
            except:
                mw = 0
                mwtxt = 'MW'
                mwcty = 0
                div = 0
            self._data[key] = [capacity / pow(10, div), div]
            self._adjust_cty[key] = QtWidgets.QDoubleSpinBox()
            self._adjust_cty[key].setRange(0, capacity / pow(10, div) * adjust_cap)
            self._adjust_cty[key].setDecimals(self._decpts)
            if self.show_multipliers:
                self._adjust_rnd[key] = QtWidgets.QDoubleSpinBox()
                self._adjust_rnd[key].setRange(0, adjust_cap)
                self._adjust_rnd[key].setDecimals(4)
            if key in adjustin.keys():
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    try:
                        self._adjust_mul[key] = adjustin[key] / capacity
                        self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
                    except:
                        self._adjust_mul[key] = 1.
                        self._adjust_rnd[key].setValue(1.)
            else:
                self._adjust_cty[key].setValue(0)
                if self.show_multipliers:
                    self._adjust_mul[key] = 0.
                    self._adjust_rnd[key].setValue(0.)
            self._adjust_cty[key].setObjectName(key)
            self.grid.addWidget(QtWidgets.QLabel(key), ctr, 0)
            self.grid.addWidget(self._adjust_cty[key], ctr, 1)
            self._adjust_txt[key] = QtWidgets.QLabel('')
            self._adjust_txt[key].setObjectName(key + 'label')
            self._adjust_txt[key].setText(mwtxt)
            self.grid.addWidget(self._adjust_txt[key], ctr, 2)
            if self.show_multipliers:
                self._adjust_cty[key].valueChanged.connect(self.adjustCap)
                self._adjust_rnd[key].setSingleStep(.1)
                self._adjust_rnd[key].setObjectName(key)
                self.grid.addWidget(self._adjust_rnd[key], ctr, 3)
                self._adjust_rnd[key].valueChanged.connect(self.adjustMult)
            ctr += 1
            if key == 'Load' and len(data) > 1:
                self.grid.addWidget(QtWidgets.QLabel('Facility'), ctr, 0)
                self.grid.addWidget(QtWidgets.QLabel('Capacity'), ctr, 1)
                if self.show_multipliers:
                    self.grid.addWidget(QtWidgets.QLabel('Multiplier'), ctr, 3)
                ctr += 1
        if prefix is None: # batch option
            note = QtWidgets.QLabel('If the input worksheet (capacity figures) contain formulae you may need to open the worksheet and save it before proceeding')
            msg_palette = QtGui.QPalette()
            msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
            note.setPalette(msg_palette)
            note.setWordWrap(True)
            self.grid.addWidget(note, ctr, 0, 2, 4)
            ctr += 2
        quit = QtWidgets.QPushButton('Quit', self)
        self.grid.addWidget(quit, ctr, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        self.grid.addWidget(show, ctr, 1)
        show.clicked.connect(self.showClicked)
        if prefix is not None:
            reset = QtWidgets.QPushButton('Reset', self)
            self.grid.addWidget(reset, ctr, 2)
            reset.clicked.connect(self.resetClicked)
        resetload = QtWidgets.QPushButton('Reset Load', self)
        self.grid.addWidget(resetload, ctr, 3)
        resetload.clicked.connect(self.resetloadClicked)
        if save_folder is not None:
            ctr += 1
            save = QtWidgets.QPushButton('Save', self)
            self.grid.addWidget(save, ctr, 0)
            save.clicked.connect(self.saveClicked)
            restore = QtWidgets.QPushButton('Restore', self)
            self.grid.addWidget(restore, ctr, 1)
            restore.clicked.connect(self.restoreClicked)
            listi = QtWidgets.QPushButton('List', self)
            self.grid.addWidget(listi, ctr, 2)
            listi.clicked.connect(self.listClicked)
            if self._batch_file is not None:
                batch = QtWidgets.QPushButton('Add to Batch', self)
                self.grid.addWidget(batch, ctr, 3)
                batch.clicked.connect(self.addtoBatch)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.niceSize(ctr)
        self.setWindowTitle('SIREN - Powermatch - Adjust generators')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def adjustMult(self):
        key = self.sender().objectName()
        if not self._ignore:
            self._adjust_mul[key] = self._adjust_rnd[key].value()
            self._adjust_cty[key].setValue(self._data[key][0] * self._adjust_rnd[key].value())
        mw, mwtxt, mwstr, div = self.setAdjValueUnits(key, self._adjust_typ[key], self._data[key][0])
        self._adjust_txt[key].setText(mwtxt)
     #   if not self._ignore:
      #      self._adjust_val[key].setText(mwstr)
        self._ignore = False
        self._reset_last = False

    def adjustCap(self):
        if self._ignore:
            return
        key = self.sender().objectName()
        if key != 'Load':
            try:
                adj = self._adjust_cty[key].value() / self._data[key][0]
            except:
                adj = 1
         #   self._adjust_rnd[key].setValue(adj)
        else:
            dimen = log10(self._data[key][0])
            if dimen > 11:
                mul = 9
            elif dimen > 8:
                mul = 6
            elif dimen > 5:
                mul = 3
            else:
                mul = 0
            adj = (self._adjust_cty[key].value() * pow(10, mul)) / self._data[key][0]
        self._adjust_mul[key] = adj
      #  self._adjust_cty[key] = self._data[key] * adj
        self._ignore = True
        self._adjust_rnd[key].setValue(round(adj, 4))
        self._ignore = False
        self._reset_last = False

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def resetClicked(self, to):
        if to is None:
            to = 0.
        else:
            to = 1.
        if self.show_multipliers:
            for key in self._adjust_rnd.keys():
                self._adjust_rnd[key].setValue(to)
        else:
            if to == 0:
                for key in self._adjust_cty.keys():
                    self._adjust_cty[key].setValue(0.)
            else:
                for key in self._adjust_cty.keys():
                    if self._reset_last and self._data[key][0] <= 1.:
                        self._adjust_cty[key].setValue(0.)
                    else:
                        self._adjust_cty[key].setValue(self._data[key][0])
        self.pfx_fld.setText('')
        self._reset_last = True

    def resetloadClicked(self, to):
        if isinstance(to, bool):
            to = 1.
        if self.show_multipliers:
            self._adjust_rnd['Load'].setValue(to)
        else:
            self._adjust_cty['Load'].setValue(self._data['Load'][0])
        self._reset_last = False

    def restoreClicked(self):
        ini_file = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Adjustments file',
                   self._save_folder, 'Preferences Files (*.ini)')[0]
        if ini_file != '':
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(ini_file)
            try:
                prefix = ini_file[ini_file.rfind('/') + 1: - 4]
            except:
                prefix = ''
            self.getIt(config, prefix)
        self._reset_last = False

    def getIt(self, config, prefix=''):
        try:
            adjustto = config.get('Powermatch', 'adjusted_capacities')
        except:
            return
        self.resetClicked(to=None)
        bits = adjustto.split(',')
        for bit in bits:
            bi = bit.split('=')
            key = bi[0]
            try:
                mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                        float(bi[1]))
                self._adjust_cty[key].setValue(mwcty)
                if self.show_multipliers:
                    self._adjust_mul[key] = float(bi[1]) / (self._data[key][0] * pow(10, self._data[key][1]))
                    self._adjust_rnd[key].setValue(round(self._adjust_mul[key], 4))
            except:
                msgbox = QtWidgets.QMessageBox()
                msgbox.setWindowTitle('SIREN - Powermatch - Adjust generators')
                msgbox.setText(f'Generator - {key} - not found in generators list and will be ignored.')
                msgbox.setIcon(QtWidgets.QMessageBox.Warning)
                msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                reply = msgbox.exec_()
                pass
        self._ignore = False
        self.pfx_fld.setText(prefix)
        self._reset_last = False

    def listClicked(self):
        if os.path.exists(self._save_folder):
            names = {}
            techs = []
            ini_files = os.listdir(self._save_folder)
            ini_files.sort()
            for ini_file in ini_files:
                if ini_file[-4:] == '.ini':
                    config = configparser.RawConfigParser()
                    try:
                        config.read(self._save_folder + ini_file)
                    except:
                        continue
                    try:
                        adjustto = config.get('Powermatch', 'adjusted_capacities')
                    except:
                        continue
                    names[ini_file[:-4]] = [0] * len(techs)
                    bits = adjustto.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        key = bi[0]
                        try:
                            mw, mwtxt, mwcty, div = self.setAdjValueUnits(key, self._adjust_typ[key],
                                                                          float(bi[1]))
                        except:
                            mwcty = 0
                        if mwcty == 0:
                            continue
                        if key not in techs:
                            techs.append(key)
                            names[ini_file[:-4]].append(0)
                        ndx = techs.index(key)
                        names[ini_file[:-4]][ndx] = mwcty
            techs.insert(0, 'Preference File')
            decpts = [1] * len(techs)
            dialog = displaytable.Table(names, title=self.sender().text(), decpts=decpts, fields=techs,
                                        save_folder=self._save_folder)
            dialog.exec_()
            chosen = dialog.getItem(0)
            self._ignore = True
            reshow = False
            config = configparser.RawConfigParser()
            config.read(self._save_folder + chosen + '.ini')
            self.getIt(config, chosen)
            del dialog
        self._reset_last = False

    def saveClicked(self):
        line = ''
        for key, value in self._adjust_cty.items():
            if self._decpts == 2:
                line += '{}={:.2f},'.format(key, value.value() * pow(10, self._data[key][1]))
            else:
                line += '{}={:.1f},'.format(key, value.value() * pow(10, self._data[key][1]))
        if line != '':
            line = 'adjusted_capacities=' + line[:-1]
            updates = {'Powermatch': ['adjustments=', line]}
            save_file = self._save_folder
            if self.pfx_fld.text() != '':
                save_file += '/' + self.pfx_fld.text()
            inifile = QtWidgets.QFileDialog.getSaveFileName(None, 'Save Adjustments to file',
                      save_file, 'Preferences Files (*.ini)')[0]
            if inifile != '':
                if inifile[-4:] != '.ini':
                    inifile = inifile + '.ini'
                SaveIni(updates, ini_file=inifile)
        self._reset_last = False

    def showClicked(self):
        self.ignoreEnter = False
        self._results = {}
        for key in list(self._adjust_cty.keys()):
            self._results[key] = self._adjust_cty[key].value() * pow(10, self._data[key][1])
        self.close()

    def getValues(self):
        return self._results

    def getPrefix(self):
        return self.pfx_fld.text()

    def addtoBatch(self):
        adj_list = {}
        for key in self._adjust_cty.keys():
            adj_list[key] = self._adjust_cty[key].value()
        wb = oxl.load_workbook(self._batch_file)
        add_msg = add_to_batch(wb, self._batch_file, adj_list)
        QtWidgets.QMessageBox.about(self, 'SIREN - Add to Batch', add_msg)
        self._reset_last = False

class setTransition(MyQDialog):
    def niceSize(window, ctr): # works for Adjustments window (probably because less that 640*480)
        height = window.frameSize().height() / 1.07
        height = 70 + ctr * 32
        width = window.frameSize().width()
        screen = QtWidgets.QDesktopWidget().availableGeometry()
        if height > (screen.height() - 70):
            height = screen.height() - 70
        if width > (screen.width() - 70):
            width = screen.width() - 70
        size = QtCore.QSize(QtCore.QSize(int(width), int(height)))
        window.resize(size)

    def __init__(self, parent, label, generators, sheet, year, loads):
        super(setTransition, self).__init__()
        self._results = None
        i = generators.rfind('/')
        generator_file = generators[i + 1:]
        self.grid = QtWidgets.QGridLayout()
        r = 0
        self.grid.addWidget(QtWidgets.QLabel(label + ' File:'), r, 0)
        file_name = QtWidgets.QLabel(generator_file)
        file_name.setStyleSheet("border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.grid.addWidget(file_name, r, 1, 1, 5)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel(label + ' Sheet:'), r, 0)
        self.sheet = QtWidgets.QLineEdit()
        if sheet[-4:].isnumeric():
            sheet = sheet[:-4] + '$YEAR$'
        else:
            sheet = sheet.replace(year, '$YEAR$')
        self.sheet.setText(sheet)
   #     self.sheet.textChanged.connect(self.shtChanged)
        self.grid.addWidget(self.sheet, r, 1, 1, 2)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Load Files:'), r, 0)
        i = loads.rfind('/')
        load_name = QtWidgets.QLabel(loads[i + 1:])
        load_name.setStyleSheet("border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.grid.addWidget(load_name, r, 1, 1, 2)
        r += 1
        note = QtWidgets.QLabel('If the input worksheet (capacity figures) contain formulae you may need to open the worksheet and save it before proceeding')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        note.setPalette(msg_palette)
        note.setWordWrap(True)
        self.grid.addWidget(note, r, 0, 2, 4)
        r += 2
        quit = QtWidgets.QPushButton('Quit', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        self.grid.addWidget(show, r, 1)
        show.clicked.connect(self.showClicked)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.niceSize(r)
        self.setWindowTitle('SIREN - Powermatch - Transition files')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def quitClicked(self):
        self.ignoreEnter = False
        self.close()

    def showClicked(self):
        self.ignoreEnter = False
        self._results = self.sheet.text()
        self.close()

    def getValues(self):
        return self._results


class powerMatch(QtWidgets.QWidget):
    log = QtCore.pyqtSignal()
    progress = QtCore.pyqtSignal()

    def get_filename(self, filename):
        if filename.find('/') == 0: # full directory in non-Windows
            return filename
        elif (sys.platform == 'win32' or sys.platform == 'cygwin') \
          and filename[1:3] == ':/': # full directory for Windows
            return filename
        elif filename[:3] == '../': # directory upwards of scenarios
            ups = filename.split('../')
            scens = self.scenarios.split('/')
            scens = scens[: -(len(ups) - 1)]
            scens.append(ups[-1])
            return '/'.join(scens)
        else: # subdirectory of scenarios
            return self.scenarios + filename

    def get_load_years(self):
        load_years = ['n/a']
        i = self.load_files.find('$YEAR$')
        if i < 0:
            return load_years
        if self.load_dir.text() != self._load_folder:
            load_files = self.get_filename(self.load_files)
        else:
            load_files = self.load_files
        i = load_files.find('$YEAR$')
        j = len(load_files) - i - 6
        files = glob.glob(load_files[:i] + '*' + load_files[i + 6:])
        for fil in files:
            load_years.append(fil[i:len(fil) - j])
        return sorted(load_years, reverse=True)

    def __init__(self, help='help.html'):
        super(powerMatch, self).__init__()
        self.help = help
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = getModelFile(sys.argv[1])
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            self.model_name = config.get('Base', 'name')
        except:
            self.model_name = ''
        self.config_file = config_file[-1][config_file[-1].rfind('/') + 1:]
        try:
            base_year = config.get('Base', 'year')
        except:
            base_year = '2012'
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.batch_template = config.get('Files', 'pmb_template')
            for key, value in parents:
                self.batch_template = self.batch_template.replace(key, value)
            self.batch_template = self.batch_template.replace('$USER$', getUser())
            if not os.path.exists(self.batch_template):
                self.batch_template = ''
        except:
            self.batch_template = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', base_year)
            self.scenarios = self.scenarios[: self.scenarios.rfind('/') + 1]
            if self.scenarios[:3] == '../':
                ups = self.scenarios.split('../')
                me = os.getcwd().split(os.sep)
                me = me[: -(len(ups) - 1)]
                me.append(ups[-1])
                self.scenarios = '/'.join(me)
        except:
            self.scenarios = ''
        try:
            self.load_files = config.get('Files', 'load')
            for key, value in parents:
                self.load_files = self.load_files.replace(key, value)
            self.load_files = self.load_files.replace('$USER$', getUser())
            self.load_files = self.get_filename(self.load_files)
            that_len = len(commonprefix([self.scenarios, self.load_files]))
            if that_len > 0:
                bits = self.scenarios[that_len:].split('/')
                pfx = ('..' + '/') * (len(bits) - 1)
                self.load_files = pfx + self.load_files[that_len + 1:]
        except:
            self.load_files = ''
        # load folder from Files section
        try:
            self._load_folder = self.load_files[:self.load_files.rfind('/')]
        except:
            self._load_folder = ''
        self.log_status = True
        try:
            rw = config.get('Windows', 'log_status')
            if rw.lower() in ['false', 'no', 'off']:
                self.log_status = False
        except:
            pass
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Jobs', 'Data', 'Results', 'Batch', 'Transition']
        ifiles = [''] * len(self.file_labels)
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.batch_new_file = False
        self.batch_prefix = False
        self.batch_3d = False
        self.more_details = False
        self.constraints = None
        self.generators = None
        self.optimisation = None
        self.adjustto = None # adjust capacity to this
        self.adjust_cap = 25
        self.adjust_gen = False
        self.change_res = True
        self.adjusted_lcoe = True
        self.corrected_lcoe = True
        self.carbon_price = 0.
        self.carbon_price_max = 200.
        self.discount_rate = 0.
        self.do_jobs = False
        self.job_charts = False
        self.load_folder = ''
        self.load_year = 'n/a'
        self.optimise_choice = 'LCOE'
        self.optimise_generations = 20
        self.optimise_mutation = 0.005
        self.optimise_population = 50
        self.optimise_stop = 0
        self.optimise_debug = False
        self.optimise_default = None
        self.optimise_multiplot = False
        self.optimise_multisurf = False
        self.optimise_multitable = False
        self.optimise_to_batch = True
        self.optimise_total_re = True
        self.remove_cost = True
        self.reserve_committed = True
        self.results_prefix = ''
        self.dispatchable = ['Biomass', 'Geothermal', 'Hydro', 'Pumped Hydro', 'Solar Thermal', 'CST'] # RE dispatchable
        self.save_tables = False
        self.show_multipliers = False
        self.show_correlation = False
        self.summary_sources = True
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        self.underlying = ['Rooftop PV'] # technologies contributing to underlying (but not operational) load
        self.operational = []
        iorder = []
        self.targets = {}
        for t in range(len(target_keys)):
            if target_keys[t] in ['re_pct', 'surplus_pct']:
                self.targets[target_keys[t]] = [target_names[t], 0., -1, 0., 0, target_fmats[t],
                                                 target_titles[t]]
            else:
                self.targets[target_keys[t]] = [target_names[t], 0., 0., -1, 0, target_fmats[t],
                                                 target_titles[t]]
        try:
            dts = config.get('Grid', 'dispatchable').split(' ')
            dispatchable = []
            for dt in dts:
                dispatchable.append(techClean(dt.replace('_', ' ').title()))
            self.dispatchable = dispatchable
        except:
            pass
        try:
            adjust_cap = config.get('Power', 'adjust_cap')
            try:
                self.adjust_cap = float(adjust_cap)
            except:
                try:
                    self.adjust_cap = eval(adjust_cap)
                except:
                    pass
            if self.adjust_cap < 0:
                self.adjust_cap = pow(10, 12)
        except:
            pass
        pref_errors = []
        try:
            items = config.items('Powermatch')
            for key, value in items:
                if key == 'batch_new_file':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_new_file = True
                elif key == 'batch_prefix':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_prefix = True
                elif key == 'batch_3d':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.batch_3d = True
                elif key[:4] == 'tml_':
                    continue
                elif key[-5:] == '_file':
                    ndx = self.file_labels.index(key[:-5].title())
                    ifiles[ndx] = value.replace('$USER$', getUser())
                elif key[-6:] == '_sheet':
                    ndx = self.file_labels.index(key[:-6].title())
                    self.isheets[ndx] = value
                elif key == 'adjust_generators':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.adjust_gen = True
                elif key == 'adjusted_capacities':
                    self.adjustto = {}
                    bits = value.split(',')
                    for bit in bits:
                        bi = bit.split('=')
                        self.adjustto[bi[0]] = float(bi[1])
                elif key == 'carbon_price':
                    try:
                        self.carbon_price = float(value)
                    except:
                        pass
                elif key == 'carbon_price_max':
                    try:
                        self.carbon_price_max = float(value)
                    except:
                        pass
                elif key == 'change_results':
                    if value.lower() in ['false', 'off', 'no']:
                        self.change_res = False
                elif key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.corrected_lcoe = False
                elif key == 'adjusted_lcoe' or key == 'corrected_lcoe':
                    if value.lower() in ['false', 'no', 'off']:
                        self.adjusted_lcoe = False
                        self.corrected_lcoe = False
                elif key == 'discount_rate':
                    try:
                        self.discount_rate = float(value)
                    except:
                        pass
                elif key == 'dispatch_order':
                    iorder = value.split(',')
                elif key == 'job_charts':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.job_charts = True
                elif key == 'jobs':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.do_jobs = True
                elif key == 'load':
                    try:
                        self.load_files = value
                        for ky, valu in parents:
                            self.load_files = self.load_files.replace(ky, valu)
                        self.load_files = self.load_files.replace('$USER$', getUser())
                        if self.load_files.rfind('/') > 0:
                            that_len = len(commonprefix([self.scenarios, self.load_files]))
                            if that_len > 0:
                                bits = self.scenarios[that_len:].split('/')
                                pfx = ('..' + '/') * (len(bits) - 1)
                                self.load_files = pfx + self.load_files[that_len + 1:]
                    except:
                        pass
                elif key == 'load_year':
                    self.load_year = value
                elif key == 'log_status':
                    if value.lower() in ['false', 'no', 'off']:
                        self.log_status = False
                elif key == 'more_details':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.more_details = True
                elif key == 'optimise_debug':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_debug = True
                elif key == 'optimise_default':
                    self.optimise_default = value
                elif key == 'optimise_choice':
                    self.optimise_choice = value
                elif key == 'optimise_generations':
                    try:
                        self.optimise_generations = int(value)
                    except:
                        pass
                elif key == 'optimise_multiplot':
                    if value.lower() in ['true', 'yes', 'on']:
                        self.optimise_multiplot = True
                    elif value.lower() in ['surf', 'tri-surf', 'trisurf']:
                        self.optimise_multiplot = True
                        self.optimise_multisurf = True
                elif key == 'optimise_multitable':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.optimise_multitable = True
                elif key == 'optimise_mutation':
                    try:
                        self.optimise_mutation = float(value)
                    except:
                        pass
                elif key == 'optimise_population':
                    try:
                        self.optimise_population = int(value)
                    except:
                        pass
                elif key == 'optimise_stop':
                    try:
                        self.optimise_stop = int(value)
                    except:
                        pass
                elif key == 'optimise_to_batch':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_to_batch = False
                elif key == 'optimise_total_re':
                    if value.lower() in ['false', 'off', 'no']:
                        self.optimise_total_re = False
                elif key[:9] == 'optimise_':
                    try:
                        bits = value.split(',')
                        t = target_keys.index(key[9:])
                        # name, weight, minimum, maximum, widget index
                        self.targets[key[9:]] = [target_names[t], float(bits[0]), float(bits[1]),
                                                float(bits[2]), 0, target_fmats[t],
                                                 target_titles[t]]
                    except:
                        pass
                elif key == 'remove_cost':
                    if value.lower() in ['false', 'off', 'no']:
                        self.remove_cost = False
                elif key == 'results_prefix':
                    self.results_prefix = value
                elif key == 'reserve_committed':
                    if value.lower() in ['false', 'off', 'no']:
                        self.reserve_committed = False
                elif key == 'save_tables':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.save_tables = True
                elif key == 'show_correlation':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_correlation = True
                elif key == 'show_multipliers':
                    if value.lower() in ['true', 'on', 'yes']:
                        self.show_multipliers = True
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
                elif key == 'summary_sources':
                    if value.lower() in ['false', 'off', 'no']:
                        self.summary_sources = False
                elif key == 'underlying':
                    self.underlying = value.split(',')
                elif key == 'operational':
                    self.operational = value.split(',')
        except:
            pref_errors.append(f"PME1: Error with '{key}' property in '{config_file}[Powermatch]'")
            pass
        self.restorewindows = False
        try:
            rw = config.get('Windows', 'restorewindows')
            if rw.lower() in ['true', 'yes', 'on']:
                self.restorewindows = True
        except:
            pass
        self.opt_progressbar = None
        self.floatstatus = None # status window
   #     self.tabs = QtGui.QTabWidget()    # Create tabs
   #     tab1 = QtGui.QWidget()
        self.grid = QtWidgets.QGridLayout()
        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.sheets = self.file_labels[:]
        del self.sheets[-2:]
        self.updated = False
        edit = [None] * D
        r = 0
        for i in range(len(self.file_labels)):
            if i == J and not self.do_jobs:
                continue
            if i == R:
                self.grid.addWidget(QtWidgets.QLabel('Results Prefix:'), r, 0)
                self.results_pfx_fld = QtWidgets.QLineEdit()
                self.results_pfx_fld.setText(self.results_prefix)
                self.results_pfx_fld.textChanged.connect(self.pfxChanged)
                self.grid.addWidget(self.results_pfx_fld, r, 1, 1, 2)
                r += 1
            self.labels[i] = QtWidgets.QLabel(self.file_labels[i] + ' File:')
            self.grid.addWidget(self.labels[i], r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
            self.files[i].setText(ifiles[i])
            self.files[i].clicked.connect(self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 5)
            button = QtWidgets.QPushButton(f'Open {self.file_labels[i]} file', self)
            self.grid.addWidget(button, r, 6)
            button.clicked.connect(self.openClicked)
            if i < D:
                r += 1
                self.grid.addWidget(QtWidgets.QLabel(self.file_labels[i] + ' Sheet:'), r, 0)
                self.sheets[i] = QtWidgets.QComboBox()
                try:
                    curfile = self.get_filename(ifiles[i])
                    ts = WorkBook()
                    ts.open_workbook(curfile)
                    ndx = 0
                    j = -1
                    for sht in ts.sheet_names():
                        j += 1
                        self.sheets[i].addItem(sht)
                        if sht == self.isheets[i]:
                            ndx = j
                    self.sheets[i].setCurrentIndex(ndx)
                    ws = ts.sheet_by_index(ndx)
                    if i == C:
                        self.getConstraints(ws)
                    elif i == G:
                        self.getGenerators(ws)
                    elif i == O:
                        self.getOptimisation(ws)
                    elif i == J:
                        self.getJobFactors(ws)
                    ts.close()
                    del ts
                except:
                    self.sheets[i].addItem(self.isheets[i])
                self.grid.addWidget(self.sheets[i], r, 1, 1, 3)
                self.sheets[i].currentIndexChanged.connect(self.sheetChanged)
                edit[i] = QtWidgets.QPushButton(self.file_labels[i], self)
                self.grid.addWidget(edit[i], r, 4, 1, 2)
                edit[i].clicked.connect(self.editClicked)
            elif i == D: # and self.load_files != '': always show this option
                r += 1
                self.grid.addWidget(QtWidgets.QLabel('Load Folder:'), r, 0)
                self.load_dir = ClickableQLabel()
                try:
                    self.load_dir.setText(self.load_files[:self.load_files.rfind('/')])
                except:
                    self.load_dir.setText('')
                self.load_dir.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
                self.load_dir.clicked.connect(self.loaddirChanged)
                self.grid.addWidget(self.load_dir, r, 1, 1, 5)
                r += 1
                self.grid.addWidget(QtWidgets.QLabel('Load Year:'), r, 0)
                self.load_years = self.get_load_years()
                self.loadCombo = QtWidgets.QComboBox()
                for choice in self.load_years:
                    self.loadCombo.addItem(choice)
                    if choice == self.load_year:
                        self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
                self.loadCombo.currentIndexChanged.connect(self.changes)
                self.grid.addWidget(self.loadCombo, r, 1)
                self.grid.addWidget(QtWidgets.QLabel("(To to use a different load year to the data file. Otherwise choose 'n/a')"), r, 2, 1, 4)
            r += 1
      #  wdth = edit[1].fontMetrics().boundingRect(edit[1].text()).width() + 9
        self.grid.addWidget(QtWidgets.QLabel('Replace Last:'), r, 0)
        if self.batch_new_file:
            msg = '(check to replace an existing Results workbook)'
        else:
            msg = '(check to replace last Results worksheet in Batch/Transition spreadsheet)'
        self.replace_last = QtWidgets.QCheckBox(msg, self)
        self.replace_last.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.replace_last, r, 1, 1, 3)
        self.grid.addWidget(QtWidgets.QLabel('Prefix facility names in Batch report:'), r, 4)
        self.batch_prefix_check = QtWidgets.QCheckBox('', self)
        if self.batch_prefix:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Checked)
        else:
            self.batch_prefix_check.setCheckState(QtCore.Qt.Unchecked)
        self.grid.addWidget(self.batch_prefix_check, r, 5)
        self.batch_prefix_check.stateChanged.connect(self.bpcchanged)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Discount Rate:'), r, 0)
        self.discount = QtWidgets.QDoubleSpinBox()
        self.discount.setRange(0, 100)
        self.discount.setDecimals(2)
        self.discount.setSingleStep(.5)
        try:
            self.discount.setValue(self.discount_rate * 100.)
        except:
            self.discount.setValue(0.)
        self.grid.addWidget(self.discount, r, 1)
        self.discount.valueChanged.connect(self.drchanged)
        self.grid.addWidget(QtWidgets.QLabel('(%. Only required if using input costs rather than reference LCOE)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Carbon Price:'), r, 0)
        self.carbon = QtWidgets.QDoubleSpinBox()
        self.carbon.setRange(0, self.carbon_price_max)
        self.carbon.setDecimals(2)
        try:
            self.carbon.setValue(self.carbon_price)
        except:
            self.carbon.setValue(0.)
        self.grid.addWidget(self.carbon, r, 1)
        self.carbon.valueChanged.connect(self.cpchanged)
        self.grid.addWidget(QtWidgets.QLabel('($/tCO2e. Use only if LCOE excludes carbon price)'), r, 2, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Adjust Generators:'), r, 0)
        self.adjust = QtWidgets.QCheckBox('(check to adjust generators capacity data)', self)
        if self.adjust_gen:
            self.adjust.setCheckState(QtCore.Qt.Checked)
        self.grid.addWidget(self.adjust, r, 1, 1, 4)
        r += 1
        self.grid.addWidget(QtWidgets.QLabel('Dispatch Order:\n(move to right\nto exclude)'), r, 0)
        self.order = ListWidget(self) #QtWidgets.QListWidget()
      #  self.order.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.order, r, 1, 1, 3)
        self.ignore = ListWidget(self) # QtWidgets.QListWidget()
      #  self.ignore.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.ignore, r, 4, 1, 3)
        r += 1
        self.log = QtWidgets.QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, r, 1, 1, 6)
        r += 1
        self.progressbar = QtWidgets.QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(20) # was 10 set to 20 to get 5% steps
        self.progressbar.setValue(0)
        self.progressbar.setStyleSheet('QProgressBar {border: 1px solid grey; border-radius: 2px; text-align: center;}' \
                                       + 'QProgressBar::chunk { background-color: #06A9D6;}')
        self.grid.addWidget(self.progressbar, r, 1, 1, 6)
        self.progressbar.setHidden(True)
        r += 1
        r += 1
        quit = QtWidgets.QPushButton('Done', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        pms = QtWidgets.QPushButton('Summary', self)
        self.grid.addWidget(pms, r, 1)
        pms.clicked.connect(self.pmClicked)
        pm = QtWidgets.QPushButton('Detail', self)
     #   pm.setMaximumWidth(wdth)
        self.grid.addWidget(pm, r, 2)
        pm.clicked.connect(self.pmClicked)
        btch = QtWidgets.QPushButton('Batch', self)
        self.grid.addWidget(btch, r, 3)
        btch.clicked.connect(self.pmClicked)
        trns = QtWidgets.QPushButton('Transition', self)
        self.grid.addWidget(trns, r, 4)
        trns.clicked.connect(self.pmClicked)
        opt = QtWidgets.QPushButton('Optimise', self)
        self.grid.addWidget(opt, r, 5)
        opt.clicked.connect(self.pmClicked)
        help = QtWidgets.QPushButton('Help', self)
        help.clicked.connect(self.helpClicked)
        if self.do_jobs:
            jobs = QtWidgets.QPushButton('Jobs', self)
     #   help.setMaximumWidth(wdth)
      #  quit.setMaximumWidth(wdth)
            self.grid.addWidget(jobs, r, 6)
            jobs.clicked.connect(self.jobsClicked)
            self.grid.addWidget(help, r + 1, 6)
        else:
     #   help.setMaximumWidth(wdth)
      #  quit.setMaximumWidth(wdth)
            self.grid.addWidget(help, r, 6)
            del headers[st_job]
        r += 1
        editini = QtWidgets.QPushButton('Preferences', self)
     #   editini.setMaximumWidth(wdth)
        self.grid.addWidget(editini, r, 0)
        editini.clicked.connect(self.editIniFile)
        do_tml = False
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                do_tml = True
        elif os.path.exists('pmtmldetail.py'):
            do_tml = True
        if do_tml:
            tmld = QtWidgets.QPushButton('TML Detail', self)
            self.grid.addWidget(tmld, r, 1)
            tmld.clicked.connect(self.tmlClicked)
        self.setOrder()
        if len(iorder) > 0:
            self.order.clear()
            self.ignore.clear()
            for gen in iorder:
                self.order.addItem(gen)
            try:
                for gen in self.generators.keys():
                    if (gen in tech_names and gen not in self.dispatchable) or gen in iorder:
                        continue
                    try:
                        chk = gen[gen.find('.') + 1:]
                        if chk in tech_names and chk not in self.dispatchable:
                            continue
                    except:
                        pass
                    self.ignore.addItem(gen)
            except:
                pass
        if self.adjust_gen and self.adjustto is None:
           self.adjustto = {}
           self.adjustto['Load'] = 0
           for gen in tech_names:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
           for gen in iorder:
               try:
                   if self.generators[gen].capacity > 0:
                       self.adjustto[gen] = self.generators[gen].capacity
               except:
                   pass
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
    #    tab1.setLayout(self.layout)
    #    self.tabs.addTab(tab1,'Parms')
    #    self.tabs.addTab(tab2,'Constraints')
    #    self.tabs.addTab(tab3,'Generators')
    #    self.tabs.addTab(tab4,'Summary')
    #    self.tabs.addTab(tab5,'Details')
        self.setWindowTitle('SIREN - powermatch (' + fileVersion() + ') - Powermatch - ' + self.model_name + ' (' + self.config_file + ')')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        if self.restorewindows:
            try:
                rw = config.get('Windows', 'powermatch_size').split(',')
                self.resize(int(rw[0]), int(rw[1]))
                mp = config.get('Windows', 'powermatch_pos').split(',')
                self.move(int(mp[0]), int(mp[1]))
            except:
                pass
        else:
            self.center()
            self.resize(int(self.sizeHint().width() * 1.2), int(self.sizeHint().height() * 1.2))
        if len(pref_errors) > 0:
            self.log_status = True
        self.show_FloatStatus() # status window
        if len(pref_errors) > 0:
            for error in pref_errors:
                self.setStatus(error)
            self.setStatus('These errors may cause issues with other properties.')
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        centerPoint = QtWidgets.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        self.setStatus('')
        for i in range(len(self.file_labels)):
            try:
                if self.files[i].hasFocus():
                    break
            except:
                pass
        if self.files[i].text() == '':
            curfile = self.scenarios[:-1]
        else:
            curfile = self.get_filename(self.files[i].text())
        if i == R:
            if self.files[i].text() == '':
                curfile = self.get_filename(self.files[D].text())
                curfile = curfile.replace('data', 'results')
                curfile = curfile.replace('Data', 'Results')
                if curfile == self.scenarios + self.files[D].text():
                    j = curfile.find(' ')
                    if j > 0:
                        jnr = ' '
                    else:
                        jnr = '_'
                    j = curfile.rfind('.')
                    curfile = curfile[:j] + jnr + 'Results' + curfile[j:]
            else:
                curfile = self.get_filename(self.files[R].text())
            newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)')[0]
        elif (i == B or i == T) and not self.batch_new_file:
            options = QtWidgets.QFileDialog.Options()
            # options |= QFileDialog.DontUseNativeDialog
            newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Open/Create and save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)', options=options)[0]
        else:
            newfile = QtWidgets.QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[i] + ' file',
                      curfile)[0]
        if newfile != '':
            if i < D:
                if i == C:
                    self.constraints = None
                elif i == G:
                    self.generators = None
                elif i == O:
                    self.optimisation = None
                elif i == J:
                    self.jobfactors = None
                ts = WorkBook()
                ts.open_workbook(newfile)
                ndx = 0
                self.sheets[i].clear()
                j = -1
                for sht in ts.sheet_names():
                    j += 1
                    self.sheets[i].addItem(sht)
                    if len(sht) >= len(self.file_labels[i]):
                        if sht[:len(self.file_labels[i])].lower() == self.file_labels[i].lower():
                            ndx = j
                self.sheets[i].setCurrentIndex(ndx)
                if i == G:
                    ws = ts.sheet_by_index(ndx)
                    self.getGenerators(ws)
                    self.setOrder()
                ts.close()
                del ts
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                if newfile.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newfile]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newfile = pfx + newfile[that_len + 1:]
                self.files[i].setText(newfile)
            if i == D and self.change_res:
                newfile = self.files[D].text()
                newfile = newfile.replace('data', 'results')
                newfile = newfile.replace('Data', 'Results')
                if newfile != self.files[D].text():
                    self.files[R].setText(newfile)
            self.updated = True

    def pfxChanged(self):
        self.results_prefix = self.results_pfx_fld.text()
     #   self.setStatus('Results filename will be ' + self.results_pfx_fld.text() + '_' + self.files[R].text())
        self.updated = True

    def sheetChanged(self, i):
        try:
            for i in range(3):
                if self.sheets[i].hasFocus():
                    break
            else:
                return
        except:
            return # probably file changed
        self.setStatus('')
        newfile = self.get_filename(self.files[i].text())
        ts = WorkBook()
        ts.open_workbook(newfile)
        ws = ts.sheet_by_name(self.sheets[i].currentText())
        self.setStatus('Sheet ' + self.sheets[i].currentText() + ' loaded')
        if i == C:
            self.getConstraints(ws)
        elif i == G:
            self.getGenerators(ws)
            self.setOrder()
        elif i == O:
            self.getOptimisation(ws)
        ts.close()
        del ts

    def loaddirChanged(self):
        curdir = self.get_filename(self.load_dir.text())
        newdir = QtWidgets.QFileDialog.getExistingDirectory(self, 'Choose Load File Folder',
                 curdir, QtWidgets.QFileDialog.ShowDirsOnly)
        if newdir != '':
            try:
                self.load_files = newdir + self.load_files[self.load_files.rfind('/'):]
            except:
                self.load_files = newdir + self.load_files
            if newdir[: len(self.scenarios)] == self.scenarios:
                self.load_dir.setText(newdir[len(self.scenarios):])
            else:
                if newdir.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newdir]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newdir = pfx + newdir[that_len + 1:]
                self.load_dir.setText(newdir)
            self.load_years = self.get_load_years()
            self.loadCombo.clear()
            for choice in self.load_years:
                self.loadCombo.addItem(choice)
                if choice == self.load_year:
                    self.loadCombo.setCurrentIndex(self.loadCombo.count() - 1)
            self.updated = True

    def helpClicked(self):
        dialog = displayobject.AnObject(QtWidgets.QDialog(), self.help,
                 title='Help for powermatch (' + fileVersion() + ')', section='powermatch')
        dialog.exec_()

    def drchanged(self):
        self.updated = True
        self.discount_rate = self.discount.value() / 100.

    def cpchanged(self):
        self.updated = True
        self.carbon_price = self.carbon.value()

    def changes(self):
        self.updated = True

    def optLoadChange(self):
        if self.optLoadMult.hasFocus():
            self.optLoad.setValue(self.optLoadMult.value() * self._optLoadv[1])
        elif self.optLoad.hasFocus():
            self.optLoadMult.setValue(self.optLoad.value() / self._optLoadv[1])
        self.updated = True

    def bpcchanged(self):
        if self.batch_prefix_check.isChecked():
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        self.updated = True

    def openClicked(self):
        bit = self.sender().text().split()
        fnr = self.file_labels.index(bit[1])
        curfile = self.get_filename(self.files[fnr].text())
        if not os.path.exists(curfile):
            if fnr == R and self.results_pfx_fld.text() != '':
                i = curfile.rfind('/')
                curfile = curfile[:i + 1] + self.results_pfx_fld.text() + '_' + curfile[i+1:]
                if not os.path.exists(curfile):
                    self.setStatus(self.file_labels[fnr] + ' not found.')
                    return
            else:
                self.setStatus(self.file_labels[fnr] + ' not found.')
                return
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            os.startfile(curfile)
        elif sys.platform == 'darwin':
            try:
                subprocess.call('open', curfile)
            except:
                try:
                    subprocess.call('open', '-a', 'Microsoft Excel', curfile)
                except:
                    self.setStatus("Can't 'launch' '" + self.file_labels[fnr] + "' file")
                    return
        elif sys.platform == 'linux2' or sys.platform == 'linux':
            subprocess.call(('xdg-open', curfile))
        self.setStatus(self.file_labels[fnr] + ' file "launched".')

    def quitClicked(self):
        if self.updated or self.order.updated or self.ignore.updated:
            updates = {}
            lines = []
            lines.append('adjust_generators=' + str(self.adjust.isChecked()))
            lines.append('adjustments=') # clean up the old way
            if self.adjustto is not None:
                line = ''
                for key, value in self.adjustto.items():
                    line += '{}={:.1f},'.format(key, value)
                if line != '':
                    lines.append('adjusted_capacities=' + line[:-1])
            line = 'batch_prefix='
            if self.batch_prefix:
                line += 'True'
            lines.append(line)
            lines.append('carbon_price=' + str(self.carbon_price))
            lines.append('discount_rate=' + str(self.discount_rate))
            line = ''
            for itm in range(self.order.count()):
                line += self.order.item(itm).text() + ','
            lines.append('dispatch_order=' + line[:-1])
            for i in range(len(self.file_labels)):
                try:
                    lines.append(self.file_labels[i].lower() + '_file=' + self.files[i].text().replace(getUser(), '$USER$'))
                except:
                    pass
            for i in range(D):
                try:
                    lines.append(self.file_labels[i].lower() + '_sheet=' + self.sheets[i].currentText())
                except:
                    pass
            line = 'load='
            if self.load_dir.text() != self._load_folder:
                if self.load_files.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, self.load_files]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        self.load_files = pfx + self.load_files[that_len + 1:]
                line += self.load_files.replace(getUser(), '$USER$')
            lines.append(line)
            line = 'load_year='
            if self.loadCombo.currentText() != 'n/a':
                line += self.loadCombo.currentText()
            lines.append(line)
            lines.append('optimise_choice=' + self.optimise_choice)
            lines.append('optimise_generations=' + str(self.optimise_generations))
            lines.append('optimise_mutation=' + str(self.optimise_mutation))
            lines.append('optimise_population=' + str(self.optimise_population))
            lines.append('optimise_stop=' + str(self.optimise_stop))
            for key, value in self.targets.items():
                line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
                lines.append(line)
            lines.append('results_prefix=' + self.results_prefix)
            updates['Powermatch'] = lines
            SaveIni(updates)
        try:
            plt.close('all')
        except:
            pass

        self.close()

    def closeEvent(self, event):
        if self.floatstatus is not None:
            self.floatstatus.exit()
        if self.restorewindows:
            updates = {}
            lines = []
            add = int((self.frameSize().width() - self.size().width()) / 2)   # need to account for border
            lines.append('powermatch_pos=%s,%s' % (str(self.pos().x() + add), str(self.pos().y() + add)))
            lines.append('powermatch_size=%s,%s' % (str(self.width()), str(self.height())))
            updates['Windows'] = lines
            SaveIni(updates)
        event.accept()

    def tmlClicked(self):
        if len(sys.argv) > 1:
            config_file = getModelFile(sys.argv[1])
        else:
            config_file = getModelFile('SIREN.ini')
        if sys.platform == 'win32' or sys.platform == 'cygwin':
            if os.path.exists('pmtmldetail.exe'):
                pid = subprocess.Popen(['pmtmldetail.exe', config_file]).pid
            elif os.path.exists('pmtmldetail.py'):
                pid = subprocess.Popen(['pmtmldetail.py', config_file], shell=True).pid
        else:
            pid = subprocess.Popen(['python3', 'pmtmldetail.py', config_file]).pid # python3
        return

    def editIniFile(self):
        if len(sys.argv) > 1:
            config_file = getModelFile(sys.argv[1])
        else:
            config_file = getModelFile('SIREN.ini')
        before = os.stat(config_file[-1]).st_mtime
        dialr = EdtDialog(config_file, section='[Powermatch]')
        dialr.exec_()
        after = os.stat(config_file[-1]).st_mtime
        if after == before:
            return
     #   self.get_config()   # refresh config values
        config = configparser.RawConfigParser()
        config.read(config_file)
        try:
            st = config.get('Powermatch', 'save_tables')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.save_tables = True
        else:
            self.save_tables = False
        try:
            st = config.get('Powermatch', 'more_details')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.more_details = True
        else:
            self.more_details = False
        try:
            st = config.get('Powermatch', 'optimise_to_batch')
        except:
            st = 'True'
        if st.lower() in ['true', 'yes', 'on']:
            self.optimise_to_batch = True
        else:
            self.optimise_to_batch = False
        try:
            st = config.get('Powermatch', 'remove_cost')
        except:
            st = 'True'
        if st.lower() in ['false', 'no', 'off']:
            self.remove_cost = False
        else:
            self.remove_cost = True
        try:
            st = config.get('Powermatch', 'show_multipliers')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.show_multipliers = True
        else:
            self.show_multipliers = False
        try:
            st = config.get('Powermatch', 'batch_new_file')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_new_file = True
            msg = '(check to replace an existing Results workbook)'
        else:
            self.batch_new_file = False
            msg = '(check to replace last Results worksheet in Batch/Transition spreadsheet)'
        self.replace_last = QtWidgets.QCheckBox(msg, self)
        try:
            st = config.get('Powermatch', 'batch_prefix')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_prefix = True
        else:
            self.batch_prefix = False
        try:
            st = config.get('Powermatch', 'batch_3d')
        except:
            st = 'False'
        if st.lower() in ['true', 'yes', 'on']:
            self.batch_3d = True
        else:
            self.batch_3d = False
        QtWidgets.QApplication.processEvents()
        self.setStatus(config_file[-1] + ' edited. Reload may be required.')

    def editClicked(self):
        def update_dictionary(it, source):
            new_keys = list(source.keys())
            # first we delete and add keys to match updated dictionary
            if it == C:
                old_keys = list(self.constraints.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.constraints[key]
                for key in new_keys:
                    self.constraints[key] = Constraint(key, '<category>',
                                            0., 1., 0, 1., 0, 0., 1., 1., 0., 1., 0., 0, 0)
                target = self.constraints
            elif it == G:
                old_keys = list(self.generators.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.generators[key]
                for key in new_keys:
                    self.generators[key] = Facility(name=key, constraint='<constraint>')
                target = self.generators
            elif it == O:
                old_keys = list(self.optimisation.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.optimisation[key]
                for key in new_keys:
                    self.optimisation[key] = Optimisation(key, 'None', None)
                target = self.optimisation
            elif it == J:
                old_keys = list(self.jobfactors.keys())
                for key in old_keys:
                    if key in new_keys:
                        del new_keys[new_keys.index(key)]
                    else:
                        del self.jobfactors[key]
                for key in new_keys:
                    self.jobfactors[key] = JobFactors(key, 'None', None)
                target = self.jobfactors
            # now update the data
            for key in list(target.keys()):
                for prop in dir(target[key]):
                    if prop[:2] != '__' and prop[-2:] != '__':
                        try:
                            if prop == 'lifetime' and source[key][prop] == 0:
                                setattr(target[key], prop, 20)
                            else:
                                setattr(target[key], prop, source[key][prop])
                        except:
                            pass

        self.setStatus('')
        msg = ''
        ts = None
        it = self.file_labels.index(self.sender().text())
        if it == C and self.constraints is not None:
            pass
        elif it == G and self.generators is not None:
            pass
        elif it == O and self.optimisation is not None:
            pass
        else:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[it].text()))
                try:
                    sht = self.sheets[it].currentText()
                except:
                    self.setStatus(self.sheets[it].currentText() + ' not found in ' \
                                     + self.file_labels[it] + ' spreadsheet.')
                    return
                ws = ts.sheet_by_name(sht)
            except:
                ws = None
        if it == C: # self.constraints
            if self.constraints is None:
                try:
                    self.getConstraints(ws)
                except:
                    return
            sp_pts = [2] * 15
            sp_pts[4] = 3 # discharge loss
            sp_pts[8] = 3 # parasitic loss
            sp_pts[11] = 3 # recharge loss
            dialog = displaytable.Table(self.constraints, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                msg = ' table updated'
        elif it == G: # generators
            if self.generators is None:
                try:
                    self.getGenerators(ws)
                except:
                    return
            sp_pts = []
            for key in self.generators.keys():
                break
            for prop in dir(self.generators[key]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop == 'name':
                        sp_pts.insert(0, 0)
                    elif prop in ['capex', 'constraint', 'fixed_om', 'order']:
                        sp_pts.append(0)
                    elif prop == 'disc_rate' or prop == 'emissions':
                        sp_pts.append(3)
                    elif prop == 'area':
                        sp_pts.append(5)
                    else:
                        sp_pts.append(2)
            dialog = displaytable.Table(self.generators, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                self.setOrder()
                msg = ' table updated'
        elif it == O: # self.optimisation
            if self.optimisation is None:
                try:
                    self.getOptimisation(ws)
                except:
                    return
            dialog = displaytable.Table(self.optimisation, title=self.sender().text(),
                     save_folder=self.scenarios, edit=True)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                for key in self.optimisation.keys():
                    if self.optimisation[key].approach == 'Discrete':
                        caps = self.optimisation[key].capacities.split()
                        self.optimisation[key].capacities = []
                        cap_max = 0.
                        for cap in caps:
                            try:
                                self.optimisation[key].capacities.append(float(cap))
                                cap_max += float(cap)
                            except:
                                pass
                        self.optimisation[key].capacity_min = 0
                        self.optimisation[key].capacity_max = round(cap_max, 3)
                        self.optimisation[key].capacity_step = None
                msg = ' table updated'
        elif it == J: # self.jobfactors
            if self.jobfactors is None:
                try:
                    self.getJobFactors(ws)
                except:
                    return
            dialog = displaytable.Table(self.jobfactors, title=self.sender().text(),
                     save_folder=self.scenarios, edit=True)
            dialog.exec_()
            if dialog.getValues() is not None:
                update_dictionary(it, dialog.getValues())
                msg = ' table updated'
        if ts is not None:
            ts.close()
            del ts
        newfile = dialog.savedfile
        if newfile is not None:
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[it].setText(newfile[len(self.scenarios):])
            else:
                self.files[it].setText(newfile)
            if msg == '':
                msg = ' table exported'
            else:
                msg += ' and exported'
        if msg != '':
            self.setStatus(self.file_labels[it] + msg)

    def getConstraints(self, ws):
        name = '<name>'
        category = '<category>'
        capacity_min = 0.
        capacity_max = 1.
        discharge_loss = 0.
        discharge_max = 1.
        discharge_start = 0
        min_run_time = 0
        parasitic_loss = 0.
        rampdown_max = 1.
        rampup_max = 1.
        recharge_loss = 0.
        recharge_max = 1.
        recharge_start = 0
        warm_time = 0
        if ws is None:
            self.constraints = {}
            self.constraints['<name>'] = Constraint(name, category, capacity_min,
                                         capacity_max, discharge_loss, discharge_max,
                                         discharge_start, min_run_time, parasitic_loss,
                                         rampdown_max, rampup_max, recharge_loss,
                                         recharge_max, recharge_start, warm_time)
            return
        col_names = ['Name', 'Category', 'Capacity Min', 'Capacity Max', 'Discharge Loss',
                     'Discharge Max', 'Discharge Start', 'Min Run Time', 'Parasitic Loss',
                     'Rampdown Max', 'Rampup Max', 'Recharge Loss', 'Recharge Max',
                     'Recharge Start', 'Warmup Time']
        col_no = [-1] * len(col_names)
        if ws.cell_value(1, 0) == 'Name' and ws.cell_value(1, 1) == 'Category':
            col = 0
            while col < ws.ncols:
                if ws.cell_value(0, col) == 'Capacity':
                    for c2 in range(col, col + 2):
                       try:
                           ndx = col_names.index('Capacity ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except:
                            pass
                    col += 2
                elif ws.cell_value(0, col) == 'Discharge':
                    for c2 in range(col, col + 3):
                       try:
                           ndx = col_names.index('Discharge ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except:
                            pass
                    col += 3
                elif ws.cell_value(0, col) == 'Recharge':
                    for c2 in range(col, col + 3):
                       try:
                           ndx = col_names.index('Recharge ' + ws.cell_value(1, c2))
                           col_no[ndx] = c2
                       except:
                            pass
                    col += 3
                else:
                    try:
                        ndx = col_names.index(ws.cell_value(1, col))
                        col_no[ndx] = col
                    except:
                        pass
                    col += 1
            strt_row = 2
        elif ws.cell_value(0, 0) == 'Name': # saved file
            for col in range(ws.ncols):
                try:
                    ndx = col_names.index(ws.cell_value(0, col))
                    col_no[ndx] = col
                except:
                    pass
            strt_row = 1
        else:
            self.setStatus('Not a ' + self.file_labels[C] + ' worksheet.')
            return
        if col_no[0] < 0:
            self.setStatus('Not a ' + self.file_labels[C] + ' worksheet.')
            return
        self.constraints = {}
        data = [name, category, capacity_min, capacity_max, discharge_loss, discharge_max,
                discharge_start, min_run_time, parasitic_loss, rampdown_max, rampup_max,
                recharge_loss, recharge_max, recharge_start, warm_time]
        for row in range(strt_row, ws.nrows):
            for ndx in range(len(col_no)):
                 if col_no[ndx] >= 0:
                     data[ndx] = ws.cell_value(row, col_no[ndx])
            self.constraints[str(data[0])] = Constraint(*data)
        return

    def getGenerators(self, ws):
        if ws is None:
            self.generators = {}
            args = {'name': '<name>', 'constraint': '<constraint>'}
            self.generators['<name>'] = Facility(**args)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not a ' + self.file_labels[G] + ' worksheet.')
            return
        args = ['name', 'order', 'constraint', 'capacity', 'lcoe', 'lcoe_cf', 'emissions', 'initial',
                'capex', 'fixed_om', 'variable_om', 'fuel', 'disc_rate', 'lifetime', 'area']
        possibles = {'name': 0}
        for col in range(ws.ncols):
            try:
                arg = ws.cell_value(0, col).lower()
            except:
                continue
            if arg in args:
                possibles[arg] = col
            elif ws.cell_value(0, col)[:9] == 'Capital':
                possibles['capex'] = col
            elif ws.cell_value(0, col)[:8] == 'Discount':
                possibles['disc_rate'] = col
            elif ws.cell_value(0, col)[:8] == 'Dispatch':
                possibles['order'] = col
            elif ws.cell_value(0, col)[:9] == 'Emissions':
                possibles['emissions'] = col
            elif ws.cell_value(0, col) == 'FOM':
                possibles['fixed_om'] = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                possibles['lcoe_cf'] = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                possibles['lcoe'] = col
            elif ws.cell_value(0, col) == 'VOM':
                possibles['variable_om'] = col
        self.generators = {}
        for row in range(1, ws.nrows):
            if ws.cell_value(row, 0) is None:
                continue
            in_args = {}
            for key, value in possibles.items():
                in_args[key] = ws.cell_value(row, value)
            self.generators[str(ws.cell_value(row, 0))] = Facility(**in_args)
        return

    def getOptimisation(self, ws):
        if ws is None:
            self.optimisation = {}
            self.optimisation['<name>'] = Optimisation('<name>', 'None', None)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        cols = ['Name', 'Approach', 'Values', 'Capacity Max', 'Capacity Min',
                'Capacity Step', 'Capacities']
        coln = [-1] * len(cols)
        for col in range(ws.ncols):
            try:
                i = cols.index(ws.cell_value(0, col))
                coln[i] = col
            except:
                pass
        if coln[0] < 0:
            self.setStatus('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        self.optimisation = {}
        for row in range(1, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is None:
                continue
            if coln[2] > 0: # values format
                self.optimisation[tech] = Optimisation(tech,
                                     ws.cell_value(row, coln[1]),
                                     ws.cell_value(row, coln[2]))
            else:
                if ws.cell_value(row, coln[1]) == 'Discrete': # fudge values format
                    self.optimisation[tech] = Optimisation(tech,
                                         ws.cell_value(row, coln[1]),
                                         ws.cell_value(row, coln[-1]))
                else:
                    self.optimisation[tech] = Optimisation(tech, '', '')
                    for col in range(1, len(coln)):
                        if coln[col] > 0:
                            attr = cols[col].lower().replace(' ', '_')
                            setattr(self.optimisation[tech], attr,
                                    ws.cell_value(row, coln[col]))
            try:
                self.optimisation[tech].capacity = self.generators[tech].capacity
            except:
                pass
        return

    def getJobFactors(self, ws):
        if ws is None:
            self.jobfactors = {}
            args = {'name': '<name>', 'year': '<years>', 'local_pct': '<lLocal_pct>', 'manufacture': '<manufacture>',
                    'install': '<install>', 'operate': '<operate>', 'dismantle': '<dismantle'}

            self.jobfactors['<name>'] = JobFactors(**args)
            return
        if ws.cell_value(0, 0) != 'Name':
            self.setStatus('Not a ' + self.file_labels[J] + ' worksheet.')
            return
        args = ['name', 'years', 'local_pct', 'manufacture', 'install', 'operate', 'dismantle']
        possibles = {'name': 0}
        for col in range(ws.ncols):
            try:
                arg = ws.cell_value(0, col).lower()
            except:
                continue
            if arg in args:
                possibles[arg] = col
            elif ws.cell_value(0, col) == 'Years':
                possibles['years'] = col
            elif ws.cell_value(0, col) == 'Local Pct' or ws.cell_value(0, col) == 'Local %age':
                possibles['local_pct'] = col
            elif ws.cell_value(0, col)[:10] == 'Manufactur':
                possibles['manufacture'] = col
            elif ws.cell_value(0, col)[:7] == 'Install' or ws.cell_value(0, col) == 'C & I':
                possibles['install'] = col
            elif ws.cell_value(0, col)[:6] == 'Operat' or ws.cell_value(0, col) == 'O & M':
                possibles['operate'] = col
            elif ws.cell_value(0, col) == 'Dismantle' or ws.cell_value(0, col)[:12] == 'Decommission':
                possibles['dismantle'] = col
        self.jobfactors = {}
        for row in range(1, ws.nrows):
            if ws.cell_value(row, 0) is None:
                continue
            in_args = {}
            for key, value in possibles.items():
                in_args[key] = ws.cell_value(row, value)
            self.jobfactors[str(ws.cell_value(row, 0))] = JobFactors(**in_args)
        return

    def getBatch(self, ws, option):
        global columns, rows, values
        def recurse(lvl):
            if lvl >= len(rows) - 1:
                return
            for i in range(len(values[lvl])):
                columns[lvl] = columns[lvl] + [values[lvl][i]] * cols[lvl+1]
                recurse(lvl + 1)

        def step_split(steps):
            bits = steps.split(',')
            if len(bits) == 1:
                bits = steps.split(';')
            try:
                strt = int(bits[0])
            except:
                return 0, 0, 0, -1
            try:
                stop = int(bits[1])
                step = int(bits[2])
                try:
                    frst = int(bits[3])
                except:
                    frst = -1
            except:
                return strt, strt, strt, frst
            return strt, stop, step, frst

        if ws is None:
            self.setStatus(self.file_labels[option] + ' worksheet missing.')
            return False
        istrt = 0
        year_row = -1
        for row in range(3):
            if ws.cell_value(row, 0) in ['Model', 'Model Label', 'Technology', 'Year']:
                istrt = row + 1
                break
        else:
            self.setStatus('Not a ' + self.file_labels[option] + ' worksheet.')
            return False
        self.batch_models = [{}] # cater for a range of capacities
        self.batch_report = [['Capacity (MW/MWh)', 1]]
        self.batch_tech = []
        istop = ws.nrows
        inrows = False
        for row in range(istrt, ws.nrows):
            tech = ws.cell_value(row, 0)
            if tech is not None and tech != '':
                if year_row < 0 and tech[:4].lower() == 'year':
                    year_row = row
                    continue
                inrows = True
                if tech[:8].lower() != 'capacity':
                    if tech.find('.') > 0:
                        tech = tech[tech.find('.') + 1:]
                    if tech != 'Total' and tech not in self.generators.keys():
                        self.setStatus(f'Unknown technology - {tech} - in {self.file_labels[option]} file.')
                        return False
                    self.batch_tech.append(ws.cell_value(row, 0))
                else:
                    self.batch_report[0][1] = row + 1
            elif inrows:
                istop = row
                break
            if tech[:5] == 'Total':
                istop = row + 1
                break
        if len(self.batch_tech) == 0:
            self.setStatus('No input technologies found in ' + self.file_labels[option] + ' worksheet (try opening and re-saving the workbook).')
            return False
        load_row = -1
        carbon_row = -1
        discount_row = -1
        for row in range(istop, ws.nrows):
            if ws.cell_value(row, 0) is not None and ws.cell_value(row, 0) != '':
                if ws.cell_value(row, 0).lower() in ['chart', 'graph', 'plot']:
                    self.batch_report.append(['Chart', row + 1])
                    break
                if ws.cell_value(row, 0).lower() in ['carbon price', 'carbon price ($/tco2e)']:
                    carbon_row = row
                if ws.cell_value(row, 0).lower() == 'discount rate' or ws.cell_value(row, 0).lower() == 'wacc':
                    discount_row = row
                if ws.cell_value(row, 0).lower() == 'load':
                    load_row = row
                self.batch_report.append([techClean(ws.cell_value(row, 0), full=True), row + 1])
        self.range_rows = {}
        for col in range(1, ws.ncols):
            model = ws.cell_value(istrt - 1, col)
            if model is None:
                break
            self.batch_models[0][col] = {'name': model}
            if option == T and year_row < 0:
                self.batch_models[0][col]['year'] = str(model)
            range_order = []
            btechs = []
            for row in range(istrt, istop):
                if row == year_row:
                    if ws.cell_value(row, col) is not None and ws.cell_value(row, col) != '':
                        self.batch_models[0][col]['year'] = str(ws.cell_value(row, col))
                    continue
                tech = ws.cell_value(row, 0)
                if tech in btechs:
                    self.setStatus(f'Duplicate technology, {tech}, found in {self.file_labels[option]} input worksheet (not yet allowed).')
                    return False
                btechs.append(tech)
                try:
                    if ws.cell_value(row, col) > 0:
                        self.batch_models[0][col][tech] = ws.cell_value(row, col)
                except:
                    if ws.cell_value(row, col) is None:
                        pass
                    elif ws.cell_value(row, col).find(',') >= 0 or ws.cell_value(row, col).find(';') >= 0:
                        try:
                            range_order.append([-1, row])
                        except:
                            pass
                        try:
                            strt, stop, step, frst = step_split(ws.cell_value(row, col))
                            range_order[-1][0] = frst
                            self.batch_models[0][col][tech] = strt
                        except:
                            pass
            if len(range_order) > 0:
                range_order.sort()
                for ro in range_order:
                    if ro[0] >= 0:
                        try:
                            self.range_rows[col].append(ro[1])
                        except:
                            self.range_rows[col] = [ro[1]]
                for ro in range_order:
                    if ro[0] < 0:
                        try:
                            self.range_rows[col].append(ro[1])
                        except:
                            self.range_rows[col] = [ro[1]]
            if carbon_row >= 0:
                if isinstance(ws.cell_value(carbon_row, col), float):
                    self.batch_models[0][col]['Carbon Price'] = ws.cell_value(carbon_row, col)
                elif isinstance(ws.cell_value(carbon_row, col), int):
                    self.batch_models[0][col]['Carbon Price'] = float(ws.cell_value(carbon_row, col))
            if discount_row >= 0:
                if isinstance(ws.cell_value(discount_row, col), float):
                    self.batch_models[0][col]['Discount Rate'] = ws.cell_value(discount_row, col)
                elif isinstance(ws.cell_value(discount_row, col), int):
                    self.batch_models[0][col]['Discount Rate'] = float(ws.cell_value(discount_row, col))
            if load_row >= 0:
                if isinstance(ws.cell_value(load_row, col), float):
                    self.batch_models[0][col]['Load'] = ws.cell_value(load_row, col)
                elif isinstance(ws.cell_value(load_row, col), int):
                    self.batch_models[0][col]['Load'] = float(ws.cell_value(load_row, col))
        if len(self.batch_models[0]) == 0:
            self.setStatus('No models found in ' + self.file_labels[option] + ' worksheet (try opening and re-saving the workbook).')
            return False
        if len(self.range_rows) == 0:
            return True
        # cater for ranges - so multiple batch_models lists
        for rcol, ranges in self.range_rows.items():
            rows = {}
            for rw in ranges:
                rows[rw] = ws.cell_value(rw, rcol)
            if len(ranges) > 1: # create sheet for each range else one sheet
                values = []
                cols = [1]
                for i in range(len(ranges) -1, 0, -1):
                    strt, stop, step, frst = step_split(rows[ranges[i]])
                    values.insert(0, [])
                    for stp in range(strt, stop + step, step):
                        values[0].append(stp)
                    cols.insert(0, cols[0] * len(values[0]))
                if cols[0] > 16384:
                    self.setStatus(f'Too many columns ({cols[0]:n}) in sheet. Maximum is 16384.')
                    return False
                columns = [[]] * len(rows)
                recurse(0)
                strt, stop, step, frst = step_split(rows[ranges[0]])
                shts = int((stop + step + step) / step)
                self.setStatus(f'{shts:n} (+1) batch sheets, {cols[0]:n} models per sheet ({(cols[0]) * shts:n} models (+1) in total)')
                my_tech = ws.cell_value(ranges[0], 0)
                tech_2 = ws.cell_value(ranges[1], 0)
              # produce new batch_models entry for first range tech
                techs = {}
                for c in range(1, len(ranges)):
                    techs[ws.cell_value(ranges[c], 0)] = c - 1
                bits = my_tech.split('.')
                strt, stop, step, frst = step_split(rows[ranges[0]]) # in case we forget
                for sht in range(strt, stop + step, step):
                    self.batch_models.append({})
                    for c2 in range(len(columns[0])):
                        self.batch_models[-1][c2] = {}
                        for key, value in self.batch_models[0][rcol].items():
                            self.batch_models[-1][c2][key] = value
                        self.batch_models[-1][c2][my_tech] = sht
                        for key, value in techs.items():
                            self.batch_models[-1][c2][key] = columns[value][c2]
                        self.batch_models[-1][c2]['name'] = f'{bits[-1]}_{sht}_{tech_2}'
            else:
                my_tech = ws.cell_value(ranges[0], 0)
                self.batch_models.append({})
                strt, stop, step, frst = step_split(rows[ranges[0]])
                c2 = -1
                for ctr in range(strt, stop + step, step):
                    c2 += 1
                    self.batch_models[-1][c2] = {}
                    if c2 == 0:
                        self.batch_models[-1][c2]['hdr'] = ws.cell_value(ranges[0], 0) # fudge to get header name
                    for key, value in self.batch_models[0][rcol].items():
                        self.batch_models[-1][c2][key] = value
                    self.batch_models[-1][c2][my_tech] = ctr
                #    for key, value in techs.items():
                 #       self.batch_models[-1][c2][key] = columns[value][c2]
                    self.batch_models[-1][c2]['name'] = f'Model {c2 + 1}'
        return True

    def setOrder(self):
        self.order.clear()
        self.ignore.clear()
        self.re_capacity = {}
        if self.generators is None:
            order = ['Storage', 'Biomass', 'PHES', 'Gas', 'CCG1', 'Other', 'Coal']
            for stn in order:
                self.order.addItem(stn)
        else:
            order = []
            zero = []
            for key, value in self.generators.items():
            #    if value.capacity == 0:
            #        continue
                if key in tech_names and key not in self.dispatchable:
                    self.re_capacity[key] = value.capacity
                    continue
                try:
                    gen = key[key.find('.') + 1:]
                    if gen in tech_names and gen not in self.dispatchable:
                        self.re_capacity[key] = value.capacity
                        continue
                except:
                    pass
                try:
                    o = int(value.order)
                    if o > 0:
                        while len(order) <= o:
                            order.append([])
                        order[o - 1].append(key)
                    elif o == 0:
                        zero.append(key)
                except:
                    pass
            order.append(zero)
            for cat in order:
                for stn in cat:
                    self.order.addItem(stn)

    def data_sources(self, sheet, sheet_row, pm_data_file, option):
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        sheet.cell(row=sheet_row, column=1).value = 'Data sources'
        sheet.cell(row=sheet_row, column=1).font = bold
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Preferences file'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        if len(sys.argv) > 1:
            config_file = getModelFile(sys.argv[1])
        else:
            config_file = getModelFile('SIREN.ini')
        sheet.cell(row=sheet_row, column=2).value = config_file[-1]
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Scenarios folder'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = self.scenarios
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Powermatch data file'
        sheet.cell(row=sheet_row, column=1).font = normal
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        sheet.cell(row=sheet_row, column=2).value = pm_data_file
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        try:
            if self.loadCombo.currentText() != 'n/a':
       #     if self.loadCombo.currentText() != 'n/a' or self.load_dir.text() != self._load_folder:
                sheet.cell(row=sheet_row, column=1).value = 'Load file'
                sheet.cell(row=sheet_row, column=1).font = normal
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sheet.cell(row=sheet_row, column=2).value = load_file
                sheet.cell(row=sheet_row, column=2).font = normal
                sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
                sheet_row += 1
        except:
            pass
        sheet.cell(row=sheet_row, column=1).value = 'Constraints worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        sheet.cell(row=sheet_row, column=2).value = str(self.files[C].text()) \
               + '.' + str(self.sheets[C].currentText())
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = 'Generators worksheet'
        sheet.cell(row=sheet_row, column=1).font = normal
        if option == T:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text()
        else:
            sheet.cell(row=sheet_row, column=2).value = self.files[G].text() \
                   + '.' + self.sheets[G].currentText()
        sheet.cell(row=sheet_row, column=2).font = normal
        sheet.merge_cells('B' + str(sheet_row) + ':M' + str(sheet_row))
        return sheet_row

    def clean_batch_sheet(self, option):
        msgbox = QtWidgets.QMessageBox()
        msgbox.setWindowTitle(f'SIREN - Powermatch {self.file_labels[option]}')
        msgbox.setText(f"{self.file_labels[option]} worksheet has more that 1,024 columns.\nSome may be invalid/empty. Would you like these to be removed")
        msgbox.setIcon(QtWidgets.QMessageBox.Warning)
        msgbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        reply = msgbox.exec_()
        if reply != QtWidgets.QMessageBox.Yes:
            return
        batch_report_file = self.get_filename(self.files[option].text())
        if os.path.exists(batch_report_file + '~'):
            os.remove(batch_report_file + '~')
        shutil.copy2(batch_report_file, batch_report_file + '~')
        wb = oxl.load_workbook(batch_report_file)
        ws = wb.worksheets[0]
        for row in range(1, 4):
            try:
                if ws.cell(row=row, column=1).value.lower() in ['model', 'model label', 'technology', 'year']:
                    break
            except:
                pass
        else:
            return # bail out
        for col in range(ws.max_column, 1, -1):
            if ws.cell(row=row, column=col).value is None:
               ws.delete_cols(col, 1)
        wb.save(batch_report_file)

    def jobsClicked(self):
      #  if self.sender().text() == 'Jobs': # jobs spreadsheet?
        option = J
        wb = WorkBook()
        try:
            wb.open_workbook(self.get_filename(self.files[T].text()))
        except FileNotFoundError:
            self.setStatus(f'{self.file_labels[T]} file not found - {self.files[T].text()}')
            return
        except Exception as e:
            self.setStatus(f'Error accessing {self.file_labels[T]} file {str(e)}')
            return
        years = {}
        rpt_sht = ''
        for sht in wb.sheet_names():
            if sht[:8] == 'Results_' and sht > rpt_sht:
                rpt_sht = sht
        if rpt_sht == '':
            wb.close()
            self.setStatus(f'No Report sheet in {self.files[T].text()}')
            return
        ws = wb.sheet_by_name(rpt_sht)
        for row in range(5):
            if ws.cell_value(row, 0) in ['Model', 'Model Label', 'Technology', 'Year']:
                model_row = row
                break
        else:
            wb.close()
            self.setStatus(f'No Model (Year) row found in {self.files[T].text()}')
            return
        techs = {}
        for row in range(model_row + 1, ws.nrows):
            if ws.cell_value(row, 0) is None or ws.cell_value(row, 0) == '' or ws.cell_value(row, 0) == '' :
                break
            if ws.cell_value(row, 0).lower().find('capacity') >= 0:
                continue
            fac = ws.cell_value(row, 0)
            if fac[:3] == 'CY_':
                fac = fac[3:]
            i = fac.find('.')
            if i >= 0:
                fac = fac[i + 1:]
            if fac in techs.keys():
                techs[fac][1].append(row)
            else:
                techs[fac] = [0, [row], 0]
        jobs = {}
        years = [ws.cell_value(model_row, 1)]
        # set up for existing
        for key in techs.keys():
            jobs[key] = []
            capacity = 0
            for row in techs[key][1]:
                try:
                    capacity += ws.cell_value(row, 1)
                except:
                    pass
            techs[key][0] = capacity
            jobs[key].append([0, 0, capacity * self.jobfactors[key].operate, 0])
        # now for transition
        for col in range(2, ws.ncols):
            if ws.cell_value(model_row, col) is None:
                break
            years.append(ws.cell_value(model_row, col))
            for key in techs.keys():
                capacity = 0
                for row in techs[key][1]:
                    try:
                        capacity += ws.cell_value(row, col)
                    except:
                        pass
                jobs[key].append([0, 0, 0, 0])
                delta = capacity - techs[key][0]
                if delta > 0:
                    jobs[key][-2][0] = delta * self.jobfactors[key].manufacture * self.jobfactors[key].local_pct
                    jobs[key][-1][1] = delta * self.jobfactors[key].install
                    jobs[key][-1][2] = delta * self.jobfactors[key].operate
                elif delta < 0:
                    jobs[key][-1][3] = -delta * self.jobfactors[key].dismantle
                    jobs[key][-1][2] = delta * self.jobfactors[key].operate
                techs[key][0] = capacity
                techs[key][2] = delta
        wb.close()
        # now open for edit to write new sheet
        cats = ['Manufacturing', 'Construction & Installation', 'Operations & Maintenance', 'Decommissioning']
        wb = oxl.load_workbook(self.get_filename(self.files[T].text()))
        rpt_sht = rpt_sht.replace('Results_', 'Jobs_')
        if rpt_sht in wb.sheetnames:
            del wb[rpt_sht]
        bs = wb.create_sheet(rpt_sht)
        normal = oxl.styles.Font(name='Arial')
        bold = oxl.styles.Font(name='Arial', bold=True)
        bs.cell(row=1, column=1).value = 'Jobs Report'
        bs.cell(row=1, column=1).font = bold
        bs.cell(row=2, column=1).value = 'Technology'
        bs.cell(row=2, column=1).font = bold
        bs.cell(row=2, column=2).value = 'Category'
        bs.cell(row=2, column=2).font = bold
        for y in range(len(years)):
            bs.cell(row=2, column=y + 3).value = years[y]
            bs.cell(row=2, column=y + 3).font = normal
            bs.cell(row=2, column=y + 3).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
        row = 2
        for key, value in jobs.items():
            for c in range(4):
                row += 1
                bs.cell(row=row, column=1).value = key
                bs.cell(row=row, column=1).font = normal
                bs.cell(row=row, column=2).value = cats[c]
                bs.cell(row=row, column=2).font = normal
                for y in range(len(years)):
                    bs.cell(row=row, column=y + 3).value = value[y][c]
                    bs.cell(row=row, column=y + 3).number_format = '#,##0'
                    bs.cell(row=row, column=y + 3).font = normal
        lst_row = row
        row += 1
        bs.cell(row=row, column=1).value = 'Total Jobs'
        bs.cell(row=row, column=1).font = bold
        for y in range(len(years)):
            bs.cell(row=row, column=y + 3).value = f'=SUM({ssCol(y + 3)}$3:{ssCol(y + 3)}${lst_row})'
            bs.cell(row=row, column=y + 3).number_format = '#,##0'
            bs.cell(row=row, column=y + 3).font = normal
        row += 1
        cat_row = [row + 1, 0]
        for c in range(len(cats)):
            row += 1
            bs.cell(row=row, column=2).value = cats[c]
            bs.cell(row=row, column=2).font = normal
            for y in range(len(years)):
                bs.cell(row=row, column=y + 3).value = f'=SUMIF($B$3:$B${lst_row},$B{row},{ssCol(y + 3)}$3:{ssCol(y + 3)}${lst_row})'
                bs.cell(row=row, column=y + 3).number_format = '#,##0'
                bs.cell(row=row, column=y + 3).font = normal
        cat_row[1] = row
        row += 1
        tec_row = [row + 1, 0]
        for key in techs.keys():
            row += 1
            bs.cell(row=row, column=1).value = key
            bs.cell(row=row, column=1).font = normal
            if self.job_charts:
                bs.cell(row=row, column=2).value = key
                bs.cell(row=row, column=2).font = normal
          #      bs.cell(row=row, column=2).number_format = ';;;' # hide
            for y in range(len(years)):
                bs.cell(row=row, column=y + 3).value = f'=SUMIF($A$3:$A${lst_row},$A{row},{ssCol(y + 3)}$3:{ssCol(y + 3)}${lst_row})'
                bs.cell(row=row, column=y + 3).number_format = '#,##0'
                bs.cell(row=row, column=y + 3).font = normal
        tec_row[1] = row
        length = 0
        for key in techs.keys():
            length = max(length, len(key))
            bs.column_dimensions['A'].width = max(length, 10) * 1.2
        length = 0
        for cat in cats:
            length = max(length, len(cat))
            bs.column_dimensions['B'].width = max(length, 10) #* 1.2
        bs.freeze_panes = 'C3'
        bs.activeCell = 'C3'
        if self.job_charts:
            cht_sht = rpt_sht.replace('Jobs_', 'Job_Charts_')
            if cht_sht in wb.sheetnames:
                del wb[cht_sht]
            cs = wb.create_sheet(cht_sht)
            min_col = 3
            max_col = len(years) + 2
            chart = BarChart(grouping='stacked', gapWidth=50)
            values = Reference(bs, min_col=2, min_row=cat_row[0], max_col=max_col, max_row=cat_row[1])
            chart.add_data(values, from_rows=True, titles_from_data=True)
            xcats = Reference(bs, min_row=2, min_col=3, max_col=max_col)
            chart.set_categories(xcats)
            chart.title = 'Jobs by Category'
            chart.x_axis.title = 'Year'
            chart.y_axis.title = 'No. of Jobs'
            chart.height = 15
            chart.width = 30
            cs.add_chart(chart, 'A1')
            chart2 = BarChart(grouping='stacked', gapWidth=50)
            values = Reference(bs, min_col=2, min_row=tec_row[0], max_col=max_col, max_row=tec_row[1])
            chart2.add_data(values, from_rows=True, titles_from_data=True)
            xcats = Reference(bs, min_row=2, min_col=3, max_col=max_col)
            chart2.set_categories(xcats)
            chart2.title = 'Jobs by Technology'
            chart2.x_axis.title = 'Year'
            chart2.y_axis.title = 'No. of Jobs'
            chart2.height = 15
            chart2.width = 30
            cs.add_chart(chart2, 'A30')
            cht_sht = f'and {cht_sht} sheets'
        else:
            cht_sht = 'sheet'
        for sheet in wb:
            wb[sheet.title].views.sheetView[0].tabSelected = False
        wb.active = bs
        wb.save(self.get_filename(self.files[T].text()))
        self.setStatus(f'{rpt_sht} {cht_sht} added to {self.files[T].text()}')
        return

    def pmClicked(self):
        def get_load_data(load_file):
            try:
                tf = open(load_file, 'r')
                lines = tf.readlines()
                tf.close()
            except:
                return None
            load_data = []
            bit = lines[0].rstrip().split(',')
            if len(bit) > 0: # multiple columns
                for b in range(len(bit)):
                    if bit[b][:4].lower() == 'load':
                        if bit[b].lower().find('kwh') > 0: # kWh not MWh
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]) * 0.001)
                        else:
                            for i in range(1, len(lines)):
                                bit = lines[i].rstrip().split(',')
                                load_data.append(float(bit[b]))
            else:
                for i in range(1, len(lines)):
                    load_data.append(float(lines[i].rstrip()))
            return load_data

        def get_batch_prefix(report_group):
            if report_group == 'Lifetime Emissions':
                return 'LES_'
            if report_group in ['Correlation To Load', 'Static Variables']:
                return ''
            bits = report_group.split(' ')
            for i in range(len(bits) -1, -1, -1):
                if bits[i][0] == '(' and bits[i][-1] == ')':
                    del bits[i]
            if len(bits) == 1:
                abr = bits[0][0] + bits[0][-1]
            else:
                abr = ''
                for bit in bits:
                    abr += bit[0]
            return abr.upper() + '_'

        col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.setStatus(self.sender().text() + ' processing started')
        if self.sender().text() == 'Detail': # detailed spreadsheet?
            option = D
        elif self.sender().text() == 'Optimise': # do optimisation?
            option = O
            self.optExit = False #??
        elif self.sender().text() == 'Batch': # do batch processsing
            option = B
        elif self.sender().text() == 'Transition': # do transition processsing
            option = T
        else:
            option = S
        if option != O:
            self.progressbar.setMinimum(0)
            self.progressbar.setMaximum(20)
            self.progressbar.setHidden(False)
            QtWidgets.QApplication.processEvents()
        err_msgs = []
        if self.constraints is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[C].text()))
                ws = ts.sheet_by_name(self.sheets[C].currentText())
                self.getConstraints(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                err_msgs.append('Constraints file not found - ' + self.files[C].text())
                self.getConstraints(None)
            except:
                err_msgs.append('Error accessing Constraints')
                self.getConstraints(None)
        if self.generators is None:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[G].text()))
                ws = ts.sheet_by_name(self.sheets[G].currentText())
                self.getGenerators(ws)
                ts.close()
                del ts
            except FileNotFoundError:
                err_msgs.append('Generators file not found - ' + self.files[G].text())
                self.getGenerators(None)
            except:
                err_msgs.append('Error accessing Generators')
                self.getGenerators(None)
        pm_data_file = self.get_filename(self.files[D].text())
        if pm_data_file[-5:] != '.xlsx': #xlsx format only
            err_msgs.append('Not a Powermatch data spreadsheet (1)')
        elif not os.path.exists(pm_data_file):
                err_msgs.append('Data file not found - ' + self.files[D].text())
        if (option == B or option == T) and len(err_msgs) == 0: # has to be xlsx workbook
            try:
                ts = WorkBook()
                bwbopen_start = time.time()
                ts.open_workbook(self.get_filename(self.files[option].text()))
                ws = ts.sheet_by_index(0)
                if ws.ncols > 1024:
                    ts.close()
                    self.clean_batch_sheet(option)
                    ts = WorkBook()
                    ts.open_workbook(self.get_filename(self.files[option].text()))
                    ws = ts.sheet_by_index(0)
                tim = time.time() - bwbopen_start
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'{self.file_labels[option]} workbook opened ({tim})')
                ok = self.getBatch(ws, option)
                ts.close()
                del ts
                if not ok:
                    return
            except FileNotFoundError:
                err_msgs.append(f'{self.file_labels[option]} file not found - {self.files[option].text()}')
            except Exception as e:
                err_msgs.append(f'Error accessing {self.file_labels[option]} file {str(e)}')
        if option == O and self.optimisation is None and len(err_msgs) == 0:
            try:
                ts = WorkBook()
                ts.open_workbook(self.get_filename(self.files[O].text()))
                ws = ts.sheet_by_name(self.sheets[O].currentText())
                self.getOptimisation(ws)
                ts.close()
                del ts
                if self.optimisation is None:
                    err_msgs.append('Not an optimisation worksheet')
            except FileNotFoundError:
                err_msgs.append('Optimisation file not found - ' + self.files[O].text())
            except:
                err_msgs.append('Error accessing Optimisation')
            if self.optimisation is None:
                self.getOptimisation(None)
        if self.do_jobs:
            if self.files[J].text() != '' and os.path.exists(self.get_filename(self.files[J].text())):
                pass
            else:
                err_msgs.append(f'{self.file_labels[J]} file not found - {self.files[J].text()}')
        if len(err_msgs) > 1:
            self.log_status = True
            self.show_FloatStatus() # status window
        if len(err_msgs) > 0:
            self.progressbar.setHidden(True)
            if len(err_msgs) == 1:
                self.setStatus(err_msgs[0])
            else:
                for error in err_msgs:
                    self.setStatus(error)
                self.setStatus('Execution aborted.')
            return
        try:
            ts = oxl.load_workbook(pm_data_file)
        except FileNotFoundError:
            self.setStatus('Data file not found - ' + self.files[D].text())
            return
        except:
            self.setStatus('Error accessing Data file - ' + self.files[D].text())
            return
        ws = ts.worksheets[0]
        top_row = ws.max_row - 8760
        if top_row < 1 or (ws.cell(row=top_row, column=1).value != 'Hour' \
                           or ws.cell(row=top_row, column=2).value != 'Period'):
            self.setStatus(f'Not a Powermatch data spreadsheet (2; {top_row})')
            self.progressbar.setHidden(True)
            return
        typ_row = top_row - 1
        gen_row = typ_row
        while typ_row > 0:
            if ws.cell(row=typ_row, column=1).value[:9] == 'Generated':
                gen_row = typ_row
            if ws.cell(row=typ_row, column=3).value in tech_names:
                break
            typ_row -= 1
        else:
            self.setStatus('No suitable data (in data file)')
            return
        do_zone = False
        zone_row = typ_row - 1
        try:
            if ws.cell(row=zone_row, column=1).value.lower() == 'zone':
                do_zone = True
                zone_techs = []
        except:
            pass
        icap_row = typ_row + 1
        while icap_row < top_row:
            if ws.cell(row=icap_row, column=1).value[:8] == 'Capacity':
                break
            icap_row += 1
        else:
            self.setStatus('No capacity data (in data file)')
            return
        year = ws.cell(row=top_row + 1, column=2).value[:4]
        pmss_details = {} # contains name, generator, capacity, fac_type, col, multiplier
        pmss_data = []
        re_order = [] # order for re technology
        dispatch_order = [] # order for dispatchable technology
        load_columns = {}
        load_col = -1
        strt_col = 3
        try:
            if self.loadCombo.currentText() != 'n/a':
                year = self.loadCombo.currentText()
                strt_col = 4
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
                pmss_details['Load'] = PM_Facility('Load', 'Load', 0, 'L', len(pmss_data), 1)
                load_columns[self.loadCombo.currentText()] = len(pmss_data)
                pmss_data.append([])
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if self.load_dir.text() != self._load_folder:
                    load_file = self.get_filename(load_file)
                pmss_data[-1] = get_load_data(load_file)
                re_order.append('Load')
        except:
            pass
        zone = ''
        for col in range(strt_col, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                continue
            key = tech_names[i]
            if key == 'Load':
                load_col = len(pmss_data)
                typ = 'L'
                capacity = 0
                fctr = 1
            else:
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                    zone_techs.append(key)
                else: # temp
                    if len(self.re_capacity) > 0 and tech_names[i] not in self.re_capacity.keys():
                        continue
                try:
                    capacity = float(ws.cell(row=icap_row, column=col).value)
                except:
                    continue
                if capacity <= 0:
                    continue
                typ = 'R'
                if do_zone:
                    fctr = 1
                elif tech_names[i] in self.re_capacity and capacity > 0:
                    fctr = self.re_capacity[tech_names[i]] / capacity
                else:
                    fctr = 1
            pmss_details[key] = PM_Facility(key, tech_names[i], capacity, typ, len(pmss_data), fctr)
            if key == 'Load':
                load_columns[year] = len(pmss_data)
            pmss_data.append([])
            re_order.append(key)
            for row in range(top_row + 1, ws.max_row + 1):
                pmss_data[-1].append(ws.cell(row=row, column=col).value)
       #     if option == O and key not in self.optimisation.keys():
       #         self.optimisation[key] = Optimisation(key, 'Range', f'{capacity} {capacity} {capacity}')
        pmss_details['Load'].capacity = sum(pmss_data[load_col])
        do_adjust = False
        if option == O:
            for itm in range(self.order.count()):
                gen = self.order.item(itm).text()
                try:
                    if self.generators[gen].capacity <= 0:
                        continue
                except KeyError as err:
                    self.setStatus('Key Error: No Generator entry for ' + str(err))
                    continue
                try:
                    if self.generators[gen].constraint in self.constraints and \
                      self.constraints[self.generators[gen].constraint].category == 'Generator':
                        typ = 'G'
                    else:
                        typ = 'S'
                except:
                    continue
                dispatch_order.append(gen)
                pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
            if self.adjust.isChecked():
                 pmss_details['Load'].multiplier = self.adjustto['Load'] / pmss_details['Load'].capacity
            self.optClicked(year, option, pmss_details, pmss_data, re_order, dispatch_order,
                            None, None)
            return
        if self.adjust.isChecked() and option != B and option != T:
            if self.adjustto is None:
                self.adjustto = {}
                self.adjustto['Load'] = 0
                if do_zone:
                    tns = zone_techs[:]
                else:
                    tns = tech_names[:]
                for gen in tns:
                    try:
                        if self.generators[gen].capacity > 0:
                            self.adjustto[gen] = self.generators[gen].capacity
                    except:
                        pass
                for i in range(self.order.count()):
                    gen = self.order.item(i).text()
                    try:
                        if self.generators[gen].capacity > 0:
                            self.adjustto[gen] = self.generators[gen].capacity
                    except:
                       pass
            generated = sum(pmss_data[load_col])
            datain = []
            datain.append(['Load', 'L', generated])
            if self.adjustto['Load'] == 0:
                self.adjustto['Load'] = generated
            for col in range(4, ws.max_column + 1):
                try:
                    valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                    i = tech_names.index(valu)
                except:
                    continue
                key = tech_names[i]
                if do_zone:
                    cell = ws.cell(row=zone_row, column=col)
                    if type(cell).__name__ == 'MergedCell':
                        pass
                    else:
                        zone = ws.cell(row=zone_row, column=col).value
                    if zone is None or zone == '':
                        zone_tech = valu
                    else:
                        zone_tech = zone + '.' + valu
                    key = zone_tech
                try:
                    typ = self.constraints[tech_names[i]].category[0]
                    if typ == '':
                        typ = 'R'
                    datain.append([key, typ, float(ws.cell(row=icap_row, column=col).value)])
                except:
                    try:
                        datain.append([key, 'R', float(ws.cell(row=icap_row, column=col).value)])
                    except:
                        pass
            for i in range(self.order.count()):
                try:
                    if self.generators[self.order.item(i).text()].capacity > 0:
                        gen = self.order.item(i).text()
                        try:
                            if self.generators[gen].constraint in self.constraints and \
                               self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                        except:
                            continue
                        datain.append([gen, typ, self.generators[gen].capacity])
                except:
                    pass
            adjust = Adjustments(self, datain, self.adjustto, self.adjust_cap, self.results_prefix,
                                 show_multipliers=self.show_multipliers, save_folder=self.scenarios,
                                 batch_file=self.get_filename(self.files[B].text()))
            adjust.exec_()
            if adjust.getValues() is None:
                self.setStatus('Execution aborted.')
                self.progressbar.setHidden(True)
                return
            self.adjustto = adjust.getValues()
            results_prefix = adjust.getPrefix()
            if results_prefix != self.results_prefix:
                self.results_prefix = results_prefix
                self.results_pfx_fld.setText(self.results_prefix)
            self.updated = True
            do_adjust = True
        ts.close()
        self.progressbar.setValue(0) # was 1
        QtWidgets.QApplication.processEvents()
        if self.files[R].text() == '':
            i = pm_data_file.rfind('/')
            if i >= 0:
                rslts_file = pm_data_file[i + 1:]
            else:
                rslts_file = pm_data_file
            rslts_file = rslts_file.replace('data', 'results')
            rslts_file = rslts_file.replace('Data', 'Results')
            if rslts_file == pm_data_file[i + 1:]:
                j = rslts_file.find(' ')
                if j > 0:
                    jnr = ' '
                else:
                    jnr = '_'
                j = rslts_file.rfind('.')
                rslts_file = rslts_file[:j] + jnr + 'Results' + rslts_file[j:]
            self.files[R].setText(rslts_file)
        else:
            rslts_file = self.get_filename(self.files[R].text())
        if self.results_prefix != '':
            j = rslts_file.rfind('/')
            if rslts_file[j + 1:j + 1 + len(self.results_prefix)] != self.results_prefix:
                rslts_file = rslts_file[: j + 1] + self.results_prefix + '_' + rslts_file[j + 1:]
        for itm in range(self.order.count()):
            gen = self.order.item(itm).text()
            try:
                if self.generators[gen].capacity <= 0:
                    continue
            except KeyError as err:
                self.setStatus('Key Error: No Generator entry for ' + str(err))
                continue
            except:
                continue
            if do_adjust:
                try:
                    if self.adjustto[gen] <= 0:
                        continue
                except:
                    pass
            try:
                if self.generators[gen].constraint in self.constraints and \
                  self.constraints[self.generators[gen].constraint].category == 'Generator':
                    typ = 'G'
                else:
                    typ = 'S'
            except:
                continue
            dispatch_order.append(gen)
            pmss_details[gen] = PM_Facility(gen, gen, self.generators[gen].capacity, typ, -1, 1)
        if option == B or option == T:
            if option == T:
                files = setTransition(self, self.file_labels[G], self.get_filename(self.files[G].text()), self.sheets[G].currentText(),
                                      self.loadCombo.currentText(), self.load_files)
                files.exec_()
                if files.getValues() is None:
                    self.setStatus('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
                gen_sheet = files.getValues()
                trn_year = ''
                newfile = self.get_filename(self.files[G].text())
                gen_book = WorkBook()
                gen_book.open_workbook(newfile)
                pmss_details['Load'].multiplier = 1
            elif self.adjust.isChecked():
                generated = sum(pmss_data[load_col])
                datain = [['Load', 'L', generated]]
                adjustto = {'Load': generated}
                adjust = Adjustments(self, datain, adjustto, self.adjust_cap, None,
                                     show_multipliers=self.show_multipliers)
                adjust.exec_()
                if adjust.getValues() is None:
                    self.setStatus('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
                adjustto = adjust.getValues()
                pmss_details['Load'].multiplier = adjustto['Load'] / pmss_details['Load'].capacity
            save_load_multiplier = pmss_details['Load'].multiplier # in case load row passed into batch
            load_varies = False
       #     start_time = time.time() # just for fun
            batch_details = {'Capacity (MW/MWh)': [st_cap, '#,##0.00'],
                             'To Meet Load (MWh)': [st_tml, '#,##0'],
                             'Generation (MWh)': [st_sub, '#,##0'],
                             'Capacity Factor': [st_cfa, '#,##0.0%'],
                             'Cost ($/Yr)': [st_cst, '#,##0'],
                             'LCOG ($/MWh)': [st_lcg, '#,##0.00'],
                             'LCOE ($/MWh)': [st_lco, '#,##0.00'],
                             'Emissions (tCO2e)': [st_emi, '#,##0'],
                             'Emissions Cost': [st_emc, '#,##0'],
                             'LCOE With CO2 ($/MWh)': [st_lcc, '#,##0.00'],
                             'Max MWh': [st_max, '#,##0'],
                             'Capital Cost': [st_cac, '#,##0'],
                             'Lifetime Cost': [st_lic, '#,##0'],
                             'Lifetime Emissions': [st_lie, '#,##0'],
                             'Lifetime Emissions Cost': [st_lec, '#,##0'],
                             'Area': [st_are, '#,###0.00']}
            if self.do_jobs:
                batch_details['Jobs'] = [st_job, '#,##0']
            batch_extra = {'RE': ['#,##0.00', ['RE %age', st_cap], ['Storage %age', st_cap], ['RE %age of Total Load', st_cap]],
                           'Load Analysis': ['#,##0', ['Total Load', st_tml], ['Shortfall', st_tml], ['Load met', st_tml], ['Load met %age', st_cap],
                           ['Largest Shortfall', st_cap], ['Storage losses', st_sub], ['Surplus', st_sub], ['Surplus %age', st_cap]],
                           'Carbon': ['#,##0.00', ['Carbon Price', st_cap], ['Carbon Cost', st_emc], ['LCOE incl. Carbon Cost', st_lcc],
                           ['Lifetime Emissions Cost', st_lec]],
                           'Correlation To Load': ['0.0000', ['RE Contribution', st_cap], ['RE plus Storage', st_cap],
                           ['To Meet Load', st_cap]],
                           'Static Variables': ['#,##0.00', ['Carbon Price', st_cap], ['Lifetime', st_cap],
                           ['Discount Rate', st_cap]],
                           'Optimisation Parameters': ['#,##0.00', ['Population size', 1], ['No. of iterations', 1],
                           ['Mutation probability', 1], ['Exit if stable', 1], ['Optimisation choice', 1],
                           ['Variable', 1], ['LCOE', 1], ['Load%', 1], ['Surplus%', 1], ['RE%', 1],
                           ['Cost', 1], ['CO2', 1]]}
                           # LCOE (incl. CO2)
         #   batch_extra['Optimisation Parameters'] = []
            batch_extra['LCOE ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                if tech == 'Total':
                    batch_extra['LCOE ($/MWh)'].append([tech + ' LCOE ($/MWh)'])
                else:
                    batch_extra['LCOE ($/MWh)'].append([tech])
            batch_extra['LCOE ($/MWh)'].append(['LCOE', st_lco])
            batch_extra['LCOE With CO2 ($/MWh)'] = ['#,##0.00']
            for tech in self.batch_tech:
                batch_extra['LCOE With CO2 ($/MWh)'].append([tech])
            batch_extra['LCOE With CO2 ($/MWh)'].append(['LCOE incl. Carbon Cost', st_lcc])
         #   batch_extra['To Meet Load (MWh)'] = ['#,##0.00', ['Total', st_tml]]
            wbopen_start = time.time()
            wb = oxl.load_workbook(self.get_filename(self.files[option].text()))
            tim = time.time() - wbopen_start
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.file_labels[option]} workbook re-opened for update ({tim})')
            batch_input_sheet = wb.worksheets[0]
            rpt_time = QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), 'yyyy-MM-dd_hhmm')
            if self.batch_new_file:
                wb.close()
                i = self.files[option].text().rfind('.')
                suffix = '_report_' + rpt_time
                batch_report_file = self.get_filename(self.files[option].text()[:i] + suffix + self.files[option].text()[i:])
                batch_report_file = QtWidgets.QFileDialog.getSaveFileName(None, f'Save {self.file_labels[option]} Report file',
                                    batch_report_file, 'Excel Files (*.xlsx)')[0]
                if batch_report_file == '':
                    self.setStatus(self.sender().text() + ' aborted')
                    return
                if batch_report_file[-5:] != '.xlsx':
                    batch_report_file += '.xlsx'
                if os.path.exists(batch_report_file) and not self.replace_last.isChecked():
                    wb = oxl.load_workbook(batch_report_file)
                    bs = wb.create_sheet('Results_' + rpt_time)
                else:
                    wb = oxl.Workbook()
                    bs = wb.active
                    bs.title = 'Results_' + rpt_time
            else:
                batch_report_file = self.get_filename(self.files[option].text())
                if self.replace_last.isChecked():
                    del_sht = ''
                    for sht in wb.sheetnames:
                        if sht[:8] == 'Results_' and sht > del_sht:
                            del_sht = sht
                    if del_sht != '':
                        del wb[del_sht]
                        del_sht = del_sht.replace('Results', 'Charts')
                        if del_sht in wb.sheetnames:
                            del wb[del_sht]
                bs = wb.create_sheet('Results_' + rpt_time)
            start_time = time.time() # just for fun
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            grey = oxl.styles.colors.Color(rgb='00f2f2f2')
            grey_fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor=grey)
            total_models = 0
            for sht in range(len(self.batch_models)):
                total_models = total_models + len(self.batch_models[sht])
            try:
                incr = 20 / total_models
            except:
                incr = .05
            prgv = incr
            prgv_int = 0
            model_row = False
            model_row_no = 0
            sht_nam_len = max(len(str(len(self.batch_models))), 2)
            if len(self.batch_models) > 1: #multiple sheets
                d3_rng1 = 0 # count of 2nd level groups in sheet
                d3_rng2 = -1 # number of cells for 2nd range
                for rcol, d3_ranges in self.range_rows.items():
                    break
                if self.batch_3d and len(d3_ranges) > 1:
                    ndx = -1
                    if '3D Summary' in wb.sheetnames:
                        ndx = wb.sheetnames.index('3D Summary')
                        del wb['3D Summary']
                    d3s = wb.create_sheet('3D Summary')
                    if ndx >= 0:
                        ndx = ndx - wb.sheetnames.index('3D Summary')
                        if ndx != 0:
                            wb.move_sheet('3D Summary', offset=ndx)
                    d3s.cell(row=1, column=1).value = '3D Summary'
                    d3s.cell(row=1, column=1).font = bold
                    d3s.cell(row=2, column=1).value = 'Choose row of interest from this pull-down list (selection values in hidden rows above it)'
                    d3s.merge_cells('A2:G2')
                    d3_selections = [['LCOE','<', 'LE'], ['LCOE incl. Carbon Cost', '<', 'CN'], ['RE %age of Total Load', '>', 'RE'],
                                     ['Load met %age', '>', 'LA'], ['Storage %age', '>', 'RE'], ['Surplus %age', '<', 'LA'],
                                     ['Total LCOG ($/MWh)', '<', 'LG'], ['RE %age', '>', 'RE']]
                    for rw in range(len(d3_selections)):
                        if self.batch_prefix:
                            d3s.cell(row=rw + 3, column=1).value = f'{d3_selections[rw][2]}_{d3_selections[rw][0]}'
                        else:
                            d3s.cell(row=rw + 3, column=1).value = d3_selections[rw][0]
                        d3s.cell(row=rw + 3, column=2).value = d3_selections[rw][1]
                        d3s.cell(row=rw + 3, column=2).alignment = oxl.styles.Alignment(horizontal='center')
                        d3s.row_dimensions[rw + 3].hidden = True
                    d3s.row_dimensions[2].hidden = False # seems to get hidden
                    formula1 = f"'3D Summary'!$A$3:$A${rw + 3}"
                    dv3 = DataValidation(type='list', formula1=formula1, allow_blank=True)
                    d3s.add_data_validation(dv3)
                    dv3_cells = f'A{rw + 4}:A{rw + 4}'
                    dv3.add(dv3_cells)
                    acolor = oxl.styles.colors.Color(rgb='00ebbd34')
                    cell_fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor=acolor)
                    d3s.cell(row=rw + 4, column=1).fill = cell_fill
                    d3s.cell(row=rw + 4, column=1).value = d3s.cell(row=3, column=1).value
                    d3s.cell(row=rw + 4, column=2).value = f'=VLOOKUP(A{rw + 4},A$3:B${rw + 3},2,0)'
                #    d3s.cell(row=rw + 3, column=2).fill = cell_fill
                    d3s.cell(row=rw + 4, column=2).alignment = oxl.styles.Alignment(horizontal='center')
                    d3s.column_dimensions['A'].width = 24
                else:
                    d3_ranges = None
            chart_groups = []
            for sht in range(len(self.batch_models)):
                sheet_start = time.time()
                if sht == 0: # normal case
                   # copy header rows to new worksheet
                   merged_cells = []
                   merge_cells = None
                   model_row = False
                   model_cols = len(self.batch_models[sht])
                   for row in range(1, self.batch_report[0][1] + 2):
                       if batch_input_sheet.cell(row=row, column=1).value in ['Model', 'Model Label', 'Technology', 'Year']:
                           model_row = True
                           model_row_no = row
                       else:
                           model_row = False
                       for col in range(1, model_cols + 2):
                           cell = batch_input_sheet.cell(row=row, column=col)
                           if type(cell).__name__ == 'MergedCell':
                               if merge_cells is None:
                                   merge_cells = [row, col - 1, col]
                               else:
                                   merge_cells[2] = col
                               continue
                           if model_row and col > 1:
                               new_cell = bs.cell(row=row, column=col, value=self.batch_models[sht][col - 1]['name'])
                           else:
                               new_cell = bs.cell(row=row, column=col, value=cell.value)
                           if cell.has_style:
                               new_cell.font = copy(cell.font)
                               new_cell.border = copy(cell.border)
                               new_cell.fill = copy(cell.fill)
                               new_cell.number_format = copy(cell.number_format)
                               new_cell.protection = copy(cell.protection)
                               new_cell.alignment = copy(cell.alignment)
                           if merge_cells is not None:
                               bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                               merged_cells.append(merge_cells)
                               merge_cells = None
                       if merge_cells is not None:
                           bs.merge_cells(start_row=row, start_column=merge_cells[1], end_row=row, end_column=merge_cells[2])
                           merged_cells.append(merge_cells)
                           merge_cells = None
                   try:
                       normal = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz)
                       bold = oxl.styles.Font(name=cell.font.name, sz=cell.font.sz, bold=True)
                   except:
                       pass
                else:
                    sheet_name = f'{sht:0{sht_nam_len}}'
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
                        if 'Charts_' + sheet_name in wb.sheetnames:
                            del wb['Charts_' + sheet_name]
                    bs = wb.create_sheet(sheet_name)
                    if model_row_no > 1:
                        title = self.batch_models[sht][0]['name']
                        tech_2 = title.split('_')
                        if len(tech_2) > 1:
                            tech_2 = tech_2[-1]
                            bits_2 = tech_2.split('.')[-1]
                            title = title.replace(tech_2, bits_2)
                            cap_2 = self.batch_models[sht][0][tech_2]
                            fst_col = 2
                            bs.cell(row=1, column=2).value = f'{title}_{cap_2}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            g = 1
                            for i in range(1, len(self.batch_models[sht])):
                                if self.batch_models[sht][i][tech_2] != cap_2:
                                    bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 1)
                                    if sht == 1:
                                        d3_rng1 += 1
                                    fst_col = i + 2
                                    cap_2 = self.batch_models[sht][i][tech_2]
                                    bs.cell(row=1, column=fst_col).value = f'{title}_{cap_2}'
                                    if g == 0:
                                        g = 1
                                    else:
                                        bs.cell(row=1, column=fst_col).fill = grey_fill
                                        g = 0
                                    bs.cell(row=1, column=fst_col).font = normal
                                    bs.cell(row=1, column=fst_col).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=fst_col, end_row=1, end_column=i + 2)
                            if sht == 1:
                                d3_rng1 += 1
                            if d3_rng2 < 0:
                                d3_rng2 = i + 2 - fst_col + 1
                        else:
                            try:
                                title = self.batch_models[sht][0]['hdr'].split('.')[-1]
                                del self.batch_models[sht][0]['hdr']
                            except:
                                pass
                            bs.cell(row=1, column=2).value = f'{title}'
                            bs.cell(row=1, column=2).font = normal
                            bs.cell(row=1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                            bs.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(self.batch_models[sht]) + 1)
                            if d3_rng2 < 0:
                                d3_rng2 = len(self.batch_models[sht])
                column = 1
                gndx = self.batch_report[0][1] # Capacity group starting row
                do_opt_parms = [False, 0, 0, 0]
                total_load_row = 0
                if self.discount_rate > 0:
                    batch_disc_row = 0
                else:
                    batch_disc_row = -1
                if self.carbon_price > 0:
                    batch_carbon_row = 0
                else:
                    batch_carbon_row = -1
                batch_lifetime = False
                batch_data_sources_row = 0
                re_tml_row = 0
                max_load_row = -1
                report_keys = []
                for g in range(len(self.batch_report)):
                    report_keys.append(self.batch_report[g][0])
                if 'Lifetime Cost' in report_keys:
                    batch_lifetime = True
                for g in range(len(self.batch_report)):
                    if self.batch_report[g][0] == 'Chart':
                        continue
                    elif self.batch_report[g][0] == 'Carbon Price':
                        batch_carbon_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0] == 'Discount Rate' or self.batch_report[g][0].lower() == 'wacc':
                        batch_disc_row = self.batch_report[g][1]
                        continue
                    elif self.batch_report[g][0].lower() == 'data sources':
                        batch_data_sources_row = gndx
                        gndx += 6
                        try:
                            if self.loadCombo.currentText() != 'n/a':
                                gndx += 1
                        except:
                            pass
                        continue
                    if self.batch_report[g][0] not in batch_details.keys() and self.batch_report[g][0] not in batch_extra.keys():
                        continue
                    self.batch_report[g][1] = gndx
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix(self.batch_report[g][0])
                    else:
                        batch_pfx = ''
                    if option == T and self.batch_report[g][0] == 'Jobs':
                        bs.cell(row=gndx, column=1).value = 'Jobs (figures indicative only)'
                    else:
                        bs.cell(row=gndx, column=1).value = self.batch_report[g][0]
                    bs.cell(row=gndx, column=1).font = bold
                    if self.batch_report[g][0] in batch_extra.keys():
                        key = self.batch_report[g][0]
                        if self.batch_report[g][0] == 'Optimisation Parameters':
                            for row in range(1, batch_input_sheet.max_row + 1):
                                if batch_input_sheet.cell(row=row, column=1).value == 'Optimisation Parameters':
                                    do_opt_parms[0] = True
                                    do_opt_parms[1] = gndx
                                    do_opt_parms[2] = row
                                    break
                            for row in range(row, batch_input_sheet.max_row + 1):
                                gndx += 1
                                if batch_input_sheet.cell(row=row, column=1).value == '':
                                    break
                            do_opt_parms[3] = row
                            continue
                        for sp in range(1, len(batch_extra[key])):
                            if batch_extra[key][sp][0] == 'Total Load':
                                total_load_row = gndx + sp
                            elif batch_extra[key][sp][0] == 'Carbon Price':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' ($/tCO2e)'
                            elif batch_extra[key][sp][0] == 'Lifetime':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0] + ' (years)'
                            elif batch_extra[key][sp][0] == 'Total incl. Carbon Cost':
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + 'LCOE incl. Carbon Cost'
                            else:
                                bs.cell(row=gndx + sp, column=1).value = batch_pfx + batch_extra[key][sp][0]
                            if batch_extra[key][sp][0] in ['RE %age of Total Load', 'Total incl. Carbon Cost'] or \
                              batch_extra[key][sp][0].find('LCOE') >= 0 and batch_extra[key][sp][0].find('Total LCOE') < 0:
                                bs.cell(row=gndx + sp, column=1).font = bold
                            else:
                                bs.cell(row=gndx + sp, column=1).font = normal
                        gndx += len(batch_extra[key]) + 1
                        if key == 'Carbon':
                            if not batch_lifetime:
                                gndx -= 1
                                tot_carb_row = gndx - 3
                            else:
                                tot_carb_row = gndx - 4
                        elif key == 'LCOE ($/MWh)':
                            tot_lco_row = gndx - 2
                        elif key == 'LCOE With CO2 ($/MWh)':
                            tot_lcc_row = gndx - 2
                    else:
                        if self.batch_report[g][0] not in batch_details.keys():
                            continue
                        if self.batch_prefix:
                            batch_pfx = get_batch_prefix(self.batch_report[g][0])
                        else:
                            batch_pfx = ''
                        for sp in range(len(self.batch_tech)):
                        #    if self.batch_report[g][0] == 'To Meet Load (MWh)' and sp == 0:
                         #       bs.cell(row=gndx + sp + 1, column=1).value = 'RE Contribution To Load'
                            if self.batch_report[g][0] != 'Capacity Factor' or self.batch_tech[sp] != 'Total':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp]
                            if self.batch_report[g][0] == 'Max MWh' and self.batch_tech[sp] == 'Total':
                                max_load_row = gndx + sp + 1
                                bs.cell(row=max_load_row, column=1).value = batch_pfx + 'Max Load'
                            elif self.batch_tech[sp] == 'Total' and self.batch_report[g][0] != 'Capacity Factor':
                                bs.cell(row=gndx + sp + 1, column=1).value = batch_pfx + self.batch_tech[sp] + ' ' + self.batch_report[g][0]
                            bs.cell(row=gndx + sp + 1, column=1).font = normal
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            batch_disc_row = gndx + sp + 2
                            bs.cell(row=batch_disc_row, column=1).value = batch_pfx + 'Discount Rate'
                            bs.cell(row=batch_disc_row, column=1).font = normal
                        if self.batch_report[g][0] == 'Capacity Factor' and self.batch_tech[-1] == 'Total':
                            gndx += len(self.batch_tech) + 1
                        else:
                            gndx += len(self.batch_tech) + 2
                        if self.batch_report[g][0] == 'Cost ($/Yr)' and batch_disc_row >= 0:
                            gndx += 1
                        if self.batch_report[g][0] == 'To Meet Load (MWh)':
                            re_tml_row = gndx - 1
                            bs.cell(row=re_tml_row, column=1).value = batch_pfx + 'RE Contribution To Load'
                            bs.cell(row=re_tml_row, column=1).font = normal
                            bs.cell(row=re_tml_row + 1, column=1).value = batch_pfx + 'Storage Contribution To Load'
                            bs.cell(row=re_tml_row + 1, column=1).font = normal
                            gndx += 2

                merge_col = 1
                last_name = ''
                # find first varying capacity to create model name
                model_key = ''
                model_nme = ''
                if sht > 0:
                    for key in self.batch_models[sht][0].keys():
                        if key == 'name':
                            continue
                        try:
                            if self.batch_models[sht][0][key] != self.batch_models[sht][1][key]:
                                model_key = key
                                bits = key.split('.')[-1].split(' ')
                                for bit in bits:
                                    model_nme += bit.strip('()')[0]
                                model_nme += '-'
                                break
                        except:
                            pass
                if option == T:
                    capex_table = {}
                    for fac in pmss_details.keys():
                        capex_table[fac] = {'cum': 0}
                for model, capacities in self.batch_models[sht].items():
                    if option == T:
                        if capacities['year'] != trn_year:
                            # get generators and load for new year
                            trn_year = capacities['year']
                            year = str(trn_year)
                            try:
                                ws = gen_book.sheet_by_name(gen_sheet.replace('$YEAR$', year))
                            except:
                                gen_book.close()
                                self.setStatus(f"No Generators sheet for year '{year}'.")
                                return
                            self.getGenerators(ws)
                            if year not in load_columns.keys():
                                load_file = self.load_files.replace('$YEAR$', year)
                                if self.load_dir.text() != self._load_folder:
                                    load_file = self.get_filename(load_file)
                                if os.path.exists(load_file):
                                    load_columns[year] = len(pmss_data)
                                    pmss_data.append([])
                                    pmss_data[-1] = get_load_data(load_file)
                                elif 'Load' not in capacities.keys() or capacities['Load'] == 0:
                                    self.setStatus(f"Missing load file - '{load_file}'")
                                    return
                                else:
                                    year = list(load_columns.keys())[-1]
                                    self.setStatus(f"Missing load file for '{trn_year}' - using '{year}'")
                                    pmss_details['Load'].col = load_columns[year]
                    for fac in pmss_details.keys():
                        if fac == 'Load' and (option == B or option == T):
                            pmss_details['Load'].capacity = sum(pmss_data[load_columns[year]])
                            pmss_details['Load'].col = load_columns[year]
                            continue
                        pmss_details[fac].multiplier = 0
                    if int(prgv) > prgv_int:
                        prgv_int = int(prgv)
                        self.progressbar.setValue(int(prgv))
                        QtWidgets.QApplication.processEvents()
                    prgv += incr
                    column += 1
                    dispatch_order = []
                    for key, capacity in capacities.items(): # cater for zones
                        if key in ['Carbon Price', 'Discount Rate', 'Load', 'Total']:
                            continue
                        if key == 'name' and model_row_no > 0:
                            if model_key != '':
                                bs.cell(row=model_row_no, column=column).value = f'{model_nme}{capacities[model_key]}'
                            elif capacity != '': # option == T:
                                bs.cell(row=model_row_no, column=column).value = f'{capacity}'
                            else:
                                bs.cell(row=model_row_no, column=column).value = f'Model {model + 1}'
                            bs.cell(row=model_row_no, column=column).font = normal
                            bs.cell(row=model_row_no, column=column).alignment = oxl.styles.Alignment(wrap_text=True,
                                    vertical='bottom', horizontal='center')
                            continue
                        if key == 'year':
                            if option == T:
                                continue
                            if capacity in load_columns.keys():
                                pmss_details['Load'].col = load_columns[capacity]
                            else:
                                load_columns[capacity] = len(pmss_data)
                                pmss_data.append([])
                                load_file = self.load_files.replace('$YEAR$', capacity)
                                # load here if no load file
                                if self.load_dir.text() != self._load_folder:
                                    load_file = self.get_filename(load_file)
                                pmss_data[-1] = get_load_data(load_file)
                                pmss_details['Load'].col = load_columns[capacity]
                            pmss_details['Load'].capacity = sum(pmss_data[pmss_details['Load'].col])
                            continue
                        if key not in re_order:
                            dispatch_order.append(key)
                        if key not in pmss_details.keys():
                            gen = key[key.find('.') + 1:]
                            if gen in re_order:
                                typ = 'R'
                            elif self.generators[gen].constraint in self.constraints and \
                              self.constraints[self.generators[gen].constraint].category == 'Generator':
                                typ = 'G'
                            else:
                                typ = 'S'
                            pmss_details[key] = PM_Facility(key, gen, capacity, typ, -1, 1)
                    for fac in pmss_details.keys():
                        if fac == 'Load':
                            continue
                        gen = pmss_details[fac].generator
                        try:
                            pmss_details[fac].multiplier = capacities[fac] * 1.0 / pmss_details[fac].capacity
                        except:
                            pass
                        if option == T:
                            if fac not in capex_table.keys():
                                capex_table[fac] = {'cum': 0}
                            if year not in capex_table[fac].keys():
                                try:
                                    capex_table[fac][year] = [self.generators[fac].capex, 0]
                                except:
                                    capex_table[fac][year] = [self.generators[fac[fac.find('.') + 1:]].capex, 0]
                            capacity = pmss_details[fac].multiplier * pmss_details[fac].capacity
                            capex_table[fac][year][1] = capacity - capex_table[fac]['cum']
                            capex_table[fac]['cum'] = capacity
                    if option == T:
                        for fac in capex_table.keys():
                            if capex_table[fac]['cum'] == 0:
                                continue
                            capex = 0
                            for key, detail in capex_table[fac].items():
                                if key == 'cum':
                                    continue
                                capex = capex + detail[0] * detail[1]
                            capex = capex / capex_table[fac]['cum']
                            try:
                                self.generators[fac].capex = round(capex)
                            except:
                                self.generators[fac[fac.find('.') + 1:]].capex = round(capex)
                    save_carbon_price = None
                    if 'Carbon Price' in capacities.keys():
                        save_carbon_price = self.carbon_price
                        self.carbon_price = capacities['Carbon Price']
                    if 'Discount Rate' in capacities.keys():
                        save_discount_rate = self.discount_rate
                        self.discount_rate = capacities['Discount Rate']
                    if 'Load' in capacities.keys() and capacities['Load'] > 0 and capacities['Load'] != pmss_details['Load'].capacity:
                        load_varies = True
                        pmss_details['Load'].multiplier = capacities['Load'] / pmss_details['Load'].capacity
                    else:
                        pmss_details['Load'].multiplier = save_load_multiplier
                    sp_data = self.doDispatch(year, option, pmss_details, pmss_data, re_order, dispatch_order,
                              pm_data_file, rslts_file, title=capacities['name'])
                    if 'Carbon Price' in capacities.keys():
                        self.carbon_price = save_carbon_price
                    # first the Facility/technology table at the top of sp_data
                    for sp in range(len(self.batch_tech) + 1):
                        if sp_data[sp][st_fac] in self.batch_tech:
                            tndx = self.batch_tech.index(sp_data[sp][st_fac]) + 1
                            for group in self.batch_report:
                                if group[0] in batch_details.keys():
                                    gndx = group[1]
                                    col = batch_details[group[0]][0]
                                    if group[0] == 'Capacity Factor' and sp_data[sp][0] == 'Total':
                                        continue
                                    if group[0] == 'Capacity Factor' and isinstance(sp_data[sp][col], str):
                                        bs.cell(row=gndx + tndx, column=column).value = float(sp_data[sp][col].strip('%')) / 100.
                                    else:
                                        bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = batch_details[group[0]][1]
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                        if sp_data[sp][st_fac] == 'Total':
                            break
                    if batch_disc_row > 1:
                         bs.cell(row=batch_disc_row, column=column).value = self.discount_rate
                         bs.cell(row=batch_disc_row, column=column).number_format = '#0.00%'
                         bs.cell(row=batch_disc_row, column=column).font = normal
                    # save details from Total row
                    for group in self.batch_report:
                        if group[0] == 'LCOE ($/MWh)':
                            try:
                                col = batch_details['LCOE ($/MWh)'][0]
                                bs.cell(row=tot_lco_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lco_row, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                bs.cell(row=tot_lco_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'LCOE With CO2 ($/MWh)':
                            try:
                                col = batch_details['LCOE With CO2 ($/MWh)'][0]
                                bs.cell(row=tot_lcc_row, column=column).value = sp_data[sp][col]
                                bs.cell(row=tot_lcc_row, column=column).number_format = batch_details['LCOE With CO2 ($/MWh)'][1]
                                bs.cell(row=tot_lcc_row, column=column).font = bold
                            except:
                                pass
                        elif group[0] == 'Carbon':
                            try:
                                bs.cell(row=tot_carb_row, column=column).value = sp_data[sp][st_emc]
                                bs.cell(row=tot_carb_row, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row, column=column).font = normal
                                bs.cell(row=tot_carb_row + 1, column=column).value = sp_data[sp][st_lcc]
                                bs.cell(row=tot_carb_row + 1, column=column).number_format = '#,##0.00'
                                bs.cell(row=tot_carb_row + 1, column=column).font = bold
                                bs.cell(row=tot_carb_row + 2, column=column).value = sp_data[sp][st_lec]
                                bs.cell(row=tot_carb_row + 2, column=column).number_format = '#,##0'
                                bs.cell(row=tot_carb_row + 2, column=column).font = normal
                            except:
                                pass
                    if 'Discount Rate' in capacities.keys():
                        self.discount_rate = save_discount_rate
                    # now the other stuff in sp_data
                    for sp in range(sp + 1, len(sp_data)):
                        if sp_data[sp][st_fac] == '':
                            continue
                        i = sp_data[sp][st_fac].find(' (')
                        if i >= 0:
                            tgt = sp_data[sp][st_fac][: i]
                        else:
                            tgt = sp_data[sp][st_fac]
                        if tgt == 'RE %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Storage %age':
                            for group in self.batch_report:
                                if group[0] == 'To Meet Load (MWh)':
                                    try:
                                        col = batch_details['To Meet Load (MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['To Meet Load (MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'LCOE':
                            for group in self.batch_report:
                                if group[0] == 'LCOE ($/MWh)':
                                    try:
                                        col = batch_details['LCOE ($/MWh)'][0]
                                        bs.cell(row=re_tml_row + 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=re_tml_row + 1, column=column).number_format = batch_details['LCOE ($/MWh)'][1]
                                        bs.cell(row=re_tml_row + 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt == 'Carbon Price':
                            for group in batch_extra['Carbon'][1:]:
                                if group[0] == 'Carbon Price':
                                    try:
                                        col = group[1]
                                        bs.cell(row=tot_carb_row - 1, column=column).value = sp_data[sp][col]
                                        bs.cell(row=tot_carb_row - 1, column=column).number_format = batch_extra['Carbon'][0]
                                        bs.cell(row=tot_carb_row - 1, column=column).font = normal
                                    except:
                                        pass
                                    break
                        elif tgt[:10] == 'Total Load':
                            for group in self.batch_report:
                                if group[0] == 'Max MWh':
                                    try:
                                        col = batch_details['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).value = sp_data[sp][col]
                                        bs.cell(row=max_load_row, column=column).number_format = batch_extra['Max MWh'][0]
                                        bs.cell(row=max_load_row, column=column).font = normal
                                    except:
                                        pass
                                    break
                        for key, details in batch_extra.items():
                            try:
                                x = [x for x in details if tgt in x][0]
                                for group in self.batch_report:
                                    if group[0] == key:
                                        gndx = group[1]
                                        break
                                else:
                                    continue
                                tndx = details.index(x)
                                col = x[1]
                                bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                if key == 'RE' or (key == 'Static Variables' and x[0] == 'Discount Rate'):
                                    pct = float(sp_data[sp][col].strip('%')) / 100.
                                    bs.cell(row=gndx + tndx, column=column).value = pct
                                    bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                else:
                                    bs.cell(row=gndx + tndx, column=column).value = sp_data[sp][col]
                                    bs.cell(row=gndx + tndx, column=column).number_format = details[0]
                                bs.cell(row=gndx + tndx, column=column).font = normal
                                if sp_data[sp][st_fac] == 'RE %age of Total Load' or \
                                  sp_data[sp][st_fac].find('LCOE') >= 0 or \
                                  sp_data[sp][st_fac].find('incl.') >= 0:
                                    bs.cell(row=gndx + tndx, column=column).font = bold
                                else:
                                    bs.cell(row=gndx + tndx, column=column).font = normal
                                if key == 'Load Analysis':
                                    if x[0] in ['Load met', 'Surplus']:
                                        tndx += 1
                                        col = batch_extra['Load Analysis'][tndx][1]
                                        pct = float(sp_data[sp][col].strip('%')) / 100.
                                        bs.cell(row=gndx + tndx, column=column).value = pct
                                        bs.cell(row=gndx + tndx, column=column).number_format = '0.0%'
                                        bs.cell(row=gndx + tndx, column=column).font = normal
                            except:
                                pass
                tim = (time.time() - sheet_start)
                if tim < 60:
                    tim = '%.1f secs' % tim
                else:
                    hhmm = tim / 60.
                    tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                timt = (time.time() - start_time)
                if timt < 60:
                    timt = '%.1f secs' % timt
                else:
                    hhmm = timt / 60.
                    timt = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
                self.setStatus(f'Processed sheet {sht + 1} of {len(self.batch_models)}; ({len(self.batch_models[sht])} models; {tim}. Total {timt})')
                QtWidgets.QApplication.processEvents()
                if total_load_row > 0:
                    if self.batch_prefix:
                        batch_pfx = get_batch_prefix('Load Analysis')
                    if option == T or load_varies:
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load'
                    else:
                        load_mult = ''
                        try:
                            mult = round(pmss_details['Load'].multiplier, 3)
                            if mult != 1:
                                load_mult = ' x ' + str(mult)
                        except:
                            pass
                        bs.cell(row=total_load_row, column=1).value = batch_pfx + 'Total Load - ' + year + load_mult
                if do_opt_parms[0]:
                    t_row = do_opt_parms[1]
                    for row in range(do_opt_parms[2], do_opt_parms[3] + 1):
                        for col in range(1, batch_input_sheet.max_column + 1):
                            cell = batch_input_sheet.cell(row=row, column=col)
                            new_cell = bs.cell(row=t_row, column=col, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                        t_row += 1
                del_rows = []
                for group in self.batch_report:
                    if group[0] in ['Generation (MWh)']:
                        # remove storage or RE
                        gndx = group[1]
                        if group[0] == 'Generation (MWh)':
                            tst = 'S'
                        else:
                            tst = 'R' # probably redundant
                        for row in range(gndx, gndx + len(self.batch_tech)):
                            try:
                                if pmss_details[bs.cell(row=row, column=1).value].fac_type == tst:
                                    del_rows.append(row)
                            except:
                                pass
                for row in sorted(del_rows, reverse=True):
                    bs.delete_rows(row, 1)
                if d3_ranges is not None and len(self.batch_models) > 1 and sht == 1: #multiple sheets
                    med_side = oxl.styles.Side(border_style='medium')
                    thin_side = oxl.styles.Side(border_style='thin')
                    border = oxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    med_border = oxl.styles.Border(left=med_side, right=med_side, top=med_side, bottom=med_side)
                    font1 = oxl.styles.Font(color='FFebbd34', bold=True)
                    font1 = oxl.styles.Font(color='FFFFFF00', italic=True, bold=True)
                    dxf1 = oxl.styles.differential.DifferentialStyle(font=font1)
                    font2 = oxl.styles.Font(italic=True)
                    dxf2 = oxl.styles.differential.DifferentialStyle(font=font2)
                    font3 = oxl.styles.Font(bold=True)
                    dxf3 = oxl.styles.differential.DifferentialStyle(font=font3)
                    fonthide = oxl.styles.Font(color='FFFFFFFF')
                   # self.setStatus(f'{bs.title}: {len(self.batch_models[sht]) + 1} columns, {bs.max_row}, {d3_rng1}, {d3_rng2}')
                    rw = d3s.max_row
                    sn = f"'{bs.title}'"
                    d3s.cell(row=rw, column=3).value = 'Row'
                    d3s.cell(row=rw, column=4).value = f'=MATCH(A{rw},INDIRECT("{sn}!$A$1:$A${bs.max_row}"),0)'
                    row_str = f'$D${rw}'
                    lege_str = f'$B${rw}'
                    rw += 1
                    d3s.cell(row=rw, column=1).value = '2nd level cells'
                    d3s.cell(row=rw, column=4).value = d3_rng2
                    d3_rng_str = f'$D${rw}'
                    rw += 1
                    d3s.cell(row=rw, column=1).value = '2nd Level'
                    d3s.cell(row=rw + 1, column=1).value = '1st column'
                    d3s.cell(row=rw + 2, column=1).value = 'Last column'
                    for s in range(d3_rng1):
                        d3s.cell(row=rw, column=s + 4).value = s + 1
                        if s == 0:
                            d3s.cell(row=rw + 1, column=s + 4).value = 2
                        else:
                            d3s.cell(row=rw + 1, column=s + 4).value = f'={ssCol(s + 3)}{rw + 2}+1'
                        d3s.cell(row=rw + 2, column=s + 4).value = f'={ssCol(s + 4)}{rw + 1}+{d3_rng_str}-1'
                        d3s.cell(row=rw + 3, column=s + 4).value = f'=ADDRESS({row_str},{ssCol(s + 4)}{rw + 1},4,1)'
                        d3s.cell(row=rw + 4, column=s + 4).value = f'=ADDRESS({row_str},{ssCol(s + 4)}{rw + 2},4,1)'
                    rw2 = rw + 7
                    hdr1 = f'INDIRECT("\'{1:0{sht_nam_len}}\'!A{d3_ranges[1]+1}")'
                    if self.batch_prefix:
                        hdr1 = f'REPLACE({hdr1},1,3,"")'
                    d3s.cell(row=rw2 - 1, column=4).value = f'={hdr1}'
                    d3s.cell(row=rw2 - 1, column=4).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                    d3s.cell(row=rw2 - 1, column=4).border = med_border
                    d3s.merge_cells(f'D{rw2 - 1}:{ssCol(d3_rng1 + 3)}{rw2 - 1}')
                    for cl in range(d3_rng1):
                        d3s.cell(row=rw2, column=cl + 4).value = f'=INDIRECT("\'{1:0{sht_nam_len}}\'!R{d3_ranges[1]+1}C"&{ssCol(cl + 4)}{rw+1},0)'
                        d3s.cell(row=rw2, column=cl + 4).number_format = '#0'
                        d3s.cell(row=rw2, column=cl + 4).border = med_border
                    d3s.cell(row=rw2, column=1).value = 'Sheet'
                    hdr2 = f'INDIRECT("\'{1:0{sht_nam_len}}\'!A{d3_ranges[0]+1}")'
                    if self.batch_prefix:
                        hdr2 = f'REPLACE({hdr2},1,3,"")'
                    d3s.cell(row=rw2 + 1, column=2).value = f'={hdr2}'
                    d3s.cell(row=rw2 + 1, column=2).alignment = oxl.styles.Alignment(wrap_text=True, vertical='center', horizontal='center',
                                                                                     textRotation=90)
                    d3s.cell(row=rw2 + 1, column=2).border = med_border
                    d3s.merge_cells(f'B{rw2 + 1}:B{rw2 + len(self.batch_models) - 1}')
                    for rs in range(1, len(self.batch_models)):
                        d3s.cell(row=rw2 + rs, column=1).value = f'{rs:0{sht_nam_len}}'
                        d3s.cell(row=rw2 + rs, column=3).value = f'=INDIRECT("\'{rs:0{sht_nam_len}}\'!B{d3_ranges[0]+1}")'
                        d3s.cell(row=rw2 + rs, column=3).number_format = '#0'
                        d3s.cell(row=rw2 + rs, column=3).border = med_border
                        for s2 in range(d3_rng1):
                            f1 = f'=MIN(INDIRECT(CONCATENATE("\'",TEXT($A{rw2 + rs},"###00"),"\'!",' + \
                                     f'TEXT({ssCol(s2 + 4)}${rw + 3},"###00"),":",TEXT({ssCol(s2 + 4)}${rw + 4},"###00"))))'
                            cels = f'INDIRECT("\'{rs:0{sht_nam_len}}\'!"&{ssCol(s2 + 4)}${rw + 3}&":"&{ssCol(s2 + 4)}${rw + 4})'
                            f1 = f'=IF({lege_str}="<",MIN({cels}),MAX({cels})'
                            d3s.cell(row=rw2 + rs, column=s2 + 4).value = f1
                            d3s.cell(row=rw2 + rs, column=s2 + 4).number_format = '#0.00'
                            d3s.cell(row=rw2 + rs, column=s2 + 4).border = border
                    # now for Len's extras
                    d3s_strt = f'D{rw2 + 1}'
                    d3s_endcol = f'{ssCol(d3_rng1 + 3)}'
                    d3s_end = f'{d3s_endcol}{rw2 + len(self.batch_models) - 1}'
                    d3s_range = f'{d3s_strt}:{d3s_end}'
                    d3s.cell(row=rw2 + rs + 1, column=1).value = 'Within range'
                    d3s.cell(row=rw2 + rs + 1, column=4).value = .05
                    d3s.cell(row=rw2 + rs + 1, column=4).number_format = '#,##0%'
                    d3s.cell(row=rw2 + rs + 2, column=1).value = 'Best range'
                    d3s.cell(row=rw2 + rs + 2, column=2).value = f'=IF(D{rw2 + rs + 3}>0,1,MATCH(1,C{rw2 + rs + 3}:{d3s_endcol}{rw2 + rs + 3},1))'
                    d3s.cell(row=rw2 + rs + 2, column=2).font = fonthide
                    d3s.cell(row=rw2 + rs + 2, column=3).value = f'=INDIRECT("R{rw2 + rs + 3}C"&3+B{rw2 + rs + 2},0)'
                    d3s.cell(row=rw2 + rs + 2, column=3).font = fonthide
                    d3s.cell(row=rw2 + rs + 2, column=4).value = f'=IF({lege_str}=">",IFERROR(MAX({d3s_range})*(1-D{rw2 + rs + 1}),' + \
                                                                 f'MAX({d3s_range})),MIN({d3s_range}))'
                    d3s.cell(row=rw2 + rs + 2, column=4).number_format = '#0.00'
                    rumin = oxl.formatting.rule.FormulaRule(formula=[f'{lege_str}="<"'], border=med_border)
                    d3s.conditional_formatting.add(f'D{rw2 + rs + 2}', rumin)
                    d3s.cell(row=rw2 + rs + 2, column=5).value = 'up to'
                    d3s.cell(row=rw2 + rs + 2, column=5).alignment = oxl.styles.Alignment(horizontal='center')
                    d3s.cell(row=rw2 + rs + 2, column=6).value = f'=IF({lege_str}=">",MAX({d3s_range}),' + \
                                                                 f'IFERROR($D${rw2 + rs + 2}*(1+$D${rw2 + rs + 1}),$D${rw2 + rs + 2}))'
                    d3s.cell(row=rw2 + rs + 2, column=6).number_format = '#0.00'
                    rumax = oxl.formatting.rule.FormulaRule(formula=[f'{lege_str}=">"'], border=med_border)
                    d3s.conditional_formatting.add(f'F{rw2 + rs + 2}', rumax)
                    rul = f'OR(AND({lege_str}="<",{d3s_strt}=$D${rw2 + rs + 2}),' + \
                          f'AND({lege_str}=">",{d3s_strt}=$F${rw2 + rs + 2}))'
                    rule1 = oxl.formatting.rule.FormulaRule(formula=[rul], font=font1)
                    d3s.conditional_formatting.add(d3s_range, rule1)
                    rule2 = oxl.formatting.rule.CellIsRule(operator='between',
                                                           formula=[f'$D${rw2 + rs + 2}',f'$F${rw2 + rs + 2}'],
                                                           font=font3)
                    d3s.conditional_formatting.add(d3s_range, rule2)
                    rule3 = ColorScaleRule(start_type='min', start_color='FF63BE7B',
                                          mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                                          end_type='max', end_color='FFF8696B')
                    d3s.conditional_formatting.add(d3s_range, rule3)
                    d3s.cell(row=rw2 + rs + 3, column=1).value = 'Sheet (1st level) with best'
                    d3s.cell(row=rw2 + rs + 3, column=2).value = f'=OFFSET($A${rw2 + 1},IF(C{rw2 + rs + 2}=0,0,C{rw2 + rs + 2}-1),0,1,1)'
                    d3s.cell(row=rw2 + rs + 4, column=1).value = '2nd level with best'
                    d3s.cell(row=rw2 + rs + 4, column=2).value = f'=MATCH(VALUE(B{rw2 + rs + 3}),D{rw2 + rs + 3}:{ssCol(d3_rng1 + 3)}{rw2 + rs + 3},0)'
                    for s2 in range(d3_rng1):
                        d3s.cell(row=rw2 + rs + 3, column=s2 + 4).value = f'=IF({lege_str}="<",IFERROR(MATCH($D${rw2 + rs + 2},{ssCol(s2 + 4)}${rw2 + 1}:' + \
                                                                          f'{ssCol(s2 + 4)}${rw2 + len(self.batch_models) - 1},0),0),' + \
                                                                          f'IFERROR(MATCH($F${rw2 + rs + 2},{ssCol(s2 + 4)}${rw2 + 1}:' + \
                                                                          f'{ssCol(s2 + 4)}${rw2 + len(self.batch_models) - 1},0),0)'
                        d3s.cell(row=rw2 + rs + 3, column=s2 + 4).font = fonthide
                        d3s.cell(row=rw2 + rs + 4, column=s2 + 4).value = f'=IF({ssCol(s2 + 4)}{rw2 + rs + 3}>0,MATCH(OFFSET({ssCol(s2 + 4)}{rw2 + 1},' + \
                                                                          f'{ssCol(s2 + 4)}{rw2 + rs + 3}-1,0),' + \
                                                                          f'INDIRECT("\'"&$B{rw2 + rs + 3}&"\'!"&{ssCol(s2 + 4)}{rw + 3}&":"&' + \
                                                                          f'{ssCol(s2 + 4)}{rw + 4}),0),"")'
                    d3s.row_dimensions[rw2 + rs + 4].hidden = True
                    d3s.cell(row=rw2 + rs + 5, column=1).value = 'Target cell with best'
                    d3s.cell(row=rw2 + rs + 5, column=2).value = f'=ADDRESS({row_str},OFFSET(D{rw + 1},0,B{rw2 + rs + 4}-1)+OFFSET(D{rw2 + rs + 4},' + \
                                                                 f'0,B{rw2 + rs + 4}-1)-1,4,1)'
                    d3s.cell(row=rw2 + rs + 6, column=1).value = 'Link to best'
                    d3s.cell(row=rw2 + rs + 6, column=2).value = f'=HYPERLINK("#\'"&B{rw2 + rs + 3}&"\'!"&B{rw2 + rs + 5},"\'"&B{rw2 + rs + 3}&"\'!"&B{rw2 + rs + 5})'
                    fontlink = oxl.styles.Font(color='FFFF0000', underline='single', name='Arial', size=10)
                    d3s.cell(row=rw2 + rs + 6, column=2).font = fontlink
                    d3s.freeze_panes = d3s_strt
                for column_cells in bs.columns:
                    length = 0
                    for cell in column_cells:
                        if cell.row < self.batch_report[0][1] - 1:
                            continue
                        try:
                            value = str(round(cell.value, 2))
                        except:
                            value = cell.value
                        if value is None:
                            continue
                        if len(value) > length:
                            length = len(value)
                    if isinstance(cell.column, int):
                        cel = ssCol(cell.column)
                    else:
                        cel = cell.column
                    bs.column_dimensions[cel].width = max(length * 1.05, 10)
                if batch_data_sources_row > 0:
                    i = self.data_sources(bs, batch_data_sources_row - len(del_rows), pm_data_file, option)
                bs.freeze_panes = 'B' + str(self.batch_report[0][1])
                bs.activeCell = 'B' + str(self.batch_report[0][1])
                for sheet in wb:
                    wb[sheet.title].views.sheetView[0].tabSelected = False
                wb.active = bs
                # check if any charts/graphs
                if self.batch_report[-1][0] == 'Chart':
                    if sht > 0 and '' not in chart_groups:
                        continue
                    bold = oxl.styles.Font(name='Arial', bold=True)
                    min_col = 2
                    max_col = len(self.batch_models[sht]) + 1
                    chs = None
                    in_chart = False
                    cht_cells = ['N', 'B']
                    cht_row = -27
                    tndx_rows = max(9, len(self.batch_tech) + 4)
                    cats = None
                    chart_group = ''
                    chart_smooth = True
                    for row in range(self.batch_report[-1][1], batch_input_sheet.max_row + 1):
                        if batch_input_sheet.cell(row=row, column=1).value is None:
                            continue
                        if batch_input_sheet.cell(row=row, column=1).value.lower() in ['chart', 'graph', 'plot']:
                            if sht > 0 and batch_input_sheet.cell(row=row, column=2).value is not None \
                              and batch_input_sheet.cell(row=row, column=2).value != '':
                                continue
                            if in_chart:
                                charts[-1].width = 20
                                charts[-1].height = 12
                                nocolor = []
                                for s in range(len(charts[-1].series)):
                                    nocolor.append(s)
                                if charts2[-1] is not None:
                                    for s in range(len(charts2[-1].series)):
                                        nocolor.append(s + len(charts[-1].series))
                                colors = PlotPalette(nocolor, lower=False, palette=48)
                                for s in range(len(charts[-1].series)):
                                    ser = charts[-1].series[s]
                                    ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                                    ser.graphicalProperties.line.solidFill = colors[s].strip('#')
                                if charts2[-1] is not None:
                                    for s in range(len(charts2[-1].series)):
                                        ser = charts2[-1].series[s]
                                        ser.marker.symbol = 'triangle'
                                        ser.graphicalProperties.line.solidFill = colors[s + len(charts[-1].series)].strip('#')
                                    charts2[-1].y_axis.crosses = 'max'
                                    charts[-1] += charts2[-1]
                                if cats is not None:
                                    charts[-1].set_categories(cats)
                                if len(charts) % 2:
                                    cht_row += 30
                                if chart_group != '':
                                    cht_col = col_letters.index(cht_cells[len(charts) % 2])
                                    chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                                    chs.cell(row=cht_row - 1, column=cht_col).font = bold
                                chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
                            in_chart = True
                            if chs is None:
                                if bs.title.find('Results') >= 0:
                                    txt = bs.title.replace('Results', 'Charts')
                                else:
                                    txt = 'Charts_' + bs.title
                                chs = wb.create_sheet(txt)
                                if sht == 0:
                                    if '3D Summary' in wb.sheetnames:
                                        ndx = wb.sheetnames.index('3D Summary')
                                        if ndx > 0:
                                            if (wb.sheetnames[ndx - 1][:8] == 'Results_' and wb.sheetnames[ndx + 1][:7] == 'Charts_') \
                                              or (wb.sheetnames[ndx + 1][:8] == 'Results_' and wb.sheetnames[ndx - 1][:7] == 'Charts_') :
                                                ndx = len(wb.sheetnames) - ndx
                                                wb.move_sheet('3D Summary', offset=ndx)
                                charts = []
                                charts2 = []
                            charts.append(LineChart())
                            charts2.append(None)
                            if batch_input_sheet.cell(row=row, column=2).value is None or len(merged_cells) == 0:
                                min_col = 2
                                max_col = len(self.batch_models[sht]) + 1
                                chart_group = ''
                            else:
                                merge_group = get_value(batch_input_sheet, row, 2)
                                for i in range(len(merged_cells) -1, -1, -1):
                                    merge_value = get_value(batch_input_sheet, merged_cells[i][0], merged_cells[i][1])
                                    if merge_value == merge_group:
                                        min_col = merged_cells[i][1]
                                        max_col = merged_cells[i][2]
                                        chart_group = merge_group
                                        break
                            if sht == 0:
                                chart_groups.append(chart_group)
                        elif not in_chart:
                            continue
                        elif batch_input_sheet.cell(row=row, column=1).value.lower()[:4] == 'line':
                            if batch_input_sheet.cell(row=row, column=2).value.lower() == 'straight':
                                chart_smooth = False
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'title':
                            charts[-1].title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'x-title':
                            charts[-1].x_axis.title = get_value(batch_input_sheet, row, 2)
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title':
                            charts[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'y-title2':
                            if charts2[-1] is None:
                                charts2[-1] = LineChart()
                                charts2[-1].x_axis.title = None
                            charts2[-1].y_axis.axId = 200
                            charts2[-1].y_axis.title = batch_input_sheet.cell(row=row, column=2).value
                        elif batch_input_sheet.cell(row=row, column=1).value.lower() in ['categories', 'y-labels', 'data', 'data2']:
                            dgrp = get_value(batch_input_sheet, row, 2)
                            if self.batch_prefix:
                                batch_pfx = get_batch_prefix(dgrp)
                            else:
                                batch_pfx = ''
                            if batch_input_sheet.cell(row=row, column=1).value.lower() == 'categories' \
                              and dgrp.lower() in ['model', 'model label', 'technology', 'year']: # models as categories
                                rw = self.batch_report[0][1] - 1
                                cats = Reference(bs, min_col=min_col, min_row=rw, max_col=max_col, max_row=rw)
                                continue
                            if dgrp.lower() in ['capacity (mw)', 'capacity (mw/mwh)']:
                                gndx = self.batch_report[0][1]
                            else:
                                for group in self.batch_report:
                                    if group[0].lower() == dgrp.lower():
                                        gndx = group[1]
                                        break
                                else:
                                     continue
                                # backup a bit in case rows deleted
                                for r in range(len(del_rows)):
                                    try:
                                        if bs.cell(row=gndx, column=1).value.lower() == group[0].lower():
                                            break
                                    except:
                                        pass
                                    gndx -= 1
                            ditm = get_value(batch_input_sheet, row, 3)
                            for tndx in range(tndx_rows):
                                if bs.cell(row=gndx + tndx, column=1).value is None:
                                    break
                                if bs.cell(row=gndx + tndx, column=1).value.lower() == f'{batch_pfx.lower()}{ditm.lower()}':
                                    if batch_input_sheet.cell(row=row, column=1).value.lower() == 'data':
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        series.smooth = chart_smooth
                                        charts[-1].append(series)
                                    elif batch_input_sheet.cell(row=row, column=1).value.lower() == 'data2':
                                        if charts2[-1] is None:
                                            charts2[-1] = LineChart()
                                        values = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                        series = Series(values)
                                        series.title = oxl.chart.series.SeriesLabel(oxl.chart.data_source.StrRef("'" + bs.title + "'!A" + str(gndx + tndx)))
                                        series.smooth = chart_smooth
                                        charts2[-1].append(series)
                                    else:
                                        cats = Reference(bs, min_col=min_col, min_row=gndx + tndx, max_col=max_col, max_row=gndx + tndx)
                                    break
                    if in_chart:
                        charts[-1].width = 20
                        charts[-1].height = 12
                        nocolor = []
                        for s in range(len(charts[-1].series)):
                            nocolor.append(s)
                        if charts2[-1] is not None:
                            for s in range(len(charts2[-1].series)):
                                nocolor.append(s + len(charts[-1].series))
                        colors = PlotPalette(nocolor, lower=False, palette=48)
                        for s in range(len(charts[-1].series)):
                            ser = charts[-1].series[s]
                            ser.marker.symbol = 'circle' #'dot', 'plus', 'triangle', 'x', 'picture', 'star', 'diamond', 'square', 'circle', 'dash', 'auto'
                            ser.graphicalProperties.line.solidFill = colors[s].strip('#')
                        if charts2[-1] is not None:
                            for s in range(len(charts2[-1].series)):
                                ser = charts2[-1].series[s]
                                ser.marker.symbol = 'triangle'
                                ser.graphicalProperties.line.solidFill = colors[s + len(charts[-1].series)].strip('#')
                            charts2[-1].y_axis.crosses = 'max'
                            charts[-1] += charts2[-1]
                        if cats is not None:
                            charts[-1].set_categories(cats)
                        if len(charts) % 2:
                            cht_row += 30
                        if chart_group != '':
                            cht_col = col_letters.index(cht_cells[len(charts) % 2])
                            chs.cell(row=cht_row - 1, column=cht_col).value = chart_group
                            chs.cell(row=cht_row - 1, column=cht_col).font = bold
                        chs.add_chart(charts[-1], cht_cells[len(charts) % 2] + str(cht_row))
            if len(self.batch_models) > 1 and len(self.batch_models[0]) == 1:
                try:
                    del wb['Results_' + rpt_time]
                    del wb['Charts_' + rpt_time]
                except:
                    pass
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'Saving {self.sender().text()} report ({total_models:,} models; {tim})')
           #     self.setStatus('Saving %s report' % (self.sender().text()))
            self.progressbar.setValue(20)
            QtWidgets.QApplication.processEvents()
            wb.save(batch_report_file)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = '%.1f secs' % tim
            else:
                hhmm = tim / 60.
                tim = f'{int(hhmm)}:{int((hhmm-int(hhmm))*60.):0>2} mins'
            self.setStatus(f'{self.sender().text()} completed ({len(self.batch_models)} sheets, {total_models:,} models; {tim}). You may need to open and save the workbook to reprocess it.')
            self.progressbar.setHidden(True)
            self.progressbar.setValue(0)
            return
        if do_adjust:
            if self.adjustto is not None:
                for fac, value in self.adjustto.items():
                    try:
                        pmss_details[fac].multiplier = value / pmss_details[fac].capacity
                    except:
                        pass
        self.doDispatch(year, option, pmss_details, pmss_data, re_order, dispatch_order,
                        pm_data_file, rslts_file)

    def doDispatch(self, year, option, pmss_details, pmss_data, re_order, dispatch_order,
                   pm_data_file, rslts_file, title=None):
        def calcLCOE(annual_output, capital_cost, annual_operating_cost, discount_rate, lifetime):
            # Compute levelised cost of electricity
            if discount_rate > 0:
                annual_cost_capital = capital_cost * discount_rate * pow(1 + discount_rate, lifetime) / \
                                      (pow(1 + discount_rate, lifetime) - 1)
            else:
                annual_cost_capital = capital_cost / lifetime
            total_annual_cost = annual_cost_capital + annual_operating_cost
            try:
                return total_annual_cost / annual_output
            except:
                return total_annual_cost

        def format_period(per):
            hr = per % 24
            day = int((per - hr) / 24)
            mth = 0
            while day > the_days[mth] - 1:
                day -= the_days[mth]
                mth += 1
            return '{}-{:02d}-{:02d} {:02d}:00'.format(year, mth+1, day+1, hr)

        def summary_totals(title=''):
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = title + 'Total'
            sp_d[st_cap] = cap_sum
            sp_d[st_tml] = tml_sum
            sp_d[st_sub] = gen_sum
            sp_d[st_cst] = cost_sum
            sp_d[st_lcg] = gs
            sp_d[st_lco] = gsw
            sp_d[st_emi] = co2_sum
            sp_d[st_emc] = co2_cost_sum
            sp_d[st_lcc] = gswc
            sp_d[st_cac] = capex_sum
            sp_d[st_lic] = lifetime_sum
            sp_d[st_lie] = lifetime_co2_sum
            sp_d[st_lec] = lifetime_co2_cost
            sp_d[st_are] = total_area
            if self.do_jobs:
                sp_d[st_job] = total_jobs
            sp_data.append(sp_d)
            if (self.carbon_price > 0 or option == B or option == T):
                sp_d = [' '] * len(headers)
                cc = co2_sum * self.carbon_price
                cl = cc * max_lifetime
                if self.adjusted_lcoe and tml_sum > 0:
                    cs = (cost_sum + cc) / tml_sum
                else:
                    if gen_sum > 0:
                        cs = (cost_sum + cc) / gen_sum
                    else:
                        cs = ''
                sp_d[st_fac] = title + 'Total incl. Carbon Cost'
                sp_d[st_cst] = cost_sum + cc
                sp_d[st_lic] = lifetime_sum + cl
                sp_data.append(sp_d)
            if tml_sum > 0:
                sp_d = [' '] * len(headers)
             #   sp_d[st_fac] = 'RE Direct Contribution to ' + title + 'Load'
                sp_d[st_fac] = 'RE %age'
                re_pct = (tml_sum - sto_sum - ff_sum) / tml_sum
                sp_d[st_cap] = '{:.1f}%'.format(re_pct * 100.)
                sp_d[st_tml] = tml_sum - ff_sum - sto_sum
                sp_data.append(sp_d)
                if sto_sum > 0:
                    sp_d = [' '] * len(headers)
                 #   sp_d[st_fac] = 'RE Contribution to ' + title + 'Load via Storage'
                    sp_d[st_fac] = 'Storage %age'
                    sp_d[st_cap] = '{:.1f}%'.format(sto_sum * 100. / tml_sum)
                    sp_d[st_tml] = sto_sum
                    sp_data.append(sp_d)
            sp_data.append([' '])
            sp_data.append([title + 'Load Analysis'])
            if sp_load != 0:
                if option == B or option == T:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = title + 'Total Load'
                    sp_d[st_tml] = sp_load
                    if title == '':
                        sp_d[st_max] = load_max
                    sp_data.append(sp_d)
                else:
                    load_mult = ''
                    try:
                        mult = round(pmss_details['Load'].multiplier, 3)
                        if mult != 1:
                            load_mult = ' x ' + str(mult)
                    except:
                        pass
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Total ' + title + 'Load - ' + year + load_mult
                    sp_d[st_tml] = sp_load
                    if title == '' or option == S:
                        sp_d[st_max] = load_max
                        sp_d[st_bal] = ' (' + format_period(load_hr)[5:] + ')'
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Shortfall'
                sp_d[st_cap] = '{:.1f}%'.format(-sf_sums[0] * 100 / sp_load)
                sp_d[st_tml] = -sf_sums[0]
                sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Load met'
                load_pct = (sp_load - sf_sums[0]) / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(load_pct * 100)
                sp_d[st_tml] = sp_load - sf_sums[0]
                sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'RE %age of Total ' + title + 'Load'
                if self.optimise_total_re:
                    re_pct = (sp_load - sf_sums[0] - ff_sum) / sp_load
                sp_d[st_cap] = '{:.1f}%'.format((sp_load - sf_sums[0] - ff_sum) * 100. / sp_load)
                sp_data.append(sp_d)
                sp_data.append(' ')
                if tot_sto_loss != 0:
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = 'Storage losses'
                    sp_d[st_sub] = tot_sto_loss
                    sp_data.append(sp_d)
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = title + 'Surplus'
                surp_pct = -sf_sums[1] / sp_load
                sp_d[st_cap] = '{:.1f}%'.format(surp_pct * 100)
                sp_d[st_sub] = -sf_sums[1]
                sp_data.append(sp_d)
            else:
                load_pct = 0
                surp_pct = 0
                re_pct = 0
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Largest Shortfall'
                sp_d[st_sub] = round(max_short[1], 2)
                sp_d[st_cfa] = ' (' + format_period(max_short[0])[5:] + ')'
                sp_data.append(sp_d)
            if option == O or option == O1:
                return load_pct, surp_pct, re_pct

        def do_detail(fac, col, ss_row):
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            col += 1
            sp_cols.append(fac)
            sp_cap.append(pmss_details[fac].capacity * pmss_details[fac].multiplier)
            if do_zone and pmss_details[fac].zone != '':
                ns.cell(row=zone_row, column=col).value = pmss_details[fac].zone
                ns.cell(row=zone_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            # facility
            try:
                ns.cell(row=what_row, column=col).value = fac[fac.find('.') + 1:]
            except:
                ns.cell(row=what_row, column=col).value = fac # gen
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            # capacity
            ns.cell(row=cap_row, column=col).value = sp_cap[-1]
            ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
            ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) \
                    + str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=col).number_format = '#,##0'
            # To meet load MWh
            ns.cell(row=tml_row, column=col).value = fac_tml[fac]
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            # subtotal MWh
            ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                    ssCol(col) + str(sum_row) +'/' + ssCol(col) + str(cap_row) + '/8760,"")'
            ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return col
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                if self.remove_cost:
                    el1 = '),0)'
                else:
                    el1 = f'),{ssCol(col)}{cap_row}*{self.generators[gen].capex}{cst_calc}+{ssCol(col)}{cap_row}*{self.generators[gen].fixed_om})'
                ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + el1
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcog_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + \
                        '>0,' + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + \
                        str(cost_row) + '/' + ssCol(col) + str(sum_row) + ',"")'
                ns.cell(row=lcog_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe > 0:
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1 - jm) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1 - jm) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcog_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcog_row, column=col).number_format = '$#,##0.00'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                if ss_row >= 0:
                    ns.cell(row=cost_row, column=col).value = '=IF(' + ssCol(col) + str(cf_row) + \
                            '>0,' + ssCol(col) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1 - jm) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1 - jm) + str(ss_row) + '/' + ssCol(col) + str(cf_row) + ',0)'
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ns.cell(row=lcog_row, column=col).value = '=IF(AND(' + ssCol(col) + str(cf_row) + '>0,' \
                        + ssCol(col) + str(cap_row) + '>0),' + ssCol(col) + str(cost_row) + '/8760/' \
                        + ssCol(col) + str(cf_row) +'/' + ssCol(col) + str(cap_row) + ',"")'
                ns.cell(row=lcog_row, column=col).number_format = '$#,##0.00'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ssCol(col) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            di = pmss_details[fac].col
            if pmss_details[fac].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = pmss_data[di][row - hrows] * \
                                                         pmss_details[fac].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            return col

        def do_detail_summary(fac, col, ss_row, dd_tml_sum, dd_re_sum):
            if do_zone and pmss_details[fac].zone != '':
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(zone_row) + \
                                                      '&"."&Detail!' + ssCol(col) + str(what_row)
            else:
                ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col) + str(what_row)
            if fac in self.generators.keys():
                gen = fac
            else:
                gen = pmss_details[fac].generator
            if self.do_jobs:
                try:
                    genj = gen[gen.find('.') + 1:]
                except:
                    genj = gen
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col) + str(cap_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # To meet load MWh
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col) + str(tml_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            dd_tml_sum += ssCol(st_tml+1) + str(ss_row) + '+'
            # subtotal MWh
            ss.cell(row=ss_row, column=st_sub+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(sum_row) + ',"")'
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            dd_re_sum += ssCol(st_sub+1) + str(ss_row) + '+'
            # CF
            ss.cell(row=ss_row, column=st_cfa+1).value = '=IF(Detail!' + ssCol(col) + str(cf_row) \
                                                  + '>0,Detail!' + ssCol(col) + str(cf_row) + ',"")'
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if gen not in self.generators.keys():
                return dd_tml_sum, dd_re_sum
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=IF(Detail!' + ssCol(col) + str(lcog_row) \
                                                      + '>0,Detail!' + ssCol(col) + str(lcog_row) + ',"")'
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = '=IF(Detail!' + ssCol(col) + str(cap_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cap_row) + '*'  \
                                                        + str(self.generators[gen].capex) + ',"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).number_format = '$#,##0.00'
                # ref cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).number_format = '#,##0.0%'
            # lifetime cost
            ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(cost_row) \
                                                    + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                                                   + '>0,Detail!' + ssCol(col) + str(max_row) + ',"")'
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            # emissions
            if self.generators[gen].emissions > 0:
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
            # area
            if self.generators[gen].area > 0:
                ss.cell(row=ss_row, column=st_are+1).value = '=Detail!' + ssCol(col) + str(cap_row) +\
                                                             '*' + str(self.generators[gen].area)
                ss.cell(row=ss_row, column=st_are+1).number_format = '#,##0.00'
            # jobs
            if self.do_jobs:
                ss.cell(row=ss_row, column=st_job+1).value = f'=IF({ssCol(st_cac+1)}{ss_row}>0,{ssCol(st_cap+1)}{ss_row}*(' + \
                                                             f'{self.jobfactors[genj].manufacture}*{self.jobfactors[genj].local_pct}+' + \
                                                             f'{self.jobfactors[genj].install}),0)+' + \
                                                             f'{self.jobfactors[genj].operate}*{ssCol(st_cap+1)}{ss_row}'
                ss.cell(row=ss_row, column=st_job+1).number_format = '#,##0'
            return dd_tml_sum, dd_re_sum

        def detail_summary_total(ss_row, title='', base_row='', back_row=''):
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Total'
            for col in range(1, len(headers) + 1):
                ss.cell(row=3, column=col).font = bold
                ss.cell(row=3, column=col).alignment = oxl.styles.Alignment(wrap_text=True, vertical='bottom', horizontal='center')
                ss.cell(row=ss_row, column=col).font = bold
            for col in [st_cap, st_tml, st_sub, st_cst, st_emi, st_emc, st_cac, st_lic, st_lie, st_lec, st_are, st_job]:
                if col == st_job and not self.do_jobs:
                    continue
                if back_row != '':
                    strt = ssCol(col, base=0) + back_row + '+'
                else:
                    strt = ''
                ss.cell(row=ss_row, column=col+1).value = '=' + strt + 'SUM(' + ssCol(col, base=0) + \
                        base_row + ':' + ssCol(col, base=0) + str(ss_row - 1) + ')'
                if col in [st_cap, st_are]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0.00'
                elif col in [st_tml, st_sub, st_emi, st_lie, st_job]:
                    ss.cell(row=ss_row, column=col+1).number_format = '#,##0'
                else:
                    ss.cell(row=ss_row, column=col+1).number_format = '$#,##0'
            ss.cell(row=ss_row, column=st_lcg+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_sub+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
            ss.cell(row=ss_row, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(ss_row) + \
                                                         '/' + ssCol(st_tml+1) + str(ss_row)
            ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
            if self.carbon_price > 0:
                ss.cell(row=ss_row, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(ss_row) + \
                    '+' + ssCol(st_emc+1) + str(ss_row) + ')/' + ssCol(st_tml+1) + str(ss_row)
                ss.cell(row=ss_row, column=st_lcc+1).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=st_lcc+1).font = bold
            last_col = ssCol(ns.max_column)
            r = 1
            if self.carbon_price > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Total incl. Carbon Cost'
                ss.cell(row=ss_row, column=st_cst+1).value = '=' + ssCol(st_cst+1) + str(ss_row - 1) + \
                        '+' + ssCol(st_emc+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ss.cell(row=ss_row, column=st_lic+1).value = '=' + ssCol(st_lic+1) + str(ss_row - r) + \
                                                             '+' + ssCol(st_lec+1) + str(ss_row - 1)
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                r += 1
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'RE %age'
            ss.cell(row=ss_row, column=st_tml+1).value = ns_tml_sum[:-1] + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' +\
                                                         ssCol(st_tml+1) + str(ss_row - r)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_re_row = ss_row
            ss_sto_row = -1
            # if storage
            if ns_sto_sum != '':
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = title + 'Storage %age'
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + ns_sto_sum[1:]
                ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ns_sto_sum[1:] + ')/' + ssCol(st_tml+1) + \
                                                             str(ss_row - r - 1)
                ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
                ss_sto_row = ss_row
            # now do the LCOE and LCOE with CO2 stuff
            if base_row == '4':
                base_col = 'C'
                if ss_sto_row >= 0:
                    for rw in range(ss_re_fst_row, ss_re_lst_row + 1):
                        if self.corrected_lcoe:
                            ss.cell(row=rw, column=st_lco+1).value = '=IF(AND(' + ssCol(st_lcg+1) + str(rw) + '<>"",' + \
                                    ssCol(st_lcg+1) + str(rw) + '>0),' + \
                                    ssCol(st_cst+1) + str(rw) + '/(' + ssCol(st_tml+1) + str(rw) + '+' + \
                                    ssCol(st_tml+1) + '$' + str(ss_sto_row) + '*' + ssCol(st_tml+1) + str(rw) + \
                                    '/' + ssCol(st_tml+1) + '$' + str(ss_re_row) + '),"")'
                        else:
                            ss.cell(row=rw, column=st_lco+1).value = '=' + \
                                    ssCol(st_cst+1) + str(rw) + '/' + ssCol(st_tml+1) + str(rw)
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            if self.corrected_lcoe:
                                ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                        ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                        ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/(' + \
                                        ssCol(st_tml+1) + str(rw) + '+' + ssCol(st_tml+1) + '$' + str(ss_sto_row) + \
                                        '*' + ssCol(st_tml+1) + str(rw) + '/' + ssCol(st_tml+1) + '$' + \
                                        str(ss_re_row) + '),"")'
                            else:
                                ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                        ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                        ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/' + \
                                        ssCol(st_tml+1) + str(rw) + ',"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                else:
                    for rw in range(ss_re_fst_row, ss_re_lst_row):
                        ss.cell(row=rw, column=st_lco+1).value = '=IF(' + ssCol(st_lcg+1) + str(rw) + '>0,' + \
                                ssCol(st_cst+1) + str(rw) + '/' + ssCol(st_tml+1) + str(rw) + '),"")'
                        ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                        if self.carbon_price > 0:
                            ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_emc+1) + str(rw) + '<>"",' + \
                                    ssCol(st_emc+1) + str(rw) + '>0),(' + \
                                    ssCol(st_cst+1) + str(rw) + ssCol(st_emc+1) + str(rw) + ')/' + \
                                    ssCol(st_tml+1) + str(rw) + '),"")'
                            ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
                for rw in range(ss_re_lst_row + 1, ss_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = f'=IF(AND({ssCol(st_tml+1)}{rw}<>"",{ssCol(st_tml+1)}{rw}>0),' + \
                        f'{ssCol(st_cst+1)}{rw}/{ssCol(st_tml+1)}{rw},{ssCol(st_lcg+1)}{rw})'
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        if self.remove_cost:
                            el3 = ',"")'
                        else:
                            el3 = f',IF({ssCol(st_lco)}{rw}>0,{ssCol(st_lco)}{rw},"")'
                        ss.cell(row=rw, column=st_lcc+1).value = '=IF(AND(' + ssCol(st_tml+1) + str(rw) + '<>"",' + \
                                    ssCol(st_tml+1) + str(rw) + '>0),(' + \
                                ssCol(st_cst+1) + str(rw) + '+' + ssCol(st_emc+1) + str(rw) + ')/' + \
                                ssCol(st_tml+1) + str(rw) + el3
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            else:
                base_col = ssCol(next_col)
                for rw in range(ul_fst_row, ul_lst_row + 1):
                    ss.cell(row=rw, column=st_lco+1).value = '=' + ssCol(st_cst+1) + str(rw) + \
                                                             '/' + ssCol(st_tml+1) + str(rw)
                    ss.cell(row=rw, column=st_lco+1).number_format = '$#,##0.00'
                    if self.carbon_price > 0:
                        ss.cell(row=rw, column=st_lcc+1).value = '=(' + ssCol(st_cst+1) + str(rw) + \
                            '+' + ssCol(st_emc+1) + str(rw) + ')/' + ssCol(st_tml+1) + str(rw)
                        ss.cell(row=rw, column=st_lcc+1).number_format = '$#,##0.00'
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = title + 'Load Analysis'
            ss.cell(row=ss_row, column=1).font = bold
            ss_row += 1
            ld_row = ss_row
            load_mult = ''
            try:
                mult = round(pmss_details['Load'].multiplier, 3)
                if mult != 1:
                    load_mult = ' x ' + str(mult)
            except:
                pass
            ss.cell(row=ss_row, column=1).value = 'Total ' + title + 'Load - ' + year + load_mult
            ss.cell(row=ss_row, column=1).font = bold
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + base_col + str(sum_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_tml+1).font = bold
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + base_col + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=st_max+1).font = bold
            ss.cell(row=ss_row, column=st_bal+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + ',MATCH(Detail!' + \
                    base_col + str(max_row) + ',Detail!' + base_col + str(hrows) + ':Detail!' + base_col + \
                    str(hrows + 8759) + ',0),0)&")"'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Shortfall'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[0] + '0",Detail!' + last_col \
                      + str(hrows) + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign > 0:
                ss.cell(row=ss_row, column=st_tml+1).value = '=' + sf_text
            else:
                ss.cell(row=ss_row, column=st_tml+1).value = '=-' + sf_text
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ss_row + 1)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Load met'
            ss.cell(row=ss_row, column=st_tml+1).value = '=SUM(' + ssCol(st_tml+1) + str(ss_row - 2) + ':' + \
                                                         ssCol(st_tml+1) + str(ss_row - 1) + ')'
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_row) + '/' + ssCol(st_tml+1) + \
                                                         str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'RE %age of Total ' + title + 'Load'
            ss.cell(row=ss_row, column=1).font = bold
            if ns_sto_sum == '':
                ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_tml+1) + str(ss_re_row) + \
                                                             '/' + ssCol(st_tml+1) + str(ld_row)
            else:
                ss.cell(row=ss_row, column=st_cap+1).value = '=(' + ssCol(st_tml+1) + str(ss_re_row) + '+' + \
                                                             ssCol(st_tml+1) + str(ss_sto_row) + ')/' + \
                                                             ssCol(st_tml+1) + str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            ss.cell(row=ss_row, column=st_cap+1).font = bold
            ss_row += 2
            if ns_loss_sum != '':
                ss.cell(row=ss_row, column=1).value = title + 'Storage Losses'
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + ns_loss_sum[1:]
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
                ss_row += 1
            ss.cell(row=ss_row, column=1).value = title + 'Surplus'
            sf_text = 'SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' + last_col \
                      + str(hrows + 8759) + ',"' + sf_test[1] + '0",Detail!' + last_col + str(hrows) \
                      + ':Detail!' + last_col + str(hrows + 8759) + ')'
            if self.surplus_sign < 0:
                ss.cell(row=ss_row, column=st_sub+1).value = '=-' + sf_text
            else:
                ss.cell(row=ss_row, column=st_sub+1).value = '=' + sf_text
            ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_cap+1).value = '=' + ssCol(st_sub+1) + str(ss_row) + '/' + ssCol(st_tml+1) + str(ld_row)
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.0%'
            max_short = [0, 0]
            for h in range(len(shortfall)):
                if shortfall[h] > max_short[1]:
                    max_short[0] = h
                    max_short[1] = shortfall[h]
            if max_short[1] > 0:
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = 'Largest ' + title + 'Shortfall:'
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + last_col + str(hrows + max_short[0])
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=st_cfa+1).value = '=" ("&OFFSET(Detail!B' + str(hrows - 1) + \
                        ',MATCH(' + ssCol(st_sub+1) + str(ss_row) + ',Detail!' + last_col + str(hrows) + \
                        ':Detail!' + last_col + str(hrows + 8759) + ',0),0)&")"'
            return ss_row, ss_re_row

    # doDispatch: The "guts" of Powermatch processing. Have a single calculation algorithm
    # for Summary, Powermatch (detail), and Optimise. The detail makes it messy
    # Note: For Batch pmss_data is reused so don't update it in doDispatch

        if self.do_jobs:
            jm = 0
        else:
            jm = 1
        the_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if self.surplus_sign < 0:
            sf_test = ['>', '<']
            sf_sign = ['+', '-']
        else:
            sf_test = ['<', '>']
            sf_sign = ['-', '+']
        sp_cols = []
        sp_cap = []
        re_tml_sum = 0. # keep tabs on how much RE is used
        start_time = time.time()
        do_zone = False # could pass as a parameter
        max_lifetime = 0
        # find max. lifetime years for all technologies selected
        for key in pmss_details.keys():
            if key == 'Load' or key == 'Total':
                continue
            if pmss_details[key].capacity * pmss_details[key].multiplier > 0:
             #   gen = key.split('.')[-1]
                gen = pmss_details[key].generator
                max_lifetime = max(max_lifetime, self.generators[gen].lifetime)
        for key in pmss_details.keys():
            if key.find('.') > 0:
                do_zone = True
                break
        underlying_facs = []
        undercol = [] * len(self.underlying)
        operational_facs = []
        fac_tml = {}
        for fac in re_order:
            if fac == 'Load':
                continue
            fac_tml[fac] = 0.
            if fac in self.operational:
              #  operational_facs.append(fac)
                continue
            if fac.find('.') > 0:
                if fac[fac.find('.') + 1:] in self.underlying:
                    underlying_facs.append(fac)
                    continue
            elif fac in self.underlying:
                underlying_facs.append(fac)
                continue
        load_col = pmss_details['Load'].col
        # find any minimum generation for generators
        committed_gen = {}
        committed_gen_tot = 0
        for gen in dispatch_order:
            if pmss_details[gen].fac_type == 'G': # generators
                try:
                    const = self.generators[gen].constraint
                except:
                    try:
                        g2 = gen[gen.find('.') + 1:]
                        const = self.generators[g2].constraint
                    except:
                        continue
                if self.constraints[const].capacity_min != 0:
                    try:
                        committed_gen[gen] = pmss_details[gen].capacity * pmss_details[gen].multiplier * \
                            self.constraints[const].capacity_min
                    except:
                        committed_gen[gen] = pmss_details[gen].capacity * \
                            self.constraints[const].capacity_min
                    if self.reserve_committed:
                        committed_gen_tot += committed_gen[gen]
        shortfall = [0] * 8760
        row_tml = [0.] * 8760
        for h in range(len(pmss_data[load_col])):
            load_h = pmss_data[load_col][h] * pmss_details['Load'].multiplier
            shortfall[h] = load_h
            re_shortfall = load_h - committed_gen_tot
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    continue
                re_shortfall -= pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                shortfall[h] -= pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
            if re_shortfall >= 0:
                alloc = 1.
            else:
                alloc = (load_h - committed_gen_tot) / (load_h - re_shortfall - committed_gen_tot)
                if alloc < 0: # don't use negative generation
                    alloc = 0
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                else:
                    row_tml[h] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier * alloc
                    fac_tml[fac] += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier * alloc
            line = ''
        fac_tml_sum = 0
        if self.corrected_lcoe:
            for fac in fac_tml.keys():
                if fac in underlying_facs:
                    continue
                fac_tml_sum += fac_tml[fac]
        if self.show_correlation:
            col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                df1 = pmss_data[col]
            else:
                tgt = []
                for h in range(len(pmss_data[col])):
                    tgt.append(pmss_data[col][h] * pmss_details['Load'].multiplier)
                df1 = tgt
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data = [['Correlation To Load']]
            corr_data.append(['RE Contribution', corr])
        else:
            corr_data = None
        if option == D:
            wb = oxl.Workbook()
            ns = wb.active
            ns.title = 'Detail'
            normal = oxl.styles.Font(name='Arial')
            bold = oxl.styles.Font(name='Arial', bold=True)
            ss = wb.create_sheet('Summary', 0)
            ns_re_sum = '=('
            ns_tml_sum = '=('
            ns_sto_sum = ''
            ns_loss_sum = ''
            ns_not_sum = ''
            cap_row = 1
            ns.cell(row=cap_row, column=2).value = 'Capacity (MW/MWh)' #headers[1].replace('\n', ' ')
            ss.row_dimensions[3].height = 40
            ss.cell(row=3, column=st_fac+1).value = headers[st_fac] # facility
            ss.cell(row=3, column=st_cap+1).value = headers[st_cap] # capacity
            ini_row = 2
            ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
            tml_row = 3
            ns.cell(row=tml_row, column=2).value = headers[st_tml].replace('\n', ' ')
            ss.cell(row=3, column=st_tml+1).value = headers[st_tml] # to meet load
            sum_row = 4
            ns.cell(row=sum_row, column=2).value = headers[st_sub].replace('\n', ' ')
            ss.cell(row=3, column=st_sub+1).value = headers[st_sub] # subtotal MWh
            cf_row = 5
            ns.cell(row=cf_row, column=2).value = headers[st_cfa].replace('\n', ' ')
            ss.cell(row=3, column=st_cfa+1).value = headers[st_cfa] # CF
            cost_row = 6
            ns.cell(row=cost_row, column=2).value = headers[st_cst].replace('\n', ' ')
            ss.cell(row=3, column=st_cst+1).value = headers[st_cst] # Cost / yr
            lcog_row = 7
            ns.cell(row=lcog_row, column=2).value = headers[st_lcg].replace('\n', ' ')
            ss.cell(row=3, column=st_lcg+1).value = headers[st_lcg] # LCOG
            ss.cell(row=3, column=st_lco+1).value = headers[st_lco] # LCOE
            emi_row = 8
            ns.cell(row=emi_row, column=2).value = headers[st_emi].replace('\n', ' ')
            ss.cell(row=3, column=st_emi+1).value = headers[st_emi] # emissions
            ss.cell(row=3, column=st_emc+1).value = headers[st_emc] # emissions cost
            ss.cell(row=3, column=st_lcc+1).value = headers[st_lcc] # LCOE with CO2
            ss.cell(row=3, column=st_max+1).value = headers[st_max] # max. MWh
            ss.cell(row=3, column=st_bal+1).value = headers[st_bal] # max. balance
            ss.cell(row=3, column=st_cac+1).value = headers[st_cac] # capital cost
            ss.cell(row=3, column=st_lic+1).value = headers[st_lic] # lifetime cost
            ss.cell(row=3, column=st_lie+1).value = headers[st_lie] # lifetime emissions
            ss.cell(row=3, column=st_lec+1).value = headers[st_lec] # lifetime emissions cost
            ss.cell(row=3, column=st_are+1).value = headers[st_are] # area
            if self.do_jobs:
                ss.cell(row=3, column=st_job+1).value = headers[st_job] # job years
            ss.cell(row=3, column=st_rlc+1 - jm).value = headers[st_rlc - jm] # reference lcoe
            ss.cell(row=3, column=st_rcf+1 - jm).value = headers[st_rcf - jm] # reference cf
            ss_row = 3
            ss_re_fst_row = 4
            fall_row = 9
            ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
            max_row = 10
            ns.cell(row=max_row, column=2).value = 'Maximum (MW/MWh)'
            hrs_row = 11
            ns.cell(row=hrs_row, column=2).value = 'Hours of usage'
            if do_zone:
                zone_row = 12
                what_row = 13
                hrows = 14
                ns.cell(row=zone_row, column=1).value = 'Zone'
            else:
                what_row = 12
                hrows = 13
            ns.cell(row=what_row, column=1).value = 'Hour'
            ns.cell(row=what_row, column=2).value = 'Period'
            ns.cell(row=what_row, column=3).value = 'Load'
            ns.cell(row=sum_row, column=3).value = '=SUM(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=sum_row, column=3).number_format = '#,##0'
            ns.cell(row=max_row, column=3).value = '=MAX(' + ssCol(3) + str(hrows) + \
                                                   ':' + ssCol(3) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=3).number_format = '#,##0.00'
            o = 4
            col = 3
            # hour, period
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=1).value = row - hrows + 1
                ns.cell(row=row, column=2).value = format_period(row - hrows)
            # and load
            load_col = pmss_details['Load'].col
            if pmss_details['Load'].multiplier == 1:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=3).value = pmss_data[load_col][row - hrows] * \
                            pmss_details['Load'].multiplier
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            # here we're processing renewables (so no storage)
            for fac in re_order:
                if fac == 'Load':
                    continue
                if fac in underlying_facs:
                    continue
                if pmss_details[fac].col <= 0:
                    continue
                ss_row += 1
                col = do_detail(fac, col, ss_row)
                ns_tml_sum, ns_re_sum = do_detail_summary(fac, col, ss_row, ns_tml_sum, ns_re_sum)
            ss_re_lst_row = ss_row
            col += 1
            shrt_col = col
            ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ssCol(shrt_col) \
                            + str(hrows) + ':' + ssCol(shrt_col) + str(hrows + 8759) + \
                            ',"' + sf_test[0] + '0")'
            ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
            ns.cell(row=what_row, column=shrt_col).value = 'Shortfall (' + sf_sign[0] \
                    + ') /\nSurplus (' + sf_sign[1] + ')'
            ns.cell(row=max_row, column=shrt_col).value = '=MAX(' + ssCol(shrt_col) + str(hrows) + \
                                           ':' + ssCol(shrt_col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=shrt_col).number_format = '#,##0.00'
            for col in range(3, shrt_col + 1):
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                for col in range(3, shrt_col + 1):
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=shrt_col).value = shortfall[row - hrows] * -self.surplus_sign
                ns.cell(row=row, column=col).number_format = '#,##0.00'
            col = shrt_col + 1
            ns.cell(row=tml_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                   ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=tml_row, column=col).number_format = '#,##0'
            ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ')'
            ns.cell(row=max_row, column=col).number_format = '#,##0.00'
            ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(hrows) + \
                                           ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=hrs_row, column=col).number_format = '#,##0'
            ns.cell(row=what_row, column=col).value = 'RE Contrib.\nto Load'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            for row in range(hrows, 8760 + hrows):
                ns.cell(row=row, column=col).value = row_tml[row - hrows]
                ns.cell(row=row, column=col).number_format = '#,##0.00'
          #  shrt_col += 1
           # col = shrt_col + 1
            ul_re_sum = ns_re_sum
            ul_tml_sum = ns_tml_sum
            nsul_sums = ['C']
            nsul_sum_cols = [3]
            for fac in underlying_facs:
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                col = do_detail(fac, col, -1)
                nsul_sums.append(ssCol(col))
                nsul_sum_cols.append(col)
            if col > shrt_col + 1: # underlying
                col += 1
                ns.cell(row=what_row, column=col).value = 'Underlying\nLoad'
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(hrows) + \
                                                         ':' + ssCol(col) + str(hrows + 8759) + ')'
                ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                for row in range(hrows, 8760 + hrows):
                    txt = '='
                    for c in nsul_sums:
                        txt += c + str(row) + '+'
                    ns.cell(row=row, column=col).value = txt[:-1]
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            next_col = col
            col += 1
        else: # O, O1, B, S, T
            sp_data = []
            sp_load = 0. # load from load curve
            hrows = 10
            load_max = 0
            load_hr = 0
            tml = 0.
            try:
                load_col = pmss_details['Load'].col
            except:
                load_col = 0
            if (option == B or option == T) and len(underlying_facs) > 0:
                # at the moment for batch or transition we won't report operational and underlying separately
                load_facs = underlying_facs[:]
                load_facs.insert(0, 'Load')
                for h in range(len(pmss_data[load_col])):
                    amt = 0
                    for fac in load_facs:
                        amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
                    sp_load += amt
                underlying_facs = []
            else:
                fac = 'Load'
                sp_load = sum(pmss_data[load_col]) * pmss_details[fac].multiplier
                for h in range(len(pmss_data[load_col])):
                    amt = pmss_data[load_col][h] * pmss_details[fac].multiplier
                    if amt > load_max:
                        load_max = amt
                        load_hr = h
            for fac in re_order:
                if fac == 'Load' or fac in underlying_facs:
                    continue
                if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                    continue
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = fac
                sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                try:
                    sp_d[st_tml] = fac_tml[fac]
                except:
                    pass
                sp_d[st_sub] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                sp_data.append(sp_d)
        if option not in [O, O1, B, T]:
            self.progressbar.setValue(6)
            QtWidgets.QApplication.processEvents()
        storage_names = []
        tot_sto_loss = 0.
        committed_gen_bal = committed_gen_tot
        for gen in dispatch_order:
         #   min_after = [0, 0, -1, 0, 0, 0] # initial, low balance, period, final, low after, period
         #  Min_after is there to see if storage is as full at the end as at the beginning
            try:
                capacity = pmss_details[gen].capacity * pmss_details[gen].multiplier
            except:
                try:
                    capacity = pmss_details[gen].capacity
                except:
                    continue
            if gen not in self.generators.keys():
                continue
            if self.generators[gen].constraint in self.constraints and \
              self.constraints[self.generators[gen].constraint].category == 'Storage': # storage
                storage_names.append(gen)
                storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                storage[0] = capacity
                if option == D:
                    ns.cell(row=cap_row, column=col + 2).value = capacity
                    ns.cell(row=cap_row, column=col + 2).number_format = '#,##0.00'
                try:
                    storage[1] = self.generators[gen].initial * pmss_details[gen].multiplier
                except:
                    storage[1] = self.generators[gen].initial
                if self.constraints[self.generators[gen].constraint].capacity_min > 0:
                    storage[2] = capacity * self.constraints[self.generators[gen].constraint].capacity_min
                if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                    storage[3] = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                else:
                    storage[3] = capacity
                recharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].recharge_max > 0:
                    recharge[0] = capacity * self.constraints[self.generators[gen].constraint].recharge_max
                else:
                    recharge[0] = capacity
                if self.constraints[self.generators[gen].constraint].recharge_loss > 0:
                    recharge[1] = self.constraints[self.generators[gen].constraint].recharge_loss
                discharge = [0., 0.] # cap, loss
                if self.constraints[self.generators[gen].constraint].discharge_max > 0:
                    discharge[0] = capacity * self.constraints[self.generators[gen].constraint].discharge_max
                if self.constraints[self.generators[gen].constraint].discharge_loss > 0:
                    discharge[1] = self.constraints[self.generators[gen].constraint].discharge_loss
                if self.constraints[self.generators[gen].constraint].parasitic_loss > 0:
                    parasite = self.constraints[self.generators[gen].constraint].parasitic_loss / 24.
                else:
                    parasite = 0.
                in_run = [False, False]
                min_run_time = self.constraints[self.generators[gen].constraint].min_run_time
                in_run[0] = True # start off in_run
                if min_run_time > 0 and self.generators[gen].initial == 0:
                    in_run[0] = False
                warm_time = self.constraints[self.generators[gen].constraint].warm_time
                storage_carry = storage[1] # self.generators[gen].initial
                if option == D:
                    ns.cell(row=ini_row, column=col + 2).value = storage_carry
                    ns.cell(row=ini_row, column=col + 2).number_format = '#,##0.00'
                storage_bal = []
                storage_can = 0.
                use_max = [0, None]
                sto_max = storage_carry
                for row in range(8760):
                    storage_loss = 0.
                    storage_losses = 0.
                    if storage_carry > 0:
                        loss = storage_carry * parasite
                        # for later: record parasitic loss
                        storage_carry = storage_carry - loss
                        storage_losses -= loss
                    if shortfall[row] < 0:  # excess generation
                        if row % 24 >= self.constraints[self.generators[gen].constraint].recharge_start:
                            if min_run_time > 0:
                                in_run[0] = False
                            if warm_time > 0:
                                in_run[1] = False
                            can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                            if can_use < 0: # can use some
                                if shortfall[row] > can_use:
                                    can_use = shortfall[row]
                                if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                    can_use = - recharge[0]
                            else:
                                can_use = 0.
                            # for later: record recharge loss
                            storage_losses += can_use * recharge[1]
                            storage_carry -= (can_use * (1 - recharge[1]))
                            shortfall[row] -= can_use
                            if corr_data is not None:
                               corr_src[row] += can_use
                        else:
                            can_use = 0
                    elif shortfall[row] > committed_gen_bal: # shortfall
                        # This is code to support delaying battery usage until a certain time
                        # to implement fully need an additional facility variable to indicate start time
                        # Ref 2024 WEM ESOO 2.5 (ESROI)
                        if row % 24 >= self.constraints[self.generators[gen].constraint].discharge_start:
                            if min_run_time > 0 and shortfall[row] > committed_gen_bal:
                                if not in_run[0]:
                                    if row + min_run_time <= 8759:
                                        for i in range(row + 1, row + min_run_time + 1):
                                            if shortfall[i] <= 0:
                                                break
                                        else:
                                            in_run[0] = True
                            if in_run[0]:
                                can_use = (shortfall[row] - committed_gen_bal) * (1 / (1 - discharge[1]))
                                can_use = min(can_use, discharge[0])
                                if can_use > storage_carry - storage[2]:
                                    can_use = storage_carry - storage[2]
                                if warm_time > 0 and not in_run[1]:
                                    in_run[1] = True
                                    can_use = can_use * (1 - warm_time)
                            else:
                                can_use = 0
                            if can_use > 0:
                                storage_loss = can_use * discharge[1]
                                storage_losses -= storage_loss
                                storage_carry -= can_use
                                can_use = can_use - storage_loss
                                shortfall[row] -= can_use
                                if corr_data is not None:
                                    corr_src[row] += can_use
                                if storage_carry < 0:
                                    storage_carry = 0
                            else:
                                can_use = 0.
                        else:
                            can_use = 0.
                    else:
                        can_use = 0.
                    if can_use < 0:
                        if use_max[1] is None or can_use < use_max[1]:
                            use_max[1] = can_use
                    elif can_use > use_max[0]:
                        use_max[0] = can_use
                    storage_bal.append(storage_carry)
                    if storage_bal[-1] > sto_max:
                        sto_max = storage_bal[-1]
                    if option == D:
                        if can_use > 0:
                            ns.cell(row=row + hrows, column=col).value = 0
                            ns.cell(row=row + hrows, column=col + 2).value = can_use * self.surplus_sign
                        else:
                            ns.cell(row=row + hrows, column=col).value = can_use * -self.surplus_sign
                            ns.cell(row=row + hrows, column=col + 2).value = 0
                        ns.cell(row=row + hrows, column=col + 1).value = storage_losses
                        ns.cell(row=row + hrows, column=col + 3).value = storage_carry
                        ns.cell(row=row + hrows, column=col + 4).value = shortfall[row] * -self.surplus_sign
                        for ac in range(5):
                            ns.cell(row=row + hrows, column=col + ac).number_format = '#,##0.00'
                            ns.cell(row=max_row, column=col + ac).value = '=MAX(' + ssCol(col + ac) + \
                                    str(hrows) + ':' + ssCol(col + ac) + str(hrows + 8759) + ')'
                            ns.cell(row=max_row, column=col + ac).number_format = '#,##0.00'
                    else:
                        tot_sto_loss += storage_losses
                        if can_use > 0:
                            storage_can += can_use
                if option == D:
                    ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 1).value = '=SUMIF(' + ssCol(col + 1) + \
                            str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + ',"<0")'
                    ns.cell(row=sum_row, column=col + 1).number_format = '#,##0'
                    ns.cell(row=sum_row, column=col + 2).value = '=SUMIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col + 2).value = '=IF(' + ssCol(col + 2) + str(cap_row) + '>0,' + \
                            ssCol(col + 2) + str(sum_row) + '/' + ssCol(col + 2) + '1/8760,"")'
                    ns.cell(row=cf_row, column=col + 2).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col + 2).value = '=COUNTIF(' + ssCol(col + 2) + \
                            str(hrows) + ':' + ssCol(col + 2) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col + 2).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 3).value = '=' + ssCol(col + 2) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 3).number_format = '#,##0.0%'
                    col += 5
                else:
                    if storage[0] == 0:
                        continue
               #     tml_tot += storage_can
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = storage[0]
                    sp_d[st_tml] = storage_can
                    sp_d[st_max] = use_max[0]
                    sp_d[st_bal] = sto_max
                    sp_data.append(sp_d)
            else: # generator
                try:
                    if self.constraints[self.generators[gen].constraint].capacity_max > 0:
                        cap_capacity = capacity * self.constraints[self.generators[gen].constraint].capacity_max
                    else:
                        cap_capacity = capacity
                except:
                    cap_capacity = capacity
                if gen in committed_gen.keys():
                    min_gen = committed_gen[gen]
                else:
                    min_gen = 0
                if option == D:
                    ns.cell(row=cap_row, column=col).value = capacity
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            short_4me = shortfall[row]
                            if self.reserve_committed:
                                try:
                                    short_4me = shortfall[row] - committed_gen_bal + committed_gen[gen]
                                except:
                                    pass
                            if short_4me >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                ns.cell(row=row + hrows, column=col).value = cap_capacity
                            elif short_4me >= min_gen:
                                shortfall[row] = shortfall[row] - short_4me
                                ns.cell(row=row + hrows, column=col).value = short_4me
                            elif short_4me < min_gen:
                                ns.cell(row=row + hrows, column=col).value = min_gen
                                shortfall[row] -= min_gen
                            else:
                                ns.cell(row=row + hrows, column=col).value = shortfall[row]
                                shortfall[row] = 0
                        else:
                            shortfall[row] -= min_gen
                            ns.cell(row=row + hrows, column=col).value = min_gen
                        ns.cell(row=row + hrows, column=col + 1).value = shortfall[row] * -self.surplus_sign
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ssCol(col) + str(hrows) + \
                            ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=IF(' + ssCol(col) + str(cap_row) + '>0,' + \
                            ssCol(col) + str(sum_row) + '/' + ssCol(col) + str(cap_row) + '/8760,"")'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.0%'
                    ns.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + \
                                str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ')'
                    ns.cell(row=max_row, column=col).number_format = '#,##0.00'
                    ns.cell(row=hrs_row, column=col).value = '=COUNTIF(' + ssCol(col) + \
                            str(hrows) + ':' + ssCol(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=hrs_row, column=col).number_format = '#,##0'
                    ns.cell(row=hrs_row, column=col + 1).value = '=' + ssCol(col) + \
                            str(hrs_row) + '/8760'
                    ns.cell(row=hrs_row, column=col + 1).number_format = '#,##0.0%'
                    col += 2
                else:
                    gen_can = 0.
                    gen_max = 0
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            short_4me = shortfall[row]
                            if self.reserve_committed:
                                try:
                                    short_4me = shortfall[row] - committed_gen_bal + committed_gen[gen]
                                except:
                                    pass
                            if short_4me >= cap_capacity:
                                shortfall[row] = shortfall[row] - cap_capacity
                                gen_can += cap_capacity
                                if cap_capacity > gen_max:
                                    gen_max = cap_capacity
                            elif short_4me >= min_gen:
                                gen_can += short_4me
                                if short_4me > gen_max:
                                    gen_max = short_4me
                                shortfall[row] = shortfall[row] - short_4me
                            elif short_4me < min_gen:
                                gen_can += min_gen
                                if min_gen > gen_max:
                                    gen_max = min_gen
                                shortfall[row] -= min_gen
                            else:
                                gen_can += shortfall[row]
                                if shortfall[row] > gen_max:
                                    gen_max = shortfall[row]
                                shortfall[row] = 0
                        else:
                            if min_gen > gen_max:
                                gen_max = min_gen
                            gen_can += min_gen
                            shortfall[row] -= min_gen
                    if capacity == 0:
                        continue
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = gen
                    sp_d[st_cap] = capacity
                    sp_d[st_tml] = gen_can
                    sp_d[st_sub] = gen_can
                    sp_d[st_max] = gen_max
                    sp_data.append(sp_d)
            if self.reserve_committed:
                try:
                    committed_gen_bal -= committed_gen[gen]
                except:
                    pass
#        if option == D: # Currently calculated elsewhere
#            if self.surplus_sign > 0:
#                maxmin = 'MIN'
#            else:
#                maxmin = 'MAX'
#            ns.cell(row=max_row, column=col-1).value = '=' + maxmin + '(' + \
#                    ssCol(col-1) + str(hrows) + ':' + ssCol(col - 1) + str(hrows + 8759) + ')'
#            ns.cell(row=max_row, column=col-1).number_format = '#,##0.00'
        if option not in [O, O1, B, T]:
            self.progressbar.setValue(8)
            QtWidgets.QApplication.processEvents()
        if corr_data is not None:
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['RE plus Storage', corr])
            col = pmss_details['Load'].col
            corr_src = []
            for h in range(len(shortfall)):
                if shortfall[h] < 0:
                    corr_src.append(pmss_data[col][h])
                else:
                    corr_src.append(pmss_data[col][h] - shortfall[h])
            try:
                corr = np.corrcoef(df1, corr_src)
                if np.isnan(corr.item((0, 1))):
                    corr = 0
                else:
                    corr = corr.item((0, 1))
            except:
                corr = 0
            corr_data.append(['To Meet Load', corr])
            for c in range(1, len(corr_data)):
                if abs(corr_data[c][1]) < 0.1:
                    corr_data[c].append('None')
                elif abs(corr_data[c][1]) < 0.3:
                    corr_data[c].append('Little if any')
                elif abs(corr_data[c][1]) < 0.5:
                    corr_data[c].append('Low')
                elif abs(corr_data[c][1]) < 0.7:
                    corr_data[c].append('Moderate')
                elif abs(corr_data[c][1]) < 0.9:
                    corr_data[c].append('High')
                else:
                    corr_data[c].append('Very high')
        if option != D:
            load_col = pmss_details['Load'].col
            cap_sum = 0.
            gen_sum = 0.
            re_sum = 0.
            tml_sum = 0.
            ff_sum = 0.
            sto_sum = 0.
            cost_sum = 0.
            co2_sum = 0.
            co2_cost_sum = 0.
            capex_sum = 0.
            lifetime_sum = 0.
            lifetime_co2_sum = 0.
            lifetime_co2_cost = 0.
            total_area = 0.
            total_jobs = 0.
            #this logic for gen2 is pretty crappy need to revisit
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                if gen in storage_names:
                    sto_sum += sp_data[sp][2]
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen in tech_names or gen2 in tech_names:
                        re_sum += sp_data[sp][st_sub]
            for sp in range(len(sp_data)):
                gen = sp_data[sp][st_fac]
                gen2 = gen
                if gen in storage_names:
                    ndx = 2
                else:
                    if gen in self.generators.keys():
                        pass
                    else:
                        try:
                            gen = gen[gen.find('.') + 1:]
                            gen2 = gen
                        except:
                            pass
                    ndx = 3
                try:
                    if sp_data[sp][st_cap] > 0:
                        cap_sum += sp_data[sp][st_cap]
                        if self.generators[gen].lcoe > 0:
                            sp_data[sp][st_cfa] = sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 # need number for now
                        else:
                            sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][ndx] / sp_data[sp][st_cap] / 8760 * 100)
                    gen_sum += sp_data[sp][st_sub]
                except:
                    pass
                try:
                    tml_sum += sp_data[sp][st_tml]
                except:
                    pass
                if gen not in self.generators.keys():
                    continue
                if self.do_jobs:
                    try:
                        genj = gen[gen.find('.') + 1:]
                    except:
                        genj = gen
                ndx = 3
                if gen in storage_names:
                    ndx = 2
                else:
                    try:
                        gen2 = gen[gen.find('.') + 1:]
                    except:
                        gen2 = gen
                    if gen not in tech_names and gen2 not in tech_names:
                        ff_sum += sp_data[sp][ndx]
                if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                  or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    capex = sp_data[sp][st_cap] * self.generators[gen].capex
                    capex_sum += capex
                    opex = sp_data[sp][st_cap] * self.generators[gen].fixed_om \
                           + sp_data[sp][ndx] * self.generators[gen].variable_om \
                           + sp_data[sp][ndx] * self.generators[gen].fuel
                    disc_rate = self.generators[gen].disc_rate
                    if disc_rate == 0:
                        disc_rate = self.discount_rate
                    lifetime = self.generators[gen].lifetime
                    sp_data[sp][st_lcg] = calcLCOE(sp_data[sp][ndx], capex, opex, disc_rate, lifetime)
                    if sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = sp_data[sp][st_lcg]
                    else:
                        sp_data[sp][st_cst] = sp_data[sp][ndx] * sp_data[sp][st_lcg]
                    if gen in tech_names or gen2 in tech_names:
                        if self.corrected_lcoe and fac_tml_sum > 0:
                            sp_data[sp][st_lco] = sp_data[sp][st_cst] / (sp_data[sp][st_tml] + (sto_sum * sp_data[sp][st_tml] / fac_tml_sum))
                        else:
                            sp_data[sp][st_lco] = sp_data[sp][st_cst] / sp_data[sp][st_tml]
                    else:
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_cac] = capex
                elif self.generators[gen].lcoe > 0:
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    if self.generators[gen].lcoe_cf > 0:
                        lcoe_cf = self.generators[gen].lcoe_cf
                    else:
                        lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_data[sp][st_cap]
                    if sp_data[sp][st_cfa] > 0:
                        sp_data[sp][st_lcg] = sp_data[sp][st_cst] / sp_data[sp][ndx]
                        sp_data[sp][st_lco] = sp_data[sp][st_lcg]
                    sp_data[sp][st_cfa] = '{:.1f}%'.format(sp_data[sp][st_cfa] * 100.)
                    cost_sum += sp_data[sp][st_cst]
                    sp_data[sp][st_rlc - jm] = self.generators[gen].lcoe
                    sp_data[sp][st_rcf - jm] = '{:.1f}%'.format(lcoe_cf * 100.)
                elif self.generators[gen].lcoe_cf == 0: # no cost facility
                    if option != T and self.remove_cost and sp_data[sp][ndx] == 0:
                        sp_data[sp][st_cst] = 0
                        continue
                    lcoe_cf = sp_data[sp][st_cfa]
                    sp_data[sp][st_cst] = 0
                    cost_sum += sp_data[sp][st_cst]
                sp_data[sp][st_lic] = sp_data[sp][st_cst] * max_lifetime
                lifetime_sum += sp_data[sp][st_lic]
                if self.generators[gen].emissions > 0 and sp_data[sp][st_tml] > 0:
                    sp_data[sp][st_emi] = sp_data[sp][ndx] * self.generators[gen].emissions
                    co2_sum += sp_data[sp][st_emi]
                    sp_data[sp][st_emc] = sp_data[sp][st_emi] * self.carbon_price
                    if sp_data[sp][st_cst] == 0:
                        sp_data[sp][st_lcc] = sp_data[sp][st_emc] / sp_data[sp][st_tml]
                    else:
                        sp_data[sp][st_lcc] = sp_data[sp][st_lco] * ((sp_data[sp][st_cst] + sp_data[sp][st_emc]) / sp_data[sp][st_cst])
                    co2_cost_sum += sp_data[sp][st_emc]
                    sp_data[sp][st_lie] = sp_data[sp][st_emi] * max_lifetime
                    lifetime_co2_sum += sp_data[sp][st_lie]
                    sp_data[sp][st_lec] = sp_data[sp][st_lie] * self.carbon_price
                    lifetime_co2_cost += sp_data[sp][st_lec]
                else:
                    sp_data[sp][st_lcc] = sp_data[sp][st_lco]
                if self.generators[gen].area > 0:
                    sp_data[sp][st_are] = sp_data[sp][st_cap] * self.generators[gen].area
                    total_area += sp_data[sp][st_are]
                if self.do_jobs:
                    jobs = 0.
                    try:
                        if sp_data[sp][st_cac] > 0:
                            jobs = (self.jobfactors[genj].manufacture * self.jobfactors[genj].local_pct + self.jobfactors[genj].install) * sp_data[sp][st_cap]
                    except:
                        pass
                    try:
                        jobs += self.jobfactors[genj].operate * sp_data[sp][st_cap] #* max_lifetime
                    except:
                        pass
                    sp_data[sp][st_job] = jobs
                    total_jobs += jobs
            sf_sums = [0., 0., 0.]
            for sf in range(len(shortfall)):
                if shortfall[sf] > 0:
                    sf_sums[0] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
                else:
                    sf_sums[1] += shortfall[sf]
                    sf_sums[2] += pmss_data[load_col][sf] * pmss_details['Load'].multiplier
            if gen_sum > 0:
                gs = cost_sum / gen_sum
            else:
                gs = ''
            if tml_sum > 0:
                gsw = cost_sum / tml_sum # LCOE
                gswc = (cost_sum + co2_cost_sum) / tml_sum
            else:
                gsw = ''
                gswc = ''
            if option == O or option == O1:
                load_pct, surp_pct, re_pct = summary_totals()
            else:
                summary_totals()
            do_underlying = False
            if len(underlying_facs) > 0:
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier > 0:
                        do_underlying = True
                        break
            if do_underlying:
                sp_data.append(' ')
                sp_data.append('Additional Underlying Load')
                for fac in underlying_facs:
                    if pmss_details[fac].capacity * pmss_details[fac].multiplier == 0:
                        continue
                    if fac in self.generators.keys():
                        gen = fac
                    else:
                        gen = pmss_details[fac].generator
                    col = pmss_details[fac].col
                    sp_d = [' '] * len(headers)
                    sp_d[st_fac] = fac
                    sp_d[st_cap] = pmss_details[fac].capacity * pmss_details[fac].multiplier
                    cap_sum += sp_d[st_cap]
                    sp_d[st_tml] = sum(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    tml_sum += sp_d[st_tml]
                    sp_d[st_sub] = sp_d[st_tml]
                    gen_sum += sp_d[st_tml]
                    sp_load += sp_d[st_tml]
                    sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_sub] / sp_d[st_cap] / 8760 * 100.)
                    sp_d[st_max] = max(pmss_data[pmss_details[fac].col]) * pmss_details[fac].multiplier
                    if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
                      or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                        capex = sp_d[st_cap] * self.generators[gen].capex
                        capex_sum += capex
                        opex = sp_d[st_cap] * self.generators[gen].fixed_om \
                               + sp_d[st_tml] * self.generators[gen].variable_om \
                               + sp_d[st_tml] * self.generators[gen].fuel
                        disc_rate = self.generators[gen].disc_rate
                        if disc_rate == 0:
                            disc_rate = self.discount_rate
                        lifetime = self.generators[gen].lifetime
                        sp_d[st_lcg] = calcLCOE(sp_d[st_tml], capex, opex, disc_rate, lifetime)
                        sp_d[st_cst] = sp_d[st_tml] * sp_d[st_lcg]
                        cost_sum += sp_d[st_cst]
                        sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cac] = capex
                    elif self.generators[gen].lcoe > 0:
                        if self.generators[gen].lcoe_cf > 0:
                            lcoe_cf = self.generators[gen].lcoe_cf
                        else:
                            lcoe_cf = sp_d[st_cfa]
                        sp_d[st_cst] = self.generators[gen].lcoe * lcoe_cf * 8760 * sp_d[st_cap]
                        cost_sum += sp_d[st_cst]
                        if sp_d[st_cfa] > 0:
                            sp_d[st_lcg] = sp_d[st_cst] / sp_d[st_tml]
                            sp_d[st_lco] = sp_d[st_lcg]
                        sp_d[st_cfa] = '{:.1f}%'.format(sp_d[st_cfa] * 100.)
                        sp_d[st_rlc - jm] = self.generators[gen].lcoe
                        sp_d[st_rcf - jm] = '{:.1f}%'.format(lcoe_cf * 100.)
                    elif self.generators[gen].lcoe_cf == 0: # no cost facility
                        sp_d[st_cst] = 0
                        sp_d[st_lcg] = 0
                        sp_d[st_lco] = 0
                        sp_d[st_rlc - jm] = self.generators[gen].lcoe
                    sp_d[st_lic] = sp_d[st_cst] * max_lifetime
                    lifetime_sum += sp_d[st_lic]
                    if self.generators[gen].emissions > 0:
                        sp_d[st_emi] = sp_d[st_tml] * self.generators[gen].emissions
                        co2_sum += sp_d[st_emi]
                        sp_d[st_emc] = sp_d[st_emi] * self.carbon_price
                        if sp_d[st_cst] > 0:
                            sp_d[st_lcc] = sp_d[st_lco] * ((sp_d[st_cst] + sp_d[st_emc]) / sp_d[st_cst])
                        else:
                            sp_d[st_lcc] = sp_d[st_emc] / sp_d[st_tml]
                        co2_cost_sum += sp_d[st_emc]
                        sp_d[st_lie] = sp_d[st_emi] * max_lifetime
                        lifetime_co2_sum += sp_d[st_lie]
                        sp_d[st_lec] = sp_d[st_lie] * self.carbon_price
                        lifetime_co2_cost += sp_d[st_lec]
                    else:
                        sp_d[st_lcc] = sp_d[st_lco]
                    if self.generators[gen].area > 0:
                        sp_d[st_are] = sp_d[st_cap] * self.generators[gen].area
                    if self.do_jobs:
                        try:
                            jobs = 0
                            if sp_data[sp][st_cac] > 0:
                                jobs = (self.jobfactors[genj].manufacture * self.jobfactors[genj].local_pct + self.jobfactors[genj].install) * sp_data[sp][st_cap]
                            jobs += self.jobfactors[genj].operate * sp_data[sp][st_cap]
                            sp_d[st_job] = jobs
                            total_jobs += jobs
                        except:
                            pass
                    sp_data.append(sp_d)
                if gen_sum > 0:
                    gs = cost_sum / gen_sum
                else:
                    gs = ''
                if tml_sum > 0:
                    gsw = cost_sum / tml_sum # LCOE
                    gswc = (cost_sum + co2_cost_sum) / tml_sum
                else:
                    gsw = ''
                    gswc = ''
                # find maximum underlying load
                if option == S:
                    load_max = 0
                    load_hr = 0
                    load_col = pmss_details['Load'].col
                    for h in range(len(pmss_data[load_col])):
                        amt = pmss_data[load_col][h] * pmss_details['Load'].multiplier
                        for fac in underlying_facs:
                            amt += pmss_data[pmss_details[fac].col][h] * pmss_details[fac].multiplier
                        if amt > load_max:
                            load_max = amt
                            load_hr = h
                if option == O or option == O1:
                    load_pct, surp_pct, re_pct = summary_totals('Underlying ')
                else:
                    summary_totals('Underlying ')
            if corr_data is not None:
                sp_data.append(' ')
                sp_data = sp_data + corr_data
            sp_data.append(' ')
            sp_data.append(['Static Variables'])
            if self.carbon_price > 0:
                sp_d = [' '] * len(headers)
                sp_d[st_fac] = 'Carbon Price ($/tCO2e)'
                sp_d[st_cap] = self.carbon_price
                sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Lifetime (years)'
            sp_d[st_cap] = max_lifetime
            sp_data.append(sp_d)
            sp_d = [' '] * len(headers)
            sp_d[st_fac] = 'Discount Rate'
            sp_d[st_cap] = '{:.2%}'.format(self.discount_rate)
            sp_data.append(sp_d)
            if option == B or option == T:
                if self.optimise_debug:
                    sp_pts = [0] * len(headers)
                    for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc - jm, st_are]:
                        sp_pts[p] = 2
                    if corr_data is not None:
                        sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
                    dialog = displaytable.Table(sp_data, title='Debug', fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=sp_pts)
                    dialog.exec_()
                return sp_data
            if option == O or option == O1:
                op_load_tot = pmss_details['Load'].capacity * pmss_details['Load'].multiplier
                if gswc != '':
                    lcoe = gswc
                elif self.adjusted_lcoe:
                    lcoe = gsw # target is lcoe
                else:
                    lcoe = gs
                if gen_sum == 0:
                    load_pct = 0
                    re_pct = 0
                multi_value = {'lcoe': lcoe, #lcoe. lower better
                    'load_pct': load_pct, #load met. 100% better
                    'surplus_pct': surp_pct, #surplus. lower better
                    're_pct': re_pct, # RE pct. higher better
                    'cost': cost_sum, # cost. lower better
                    'co2': co2_sum} # CO2. lower better
                if option == O:
                    if multi_value['lcoe'] == '':
                        multi_value['lcoe'] = 0
                    return multi_value, sp_data, None
                else:
                    extra = [gsw, op_load_tot, sto_sum, re_sum, re_pct, sf_sums]
                    return multi_value, sp_data, extra
        #    list(map(list, list(zip(*sp_data))))
            span = None
            if self.summary_sources: # want data sources
                sp_data.append(' ')
                sp_data.append('Data sources')
                span = 'Data sources'
                if len(sys.argv) > 1:
                    config_file = getModelFile(sys.argv[1])
                else:
                   config_file = getModelFile('SIREN.ini')
                sp_data.append(['Preferences file', config_file[-1]])
                sp_data.append(['Scenarios folder', self.scenarios])
                if pm_data_file[: len(self.scenarios)] == self.scenarios:
                    pm_data_file = pm_data_file[len(self.scenarios):]
                sp_data.append(['Powermatch data file', pm_data_file])
                load_file = self.load_files.replace('$YEAR$', self.loadCombo.currentText())
                if load_file[: len(self.scenarios)] == self.scenarios:
                    load_file = load_file[len(self.scenarios):]
                sp_data.append(['Load file', load_file])
                sp_data.append(['Constraints worksheet', str(self.files[C].text()) \
                                + '.' + str(self.sheets[C].currentText())])
                sp_data.append(['Generators worksheet', str(self.files[G].text()) \
                                + '.' + str(self.sheets[G].currentText())])
            sp_pts = [0] * len(headers)
            for p in [st_cap, st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc - jm, st_are]:
                sp_pts[p] = 2
            if corr_data is not None:
                sp_pts[st_cap] = 3 # compromise between capacity (2) and correlation (4)
            self.setStatus(self.sender().text() + ' completed')
            if title is not None:
                atitle = title
            elif self.results_prefix != '':
                atitle = self.results_prefix + '_' + self.sender().text()
            else:
                atitle = self.sender().text()
            dialog = displaytable.Table(sp_data, title=atitle, fields=headers,
                     save_folder=self.scenarios, sortby='', decpts=sp_pts,
                     span=span)
            dialog.exec_()
            self.progressbar.setValue(20)
            self.progressbar.setHidden(True)
            self.progressbar.setValue(0)
            return # finish if not detailed spreadsheet
        col = next_col + 1
        is_storage = False
        ss_sto_rows = []
        ss_st_row = -1
        for gen in dispatch_order:
            ss_row += 1
            try:
                if self.constraints[self.generators[gen].constraint].category == 'Storage':
                    ss_sto_rows.append(ss_row)
                    nc = 2
                    ns.cell(row=what_row, column=col).value = 'Charge\n' + gen
                    ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    ns.cell(row=what_row, column=col + 1).value = gen + '\nLosses'
                    ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    is_storage = True
                    ns_sto_sum += '+' + ssCol(st_tml+1) + str(ss_row)
                    ns_loss_sum += '+Detail!' + ssCol(col + 1) + str(sum_row)
                else:
                    nc = 0
                    is_storage = False
                    ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            except KeyError as err:
                msg = 'Key Error: No Constraint for ' + gen
                if title is not None:
                    msg += ' (model ' + title + ')'
                self.setStatus(msg)
                nc = 0
                is_storage = False
                ns_not_sum += '-' + ssCol(st_tml+1) + str(ss_row)
            ns.cell(row=what_row, column=col + nc).value = gen
            ss.cell(row=ss_row, column=st_fac+1).value = '=Detail!' + ssCol(col + nc) + str(what_row)
            # facility
            ss.cell(row=ss_row, column=st_cap+1).value = '=Detail!' + ssCol(col + nc) + str(cap_row)
            # capacity
            ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00'
            # tml
            ss.cell(row=ss_row, column=st_tml+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
            ss.cell(row=ss_row, column=st_tml+1).number_format = '#,##0'
            # subtotal
            try:
                if self.constraints[self.generators[gen].constraint].category != 'Storage':
                    ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                    ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            except KeyError as err:
                ss.cell(row=ss_row, column=st_sub+1).value = '=Detail!' + ssCol(col + nc) + str(sum_row)
                ss.cell(row=ss_row, column=st_sub+1).number_format = '#,##0'
            # cf
            ss.cell(row=ss_row, column=st_cfa+1).value = '=Detail!' + ssCol(col + nc) + str(cf_row)
            ss.cell(row=ss_row, column=st_cfa+1).number_format = '#,##0.0%'
            if self.generators[gen].capex > 0 or self.generators[gen].fixed_om > 0 \
              or self.generators[gen].variable_om > 0 or self.generators[gen].fuel > 0:
                disc_rate = self.generators[gen].disc_rate
                if disc_rate == 0:
                    disc_rate = self.discount_rate
                if disc_rate == 0:
                    cst_calc = '/' + str(self.generators[gen].lifetime)
                else:
                    pwr_calc = 'POWER(1+' + str(disc_rate) + ',' + str(self.generators[gen].lifetime) + ')'
                    cst_calc = '*' + str(disc_rate) + '*' + pwr_calc + '/SUM(' + pwr_calc + ',-1)'
                if self.remove_cost:
                    el2 = '),0)'
                else:
                    el2 = f'),{ssCol(col + nc)}{cap_row}*{self.generators[gen].capex}{cst_calc}+' + \
                          f'{ssCol(col + nc)}{cap_row}*{self.generators[gen].fixed_om})'
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(cap_row) + '*' + str(self.generators[gen].capex) + \
                        cst_calc + '+' + ssCol(col + nc) + str(cap_row) + '*' + \
                        str(self.generators[gen].fixed_om) + '+' + ssCol(col + nc) + str(sum_row) + '*(' + \
                        str(self.generators[gen].variable_om) + '+' + str(self.generators[gen].fuel) + el2
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
              #  ns.cell(row=lcog_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + \
               #         '>0,' + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + \
                #        str(cost_row) + '/' + ssCol(col + nc) + str(sum_row) + ',"")'
                ns.cell(row=lcog_row, column=col + nc).value = f'=IF({ssCol(col + nc)}{cap_row}>0,' + \
                        f'IF({ssCol(col + nc)}{cf_row}>0,{ssCol(col + nc)}{cost_row}/{ssCol(col + nc)}{sum_row},' +\
                        f'IF({ssCol(col + nc)}{cost_row}>0,{ssCol(col + nc)}{cost_row},"")),"")'
                ns.cell(row=lcog_row, column=col + nc).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # capital cost
                ss.cell(row=ss_row, column=st_cac+1).value = f'=IF(AND(Detail!{ssCol(col + nc)}{cap_row}>0,' + \
                        f'Detail!{ssCol(col + nc)}{cost_row}>0),Detail!{ssCol(col + nc)}{cap_row}*{self.generators[gen].capex},"")'
                ss.cell(row=ss_row, column=st_cac+1).number_format = '$#,##0'
            elif self.generators[gen].lcoe > 0:
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1 - jm) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1 - jm) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcog_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcog_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1 - jm).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1 - jm).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).number_format = '#,##0.0%'
            elif self.generators[gen].lcoe_cf == 0: # no cost facility
                ns.cell(row=cost_row, column=col + nc).value = '=IF(' + ssCol(col + nc) + str(cf_row) + \
                        '>0,' + ssCol(col + nc) + str(sum_row) + '*Summary!' + ssCol(st_rlc + 1 - jm) + str(ss_row) + \
                        '*Summary!' + ssCol(st_rcf + 1 - jm) + str(ss_row) + '/' + ssCol(col + nc) + str(cf_row) + ',0)'
                ns.cell(row=cost_row, column=col + nc).number_format = '$#,##0'
                # cost / yr
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(cost_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_cst+1).value = '=Detail!' + ssCol(col + nc) + str(cost_row)
                ss.cell(row=ss_row, column=st_cst+1).number_format = '$#,##0'
                ns.cell(row=lcog_row, column=col + nc).value = '=IF(AND(' + ssCol(col + nc) + str(cf_row) + '>0,' \
                            + ssCol(col + nc) + str(cap_row) + '>0),' + ssCol(col + nc) + str(cost_row) + '/8760/' \
                            + ssCol(col + nc) + str(cf_row) + '/' + ssCol(col + nc) + str(cap_row)+  ',"")'
                ns.cell(row=lcog_row, column=col + nc).number_format = '$#,##0.00'
                # lcog
                ss.cell(row=ss_row, column=st_lcg+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lcg+1).number_format = '$#,##0.00'
                # lcoe
                ss.cell(row=ss_row, column=st_lco+1).value = '=Detail!' + ssCol(col + nc) + str(lcog_row)
                ss.cell(row=ss_row, column=st_lco+1).number_format = '$#,##0.00'
                # ref lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).value = self.generators[gen].lcoe
                ss.cell(row=ss_row, column=st_rlc+1 - jm).number_format = '$#,##0.00'
                # ref cf
                if self.generators[gen].lcoe_cf == 0:
                    ss.cell(row=ss_row, column=st_rcf+1 - jm).value = '=' + ssCol(st_cfa+1) + str(ss_row)
                else:
                    ss.cell(row=ss_row, column=st_rcf+1 - jm).value = self.generators[gen].lcoe_cf
                ss.cell(row=ss_row, column=st_rcf+1 - jm).number_format = '#,##0.0%'
            if self.generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col + nc).value = '=' + ssCol(col + nc) + str(sum_row) \
                        + '*' + str(self.generators[gen].emissions)
                ns.cell(row=emi_row, column=col + nc).number_format = '#,##0'
                # emissions
                if self.remove_cost:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=IF(Detail!' + ssCol(col + nc) + str(sum_row) \
                            + '>0,Detail!' + ssCol(col + nc) + str(emi_row) + ',"")'
                else:
                    ss.cell(row=ss_row, column=st_emi+1).value = '=Detail!' + ssCol(col + nc) + str(emi_row)
                ss.cell(row=ss_row, column=st_emi+1).number_format = '#,##0'
                if self.carbon_price > 0:
                    ss.cell(row=ss_row, column=st_emc+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                                 ssCol(st_emi+1) + str(ss_row) + '*carbon_price,"")'
                    ss.cell(row=ss_row, column=st_emc+1).number_format = '$#,##0'
            # max mwh
            ss.cell(row=ss_row, column=st_max+1).value = '=Detail!' + ssCol(col + nc) + str(max_row)
            ss.cell(row=ss_row, column=st_max+1).number_format = '#,##0.00'
            # max balance
            if nc > 0: # storage
                ss.cell(row=ss_row, column=st_bal+1).value = '=Detail!' + ssCol(col + nc + 1) + str(max_row)
                ss.cell(row=ss_row, column=st_bal+1).number_format = '#,##0.00'
            ns.cell(row=what_row, column=col + nc).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + nc + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if is_storage:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col + 2) + str(cost_row) \
                                                        + '>0,Detail!' + ssCol(col + 2) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                # ns.cell(row=what_row, column=col + 1).value = gen
                ns.cell(row=what_row, column=col + 3).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 3).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=what_row, column=col + 4).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 4).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 4).value = '=COUNTIF(' + ssCol(col + 4) \
                        + str(hrows) + ':' + ssCol(col + 4) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 4).number_format = '#,##0'
                col += 5
            else:
                # lifetime cost
                ss.cell(row=ss_row, column=st_lic+1).value = '=IF(Detail!' + ssCol(col) + str(cost_row) \
                                                        + '>0,Detail!' + ssCol(col) + str(cost_row) + '*lifetime,"")'
                ss.cell(row=ss_row, column=st_lic+1).number_format = '$#,##0'
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ssCol(col + 1) \
                        + str(hrows) + ':' + ssCol(col + 1) + str(hrows + 8759) + \
                        ',"' + sf_test[0] + '0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
            ss.cell(row=ss_row, column=st_lie+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lie+1).number_format = '#,##0'
            ss.cell(row=ss_row, column=st_lec+1).value = '=IF(AND(' + ssCol(st_emi+1) + str(ss_row) + '<>"",' + \
                                                         ssCol(st_emi+1) + str(ss_row) + '>0),' + \
                                                         ssCol(st_emc+1) + str(ss_row) + '*lifetime,"")'
            ss.cell(row=ss_row, column=st_lec+1).number_format = '$#,##0'
            # area
            if self.generators[gen].area > 0:
                ss.cell(row=ss_row, column=st_are+1).value = '=Detail!' + ssCol(col) + str(cap_row) +\
                                                             '*' + str(self.generators[gen].area)
                ss.cell(row=ss_row, column=st_are+1).number_format = '#,##0.00'
            # jobs
            if self.do_jobs:
                try:
                    genj = gen[gen.find('.') + 1:]
                except:
                    genj = gen
                try:
                    ss.cell(row=ss_row, column=st_job+1).value = f'=IF({ssCol(st_cac+1)}{ss_row}>0,{ssCol(st_cap+1)}{ss_row}*(' + \
                                                                 f'{self.jobfactors[genj].manufacture}*{self.jobfactors[genj].local_pct}+' + \
                                                                 f'{self.jobfactors[genj].install}),0)+' + \
                                                                 f'{self.jobfactors[genj].operate}*{ssCol(st_cap+1)}{ss_row}'
                    ss.cell(row=ss_row, column=st_job+1).number_format = '#,##0'
                except:
                    pass
        if is_storage:
            ns.cell(row=emi_row, column=col - 2).value = '=MIN(' + ssCol(col - 2) + str(hrows) + \
                    ':' + ssCol(col - 2) + str(hrows + 8759) + ')'
            ns.cell(row=emi_row, column=col - 2).number_format = '#,##0.00'
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            sum_value = 0
            do_sum = False
            do_cost = False
            for cell in column_cells:
                if cell.row >= hrows:
                    if do_sum:
                        try:
                            sum_value += abs(cell.value)
                        except:
                            pass
                    else:
                        try:
                            value = str(round(cell.value, 2))
                            if len(value) > length:
                                length = len(value) + 2
                        except:
                            pass
                elif cell.row > 0:
                    if str(cell.value)[0] != '=':
                        values = str(cell.value).split('\n')
                        for value in values:
                            if cell.row == cost_row:
                                valf = value.split('.')
                                alen = int(len(valf[0]) * 1.6)
                            else:
                                alen = len(value) + 2
                            if alen > length:
                                length = alen
                    else:
                        if cell.row == cost_row or cell.row == lcog_row:
                            length = max(length, 15)
                        if cell.row == cost_row:
                            do_cost = True
                        if cell.value[1:4] == 'SUM':
                            do_sum = True
            if sum_value > 0:
                alen = len(str(int(sum_value))) * 1.5
                if do_cost:
                    alen = int(alen * 1.5)
                if alen > length:
                    length = alen
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            ns.column_dimensions[cel].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        ns.column_dimensions['B'].width = 21
        st_row = hrows + 8760
        st_col = col
        for row in range(1, st_row):
            for col in range(1, st_col):
                try:
                    ns.cell(row=row, column=col).font = normal
                except:
                    pass
        self.progressbar.setValue(12)
        QtWidgets.QApplication.processEvents()
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        if self.results_prefix != '':
            ss.cell(row=1, column=1).value = 'Powermatch - ' + self.results_prefix + ' Summary'
        else:
            ss.cell(row=1, column=1).value = 'Powermatch - Summary'
        ss.cell(row=1, column=1).font = bold
        ss_lst_row = ss_row + 1
        ss_row, ss_re_row = detail_summary_total(ss_row, base_row='4')
        if len(nsul_sum_cols) > 1: # if we have underlying there'll be more than one column
            ss_row += 2
            ss.cell(row=ss_row, column=1).value = 'Additional Underlying Load'
            ss.cell(row=ss_row, column=1).font = bold
            base_row = str(ss_row + 1)
            for col in nsul_sum_cols[1:]:
                ss_row += 1
                ul_tml_sum, ul_re_sum = do_detail_summary(fac, col, ss_row, ul_tml_sum, ul_re_sum)
            ul_fst_row = int(base_row)
            ul_lst_row = ss_row
            ns_re_sum = ul_re_sum
            ns_tml_sum = ul_tml_sum
            ss_row, ss_re_row = detail_summary_total(ss_row, title='Underlying ', base_row=base_row,
                                          back_row=str(ss_lst_row))
        wider = [ssCol(st_lcg + 1), ssCol(st_lco + 1), ssCol(st_lcc + 1), ssCol(st_cac + 1), ssCol(st_lic + 1)]
        for column_cells in ss.columns:
            length = 0
            value = ''
            for cell in column_cells:
                if str(cell.value)[0] != '=':
                    values = str(cell.value).split('\n')
                    for value in values:
                        if len(value) + 1 > length:
                            length = len(value) + 1
            if isinstance(cell.column, int):
                cel = ssCol(cell.column)
            else:
                cel = cell.column
            if cel in wider:
                ss.column_dimensions[cel].width = max(length, 10) * 1.5
            else:
                ss.column_dimensions[cel].width = max(length, 10) * 1.2

        if corr_data is not None:
            ss_row += 2
            for corr in corr_data:
                ss.cell(row=ss_row, column=1).value = corr[0]
                if len(corr) > 1:
                    ss.cell(row=ss_row, column=2).value = corr[1]
                    ss.cell(row=ss_row, column=2).number_format = '#0.0000'
                    ss.cell(row=ss_row, column=3).value = corr[2]
                ss_row += 1
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Static Variables'
        ss.cell(row=ss_row, column=1).font = bold
        if self.carbon_price > 0:
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = 'Carbon Price ($/tCO2e)'
            ss.cell(row=ss_row, column=st_cap+1).value = self.carbon_price
            ss.cell(row=ss_row, column=st_cap+1).number_format = '$#,##0.00'
            attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
            carbon_cell = oxl.workbook.defined_name.DefinedName('carbon_price', attr_text=attr_text)
            try: # openpyxl 3.1.x
                wb.defined_names['carbon_price'] = carbon_cell
            except: # openpyxl 3.1.x
                wb.defined_names.append(carbon_cell)
        ss_row += 1
        attr_text = 'Summary!$' + ssCol(st_cap+1) + '$' + str(ss_row)
        lifetime_cell = oxl.workbook.defined_name.DefinedName('lifetime', attr_text=attr_text)
        try:
            wb.defined_names['lifetime'] = lifetime_cell
        except:
            wb.defined_names.append(lifetime_cell)
        ss.cell(row=ss_row, column=1).value = 'Lifetime (years)'
        ss.cell(row=ss_row, column=st_cap+1).value = max_lifetime
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Discount Rate'
        ss.cell(row=ss_row, column=st_cap+1).value = self.discount_rate
        ss.cell(row=ss_row, column=st_cap+1).number_format = '#,##0.00%'
        ss_row += 2
        ss_row = self.data_sources(ss, ss_row, pm_data_file, option)
        self.progressbar.setValue(14)
        QtWidgets.QApplication.processEvents()
        for row in range(1, ss_row + 1):
            for col in range(1, len(headers) + 1):
                try:
                    if ss.cell(row=row, column=col).font.name != 'Arial':
                        ss.cell(row=row, column=col).font = normal
                except:
                    pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        if self.save_tables:
            gens = []
            cons = []
            for fac in re_order:
                if fac == 'Load':
                    continue
                if pmss_details[fac].multiplier <= 0:
                    continue
                if fac.find('.') > 0:
                    gens.append(fac[fac.find('.') + 1:])
                else:
                    gens.append(fac)
                cons.append(self.generators[pmss_details[fac].generator].constraint)
            for gen in dispatch_order:
                gens.append(gen)
                cons.append(self.generators[gen].constraint)
            gs = wb.create_sheet(self.sheets[G].currentText())
            fields = []
            col = 1
            row = 1
            if hasattr(self.generators[list(self.generators.keys())[0]], 'name'):
                fields.append('name')
                gs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.generators[list(self.generators.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        txt = prop.replace('_', ' ').title()
                        txt = txt.replace('Cf', 'CF')
                        txt = txt.replace('Lcoe', 'LCOE')
                        txt = txt.replace('Om', 'OM')
                        gs.cell(row=row, column=col).value = txt
                        if prop == 'capex':
                            txt = txt + txt
                        gs.column_dimensions[ssCol(col)].width = max(len(txt) * 1.4, 10)
                        col += 1
            nme_width = 4
            con_width = 4
            for key, value in self.generators.items():
                if key in gens:
                    row += 1
                    col = 1
                    for field in fields:
                        gs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'constraint']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > con_width:
                                    con_width = len(txt)
                                    gs.column_dimensions[ssCol(col)].width = con_width * 1.4
                        elif field in ['capex', 'fixed_om']:
                            gs.cell(row=row, column=col).number_format = '$#,##0'
                        elif field in ['lcoe', 'variable_om', 'fuel']:
                            gs.cell(row=row, column=col).number_format = '$#,##0.00'
                        elif field in ['disc_rate']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00%'
                        elif field in ['capacity', 'lcoe_cf', 'initial']:
                            gs.cell(row=row, column=col).number_format = '#,##0.00'
                        elif field in ['emissions']:
                            gs.cell(row=row, column=col).number_format = '#,##0.000'
                        elif field in ['lifetime', 'order']:
                            gs.cell(row=row, column=col).number_format = '#,##0'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    gs.cell(row=row, column=col).font = normal
            gs.freeze_panes = 'B2'
            gs.activeCell = 'B2'
            fields = []
            col = 1
            row = 1
            cs = wb.create_sheet(self.sheets[C].currentText())
            if hasattr(self.constraints[list(self.constraints.keys())[0]], 'name'):
                fields.append('name')
                cs.cell(row=row, column=col).value = 'Name'
                col += 1
            for prop in dir(self.constraints[list(self.constraints.keys())[0]]):
                if prop[:2] != '__' and prop[-2:] != '__':
                    if prop != 'name':
                        fields.append(prop)
                        if prop == 'warm_time':
                            cs.cell(row=row, column=col).value = 'Warmup Time'
                        else:
                            cs.cell(row=row, column=col).value = prop.replace('_', ' ').title()
                        cs.column_dimensions[ssCol(col)].width = max(len(prop) * 1.4, 10)
                        col += 1
            nme_width = 4
            cat_width = 4
            for key, value in self.constraints.items():
                if key in cons:
                    row += 1
                    col = 1
                    for field in fields:
                        cs.cell(row=row, column=col).value = getattr(value, field)
                        if field in ['name', 'category']:
                            txt = getattr(value, field)
                            if field == 'name':
                                if len(txt) > nme_width:
                                    nme_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = nme_width * 1.4
                            else:
                                if len(txt) > cat_width:
                                    cat_width = len(txt)
                                    cs.column_dimensions[ssCol(col)].width = cat_width * 1.4
                        elif field == 'warm_time':
                            cs.cell(row=row, column=col).number_format = '#0.00'
                        elif field != 'category':
                            cs.cell(row=row, column=col).number_format = '#,##0%'
                        col += 1
            for row in range(1, row + 1):
                for col in range(1, len(fields) + 1):
                    try:
                        cs.cell(row=row, column=col).font = normal
                    except:
                        pass
            cs.freeze_panes = 'B2'
            cs.activeCell = 'B2'
        wb.save(rslts_file)
        self.progressbar.setValue(20)
        QtWidgets.QApplication.processEvents()
        j = rslts_file.rfind('/')
        rslts_file = rslts_file[j + 1:]
        msg = '%s created (%.2f seconds)' % (rslts_file, time.time() - start_time)
        msg = '%s created.' % rslts_file
        self.setStatus(msg)
        self.progressbar.setHidden(True)
        self.progressbar.setValue(0)

    def show_ProgressBar(self, maximum, msg, title):
        if self.opt_progressbar is None:
            self.opt_progressbar = ProgressBar(maximum=maximum, msg=msg, title=title)
            self.opt_progressbar.setWindowModality(QtCore.Qt.WindowModal)
            self.opt_progressbar.show()
            self.opt_progressbar.setVisible(False)
            self.activateWindow()
        else:
            self.opt_progressbar.barRange(0, maximum, msg=msg)

    def show_FloatStatus(self):
        if not self.log_status:
            return
        if self.floatstatus is None:
            self.floatstatus = FloatStatus(self, self.scenarios, None, program='Powermatch')
            self.floatstatus.setWindowModality(QtCore.Qt.WindowModal)
            self.floatstatus.setWindowFlags(self.floatstatus.windowFlags() |
                         QtCore.Qt.WindowSystemMenuHint |
                         QtCore.Qt.WindowMinMaxButtonsHint)
            self.floatstatus.procStart.connect(self.getStatus)
            self.floatstatus.show()
            self.activateWindow()

    def setStatus(self, text):
        if self.log.text() == text:
            return
        self.log.setText(text)
        if text == '':
            return
        if self.floatstatus and self.log_status:
            self.floatstatus.log(text)
            QtWidgets.QApplication.processEvents()

    @QtCore.pyqtSlot(str)
    def getStatus(self, text):
        if text == 'goodbye':
            self.floatstatus = None

    def exit(self):
        self.updated = False
        self.order.updated = False
        self.ignore.updated = False
        if self.floatstatus is not None:
            self.floatstatus.exit()
        self.close()

    def optClicked(self, in_year, in_option, in_pmss_details, in_pmss_data, in_re_order,
                   in_dispatch_order, pm_data_file, rslts_file):

        def create_starting_population(individuals, chromosome_length):
            # Set up an initial array of all zeros
            population = np.zeros((individuals, chromosome_length))
            # Loop through each row (individual)
            for i in range(individuals):
                # Choose a random number of ones to create but at least one 1
                ones = random.randint(1, chromosome_length)
                # Change the required number of zeros to ones
                population[i, 0:ones] = 1
                # Sfuffle row
                np.random.shuffle(population[i])
            return population

        def select_individual_by_tournament(population, *argv):
            # Get population size
            population_size = len(population)

            # Pick individuals for tournament
            fighter = [0, 0]
            fighter_fitness = [0, 0]
            fighter[0] = random.randint(0, population_size - 1)
            fighter[1] = random.randint(0, population_size - 1)

            # Get fitness score for each
            if len(argv) == 1:
                fighter_fitness[0] = argv[0][fighter[0]]
                fighter_fitness[1] = argv[0][fighter[1]]
            else:
                for arg in argv:
                    min1 = min(arg)
                    max1 = max(arg)
                    for f in range(len(fighter)):
                        try:
                            fighter_fitness[f] += (arg[f] - min1) / (max1 - min1)
                        except:
                            pass
            # Identify individual with lowest score
            # Fighter 1 will win if score are equal
            if fighter_fitness[0] <= fighter_fitness[1]:
                winner = fighter[0]
            else:
                winner = fighter[1]

            # Return the chromsome of the winner
            return population[winner, :]

        def breed_by_crossover(parent_1, parent_2, points=2):
            # Get length of chromosome
            chromosome_length = len(parent_1)

            # Pick crossover point, avoiding ends of chromsome
            if points == 1:
                crossover_point = random.randint(1,chromosome_length-1)
            # Create children. np.hstack joins two arrays
                child_1 = np.hstack((parent_1[0:crossover_point],
                                     parent_2[crossover_point:]))
                child_2 = np.hstack((parent_2[0:crossover_point],
                                     parent_1[crossover_point:]))
            else: # only do 2 at this    stage
                crossover_point_1 = random.randint(1, chromosome_length - 2)
                crossover_point_2 = random.randint(crossover_point_1 + 1, chromosome_length - 1)
                child_1 = np.hstack((parent_1[0:crossover_point_1],
                                     parent_2[crossover_point_1:crossover_point_2],
                                     parent_1[crossover_point_2:]))
                child_2 = np.hstack((parent_2[0:crossover_point_1],
                                     parent_1[crossover_point_1:crossover_point_2],
                                     parent_2[crossover_point_2:]))
            # Return children
            return child_1, child_2

        def randomly_mutate_population(population, mutation_probability):
            # Apply random mutation
            random_mutation_array = np.random.random(size=(population.shape))
            random_mutation_boolean = random_mutation_array <= mutation_probability
        #    random_mutation_boolean[0][:] = False # keep the best multi and lcoe
       #     random_mutation_boolean[1][:] = False
            population[random_mutation_boolean] = np.logical_not(population[random_mutation_boolean])
            # Return mutation population
            return population

        def calculate_fitness(population):
            lcoe_fitness_scores = [] # scores = LCOE values
            multi_fitness_scores = [] # scores = multi-variable weight
            multi_values = [] # values for each of the six variables
            if len(population) == 1:
                option = O1
            else:
                option = O
            if self.debug:
                self.popn += 1
                self.chrom = 0
            for chromosome in population:
                # now get random amount of generation per technology (both RE and non-RE)
                for fac, value in opt_order.items():
                    try:
                        capacity = value[2]
                        for c in range(value[0], value[1]):
                            if chromosome[c]:
                                capacity = capacity + capacities[c]
                        try:
                            pmss_details[fac].multiplier = capacity / pmss_details[fac].capacity
                        except:
                            print('PME1:', fac, capacity, pmss_details[fac].capacity)
                    except:
                        print('PME2:', fac, capacity, pmss_details[fac].capacity)
                multi_value, op_data, extra = self.doDispatch(year, option, pmss_details, pmss_data, re_order,
                                              dispatch_order, pm_data_file, rslts_file)
                if multi_value['load_pct'] < self.targets['load_pct'][3]:
                    if multi_value['load_pct'] == 0:
                        print('PME3:', multi_value['lcoe'], self.targets['load_pct'][3], multi_value['load_pct'])
                        lcoe_fitness_scores.append(1)
                    else:
                        try:
                            lcoe_fitness_scores.append(pow(multi_value['lcoe'],
                                self.targets['load_pct'][3] / multi_value['load_pct']))
                        except OverflowError as err:
                            self.setStatus(f"Overflow error: {err}; POW({multi_value['lcoe']:,}, " \
                                         + f"{self.targets['load_pct'][3] / multi_value['load_pct']:,}) " \
                                         + f"({self.targets['load_pct'][3]:,} / {multi_value['load_pct']:,} )")
                        except:
                            pass
                else:
                    lcoe_fitness_scores.append(multi_value['lcoe'])
                multi_values.append(multi_value)
                multi_fitness_scores.append(calc_weight(multi_value))
                if self.debug:
                    self.chrom += 1
                    line = str(self.popn) + ',' + str(self.chrom) + ','
                    for fac, value in opt_order.items():
                        try:
                            line += str(pmss_details[fac].capacity * pmss_details[fac].multiplier) + ','
                        except:
                            line += ','
                    for key in self.targets.keys():
                        try:
                            line += '{:.3f},'.format(multi_value[key])
                        except:
                            line += multi_value[key] + ','
                    line += '{:.5f},'.format(multi_fitness_scores[-1])
                    self.db_file.write(line + '\n')
            # alternative approach to calculating fitness
            multi_fitness_scores1 = []
            maxs = {}
            mins = {}
            tgts = {}
            for key in multi_value.keys():
                if key[-4:] == '_pct':
                    tgts[key] = abs(self.targets[key][2] - self.targets[key][3])
                else:
                    maxs[key] = 0
                    mins[key] = -1
                    for popn in multi_values:
                        try:
                            tgt = abs(self.targets[key][2] - popn[key])
                        except:
                            continue
                        if tgt > maxs[key]:
                            maxs[key] = tgt
                        if mins[key] < 0 or tgt < mins[key]:
                            mins[key] = tgt
            for popn in multi_values:
                weight = 0
                for key, value in multi_value.items():
                    if self.targets[key][1] <= 0:
                        continue
                    try:
                        tgt = abs(self.targets[key][2] - popn[key])
                    except:
                        continue
                    if key[-4:] == '_pct':
                        if tgts[key] != 0:
                            if tgt > tgts[key]:
                                weight += 1 * self.targets[key][1]
                            else:
                                try:
                                    weight += 1 - ((tgt / tgts[key]) * self.targets[key][1])
                                except:
                                    pass
                    else:
                        try:
                            weight += 1 - (((maxs[key] - tgt) / (maxs[key] - mins[key])) \
                                      * self.targets[key][1])
                        except:
                            pass
                multi_fitness_scores1.append(weight)
            if len(population) == 1: # return the table for best chromosome
                return op_data, multi_values
            else:
                return lcoe_fitness_scores, multi_fitness_scores, multi_values

        def optQuitClicked(event):
            self.optExit = True
            optDialog.close()

        def chooseClicked(event):
            self.opt_choice = self.sender().text()
            chooseDialog.close()

        def calc_weight(multi_value, calc=0):
            weight = [0., 0.]
            if calc == 0:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    elif value[2] > value[3]: # wants higher target
                        if multi_value[key] > value[2]: # high no weight
                            w = 0.
                        elif multi_value[key] < value[3]: # low maximum weight
                            w = 2.
                        else:
                            w = 1 - (multi_value[key] - value[3]) / (value[2] - value[3])
                    else: # lower target
                        if multi_value[key] == -1 or multi_value[key] > value[3]: # high maximum weight
                            w = 2.
                        elif multi_value[key] < value[2]: # low no weight
                            w = 0.
                        else:
                            w = multi_value[key] / (value[3] - value[2])
                    weight[0] += w * value[1]
            elif calc == 1:
                for key, value in self.targets.items():
                    if multi_value[key] == '':
                        continue
                    if value[1] <= 0:
                        continue
                    if multi_value[key] < 0:
                        w = 1
                    elif value[2] == value[3]: # wants specific target
                        if multi_value[key] == value[2]:
                            w = 0
                        else:
                            w = 1
                    else: # target range
                        w = min(abs(value[2] - multi_value[key]) / abs(value[2] - value[3]), 1)
                    weight[1] += w * value[1]
            return weight[calc]

        def plot_multi(multi_scores, multi_best, multi_order, title):
            data = [[], [], []]
            max_amt = [0., 0.]
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
            pwr_chr = ['', '']
            divisor = [1., 1.]
            pwr_chrs = ' KMBTPEZY'
            for m in range(2):
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if max_amt[m] > pow(10, pwr * 3):
                        pwr_chr[m] = pwr_chrs[pwr]
                        divisor[m] = 1. * pow(10, pwr * 3)
                        break
            self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
            self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
            for multi in multi_best:
                for axis in range(3): # only three axes in plot
                    if multi_order[axis] == 'cost':
                        data[axis].append(multi[multi_order[axis]] / divisor[0]) # cost
                    elif multi_order[axis] == 'co2':
                        data[axis].append(multi[multi_order[axis]] / divisor[1]) # co2
                    elif multi_order[axis][-4:] == '_pct': # percentage
                        data[axis].append(multi[multi_order[axis]] * 100.)
                    else:
                        data[axis].append(multi[multi_order[axis]])
            # create colour map
            colours = multi_scores[:]
            cmax = max(colours)
            cmin = min(colours)
            if cmin == cmax:
                return
            for c in range(len(colours)):
                try:
                    colours[c] = (colours[c] - cmin) / (cmax - cmin)
                except:
                    pass
            scolours = sorted(colours)
            cvals  = [-1., 0, 1]
            colors = ['green' ,'orange', 'red']
            norm = plt.Normalize(vmin=min(cvals), vmax=max(cvals))
            tuples = list(zip(map(norm,cvals), colors))
            cmap = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples)
            fig = plt.figure(title + QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(), '_yyyy-MM-dd_hhmm'))
            mx = plt.axes(projection='3d')
            plt.title('\n' + title.title() + '\n')
            try:
                for i in range(len(data[0])):
                    mx.scatter3D(data[2][i], data[1][i], data[0][i], picker=True, color=cmap(colours[i]), cmap=cmap)
                    if title[:5] == 'start':
                        mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(i+1))
                    else:
                        j = scolours.index(colours[i])
                        if j < 10:
                            mx.text(data[2][i], data[1][i], data[0][i], '%s' % str(j+1))
            except:
                return
            if self.optimise_multisurf:
                cvals_r  = [-1., 0, 1]
                colors_r = ['red' ,'orange', 'green']
                norm_r = plt.Normalize(min(cvals_r), max(cvals_r))
                tuples_r = list(zip(map(norm_r, cvals_r), colors_r))
                cmap_r = matplotlib.colors.LinearSegmentedColormap.from_list('', tuples_r)
                # https://www.fabrizioguerrieri.com/blog/surface-graphs-with-irregular-dataset/
                triang = mtri.Triangulation(data[2], data[1])
                mx.plot_trisurf(triang, data[0], cmap=cmap_r)
            mx.xaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[2]][5]))
            mx.yaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[1]][5]))
            mx.zaxis.set_major_formatter(FormatStrFormatter(self.targets[multi_order[0]][5]))
            mx.set_xlabel(self.targets[multi_order[2]][6])
            mx.set_ylabel(self.targets[multi_order[1]][6])
            mx.set_zlabel(self.targets[multi_order[0]][6])
            zp = ZoomPanX()
            f = zp.zoom_pan(mx, base_scale=1.2, annotate=True)
            plt.show(block=True)
            if zp.datapoint is not None: # user picked a point
                if zp.datapoint[0][0] < 0: # handle problem in matplotlib sometime after version 3.0.3
                    best = [0, 999]
                    for p in range(len(multi_best)):
                        diff = 0
                        for v in range(3):
                            key = multi_order[v]
                            valu = multi_best[p][key]
                            if key[-4:] == '_pct':
                                valu = valu * 100.
                            diff += abs((valu - zp.datapoint[0][v + 1]) / valu)
                        if diff < best[1]:
                            best = [p, diff]
                    zp.datapoint = [[best[0]]]
                    for v in range(3):
                        key = multi_order[v]
                        zp.datapoint[0].append(multi_best[p][key])
                if self.more_details:
                    for p in zp.datapoint:
                        msg = 'iteration ' + str(p[0]) + ': '
                        mult = []
                        for i in range(3):
                            if self.targets[multi_order[i]][6].find('%') > -1:
                                mult.append(100.)
                            else:
                                mult.append(1.)
                            msg += self.targets[multi_order[i]][6].replace('%', '%%') + ': ' + \
                                   self.targets[multi_order[i]][5] + '; '
                        msg = msg % (p[1] * mult[0], p[2] * mult[1], p[3] * mult[2])
                        self.setStatus(msg)
            return zp.datapoint

        def show_multitable(best_score_progress, multi_best, multi_order, title):
            def pwr_chr(amt):
                pwr_chrs = ' KMBTPEZY'
                pchr = ''
                divisor = 1.
                for pwr in range(len(pwr_chrs) - 1, -1, -1):
                    if amt > pow(10, pwr * 3):
                        pchr = pwr_chrs[pwr]
                        divisor = 1. * pow(10, pwr * 3)
                        break
                return amt / divisor, pchr

            def opt_fmat(amt, fmat):
                tail = ' '
                p = fmat.find('pwr_chr')
                if p > 0:
                    amt, tail = pwr_chr(amt)
                    fmat = fmat.replace('pwr_chr', '')
              #  d = fmat.find('d')
             #   if d > 0:
              #      amt = amt * 100.
                fmat = fmat.replace('.1f', '.2f')
                i = fmat.find('%')
                fmt = fmat[:i] + '{:> 8,' + fmat[i:].replace('%', '').replace('d', '.2%') + '}' + tail
                return fmt.format(amt)

            best_table = []
            best_fmate = []
            for b in range(len(multi_best)):
                bl = list(multi_best[b].values())
                bl.insert(0, best_score_progress[b])
                bl.insert(0, b + 1)
                best_table.append(bl)
                best_fmate.append([b + 1])
                best_fmate[-1].insert(1, best_score_progress[b])
                for f in range(2, len(best_table[-1])):
                    if target_keys[f - 2] in ['load_pct', 'surplus_pct', 're_pct']:
                        best_fmate[-1].append(opt_fmat(bl[f], '%d%%'))
                    else:
                        best_fmate[-1].append(opt_fmat(bl[f], target_fmats[f - 2]))
            fields = target_names[:]
            fields.insert(0, 'weight')
            fields.insert(0, 'iteration')
            dialog = displaytable.Table(best_fmate, fields=fields, txt_align='R', decpts=[0, 4],
                     title=title, sortby='weight')
            dialog.exec_()
            b = int(dialog.getItem(0)) - 1
            del dialog
            pick = [b]
            for fld in multi_order[:3]:
                i = target_keys.index(fld)
                pick.append(best_table[b][i + 2])
            return [pick]

#       optClicked mainline starts here
        year = in_year
        option = in_option
        pmss_details = dict(in_pmss_details)
        pmss_data = in_pmss_data[:]
        re_order = in_re_order[:]
        dispatch_order = in_dispatch_order[:]
        if self.optimise_debug:
            self.debug = True
        else:
            self.debug = False
        missing = []
        for fac in re_order:
            if fac == 'Load':
                continue
            if fac not in self.optimisation.keys():
                missing.append(fac)
        for gen in dispatch_order:
            if gen not in self.optimisation.keys():
                missing.append(gen)
        if len(missing) > 0:
            bad = False
            if self.optimise_default is not None:
                defaults = self.optimise_default.split(',')
                if len(defaults) < 1 or len(defaults) > 3:
                    bad = True
                else:
                    try:
                        for miss in missing:
                            if len(defaults) == 2:
                                minn = 0
                            else:
                                if defaults[0][-1] == 'd':
                                    minn = pmss_details[miss].capacity * float(defaults[0][:-1])
                                elif defaults[0][-1] == 'c':
                                    minn = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[0][:-1])
                                else:
                                    minn = float(defaults[0])
                            if defaults[-2][-1] == 'd':
                                maxx = pmss_details[miss].capacity * float(defaults[-2][:-1])
                            elif defaults[-2][-1] == 'c':
                                maxx = pmss_details[miss].capacity * pmss_details[miss].multiplier * float(defaults[-2][:-1])
                            else:
                                maxx = float(defaults[-2])
                            if len(defaults) == 3:
                                if defaults[-1][-1].lower() == 'd':
                                    step = capacity * float(defaults[-1][:-1])
                                elif defaults[-1][-1].lower() == 'c':
                                    step = capacity * multiplier * float(defaults[-1][:-1])
                                else:
                                    step = float(defaults[-1])
                            else:
                                step = (maxx - minn) / float(defaults[-1])
                            self.optimisation[miss] =  Optimisation(miss, 'None', None)
                            self.optimisation[miss].approach = 'Range'
                            self.optimisation[miss].capacity_min = minn
                            self.optimisation[miss].capacity_max = maxx
                            self.optimisation[miss].capacity_step = step
                    except:
                        bad = True
                check = ''
                for miss in missing:
                    check += miss + ', '
                check = check[:-2]
                if bad:
                    self.setStatus('Key Error: Missing Optimisation entries for: ' + check)
                    return
                self.setStatus('Missing Optimisation entries added for: ' + check)
        self.optExit = False
        self.setStatus('Optimise processing started')
        err_msg = ''
        optDialog = QtWidgets.QDialog()
        grid = QtWidgets.QGridLayout()
        rw = 0
        self._optLoadv = [pmss_details['Load'].multiplier, pmss_details['Load'].capacity]
        grid.addWidget(QtWidgets.QLabel('Adjust load'), rw, 0)
        self.optLoadMult = QtWidgets.QDoubleSpinBox()
        self.optLoadMult.setRange(-1, self.adjust_cap)
        self.optLoadMult.setDecimals(4)
        self.optLoadMult.setSingleStep(.1)
        grid.addWidget(self.optLoadMult, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Multiplier gives Load (MWh)'), rw, 2, 1, 3)
        self.optLoadMult.setValue(self._optLoadv[0])
        self.optLoad = QtWidgets.QDoubleSpinBox()
        self.optLoad.setRange(1, self._optLoadv[1] * self.adjust_cap)
        self.optLoad.setDecimals(0)
        self.optLoad.setSingleStep(1000)
        self.optLoad.setValue(self._optLoadv[0] * self._optLoadv[1])
        self.optLoadMult.valueChanged.connect(self.optLoadChange)
        self.optLoad.valueChanged.connect(self.optLoadChange)
        grid.addWidget(self.optLoad, rw, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Population size'), rw, 0)
        optPopn = QtWidgets.QSpinBox()
        optPopn.setRange(10, 500)
        optPopn.setSingleStep(10)
        optPopn.setValue(self.optimise_population)
        optPopn.valueChanged.connect(self.changes)
        grid.addWidget(optPopn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Size of population'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('No. of iterations'), rw, 0, 1, 3)
        optGenn = QtWidgets.QSpinBox()
        optGenn.setRange(10, 500)
        optGenn.setSingleStep(10)
        optGenn.setValue(self.optimise_generations)
        optGenn.valueChanged.connect(self.changes)
        grid.addWidget(optGenn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Number of iterations (generations)'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Mutation probability'), rw, 0)
        optMutn = QtWidgets.QDoubleSpinBox()
        optMutn.setRange(0, 1)
        optMutn.setDecimals(4)
        optMutn.setSingleStep(0.001)
        optMutn.setValue(self.optimise_mutation)
        optMutn.valueChanged.connect(self.changes)
        grid.addWidget(optMutn, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Add in mutation'), rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Exit if stable'), rw, 0)
        optStop = QtWidgets.QSpinBox()
        optStop.setRange(0, 50)
        optStop.setSingleStep(10)
        optStop.setValue(self.optimise_stop)
        optStop.valueChanged.connect(self.changes)
        grid.addWidget(optStop, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Exit if LCOE/weight remains the same after this many iterations'),
                       rw, 2, 1, 3)
        rw += 1
        grid.addWidget(QtWidgets.QLabel('Optimisation choice'), rw, 0)
        optCombo = QtWidgets.QComboBox()
        choices = ['LCOE', 'Multi', 'Both']
        for choice in choices:
            optCombo.addItem(choice)
            if choice == self.optimise_choice:
                optCombo.setCurrentIndex(optCombo.count() - 1)
        grid.addWidget(optCombo, rw, 1)
        grid.addWidget(QtWidgets.QLabel('Choose type of optimisation'),
                       rw, 2, 1, 3)
        rw += 1
        # for each variable name
        grid.addWidget(QtWidgets.QLabel('Variable'), rw, 0)
        grid.addWidget(QtWidgets.QLabel('Weight'), rw, 1)
        grid.addWidget(QtWidgets.QLabel('Better'), rw, 2)
        grid.addWidget(QtWidgets.QLabel('Worse'), rw, 3)
        rw += 1
        ndx = grid.count()
        for key in self.targets.keys():
            self.targets[key][4] = ndx
            ndx += 4
        for key, value in self.targets.items():
            if value[2] == value[3]:
                ud = '(=)'
            elif value[2] < 0:
                ud = '(<html>&uarr;</html>)'
            elif value[3] < 0 or value[3] > value[2]:
                ud = '(<html>&darr;</html>)'
            else:
                ud = '(<html>&uarr;</html>)'
            grid.addWidget(QtWidgets.QLabel(value[0] + ': ' + ud), rw, 0)
            weight = QtWidgets.QDoubleSpinBox()
            weight.setRange(0, 1)
            weight.setDecimals(2)
            weight.setSingleStep(0.05)
            weight.setValue(value[1])
            grid.addWidget(weight, rw, 1)
            if key[-4:] == '_pct':
                minim = QtWidgets.QDoubleSpinBox()
                minim.setRange(-.1, 1.)
                minim.setDecimals(2)
                minim.setSingleStep(0.1)
                minim.setValue(value[2])
                grid.addWidget(minim, rw, 2)
                maxim = QtWidgets.QDoubleSpinBox()
                maxim.setRange(-.1, 1.)
                maxim.setDecimals(2)
                maxim.setSingleStep(0.1)
                maxim.setValue(value[3])
                grid.addWidget(maxim, rw, 3)
            else:
                minim = QtWidgets.QLineEdit()
                minim.setValidator(QtGui.QDoubleValidator())
                minim.validator().setDecimals(2)
                minim.setText(str(value[2]))
                grid.addWidget(minim, rw, 2)
                maxim = QtWidgets.QLineEdit()
                maxim.setValidator(QtGui.QDoubleValidator())
                maxim.validator().setDecimals(2)
                maxim.setText(str(value[3]))
                grid.addWidget(maxim, rw, 3)
            rw += 1
        quit = QtWidgets.QPushButton('Quit', self)
        grid.addWidget(quit, rw, 0)
        quit.clicked.connect(optQuitClicked)
        show = QtWidgets.QPushButton('Proceed', self)
        grid.addWidget(show, rw, 1)
        show.clicked.connect(optDialog.close)
        optDialog.setLayout(grid)
        optDialog.setWindowTitle('Choose Optimisation Parameters')
        optDialog.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        optDialog.exec_()
        if self.optExit: # a fudge to exit
            self.setStatus('Execution aborted.')
            return
        # check we have optimisation entries for generators and storage
        # update any changes to targets
        self.optimise_choice = optCombo.currentText()
        for key in self.targets.keys():
            weight = grid.itemAt(self.targets[key][4] + 1).widget()
            self.targets[key][1] = weight.value()
            minim = grid.itemAt(self.targets[key][4] + 2).widget()
            try:
                self.targets[key][2] = minim.value()
            except:
                self.targets[key][2] = float(minim.text())
            maxim = grid.itemAt(self.targets[key][4] + 3).widget()
            try:
                self.targets[key][3] = maxim.value()
            except:
                self.targets[key][3] = float(maxim.text())
        # might want to save load value if changed
        pmss_details['Load'].multiplier = self.optLoadMult.value()
        updates = {}
        lines = []
        lines.append('optimise_choice=' + self.optimise_choice)
        lines.append('optimise_generations=' + str(self.optimise_generations))
        lines.append('optimise_mutation=' + str(self.optimise_mutation))
        lines.append('optimise_population=' + str(self.optimise_population))
        lines.append('optimise_stop=' + str(self.optimise_stop))
        if self.optimise_choice == 'LCOE':
            do_lcoe = True
            do_multi = False
        elif self.optimise_choice == 'Multi':
            do_lcoe = False
            do_multi = True
        else:
            do_lcoe = True
            do_multi = True
        multi_order = []
        for key, value in self.targets.items():
            line = 'optimise_{}={:.2f},{:.2f},{:.2f}'.format(key, value[1], value[2], value[3])
            lines.append(line)
            multi_order.append('{:.2f}{}'.format(value[1], key))
        updates['Powermatch'] = lines
        SaveIni(updates)
        multi_order.sort(reverse=True)
    #    multi_order = multi_order[:3] # get top three weighted variables - but I want them all
        multi_order = [o[4:] for o in multi_order]
        self.adjust_gen = True
        orig_load = []
        load_col = -1
        orig_tech = {}
        orig_capacity = {}
        opt_order = {} # rely on it being processed in added order
        # each entry = [first entry in chrom, last entry, minimum capacity]
        # first get original renewables generation from data sheet
        for fac in re_order:
            if fac == 'Load':
                continue
            opt_order[fac] = [0, 0, 0]
        # now add scheduled generation
        for gen in dispatch_order:
            opt_order[gen] = [0, 0, 0]
        capacities = []
        for gen in opt_order.keys():
            opt_order[gen][0] = len(capacities) # first entry
            try:
                if self.optimisation[gen].approach == 'Discrete':
                    capacities.extend(self.optimisation[gen].capacities)
                    opt_order[gen][1] = len(capacities) # last entry
                elif self.optimisation[gen].approach == 'Range':
                    if self.optimisation[gen].capacity_max == self.optimisation[gen].capacity_min:
                        if self.optimisation[gen].capacity_max is None or self.optimisation[gen].capacity_max == 0:
                            opt_order[gen][1] = opt_order[gen][0]
                            continue
                        capacities.extend([0])
                        opt_order[gen][1] = len(capacities)
                        opt_order[gen][2] = self.optimisation[gen].capacity_min
                        continue
                    ctr = int((self.optimisation[gen].capacity_max - self.optimisation[gen].capacity_min) / \
                              self.optimisation[gen].capacity_step)
                    if ctr < 1:
                        self.setStatus("Error with Optimisation table entry for '" + gen + "'")
                        return
                    capacities.extend([self.optimisation[gen].capacity_step] * ctr)
                    tot = self.optimisation[gen].capacity_step * ctr + self.optimisation[gen].capacity_min
                    if tot < self.optimisation[gen].capacity_max:
                        capacities.append(self.optimisation[gen].capacity_max - tot)
                    opt_order[gen][1] = len(capacities)
                    opt_order[gen][2] = self.optimisation[gen].capacity_min
                else:
                    opt_order[gen][1] = len(capacities)
            except KeyError as err:
                self.setStatus('Key Error: No Optimisation entry for ' + str(err))
            except ZeroDivisionError as err:
                self.setStatus('Zero capacity: ' + gen + ' ignored')
            except:
                err = str(sys.exc_info()[0]) + ',' + str(sys.exc_info()[1]) + ',' + gen + ',' \
                      + str(opt_order[gen])
                self.setStatus('Error: ' + str(err))
                return
        # chromosome = [1] * int(len(capacities) / 2) + [0] * (len(capacities) - int(len(capacities) / 2))
        # we have the original data - from here down we can do our multiple optimisations
        # Set general parameters
        self.setStatus('Optimisation choice is ' + self.optimise_choice)
        chromosome_length = len(capacities)
        self.setStatus(f'Chromosome length: {chromosome_length}; {pow(2, chromosome_length):,} permutations')
        self.optimise_population = optPopn.value()
        population_size = self.optimise_population
        self.optimise_generations = optGenn.value()
        maximum_generation = self.optimise_generations
        self.optimise_mutation = optMutn.value()
        self.optimise_stop = optStop.value()
        lcoe_scores = []
        multi_scores = []
        multi_values = []
     #   if do_lcoe:
      #      lcoe_target = 0. # aim for this LCOE
        if do_multi:
            multi_best = [] # list of six variables for best weight
            multi_best_popn = [] # list of chromosomes for best weight
        self.show_ProgressBar(maximum=optGenn.value(), msg='Process iterations', title='SIREN - Powermatch Progress')
        self.opt_progressbar.setVisible(True)
        start_time = time.time()
        # Create starting population
        self.opt_progressbar.barProgress(1, 'Processing iteration 1')
        QtCore.QCoreApplication.processEvents()
        population = create_starting_population(population_size, chromosome_length)
        # calculate best score(s) in starting population
        # if do_lcoe best_score = lowest non-zero lcoe
        # if do_multi best_multi = lowest weight and if not do_lcoe best_score also = best_weight
        if self.debug:
            filename = self.scenarios + 'opt_debug_' + \
                       QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                       'yyyy-MM-dd_hhmm') + '.csv'
            self.db_file = open(filename, 'w')
            line0 = 'Popn,Chrom,'
            line1 = 'Weights,,'
            line2 = 'Targets,,'
            line3 = 'Range,' + str(population_size) + ','
            for gen, value in opt_order.items():
                 line0 += gen + ','
                 line1 += ','
                 line2 += ','
                 line3 += ','
            for key in self.targets.keys():
                 line0 += key + ','
                 line1 += str(self.targets[key][1]) + ','
                 line2 += str(self.targets[key][2]) + ','
                 if key[-4:] == '_pct':
                     line3 += str(abs(self.targets[key][2] - self.targets[key][3])) + ','
                 else:
                     line3 += ','
            line0 += 'Score'
            self.db_file.write(line0 + '\n' + line1 + '\n' + line2 + '\n' + line3 + '\n')
            self.popn = 0
            self.chrom = 0
        lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
        if do_lcoe:
            try:
                best_score = np.min(lcoe_scores)
            except:
                print('PME4:', lcoe_scores)
            best_ndx = lcoe_scores.index(best_score)
            lowest_chrom = population[best_ndx]
            self.setStatus('Starting LCOE: $%.2f' % best_score)
        if do_multi:
            if self.more_details: # display starting population ?
                pick = plot_multi(multi_scores, multi_values, multi_order, 'starting population')
            # want maximum from first round to set base upper limit
            for key in self.targets.keys():
                if self.targets[key][2] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][2] = setit
                if self.targets[key][3] < 0: # want a maximum from first round
                    setit = 0
                    for multi in multi_values:
                        setit = max(multi[key], setit)
                    self.targets[key][3] = setit
            # now we can find the best weighted result - lowest is best
            best_multi = np.min(multi_scores)
            best_mndx = multi_scores.index(best_multi)
            multi_lowest_chrom = population[best_mndx]
            multi_best_popn.append(multi_lowest_chrom)
            multi_best.append(multi_values[best_mndx])
            self.setStatus('Starting Weight: %.4f' % best_multi)
            multi_best_weight = best_multi
            best_multi_progress = [best_multi]
            if not do_lcoe:
                best_score = best_multi
            last_multi_score = best_multi
            lowest_multi_score = best_multi
            mud = '='
        # Add starting best score to progress tracker
        best_score_progress = [best_score]
        best_ctr = 1
        last_score = best_score
        lowest_score = best_score
        lud = '='
        # Now we'll go through the generations of genetic algorithm
        if do_lcoe:
            self.setStatus('Final LCOE: $%.2f' % best_score)
            fig = 'optimise_lcoe'
            titl = 'Optimise LCOE using Genetic Algorithm'
            ylbl = 'Best LCOE ($/MWh)'
        else:
            fig = 'optimise_multi'
            titl = 'Optimise Multi using Genetic Algorithm'
            ylbl = 'Best Weight'
        matplotlib.rcParams['savefig.directory'] = self.scenarios
        plt.ion()
        fig2 = plt.figure(fig + QtCore.QDateTime.toString(QtCore.QDateTime.currentDateTime(),
                   '_yyyy-MM-dd_hhmm'))
        lx = plt.subplot(111)
        x = list(range(1, len(best_score_progress) + 1))
        iteration = len(x)
        plt.title(titl)
        for generation in range(1, maximum_generation):
            lcoe_status = ''
            multi_status = ''
            if do_lcoe:
                lcoe_status = ' %s $%.2f ;' % (lud, best_score)
            if do_multi:
                multi_status = ' %s %.4f ;' % (mud, best_multi)
            tim = (time.time() - start_time)
            if tim < 60:
                tim = ' (%s%s %.1f secs)' % (lcoe_status, multi_status, tim)
            else:
                tim = ' (%s%s %.2f mins)' % (lcoe_status, multi_status, tim / 60.)
            self.opt_progressbar.barProgress(generation + 1,
                f'Processing iteration {generation + 1} of {maximum_generation} {tim}')
            QtWidgets.QApplication.processEvents()
            if not self.opt_progressbar.be_open:
                break
        # Create an empty list for new population
            new_population = []
        # Using elitism approach include best individual
            if do_lcoe:
                new_population.append(lowest_chrom)
            if do_multi:
                new_population.append(multi_lowest_chrom)
            # Create new population generating two children at a time
            if do_lcoe:
                if do_multi:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores,
                                                                   multi_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
                else:
                    for i in range(int(population_size/2)):
                        parent_1 = select_individual_by_tournament(population, lcoe_scores)
                        parent_2 = select_individual_by_tournament(population, lcoe_scores)
                        child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                        new_population.append(child_1)
                        new_population.append(child_2)
            else:
                for i in range(int(population_size/2)):
                    parent_1 = select_individual_by_tournament(population, multi_scores)
                    parent_2 = select_individual_by_tournament(population, multi_scores)
                    child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                    new_population.append(child_1)
                    new_population.append(child_2)
            # get back to original size (after elitism adds)
            if do_lcoe:
                new_population.pop()
            if do_multi:
                new_population.pop()
            # Replace the old population with the new one
            population = np.array(new_population)
            if self.optimise_mutation > 0:
                population = randomly_mutate_population(population, self.optimise_mutation)
            # Score best solution, and add to tracker
            lcoe_scores, multi_scores, multi_values = calculate_fitness(population)
            if do_lcoe:
                best_lcoe = np.min(lcoe_scores)
                best_ndx = lcoe_scores.index(best_lcoe)
                best_score = best_lcoe
            # now we can find the best weighted result - lowest is best
            if do_multi:
                best_multi = np.min(multi_scores)
                best_mndx = multi_scores.index(best_multi)
                multi_lowest_chrom = population[best_mndx]
                multi_best_popn.append(multi_lowest_chrom)
                multi_best.append(multi_values[best_mndx])
           #     if multi_best_weight > best_multi:
                multi_best_weight = best_multi
                if not do_lcoe:
                    best_score = best_multi
                best_multi_progress.append(best_multi)
            best_score_progress.append(best_score)
            if best_score < lowest_score:
                lowest_score = best_score
                if do_lcoe:
                    lowest_chrom = population[best_ndx]
                else: #(do_multi only)
                    multi_lowest_chrom = population[best_mndx]
            if self.optimise_stop > 0:
                if best_score == last_score:
                    best_ctr += 1
                    if best_ctr >= self.optimise_stop:
                        break
                else:
                    last_score = best_score
                    best_ctr = 1
            last_score = best_score
            if do_lcoe:
                if best_score == best_score_progress[-2]:
                    lud = '='
                elif best_score < best_score_progress[-2]:
                    lud = '<html>&darr;</html>'
                else:
                    lud = '<html>&uarr;</html>'
            if do_multi:
                if best_multi == last_multi_score:
                    mud = '='
                elif best_multi < last_multi_score:
                    mud = '<html>&darr;</html>'
                else:
                    mud = '<html>&uarr;</html>'
                last_multi_score = best_multi
            # Plot progress
            iteration += 1
            x.append(iteration)
            lx.plot(x, best_score_progress)
            xlabel = f'Optimise Cycle ({iteration} iterations)'
            lx.set_xlabel(xlabel)
            lx.set_ylabel(ylbl)
            plt.draw()
        if self.debug:
            try:
                self.db_file.close()
                optimiseDebug(self.db_file.name)
                os.remove(self.db_file.name)
            except:
                pass
            self.debug = False
        self.opt_progressbar.setVisible(False)
        self.opt_progressbar.close()
        tim = (time.time() - start_time)
        if tim < 60:
            tim = '%.1f secs)' % tim
        else:
            tim = '%.2f mins)' % (tim / 60.)
        msg = 'Optimise completed (%0d iterations; %s' % (generation + 1, tim)
        if best_score > lowest_score:
            msg += ' Try more iterations.'
        # we'll keep two or three to save re-calculating_fitness
        op_data = [[], [], [], [], []]
        score_data = [None, None, None, None, None]
        if do_lcoe:
            op_data[0], score_data[0] = calculate_fitness([lowest_chrom])
        if do_multi:
            op_data[1], score_data[1] = calculate_fitness([multi_lowest_chrom])
        self.setStatus(msg)
        QtWidgets.QApplication.processEvents()
        self.progressbar.setHidden(True)
        self.progressbar.setValue(0)
        # GA has completed required generation
        if do_multi:
            self.setStatus('Final Weight: %.4f' % multi_best_weight)
        pick = None
        pickf = None
        if do_multi:
            if self.optimise_multiplot:
                pick = plot_multi(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                if self.more_details:
                    pickf = plot_multi(multi_scores, multi_values, multi_order, 'final iteration')
            if self.optimise_multitable:
                pick2 = show_multitable(best_multi_progress, multi_best, multi_order, 'best of each iteration')
                try:
                    pick = pick + pick2
                except:
                    pick = pick2
                if self.more_details:
                    pick2 = show_multitable(multi_scores, multi_values, multi_order, 'final iteration')
                    try:
                        pickf = pickf + pick2
                    except:
                        pickf = pick2
        if self.do_jobs:
            jm = 0
        else:
            jm = 1
        op_pts = [0] * len(headers)
        for p in [st_lcg, st_lco, st_lcc, st_max, st_bal, st_rlc - jm, st_are]:
            op_pts[p] = 2
        op_pts[st_cap] = 3
        if self.more_details:
            if do_lcoe:
                list(map(list, list(zip(*op_data[0]))))
                dialog = displaytable.Table(op_data[0], title=self.sender().text(), fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
            if do_multi:
                list(map(list, list(zip(*op_data[1]))))
                dialog = displaytable.Table(op_data[1], title='Multi_' + self.sender().text(), fields=headers,
                         save_folder=self.scenarios, sortby='', decpts=op_pts)
                dialog.exec_()
                del dialog
        # now I'll display the resulting capacities for LCOE, lowest weight, picked
        # now get random amount of generation per technology (both RE and non-RE)
        its = {}
        for fac, value in opt_order.items():
            its[fac] = []
        chrom_hdrs = []
        chroms = []
        ndxes = []
        if do_lcoe:
            chrom_hdrs = ['Lowest LCOE']
            chroms = [lowest_chrom]
            ndxes = [0]
        if do_multi:
            chrom_hdrs.append('Lowest Weight')
            chroms.append(multi_lowest_chrom)
            ndxes.append(1)
        if pickf is not None:
            for p in range(len(pickf)):
                if pick is None:
                    pick = [pickf[f]]
                else:
                    pick.append(pickf[p][:])
                pick[-1][0] = len(multi_best_popn)
                multi_best_popn.append(population[pickf[p][0]])
        if pick is not None:
            # at present I'll calculate the best weight for the chosen picks. Could actually present all for user choice
            if len(pick) <= 3:
                multi_lowest_chrom = multi_best_popn[pick[0][0]]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + self.sender().text(), fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
                if len(pick) >= 2:
                    multi_lowest_chrom = multi_best_popn[pick[1][0]]
                    op_data[3], score_data[3] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[3]))))
                        dialog = displaytable.Table(op_data[3], title='Pick_' + self.sender().text(), fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 2nd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(3)
                if len(pick) == 3:
                    multi_lowest_chrom = multi_best_popn[pick[2][0]]
                    op_data[4], score_data[4] = calculate_fitness([multi_lowest_chrom])
                    if self.more_details:
                        list(map(list, list(zip(*op_data[4]))))
                        dialog = displaytable.Table(op_data[4], title='Pick_' + self.sender().text(), fields=headers,
                                 save_folder=self.scenarios, sortby='', decpts=op_pts)
                        dialog.exec_()
                        del dialog
                    chrom_hdrs.append('Your 3rd pick')
                    chroms.append(multi_lowest_chrom)
                    ndxes.append(4)
            else:
                picks = []
                for pck in pick:
                    picks.append(multi_best_popn[pck[0]])
                a, b, c = calculate_fitness(picks)
                best_multi = np.min(b)
                best_mndx = b.index(best_multi)
                multi_lowest_chrom = picks[best_mndx]
                op_data[2], score_data[2] = calculate_fitness([multi_lowest_chrom])
                if self.more_details:
                    list(map(list, list(zip(*op_data[2]))))
                    dialog = displaytable.Table(op_data[2], title='Pick_' + self.sender().text(), fields=headers,
                             save_folder=self.scenarios, sortby='', decpts=op_pts)
                    dialog.exec_()
                    del dialog
                chrom_hdrs.append('Your pick')
                chroms.append(multi_lowest_chrom)
                ndxes.append(2)
        for chromosome in chroms:
            for fac, value in opt_order.items():
                capacity = opt_order[fac][2]
                for c in range(value[0], value[1]):
                    if chromosome[c]:
                        capacity = capacity + capacities[c]
                its[fac].append(capacity / pmss_details[fac].capacity)
        chooseDialog = QtWidgets.QDialog()
        hbox = QtWidgets.QHBoxLayout()
        grid = [QtWidgets.QGridLayout()]
        label = QtWidgets.QLabel('<b>Facility</b>')
        label.setAlignment(QtCore.Qt.AlignCenter)
        grid[0].addWidget(label, 0, 0)
        for h in range(len(chrom_hdrs)):
            grid.append(QtWidgets.QGridLayout())
            label = QtWidgets.QLabel('<b>' + chrom_hdrs[h] + '</b>')
            label.setAlignment(QtCore.Qt.AlignCenter)
            grid[-1].addWidget(label, 0, 0, 1, 3)
        rw = 1
        for key, value in its.items():
            grid[0].addWidget(QtWidgets.QLabel(key), rw, 0)
            if pmss_details[key].fac_type == 'S':
                typ = ' MWh'
            else:
                typ = ' MW'
            if self.show_multipliers:
                for h in range(len(chrom_hdrs)):
                    label = QtWidgets.QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0)
                    label = QtWidgets.QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 1)
                    label = QtWidgets.QLabel('({:.2f})'.format(value[h]))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 2)
            else:
                for h in range(len(chrom_hdrs)):
                    label = QtWidgets.QLabel('{:,.1f}'.format(value[h] * pmss_details[key].capacity))
                    label.setAlignment(QtCore.Qt.AlignRight)
                    grid[h + 1].addWidget(label, rw, 0, 1, 2)
                    label = QtWidgets.QLabel(typ)
                    label.setAlignment(QtCore.Qt.AlignLeft)
                    grid[h + 1].addWidget(label, rw, 2)
            rw += 1
        max_amt = [0., 0.]
        if do_lcoe:
            max_amt[0] = score_data[0][0]['cost']
            max_amt[1] = score_data[0][0]['co2']
        if do_multi:
            for multi in multi_best:
                max_amt[0] = max(max_amt[0], multi['cost'])
                max_amt[1] = max(max_amt[1], multi['co2'])
        pwr_chr = ['', '']
        divisor = [1., 1.]
        pwr_chrs = ' KMBTPEZY'
        for m in range(2):
            for pwr in range(len(pwr_chrs) - 1, -1, -1):
                if max_amt[m] > pow(10, pwr * 3):
                    pwr_chr[m] = pwr_chrs[pwr]
                    divisor[m] = 1. * pow(10, pwr * 3)
                    break
        self.targets['cost'][5] = self.targets['cost'][5].replace('pwr_chr', pwr_chr[0])
        self.targets['co2'][5] = self.targets['co2'][5].replace('pwr_chr', pwr_chr[1])
        for key in multi_order:
            lbl = QtWidgets.QLabel('<i>' + self.targets[key][0] + '</i>')
            lbl.setAlignment(QtCore.Qt.AlignCenter)
            grid[0].addWidget(lbl, rw, 0)
            for h in range(len(chrom_hdrs)):
                if key == 'cost':
                    amt = score_data[ndxes[h]][0][key] / divisor[0] # cost
                elif key == 'co2':
                    amt = score_data[ndxes[h]][0][key] / divisor[1] # co2
                elif key[-4:] == '_pct': # percentage
                    amt = score_data[ndxes[h]][0][key] * 100.
                else:
                    amt = score_data[ndxes[h]][0][key]
                txt = '<i>' + self.targets[key][5] + '</i>'
                try:
                    label = QtWidgets.QLabel(txt % amt)
                except:
                    label = QtWidgets.QLabel('?')
                    print('PME5:', key, txt, amt)
                label.setAlignment(QtCore.Qt.AlignCenter)
                grid[h + 1].addWidget(label, rw, 0, 1, 3)
            rw += 1
        cshow = QtWidgets.QPushButton('Quit', self)
        grid[0].addWidget(cshow)
        cshow.clicked.connect(chooseDialog.close)
        for h in range(len(chrom_hdrs)):
            button = QtWidgets.QPushButton(chrom_hdrs[h], self)
            grid[h + 1].addWidget(button, rw, 0, 1, 3)
            button.clicked.connect(chooseClicked) #(chrom_hdrs[h]))
        for gri in grid:
            frame = QtWidgets.QFrame()
            frame.setFrameStyle(QtWidgets.QFrame.Box)
            frame.setLineWidth(1)
            frame.setLayout(gri)
            hbox.addWidget(frame)
        scroll = QtWidgets.QScrollArea()
     #   scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
     #   scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        scroll.setWidgetResizable(True)
        widget = QtWidgets.QWidget()
        widget.setLayout(hbox)
        scroll.setWidget(widget)
        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(scroll)
        chooseDialog.setLayout(vbox)
        chooseDialog.setWindowTitle('Choose Optimal Generator Mix')
        chooseDialog.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
     #  this is a big of a kluge but I couldn't get it to behave
        self.opt_choice = ''
        chooseDialog.exec_()
        del chooseDialog
        plt.close('all')
        try:
            h = chrom_hdrs.index(self.opt_choice)
        except:
            return
        op_data[h], score_data[h] = calculate_fitness([chroms[h]]) # make it current
        msg = chrom_hdrs[h] + ': '
        for key in multi_order[:3]:
            msg += self.targets[key][0] + ': '
            if key == 'cost':
                amt = score_data[h][0][key] / divisor[0] # cost
            elif key == 'co2':
                amt = score_data[h][0][key] / divisor[1] # co2
            elif key[-4:] == '_pct': # percentage
                amt = score_data[h][0][key] * 100.
            else:
                amt = score_data[h][0][key]
            txt = self.targets[key][5]
            txt = txt % amt
            msg += txt + '; '
        self.setStatus(msg)
        list(map(list, list(zip(*op_data[h]))))
        op_data[h].append(' ')
        op_data[h].append('Optimisation Parameters')
        op_op_prm = len(op_data[h])
        op_data[h].append(['Population size', str(self.optimise_population)])
        op_data[h].append(['No. of iterations', str(self.optimise_generations)])
        op_data[h].append(['Mutation probability', '%0.4f' % self.optimise_mutation])
        op_data[h].append(['Exit if stable', str(self.optimise_stop)])
        op_data[h].append(['Optimisation choice', self.optimise_choice])
        op_data[h].append(['Variable', 'Weight', 'Better', 'Worse'])
        for i in range(len(target_keys)):
            op_data[h].append([])
            for j in range(4):
                if j == 0:
                    op_data[h][-1].append(self.targets[target_keys[i]][j])
                else:
                    op_data[h][-1].append('{:.2f}'.format(self.targets[target_keys[i]][j]))
        op_max_row = len(op_data[h])
        for key in self.optimisation.keys():
            op_data[h].append(['Max. ' + key, self.optimisation[key].capacity_max])
        dialog = displaytable.Table(op_data[h], title='Chosen_' + self.sender().text(), fields=headers,
                 save_folder=self.scenarios, sortby='', decpts=op_pts)
        dialog.exec_()
        del dialog
        if self.optimise_to_batch:
            msgbox = QtWidgets.QMessageBox()
            msgbox.setWindowTitle('SIREN - Add to Batch')
            msgbox.setText("Press 'Yes' to add to Batch file")
            msgbox.setIcon(QtWidgets.QMessageBox.Question)
            msgbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            reply = msgbox.exec_()
            if reply == QtWidgets.QMessageBox.Yes:
                if self.files[B].text() != '':
                    newfile = self.get_filename(self.files[B].text())
                else:
                    curfile = self.scenarios[:-1]
                    newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Create and save ' + self.file_labels[B] + ' file',
                              curfile, 'Excel Files (*.xlsx)')[0]
                    if newfile == '':
                        return
                if os.path.exists(newfile):
                    wb = oxl.load_workbook(newfile)
                elif self.batch_template == '':
                    return
                else:
                    wb = oxl.load_workbook(self.batch_template)   #copy batch
                    if newfile[: len(self.scenarios)] == self.scenarios:
                        self.files[B].setText(newfile[len(self.scenarios):])
                    else:
                        if newfile.rfind('/') > 0:
                            that_len = len(commonprefix([self.scenarios, newfile]))
                            if that_len > 0:
                                bits = self.scenarios[that_len:].split('/')
                                pfx = ('..' + '/') * (len(bits) - 1)
                                newfile = pfx + newfile[that_len + 1:]
                        if newfile[-5:] != '.xlsx':
                            newfile += '.xlsx'
                        self.files[B].setText(newfile)
                if wb.worksheets[0].max_column > 1024:
                    self.clean_batch_sheet(B)
                    ds = oxl.load_workbook(self.get_filename(self.files[B].text()))
                adj_list = {}
                for o_r in range(len(op_data[h])):
                    if isinstance(op_data[h][o_r], list):
                        if op_data[h][o_r][0][:10] == 'Total Load':
                            adj_list['Load'] = op_data[h][o_r][2]
                        else:
                            try:
                                adj_list[op_data[h][o_r][0]] = op_data[h][o_r][1]
                            except:
                                adj_list[op_data[h][o_r][0]] = ''
                    elif op_data[h][o_r] != '':
                        adj_list[op_data[h][o_r]] = ''
                add_msg = add_to_batch(wb, self.get_filename(self.files[B].text()), adj_list)
                ##QtWidgets.QMessageBox.about(self, 'SIREN - Add to Batch', add_msg)
                self.setStatus(add_msg)
        if self.adjust.isChecked():
            self.adjustto = {}
            for fac, value in sorted(pmss_details.items()):
                self.adjustto[fac] = value.capacity * value.multiplier
        return

if "__main__" == __name__:
    app = QtWidgets.QApplication(sys.argv)
    setFontSize(app)
    ex = powerMatch()
    app.exec_()
    app.deleteLater()
    sys.exit()
