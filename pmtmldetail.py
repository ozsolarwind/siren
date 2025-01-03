#!/usr/bin/python3
#
#  Copyright (C) 2023-2024 Angus King
#
#  pmtmldetail.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#
# Note: Batch process is all rather messy.
import os
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
import displayobject
from credits import fileVersion
import openpyxl as oxl
from senutils import ClickableQLabel, getParents, getUser, ListWidget, ssCol, WorkBook
from editini import EdtDialog, SaveIni
from getmodels import getModelFile, commonprefix
import configparser  # decode .ini file

# same order as self.file_labels
R = 0 # Results - xlsx
D = 1 # More detail


class TMLDetail(QtWidgets.QWidget):
    log = QtCore.pyqtSignal()
    progress = QtCore.pyqtSignal()

    def get_filename(self, filename):
        if filename.find('/') == 0: # full directory in non-Windows
            return filename
        elif (sys.platform == 'win32' or sys.platform == 'cygwin') \
          and filename[1:2] == ':/': # full directory for Windows
            return filename
        elif filename[:3] == '../': # directory upwards of scenarios
            ups = filename.split('../')
            scens = self.scenarios.split('/')
            scens = scens[: -(len(ups) - 1)]
            scens.append(ups[-1])
            return '/'.join(scens)
        else: # subdirectory of scenarios
            return self.scenarios + filename

    def __init__(self, help='help.html'):
        super(TMLDetail, self).__init__()
        self.help = help
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            base_year = config.get('Base', 'year')
        except:
            base_year = '2012'
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', base_year)
            self.scenarios = self.scenarios[: self.scenarios.rfind('/') + 1]
            if self.scenarios[:3] == '../':
                ups = self.scenarios.split('../')
                me = os.getcwd().split(os.sep)
                me = me[: -(len(ups) - 1)]
                me.append(ups[-1])
                self.scenarios = '/'.join(me)
        except:
            self.scenarios = ''
        self.file_labels = ['Results', 'Tml Detail']
        ifiles = [''] * len(self.file_labels)
        self.isheet = 'Detail'
        self.surplus_sign = 1 # Note: Preferences file has it called shortfall_sign
        # it's easier for the user to understand while for the program logic surplus is easier
        try:
            items = config.items('Powermatch')
            for key, value in items:
                if key[-5:] == '_file':
                    if key[:-5].title() in self.file_labels:
                        ndx = self.file_labels.index(key[:-5].title())
                        ifiles[ndx] = value.replace('$USER$', getUser())
                    elif key[:-5].title().replace('_', ' ') in self.file_labels:
                        ndx = self.file_labels.index(key[:-5].title().replace('_', ' '))
                        ifiles[ndx] = value.replace('$USER$', getUser())
                elif key == 'tml_results_sheet':
                    self.isheet = value
                elif key == 'shortfall_sign':
                    if value[0] == '+' or value[0].lower() == 'p':
                        self.surplus_sign = -1
        except:
            print('PME1: Error with', key)
        self.restorewindows = False
        try:
            rw = config.get('Windows', 'restorewindows')
            if rw.lower() in ['true', 'yes', 'on']:
                self.restorewindows = True
        except:
            pass
        self.grid = QtWidgets.QGridLayout()
        self.labels = [None] * len(self.file_labels)
        self.files = [None] * len(self.file_labels)
        self.updated = False
        r = 0
        for i in range(len(self.file_labels)):
            self.labels[i] = QtWidgets.QLabel(self.file_labels[i] + ' File:')
            self.grid.addWidget(self.labels[i], r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
            self.files[i].setText(ifiles[i])
            self.files[i].clicked.connect(self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 5)
            if i < D: # only for Results file
                r += 1
                self.grid.addWidget(QtWidgets.QLabel('Results Sheet:'), r, 0)
                self.sheet = QtWidgets.QComboBox()
                try:
                    curfile = self.get_filename(ifiles[i])
                    ts = WorkBook()
                    ts.open_workbook(curfile)
                    ndx = 0
                    j = -1
                    for sht in ts.sheet_names():
                        j += 1
                        self.sheet.addItem(sht)
                        if sht == self.isheet:
                            ndx = j
                    self.sheet.setCurrentIndex(ndx)
                    ws = ts.sheet_by_index(ndx)
                    ts.close()
                    del ts
                except:
                    self.sheet.addItem(self.isheet)
                self.grid.addWidget(self.sheet, r, 1, 1, 3)
                self.sheet.currentIndexChanged.connect(self.sheetChanged)
            r += 1
        r += 1
        self.log = QtWidgets.QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, r, 1, 1, 6)
        r += 1
        r += 1
        quit = QtWidgets.QPushButton('Done', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        pm = QtWidgets.QPushButton('TML Detail', self)
        self.grid.addWidget(pm, r, 1)
        pm.clicked.connect(self.pmClicked)
        editini = QtWidgets.QPushButton('Preferences', self)
        self.grid.addWidget(editini, r, 2)
        editini.clicked.connect(self.editIniFile)
        help = QtWidgets.QPushButton('Help', self)
        self.grid.addWidget(help, r, 3)
        help.clicked.connect(self.helpClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - pmtmldetail (' + fileVersion() + ') - TML Detail')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        self.center()
        self.resize(int(self.sizeHint().width() * 1.2), int(self.sizeHint().height() * 1.2))
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        centerPoint = QtWidgets.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        for i in range(len(self.file_labels)):
            if self.files[i].hasFocus():
                break
        if self.files[i].text() == '':
            curfile = self.scenarios[:-1]
        else:
            curfile = self.get_filename(self.files[i].text())
        if i == R:
            newfile = QtWidgets.QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[R] + ' file',
                      curfile)[0]
        else: # i == D:
            if self.files[i].text() == '':
                curfile = self.get_filename(self.files[R].text())
                curfile = curfile.replace('Results', 'TML_Detail')
                curfile = curfile.replace('results', 'tml_detail')
                if curfile == self.scenarios + self.files[D].text():
                    j = curfile.find(' ')
                    if j > 0:
                        jnr = ' '
                    else:
                        jnr = '_'
                    j = curfile.rfind('.')
                    curfile = curfile[:j] + jnr + 'TML_Detail' + curfile[j:]
            else:
                curfile = self.get_filename(self.files[D].text())
            newfile = QtWidgets.QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[D] + ' file',
                      curfile, 'Excel Files (*.xlsx)')[0]
        if newfile != '':
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                if newfile.rfind('/') > 0:
                    that_len = len(commonprefix([self.scenarios, newfile]))
                    if that_len > 0:
                        bits = self.scenarios[that_len:].split('/')
                        pfx = ('..' + '/') * (len(bits) - 1)
                        newfile = pfx + newfile[that_len + 1:]
                self.files[i].setText(newfile)
            if i == R: # if results need to change details
                rsltfile = self.files[i].text().replace('Results', 'TML_Detail')
                rsltfile = rsltfile.replace('results', 'tml_detail')
                self.files[D].setText(rsltfile)
            self.updated = True

    def sheetChanged(self):
        ts = WorkBook()
        ts.open_workbook(newfile)
        ws = ts.sheet_by_name(self.sheet.currentText())
        ts.close()
        del ts

    def helpClicked(self):
        dialog = displayobject.AnObject(QtWidgets.QDialog(), self.help,
                 title='Help for pmtmldetail (' + fileVersion() + ')', section='pmtmldetail')
        dialog.exec_()

    def changes(self):
        self.updated = True

    def quitClicked(self):
        if self.updated or self.order.updated or self.ignore.updated:
            updates = {}
            lines = []
            for i in range(len(self.file_labels)):
                lines.append(self.file_labels[i].lower().replace(' ', '_') + '_file=' + self.files[i].text().replace(getUser(), '$USER$'))
            lines.append('tml_results_sheet=' + self.sheet.currentText())
            updates['Powermatch'] = lines
            SaveIni(updates)
        self.close()

    def closeEvent(self, event):
        event.accept()

    def editIniFile(self):
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        before = os.stat(config_file).st_mtime
        dialr = EdtDialog(config_file, section='[Powermatch]')
        dialr.exec_()
        after = os.stat(config_file).st_mtime
        if after == before:
            return
        config = configparser.RawConfigParser()
        config.read(config_file)
        self.log.setText(config_file + ' edited. Reload may be required.')

    def pmClicked(self):
        err_msg = ''
        pm_report_file = self.get_filename(self.files[R].text())
        if pm_report_file[-5:] != '.xlsx': #xlsx format only
            self.log.setText('Not a Powermatch data spreadsheet')
            return
        try:
            wb = oxl.load_workbook(pm_report_file)
        except FileNotFoundError:
            self.log.setText('Report file not found - ' + self.files[R].text())
            return
        except:
            self.log.setText('Error accessing Report file - ' + self.files[R].text())
            return
        try:
            ws = wb[self.sheet.currentText()]
        except:
            wb.close()
            self.log.setText('Invalid worksheet - ' +  self.sheet.currentText())
            return
        top_row = ws.max_row - 8760
        ws.row_dimensions[top_row].height = 15
        if top_row < 1 or (ws.cell(row=top_row, column=1).value != 'Hour' \
                           or ws.cell(row=top_row, column=2).value != 'Period' \
                           or ws.cell(row=top_row, column=3).value != 'Load'):
            wb.close()
            self.log.setText('Not a Powermatch detail spreadsheet')
            return
        normal = oxl.styles.Font(name='Arial')
        for sht in wb.sheetnames:
            if sht != self.sheet.currentText():
                del wb[sht]
        ws = wb[self.sheet.currentText()]
        ws.title = 'TML Detail'
        del_rows = ['Initial Capacity', 'CF', 'Cost ($/yr)', 'LCOG Cost ($/MWh)', 'Emissions (tCO2e)']
        zone_row = -1
        for row in range(top_row -1, 0, -1):
            if ws.cell(row=row, column=2).value in del_rows:
                top_row -= 1
                ws.delete_rows(row, 1)
        if ws.cell(row=top_row -1, column=1).value == 'Zone':
            zone_row = top_row - 1
            split_row = top_row - 1
        else:
            split_row = top_row
        ws.insert_rows(split_row)
        if zone_row > 0:
            zone_row += 1
        top_row += 1
        ws.cell(row=split_row, column=1).value = 'Split'
        ws.cell(row=split_row, column=1).font = normal
        subt_row = -1
        shfl_row = -1
        max_row = -1
        use_row = -1
        for row in range(1, split_row):
           if ws.cell(row=row, column=2).value == 'Subtotal (MWh)':
               subt_row = row
           elif ws.cell(row=row, column=2).value == 'Shortfall periods':
               shfl_row = row
           elif ws.cell(row=row, column=2).value == 'Maximum (MW/MWh)':
               max_row = row
           elif ws.cell(row=row, column=2).value == 'Hours of usage':
               use_row = row
        techs = [[[], []], [[], []]] # RE [value, column], Storage [value, column]
        grp_techs = {}
        t = 0 # RE
        short_col = -1
        under_col = -1
        for col in range(4, ws.max_column + 1):
            if ws.cell(row=top_row, column=col).value[:9] == 'Shortfall':
                short_col = col
                t = 1 # now storage
            elif t == 0:
                techs[t][0].append(ws.cell(row=top_row, column=col).value)
                techs[t][1].append(col)
                if zone_row > 0:
                    if techs[t][0][-1] in grp_techs.keys():
                        grp_techs[techs[t][0][-1]].append(col - 4)
                    else:
                        grp_techs[techs[t][0][-1]] = [col - 4]
                ws.cell(row=split_row, column=col).value = 'Generation'
                ws.cell(row=split_row, column=col).font = normal
                ws.cell(row=split_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
            elif ws.cell(row=top_row, column=col).value[:6] == 'Charge':
                techs[t][0].append(ws.cell(row=top_row, column=col).value)
                techs[t][1].append(col)
            elif  ws.cell(row=top_row, column=col).value[:10] == 'Underlying':
                under_col = col
        if under_col > 0: # adjust the rows
            his_formula = ws.cell(row=top_row + 1, column=under_col).value
            his_row = ''
            in_digit = False
            for c in his_formula:
                if in_digit:
                    if c.isdigit():
                        his_row += c
                    else:
                        break
                elif c.isdigit():
                    his_row += c
                    in_digit = True
            if his_row != top_row + 1:
                for row in range(top_row + 1, ws.max_row + 1):
                    ws.cell(row=row, column=under_col).value = his_formula.replace(his_row, str(row))
        ws.row_dimensions[top_row].height = 30
        col = ws.max_column + 1
        col_s = col
        # add columns for totals of other techs, e.g. Rooftop PV, Gas, ...
        splits = ['TML', 'Charge', 'Surplus']
        for s in range(3):
            if s == 1 and len(techs[1][0]) == 0:
                continue
            for t in range(len(techs[0][0])):
                ws.cell(row=split_row, column=col).value = splits[s]
                ws.cell(row=split_row, column=col).font = normal
                ws.cell(row=split_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                if zone_row > 0:
                    ws.cell(row=zone_row, column=col).value = ws.cell(row=zone_row, column=techs[0][1][t]).value
                    ws.cell(row=zone_row, column=col).font = normal
                    ws.cell(row=zone_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ws.cell(row=top_row, column=col).value = ws.cell(row=top_row, column=techs[0][1][t]).value
                ws.cell(row=top_row, column=col).font = normal
                ws.cell(row=top_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                for row in range(top_row + 1, ws.max_row + 1):
                    if s == 0: # TML
                        ws.cell(row=row, column=col).value = '=IF($' + ssCol(short_col) + str(row) + '>0,$C' + \
                                str(row) + '/($C' + str(row) + '+$' + ssCol(short_col) + str(row) + ')*' + \
                                ssCol(techs[0][1][t]) + str(row) + ',' + ssCol(techs[0][1][t]) + str(row) + ')'
                        ws.cell(row=row, column=col).font = normal
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                    elif s == 1: # Charge
                        chgs = '('
                        for c in techs[1][1]:
                            chgs += '$' + ssCol(c) + str(row) + '+'
                        chgs = chgs[:-1] + ')'
                        ws.cell(row=row, column=col).value = '=IF(' + chgs + '>0,' + ssCol(techs[0][1][t]) + \
                                str(row) + '/SUM($' + ssCol(techs[0][1][0]) + str(row) + ':$' + \
                                ssCol(techs[0][1][-1]) + str(row) + ')*' + chgs + ',0)'
                        ws.cell(row=row, column=col).font = normal
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                    elif s == 2: # Surplus
                        sur = '=' + ssCol(techs[0][1][t]) + str(row) + '-' + ssCol(col_s + t) + str(row)
                        if len(techs[1][0]) > 0:
                            sur += '-' + ssCol(col_s + len(techs[0][0]) + t + len(grp_techs)) + str(row)
                        ws.cell(row=row, column=col).value = sur
                        ws.cell(row=row, column=col).font = normal
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                col += 1
            if zone_row > 0:
                ref_col = col - len(techs[0][0])
                for key, value in grp_techs.items():
                    ws.cell(row=split_row, column=col).value = splits[s]
                    ws.cell(row=split_row, column=col).font = normal
                    ws.cell(row=split_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                            vertical='bottom', horizontal='center')
                    ws.cell(row=top_row, column=col).value = key
                    ws.cell(row=top_row, column=col).font = normal
                    ws.cell(row=top_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                    for row in range(top_row + 1, ws.max_row + 1):
                        chgs = '='
                        for c in value:
                            chgs += ssCol(ref_col + c) + str(row) + '+'
                        chgs = chgs[:-1]
                        ws.cell(row=row, column=col).value = chgs
                        ws.cell(row=row, column=col).font = normal
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
                    col += 1
        if subt_row >= 0:
            for col in range(3, ws.max_column + 1):
                ws.cell(row=subt_row, column=col).value = '=SUM(' + ssCol(col) + str(top_row + 1) + ':' + \
                        ssCol(col) + str(ws.max_row) + ')'
                ws.cell(row=subt_row, column=col).font = normal
                ws.cell(row=subt_row, column=col).number_format = '#,##0'
        if shfl_row >= 0:
            for col in range(3, ws.max_column + 1):
                if ws.cell(row=shfl_row, column=col).value is None or ws.cell(row=shfl_row, column=col).value == '':
                    continue
                ws.cell(row=shfl_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(top_row + 1) + ':' + \
                        ssCol(col) + str(ws.max_row) + ',"<0")'
                ws.cell(row=shfl_row, column=col).font = normal
                ws.cell(row=shfl_row, column=col).number_format = '#,##0'
        if max_row >= 0:
            for col in range(3, ws.max_column + 1):
                ws.cell(row=max_row, column=col).value = '=MAX(' + ssCol(col) + str(top_row + 1) + ':' + \
                        ssCol(col) + str(ws.max_row) + ')'
                ws.cell(row=max_row, column=col).font = normal
                ws.cell(row=max_row, column=col).number_format = '#,##0.00'
        if use_row >= 0:
            for col in range(3, ws.max_column + 1):
                ws.cell(row=use_row, column=col).value = '=COUNTIF(' + ssCol(col) + str(top_row + 1) + ':' + \
                        ssCol(col) + str(ws.max_row) + ',">0")'
                ws.cell(row=use_row, column=col).font = normal
                ws.cell(row=use_row, column=col).number_format = '#,##0'
        ws.freeze_panes = 'C' + str(top_row + 1)
        ws.activeCell = 'C' + str(top_row + 1)
        wb.save(self.get_filename(self.files[D].text()))
        self.log.setText('File created - ' +  self.files[D].text())

    def exit(self):
        self.updated = False
        self.close()

if "__main__" == __name__:
    app = QtWidgets.QApplication(sys.argv)
    ex = TMLDetail()
    app.exec_()
    app.deleteLater()
    sys.exit()
