#!/usr/bin/python3
#
#  Copyright (C) 2015-2023 Sustainable Energy Now Inc., Angus King
#
#  senutils.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

import csv
import math
import openpyxl as oxl
# from openpyxl.formula import Tokenizer
import os
try:
    import pwd
except:
    pass
import sys
from PyQt5 import QtCore, QtWidgets
import xlrd
if xlrd.__version__[:2] [0] == '1.': # if xlsx files still supported
    if sys.version_info[1] >= 9: # python 3.9 onwards
        xlrd.xlsx.ensure_elementtree_imported(False, None)
        xlrd.xlsx.Element_has_iter = True


class ClickableQLabel(QtWidgets.QLabel):
    clicked = QtCore.pyqtSignal()

    def __init(self, parent):
        QLabel.__init__(self, parent)

    def mousePressEvent(self, event):
        QtWidgets.QApplication.widgetAt(event.globalPos()).setFocus()
        self.clicked.emit()

# class to support listwidget drag and drop between two lists
# also supports using keys where drag and drop not working (e.g. Ubuntu 23.04)
class ListWidget(QtWidgets.QListWidget):
    def decode_data(self, bytearray):
        data = []
        ds = QtCore.QDataStream(bytearray)
        while not ds.atEnd():
            row = ds.readInt32()
            column = ds.readInt32()
            map_items = ds.readInt32()
            for i in range(map_items):
                key = ds.readInt32()
                value = QtCore.QVariant()
                ds >> value
                data.append(value.value())
        return data

    def __init__(self, parent=None):
        super(ListWidget, self).__init__(parent)
        self.setDragDropMode(self.DragDrop)
        self.setSelectionMode(self.ExtendedSelection)
        self.setAcceptDrops(True)
        self._other = None
        for child in self.parent().children():
            if isinstance(child, ListWidget) and child != self: # will work if more than one ListWidget
                self._other = child
                self.setObjectName('Exclude')
                child._other = self
                child.setObjectName('Include')

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            super(ListWidget, self).dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
        else:
            super(ListWidget, self).dragMoveEvent(event)

    def dropEvent(self, event):
        if event.source() == self:
            event.setDropAction(QtCore.Qt.MoveAction)
            QtWidgets.QListWidget.dropEvent(self, event)
        else:
            ba = event.mimeData().data('application/x-qabstractitemmodeldatalist')
            data_items = self.decode_data(ba)
            event.setDropAction(QtCore.Qt.MoveAction)
            event.source().deleteItems(data_items)
            super(ListWidget, self).dropEvent(event)

    def deleteItems(self, items):
        for row in range(self.count() -1, -1, -1):
            if self.item(row).text() in items:
             #   r = self.row(item)
                self.takeItem(row)

    def keyPressEvent(self, event):
        if self.currentRow() < 0:
            return
        action = ''
        try:
            if event.key() == 16777223:
                action = 'Delete'
            elif event.key() == 16777235:
                action = 'Up'
            elif event.key() == 16777237:
                action = 'Down'
            elif event.key() == 16777234:
                if self.objectName == 'Include':
                    return
                action = 'Shift'
            elif event.key() == 16777236:
                if self.objectName == 'Exclude':
                    return
                action = 'Shift'
            elif chr(event.key()) == 'U':
                action = 'Up'
            elif chr(event.key()) == 'D':
                action = 'Down'
            elif chr(event.key()) == '+' or chr(event.key()) == '=' or chr(event.key()) == 'I' or chr(event.key()) == 'L':
                if self.objectName == 'Include':
                    return
                action = 'Shift'
            elif chr(event.key()) == '-' or chr(event.key()) == 'E' or chr(event.key()) == 'R':
                if self.objectName == 'Exclude':
                    return
                action = 'Shift'
        except:
            return
        if action == 'Shift':
            background = self.currentItem().background()
            self._other.addItem(self.currentItem().text())
            self._other.item(self._other.count() - 1).setBackground(background)
            self.takeItem(self.currentRow())
        elif action == 'Up':
            if self.currentRow() > 0:
                background = self.currentItem().background()
                self.insertItem(self.currentRow() - 1, self.currentItem().text())
                self.takeItem(self.currentRow())
                self.setCurrentRow(self.currentRow() - 1)
                self.currentItem().setBackground(background)
        elif action == 'Down':
            if self.currentRow() < self.count() - 1:
                background = self.currentItem().background()
                self.insertItem(self.currentRow() + 2, self.currentItem().text())
                row = self.currentRow()
                self.takeItem(self.currentRow())
                self.setCurrentRow(row + 1)
                self.currentItem().setBackground(background)


# Class to support input file as .csv, .xls, or .xlsx
class WorkBook(object):
    def __init__(self):
        self._book = None
        self._data_only = None
        self._sheet = None
        self._sheet_names = []
        self._type = None
        self._nrows = 0
        self._ncols = 0

    def open_workbook(self, filename=None, on_demand=True, data_only=True, filetype=None):
        if not os.path.exists(filename):
            raise Exception('File not found')
        if filetype is None:
            self._type = filename[filename.rfind('.') + 1:]
        else:
            self._type = filetype
        self._data_only = data_only
        try:
            if self._type == 'xls':
                self._book = xlrd.open_workbook(filename, on_demand=on_demand)
                self._sheet_names = self._book.sheet_names()
            elif self._type == 'xlsx':
                self._book = oxl.load_workbook(filename, data_only=data_only)
                self._sheet_names = self._book.sheetnames
            elif self._type == 'csv':
                csv_file = open(filename, newline='')
                things = csv.reader(csv_file)
                self._sheet_names = ['n/a']
                self._worksheet = []
                for row in things:
                    self._worksheet.append([])
                    for cell in row:
                        if len(cell) > 0:
                            if cell[0] == '-':
                                minus = '-'
                                cell = cell[1:]
                            else:
                                minus = ''
                            if cell.isdigit():
                                 self._worksheet[-1].append(int(minus + cell))
                            elif cell.replace(',', '').isdigit():
                                 self._worksheet[-1].append(int(minus + cell.replace(',', '')))
                            elif cell.replace(',', '').replace('.', '').isdigit():
                                 try:
                                     self._worksheet[-1].append(float(minus + cell.replace(',', '')))
                                 except:
                                     self._worksheet[-1].append(minus + cell)
                            else:
                                 self._worksheet[-1].append(minus + cell)
                        else:
                             self._worksheet[-1].append('')
                csv_file.close()
                self.nrows = len(self._worksheet)
                self.ncols = len(self._worksheet[0])
            else:
                raise Exception(f"Error with filetype - '{self._type}'")
        except Exception as err:
            if isinstance(err, Exception) and err.args[0][:19] == 'Error with filetype':
                raise
            else:
                raise Exception('Error opening file')

    def release_resources(self):
        if self._type == 'xls':
            try:
                self._book.release_resources()
            except:
                pass

    def sheet_names(self):
        return self._sheet_names[:]

    def sheet_by_index(self, sheetx):
        return self.get_sheet(sheetx)

    def sheet_by_name(self, sheet_name):
        try:
            sheetx = self._sheet_names.index(sheet_name)
        except ValueError:
            raise Exception('No sheet named <%r>' % sheet_name)
        return self.sheet_by_index(sheetx)

    def get_sheet(self, sheetx):
        self._sheet = self.WorkSheet(sheetx, self._type, self._data_only)
        try:
            if self._type == 'xls':
                self._sheet._sheet = self._book.sheet_by_index(sheetx)
                self._sheet.name = self._book.sheet_names()[sheetx]
                self._sheet.nrows = self._sheet._sheet.nrows
                self._sheet.ncols = self._sheet._sheet.ncols
            elif self._type == 'xlsx':
                self._sheet._sheet = self._book.worksheets[sheetx]
                self._sheet.name = self._book.sheetnames[sheetx]
                self._sheet.nrows = self._sheet._sheet.max_row
                self._sheet.ncols = self._sheet._sheet.max_column
            else: #if self._type == 'csv':
              #  self._sheet = self._book.worksheets[sheetx]
                self._sheet.name = 'n/a'
                self._sheet._sheet = self._worksheet
                self._sheet.nrows = len(self._sheet._sheet)
                try:
                    self._sheet.ncols = len(self._sheet._sheet[0])
                except:
                    self._sheet.ncols = 0
        except:
            raise Exception('Error accessing sheet')
        return self._sheet

    class WorkSheet(object):
        def __init__(self, sheet, typ, data_only):
            self.name = sheet
            self._sheet = None
            self.nrows = 0
            self.ncols = 0
            self._type = typ
            self._data_only = data_only

        def cell_value(self, row, col):
            if self._type == 'xls':
                return self._sheet.cell_value(row, col)
            elif self._type == 'xlsx':
                if self._data_only:
                    return self._sheet.cell(row=row + 1, column=col + 1).value
                else:
                    return self._sheet.cell(row=row + 1, column=col + 1).value
                    # sometime in the future
                    # if self._sheet.cell(row=row + 1, column=col + 1).data_type == 'f':
                    #     tok = Tokenizer(self._sheet.cell(row=row + 1, column=col + 1).value)
                    #     print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))
                    #     return self._sheet.cell(row=row + 1, column=col + 1).value
                    # else:
            else: #if self._type == 'csv':
                return self._sheet[row][col]

        def cell_type(self, row, col):
            if self._type == 'xlsx':
                return self._sheet.cell(row=row + 1, column=col + 1).data_type
            else:
                return None

#        def cell_write(self, row, col, value):
#            if self._type == 'xls':
#                self._sheet.write(row, col, value)
#            elif self._type == 'xlsx':
#                self._sheet.cell(row=row + 1, column=col + 1).value = value
#            elif self._type == 'csv':
#                self._sheet[row][col] = value


#
# replace parent string in filenames
def getParents(aparents):
    parents = []
    for key, value in aparents:
        for key2, value2 in aparents:
            if key2 == key:
                continue
            value = value.replace(key2, value2)
        for key2, value2 in parents:
            if key2 == key:
                continue
            value = value.replace(key2, value2)
        parents.append((key, value))
    return parents

#
# return current userid
def getUser():
    if sys.platform == 'win32' or sys.platform == 'cygwin':   # windows
        return os.environ.get("USERNAME")
    elif sys.platform == 'darwin':   # osx64
        return pwd.getpwuid(os.geteuid()).pw_name
    elif sys.platform == 'linux' or sys.platform == 'linux2':   # linux
        return pwd.getpwuid(os.geteuid()).pw_name
    else:
        return os.environ.get("USERNAME")

#
# clean up tech names
def techClean(tech, full=False):
    cleantech = tech.replace('_', ' ').title()
    cleantech = cleantech.replace('Bess', 'BESS')
    cleantech = cleantech.replace('Bm', 'BM')
    cleantech = cleantech.replace('Ccgt', 'CCGT')
    cleantech = cleantech.replace('Ccg', 'CCG')
    cleantech = cleantech.replace('Cst', 'CST')
    cleantech = cleantech.replace('Lng', 'LNG')
    cleantech = cleantech.replace('Ocgt', 'OCGT')
    cleantech = cleantech.replace('Ocg', 'OCG')
    cleantech = cleantech.replace('Phs', 'PHS')
    cleantech = cleantech.replace('Pv', 'PV')
    if full:
        alll = [['Cf', 'CF'], ['Co2', 'CO2'], ['hi ', 'HI '], ['Lcog', 'LCOG'],
                ['Lcoe', 'LCOE'], ['Mw', 'MW'], ['ni ', 'NI '], ['Npv', 'NPV'],
                ['Re', 'RE'], ['Tco2E', 'tCO2e'], ['REference', 'Reference']]
        for each in alll:
            cleantech = cleantech.replace(each[0], each[1])
        # fudge
        cleantech = cleantech.replace('REc', 'Rec')
    return cleantech

#
# add another windspeed height
def extrapolateWind(wind_file, tgt_height, law='logarithmic', replace=False, spread=None):
    if tgt_height < 60:
        return False
    if not os.path.exists(wind_file):
        if replace:
            return False
        else:
            return None
    if wind_file[-4:] != '.srw':
        if replace:
            return False
        else:
            return None
    tf = open(wind_file, 'r')
    lines = tf.readlines()
    tf.close()
    fst_row = 5
    units = lines[3].rstrip(',\n').split(',')
    hghts = lines[4].rstrip(',\n').split(',')
    col = -1
    heights_ms = []
    heights_dirn = []
    for j in range(len(units)):
        if units[j] == 'm/s':
             heights_ms.append([int(hghts[j]), j])
             if spread is not None:
                 if tgt_height in range(heights_ms[-1][0] - spread, heights_ms[-1][0] + spread):
                     return None
             if heights_ms[-1][0] == tgt_height:
                 if replace:
                     return False
                 else:
                     return None
        elif units[j] == 'degrees':
             heights_dirn.append([int(hghts[j]), j])
    lines[2] = lines[2].rstrip(',\n') + ',Direction,Speed\n'
    lines[3] = lines[3].rstrip(',\n') + ',degrees,m/s\n'
    lines[4] = lines[4].rstrip(',\n') + ',' + str(tgt_height) + ',' + str(tgt_height) + '\n'
    heights_ms.sort(key=lambda x: x[0], reverse=True)
    heights_dirn.sort(key=lambda x: x[0], reverse=True)
    height = float(heights_ms[0][0])
    col = heights_ms[0][1]
    height0 = float(heights_ms[1][0])
    if height0 == height:
        if replace:
            return False
        else:
            return None
    col0 = heights_ms[1][1]
    cold = heights_dirn[0][1]
    for i in range(fst_row, len(lines)):
        bits = lines[i].rstrip(',\n').split(',')
        speed = float(bits[col])
        speed0 = float(bits[col0])
        if speed0 >= speed:
            alpha = 1. / 7. # one-seventh power law
        else:
         #   alpha = (math.log(speed)-math.log(speed0))/(math.log(height)-math.log(height0))
            alpha = math.log(speed / speed0) / math.log(height / height0)
        z0 = math.exp(((pow(height0, alpha) * math.log(height)) - pow(height, alpha) * math.log(height0)) \
                      / ( pow(height0, alpha) - pow(height, alpha)))
        # z1 = math.exp(((pow(height, alpha) * math.log(height0)) - pow(height0, alpha) * math.log(height)) \
        #               / ( pow(height, alpha) - pow(height0, alpha)))
        if z0 < 1e-308:
            z0 = 0.03
        elif z0 < 1e-302:
            print('(311)', i, wind_file, z0)
            z0 = 0.03
        if law.lower()[0] == 'l': # law == 'logarithmic'
            speedz = math.log(tgt_height / z0) / math.log(height0 / z0) * speed0
            lines[i] = lines[i].strip() + ',' + bits[cold] + ',' + str(round(speedz, 4)) + '\n'
        else: # law == 'hellmann'
            speeda = pow(tgt_height / height0, alpha) * speed0
            lines[i] = lines[i].strip() + ',' + bits[cold] + ',' + str(round(speeda, 4)) + '\n'
    if replace:
        if os.path.exists(wind_file + '~'):
            os.remove(wind_file + '~')
        os.rename(wind_file, wind_file + '~')
        nf = open(wind_file, 'w')
        for line in lines:
            nf.write(line)
        nf.close()
        return True
    else:
        return lines
        array = [] # this doesn't work yet
        for i in range(4):
            bits = lines[i].split(',')
            array.append(bits)
        for i in range(fst_row, len(lines)):
            bits = lines[i].split(',')
            for j in range(len(bits)):
                bits[j] = float(bits[j])
            array.append(bits)
        return array

# split a string
def strSplit(string, char=',', dropquote=True):
    last = 0
    splits = []
    inQuote = None
    for i, letter in enumerate(string):
        if inQuote:
            if (letter == inQuote):
                inQuote = None
                if dropquote:
                    splits.append(string[last:i])
                    last = i + 1
                    continue
        elif (letter == '"' or letter == "'"):
            inQuote = letter
            if dropquote:
                last += 1
        elif letter == char:
            if last != i:
                splits.append(string[last:i])
            last = i + 1
    if last < len(string):
        splits.append(string[last:])
    return splits
