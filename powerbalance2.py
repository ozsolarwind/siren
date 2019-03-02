#!/usr/bin/python
#
#  Copyright (C) 2018-2019 Sustainable Energy Now Inc., Angus King
#
#  powerbalance2.py - This file is possibly part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

import os
import sys
import time
from PyQt4 import QtCore, QtGui
import displayobject
import displaytable
from credits import fileVersion
import openpyxl as oxl
# from openpyxl.utils import get_column_letter
from parents import getParents
from senuser import getUser
from editini import SaveIni
import xlrd
import ConfigParser  # decode .ini file

tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def ss_col(col, base=1):
    if base == 1:
        col -= 1
    c1 = col // 26
    c2 = col % 26
    return (col_letters[c1] + col_letters[c2 + 1]).strip()

class ClickableQLabel(QtGui.QLabel):
    def __init(self, parent):
        QLabel.__init__(self, parent)

    def mousePressEvent(self, event):
        QtGui.QApplication.widgetAt(event.globalPos()).setFocus()
        self.emit(QtCore.SIGNAL('clicked()'))


class Constraint:
    def __init__(self, name, category, capacity_min, capacity_max, rampup_max, rampdown_max,
                 recharge_max, recharge_loss, discharge_max, discharge_loss, parasitic_loss):
        self.name = name
        self.category = category
        try:
            self.capacity_min = float(capacity_min) # minimum run_rate for generator; don't drain below for storage
        except:
            self.capacity_min = 0.
        try:
            self.capacity_max = float(capacity_max) # maximum run_rate for generator; don't drain more than this for storage
        except:
            self.capacity_max = 1.
        try:
            self.recharge_max = float(recharge_max) # can't charge more than this per hour
        except:
            self.recharge_max = 1.
        try:
            self.recharge_loss = float(recharge_loss)
        except:
            self.recharge_loss = 0.
        try:
            self.discharge_max = float(discharge_max) # can't discharge more than this per hour
        except:
            self.discharge_max = 1.
        try:
            self.discharge_loss = float(discharge_loss)
        except:
            self.discharge_loss = 0.
        try:
            self.parasitic_loss = float(parasitic_loss) # daily parasitic loss / hourly ?
        except:
            self.parasitic_loss = 0.
        try:
            self.rampup_max = float(rampup_max)
        except:
            self.rampup_max = 1.
        try:
            self.rampdown_max = float(rampdown_max)
        except:
            self.rampdown_max = 1.


class Facility:
    def __init__(self, name, order, constraint, capacity, lcoe, lcoe_cf, emissions, initial=None):
        self.name = name
        self.constraint = constraint
        try:
            self.order = int(order)
        except:
            self.order = 0.
        try:
            self.capacity = float(capacity)
        except:
            self.capacity = 0.
        try:
            self.lcoe = float(lcoe)
        except:
            self.lcoe = 0.
        try:
            self.lcoe_cf = float(lcoe_cf)
        except:
            self.lcoe_cf = 0.
        try:
            self.emissions = float(emissions)
        except:
            self.emissions = 0.
        try:
            self.initial = float(initial)
        except:
            self.initial = 0.


class Adjustments(QtGui.QDialog):
    def __init__(self, data):
        super(Adjustments, self).__init__()
        self.adjusts = {}
        self.checkbox = {}
        self.results = None
        self.grid = QtGui.QGridLayout()
        ctr = 0
        for key, capacity in data:
            if key != 'Load' and capacity is None:
                continue
            self.adjusts[key] = QtGui.QDoubleSpinBox()
            self.adjusts[key].setRange(0, 25)
            self.adjusts[key].setValue(1.)
            self.adjusts[key].setDecimals(2)
            self.adjusts[key].setSingleStep(.1)
            self.grid.addWidget(QtGui.QLabel(key), ctr, 0)
            self.grid.addWidget(self.adjusts[key], ctr, 1)
            if key != 'Load':
                self.grid.addWidget(QtGui.QLabel(str(capacity) + 'MW'), ctr, 2)
            ctr += 1
        quit = QtGui.QPushButton('Quit', self)
        self.grid.addWidget(quit, ctr, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtGui.QPushButton('Proceed', self)
        self.grid.addWidget(show, ctr, 1)
        show.clicked.connect(self.showClicked)
        frame = QtGui.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtGui.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtGui.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - PB2 - Adjust renewables')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtGui.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def closeEvent(self, event):
        event.accept()

    def quitClicked(self):
        self.close()

    def showClicked(self):
        self.results = {}
        for key in self.adjusts.keys():
            self.results[key] = round(self.adjusts[key].value(), 2)
        self.close()

    def getValues(self):
        return self.results

class powerBalance(QtGui.QWidget):

    def __init__(self, help='help.html'):
        super(powerBalance, self).__init__()
        self.help = help
        config = ConfigParser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = 'SIREN.ini'
        config.read(config_file)
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            base_year = config.get('Base', 'year')
        except:
            base_year = '2012'
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', base_year)
            self.scenarios = self.scenarios[: self.scenarios.rfind('/') + 1]
            if self.scenarios[:3] == '../':
                ups = self.scenarios.split('../')
                me = os.getcwd().split(os.sep)
                me = me[: -(len(ups) - 1)]
                me.append(ups[-1])
                self.scenarios = '/'.join(me)
        except:
            self.scenarios = ''
        self.file_labels = ['Constraints', 'Generators', 'Data', 'Results']
        self.ifiles = [''] * 4
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.adjust_re = False
        self.change_res = True
        self.details = True
        try:
             items = config.items('Powerbalance')
             for key, value in items:
                 if key == 'adjust_renewables':
                     if value.lower() in ['true', 'on', 'yes']:
                         self.adjust_re = True
                 elif key == 'change_results':
                     if value.lower() in ['false', 'off', 'no']:
                         self.change_res = False
                 elif key[-5:] == '_file':
                     ndx = self.file_labels.index(key[:-5].title())
                     self.ifiles[ndx] = value
                 elif key[-6:] == '_sheet':
                     ndx = self.file_labels.index(key[:-6].title())
                     self.isheets[ndx] = value
                 elif key == 'renewables_details':
                     if value.lower() in ['false', 'off', 'no']:
                         self.details = False
        except:
            pass
        self.initUI()

    def initUI(self):
        self.grid = QtGui.QGridLayout()
        self.files = [None] * 4
        self.sheets = self.file_labels[:]
        del self.sheets[-2:]
        self.updated = False
        r = 0
        for i in range(4):
            self.grid.addWidget(QtGui.QLabel(self.file_labels[i] + ' File:'), r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setFrameStyle(6)
            self.files[i].setText(self.ifiles[i])
            self.connect(self.files[i], QtCore.SIGNAL('clicked()'), self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 3)
            if i < 2:
                r += 1
                self.grid.addWidget(QtGui.QLabel(self.file_labels[i] + ' Sheet:'), r, 0)
                self.sheets[i] = QtGui.QComboBox()
                self.sheets[i].addItem(self.isheets[i])
                self.grid.addWidget(self.sheets[i], r, 1, 1, 2)
            r += 1
        self.grid.addWidget(QtGui.QLabel('Adjust Renewables:'), 6, 0)
        self.adjust = QtGui.QCheckBox('(check to adjust/multiply renewable data)', self)
        if self.adjust_re:
            self.adjust.setCheckState(QtCore.Qt.Checked)
        self.grid.addWidget(self.adjust, 6, 1, 1, 3)
        self.grid.addWidget(QtGui.QLabel('Dispatch Order:'), 7, 0)
        self.order = QtGui.QListWidget()
        self.order.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.order, 7, 1, 1, 2)
        self.log = QtGui.QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, 8, 1, 1, 4)
        self.progressbar = QtGui.QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(10)
        self.progressbar.setValue(0)
        self.progressbar.setStyleSheet('QProgressBar {border: 1px solid grey; border-radius: 2px; text-align: center;}' \
                                       + 'QProgressBar::chunk { background-color: #6891c6;}')
        self.grid.addWidget(self.progressbar, 9, 1, 1, 4)
        self.progressbar.setHidden(True)
        quit = QtGui.QPushButton('Done', self)
        self.grid.addWidget(quit, 10, 0)
        quit.clicked.connect(self.quitClicked)
        QtGui.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        edit = [None, None]
        for i in range(2):
            edit[i] = QtGui.QPushButton(self.file_labels[i], self)
            self.grid.addWidget(edit[i], 10, 1 + i)
            edit[i].clicked.connect(self.editClicked)
        wdth = edit[1].fontMetrics().boundingRect(edit[1].text()).width() + 9
        pb = QtGui.QPushButton('PowerBalance', self)
     #   pb.setMaximumWidth(wdth)
        self.grid.addWidget(pb, 10, 3)
        pb.clicked.connect(self.pbClicked)
        help = QtGui.QPushButton('Help', self)
        help.setMaximumWidth(wdth)
        quit.setMaximumWidth(wdth)
        self.grid.addWidget(help, 10, 5)
        help.clicked.connect(self.helpClicked)
        QtGui.QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
        try:
            ts = xlrd.open_workbook(self.scenarios + str(self.files[1].text()))
            ws = ts.sheet_by_name('Generators')
            self.setOrder(self.getGenerators(ws))
            ts.release_resources()
            del ts
        except:
            pass
        frame = QtGui.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtGui.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtGui.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - powerbalance2 (' + fileVersion() + ') - PowerBalance 2.0')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        self.center()
        self.resize(int(self.sizeHint().width()* 1.07), int(self.sizeHint().height() * 1.07))
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtGui.QApplication.desktop().screenNumber(QtGui.QApplication.desktop().cursor().pos())
        centerPoint = QtGui.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        self.log.setText('')
        for i in range(4):
            if self.files[i].hasFocus():
                break
        curfile = self.scenarios + self.files[i].text()
        if i == 3:
            if self.files[i].text() == '':
                curfile = self.scenarios + self.files[2].text()
                curfile = curfile.replace('data', 'results')
                curfile = curfile.replace('Data', 'Results')
            newfile = str(QtGui.QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)'))
        else:
            newfile = str(QtGui.QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[i] + ' file',
                      curfile))
        if newfile != '':
            if i < 2:
                ts = xlrd.open_workbook(newfile, on_demand = True)
                ndx = 0
                self.sheets[i].clear()
                j = -1
                for sht in ts.sheet_names():
                    j += 1
                    self.sheets[i].addItem(str(sht))
                    if str(sht) == self.file_labels[i]:
                        ndx = j
                self.sheets[i].setCurrentIndex(ndx)
                if i == 1:
                    ws = ts.sheet_by_index(ndx)
                    self.setOrder(self.getGenerators(ws))
                ts.release_resources()
                del ts
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                self.files[i].setText(newfile)
            if i == 2 and self.change_res:
                newfile = str(self.files[2].text())
                newfile = newfile.replace('data', 'results')
                newfile = newfile.replace('Data', 'Results')
                if newfile != str(self.files[2].text()):
                    self.files[3].setText(newfile)
            self.updated = True

    def helpClicked(self):
        dialog = displayobject.AnObject(QtGui.QDialog(), self.help,
                 title='Help for powerbalance2 (' + fileVersion() + ')', section='powerbalance')
        dialog.exec_()

    def quitClicked(self):
        if self.updated:
            updates = {}
            lines = []
            lines.append('adjust_renewables=' + str(self.adjust.isChecked()))
            for i in range(len(self.file_labels)):
                lines.append(str(self.file_labels[i].lower()) + '_file=' + str(self.files[i].text()))
            for i in range(2):
                lines.append(str(self.file_labels[i].lower()) + '_sheet=' + str(self.sheets[i].currentText()))
            updates['Powerbalance'] = lines
            SaveIni(updates)
        self.close()

    def editClicked(self):
        self.log.setText('')
        it = self.file_labels.index(self.sender().text())
        try:
            if str(self.files[it].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[it].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[it].text()))
            try:
                sht = self.sheets[it].currentText()
            except:
                self.log.setText(self.sheets[it].currentText() + ' not found in ' + self.file_labels[it] + ' spreadsheet.')
                return
            ws = ts.sheet_by_name(sht)
        except:
            ts = None
            ws = None
        if it == 0: # constraints
            try:
                constraints = self.getConstraints(ws)
            except:
                return
            sp_pts = [2] * 11
            sp_pts[4] = 3 # discharge loss
            sp_pts[6] = 3 # parasitic loss
            sp_pts[9] = 3 # recharge loss
            dialog = displaytable.Table(constraints, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                constraints = dialog.getValues()
        else:
            try:
                generators = self.getGenerators(ws)
            except:
                return
            sp_pts = [2] * 8
            sp_pts[7] = 0 # dispatch order column
            dialog = displaytable.Table(generators, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                newgenerators = dialog.getValues()
                self.setOrder(generators)
        if ts is not None:
            ts.release_resources()
            del ts
        newfile = dialog.savedfile
        if newfile is not None:
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[it].setText(newfile[len(self.scenarios):])
            else:
                self.files[it].setText(newfile)
            self.log.setText(self.file_labels[it] + ' spreadsheet changed.')

    def getConstraints(self, ws):
        if ws is None:
            constraints = {}
            constraints['<name>'] = Constraint('<name>', '<category>', 0., 1.,
                                              1., 1., 1., 0., 1., 0., 0.)
            return constraints
        if ws.cell_value(1, 0) == 'Name' and ws.cell_value(1, 1) == 'Category':
            cat_col = 1
            for col in range(ws.ncols):
                if ws.cell_value(0, col)[:8] == 'Capacity':
                    cap_col = [col, col + 1]
                elif ws.cell_value(0, col)[:9] == 'Ramp Rate':
                    ramp_col = [col, col + 1]
                elif ws.cell_value(0, col)[:8] == 'Recharge':
                    rec_col = [col, col + 1]
                elif ws.cell_value(0, col)[:9] == 'Discharge':
                    dis_col = [col, col + 1]
                elif ws.cell_value(1, col)[:9] == 'Parasitic':
                    par_col = col
            strt_row = 2
        elif ws.cell_value(0, 0) == 'Name': # saved file
            cap_col = [-1, -1]
            ramp_col = [-1, -1]
            rec_col = [-1, -1]
            dis_col = [-1, -1]
            for col in range(ws.ncols):
                if ws.cell_value(0, col)[:8] == 'Category':
                    cat_col = col
                elif ws.cell_value(0, col)[:8] == 'Capacity':
                    if ws.cell_value(0, col)[-3:] == 'Min':
                        cap_col[0] = col
                    else:
                        cap_col[1] = col
                elif ws.cell_value(0, col)[:6] == 'Rampup':
                    ramp_col[0] = col
                elif ws.cell_value(0, col)[:8] == 'Rampdown':
                    ramp_col[1] = col
                elif ws.cell_value(0, col)[:8] == 'Recharge':
                    if ws.cell_value(0, col)[-3:] == 'Max':
                        rec_col[0] = col
                    else:
                        rec_col[1] = col
                elif ws.cell_value(0, col)[:9] == 'Discharge':
                    if ws.cell_value(0, col)[-3:] == 'Max':
                        dis_col[0] = col
                    else:
                        dis_col[1] = col
                elif ws.cell_value(0, col)[:9] == 'Parasitic':
                    par_col = col
            strt_row = 1
        else:
            self.log.setText('Not a ' + self.file_labels[it] + ' worksheet.')
            return
        try:
            cat_col = cat_col
        except:
            self.log.setText('Not a ' + self.file_labels[it] + ' worksheet.')
            return
        constraints = {}
        for row in range(strt_row, ws.nrows):
            constraints[str(ws.cell_value(row, 0))] = Constraint(str(ws.cell_value(row, 0)), str(ws.cell_value(row, cat_col)),
                                     ws.cell_value(row, cap_col[0]), ws.cell_value(row, cap_col[1]),
                                     ws.cell_value(row, ramp_col[0]), ws.cell_value(row, ramp_col[1]),
                                     ws.cell_value(row, rec_col[0]), ws.cell_value(row, rec_col[1]),
                                     ws.cell_value(row, dis_col[0]), ws.cell_value(row, dis_col[1]),
                                     ws.cell_value(row, par_col))
        return constraints

    def getGenerators(self, ws):
        if ws is None:
            generators = {}
            generators['<name>'] = Facility('<name>', 0, '<constraint>', 0., 0., 0., 0.)
            return generators
        if ws.cell_value(0, 0) != 'Name':
            self.log.setText('Not a ' + self.file_labels[1] + ' worksheet.')
            return
        for col in range(ws.ncols):
            if ws.cell_value(0, col)[:8] == 'Dispatch' or ws.cell_value(0, col) == 'Order':
                ord_col = col
            elif ws.cell_value(0, col) == 'Constraint':
                con_col = col
            elif ws.cell_value(0, col)[:8] == 'Capacity':
                cap_col = col
            elif ws.cell_value(0, col)[:7] == 'Initial':
                ini_col = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                lcc_col = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                lco_col = col
            elif ws.cell_value(0, col) == 'Emissions':
                emi_col = col
        try:
            lco_col = lco_col
        except:
            self.log.setText('Not a ' + self.file_labels[1] + ' worksheet.')
            return
        generators = {}
        for row in range(1, ws.nrows):
            generators[str(ws.cell_value(row, 0))] = Facility(str(ws.cell_value(row, 0)),
                                     ws.cell_value(row, ord_col), str(ws.cell_value(row, con_col)),
                                     ws.cell_value(row, cap_col), ws.cell_value(row, lco_col),
                                     ws.cell_value(row, lcc_col), ws.cell_value(row, emi_col),
                                     initial=ws.cell_value(row, ini_col))
        return generators

    def setOrder(self, generators=None):
        self.order.clear()
        if generators is None:
            order = ['Storage', 'Biomass', 'PHS', 'Gas', 'CCG1', 'Other', 'Coal']
            for stn in order:
                self.order.addItem(stn)
        else:
            order = []
            for key, value in generators.iteritems():
                try:
                    o = int(value.order)
                    if o > 0:
                        while len(order) <= o:
                            order.append([])
                        order[o - 1].append(key)
                except:
                    pass
            for cat in order:
                for stn in cat:
                    self.order.addItem(stn)

    def pbClicked(self):
        self.log.setText('PowerBalance processing started')
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(10)
        self.progressbar.setHidden(False)
        details = self.details
        try:
            if str(self.files[0].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[0].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[0].text()))
            ws = ts.sheet_by_name(self.sheets[0].currentText())
            constraints = self.getConstraints(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Constraints.')
            self.progressbar.setHidden(True)
            return
        try:
            if str(self.files[1].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[1].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[1].text()))
            ws = ts.sheet_by_name(self.sheets[1].currentText())
            generators = self.getGenerators(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Generators.')
            self.progressbar.setHidden(True)
            return
        start_time = time.time()
        clock_start = time.clock()
        re_capacities = [0.] * len(tech_names)
        if str(self.files[2].text()).find('/') >= 0:
            pb_data_file = str(self.files[2].text())
        else:
            pb_data_file = self.scenarios + str(self.files[2].text())
        re_capacity = 0.
        data = []
        load = []
        shortfall = []
        if pb_data_file[-4:] == '.xls': #xls format
            xlsx = False
            details = False
            ts = xlrd.open_workbook(pb_data_file)
            ws = ts.sheet_by_index(0)
            if ws.cell_value(0, 0) != 'Hourly Shortfall Table' \
              or ws.cell_value(0, 4) != 'Generation Summary Table':
                self.log.setText('not a pb1 spreadsheet')
                self.progressbar.setHidden(True)
                return
            for row in range(20):
                if ws.cell_value(row, 5) == 'Total':
                    re_capacity = ws.cell_value(row, 6)
                    re_generation = ws.cell_value(row, 8)
                    break
            for row in range(2, ws.nrows):
                shortfall.append(ws.cell_value(row, 2))
        else: # xlsx format
            xlsx = True
            shortfall = [0.] * 8760
            if details:
                cols = []
            ts = oxl.load_workbook(pb_data_file)
            ws = ts.active
            top_row = ws.max_row - 8760
            if ws.cell(row=top_row, column=1).value != 'Hour' or ws.cell(row=top_row, column=2).value != 'Period':
                self.log.setText('not a pb data spreadsheet')
                self.progressbar.setHidden(True)
                return
            typ_row = top_row - 1
            while typ_row > 0:
                if ws.cell(row=typ_row, column=3).value in tech_names:
                    break
                typ_row -= 1
            else:
                self.log.setText('no suitable data')
                return
            icap_row = typ_row + 1
            while icap_row < top_row:
                if ws.cell(row=icap_row, column=1).value == 'Capacity (MW)':
                    break
                icap_row += 1
            else:
                self.log.setText('no capacity data')
                return
            adjustby = None
            if self.adjust.isChecked():
                adjustin = []
                for col in range(3, ws.max_column + 1):
                    try:
                        valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                        i = tech_names.index(valu)
                    except:
                        break
                    adjustin.append([tech_names[i], ws.cell(row=icap_row, column=col).value])
                adjust = Adjustments(adjustin)
                adjust.exec_()
                adjustby = adjust.getValues()
                if adjustby is None:
                    self.log.setText('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
            load_col = -1
            det_col = 3
            self.progressbar.setValue(1)
            for col in range(3, ws.max_column + 1):
                try:
                    valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                    i = tech_names.index(valu)
                except:
                    break # ?? or continue??
                if tech_names[i] != 'Load':
                    if ws.cell(row=icap_row, column=col).value <= 0:
                        continue
                data.append([])
                try:
                    multiplier = adjustby[tech_names[i]]
                except:
                    multiplier = 1
                if details:
                    cols.append(i)
                    try:
                        re_capacities[i] = ws.cell(row=icap_row, column=col).value * multiplier
                    except:
                        pass
                try:
                    re_capacity += ws.cell(row=icap_row, column=col).value * multiplier
                except:
                    pass
                for row in range(top_row + 1, ws.max_row + 1):
                    data[-1].append(ws.cell(row=row, column=col).value * multiplier)
                    try:
                        if tech_names[i] == 'Load':
                            load_col = len(data) - 1
                            shortfall[row - top_row - 1] += ws.cell(row=row, column=col).value * multiplier
                        else:
                            shortfall[row - top_row - 1] -= ws.cell(row=row, column=col).value * multiplier
                    except:
                        pass
            ts.close()
        if str(self.files[3].text()) == '':
            i = pb_data_file.find('/')
            if i >= 0:
                data_file = pb_data_file[i + 1:]
            else:
                data_file = pb_data_file
            data_file = data_file.replace('data', 'results')
        else:
            if str(self.files[3].text()).find('/') >= 0:
                data_file = str(self.files[3].text())
            else:
                data_file = self.scenarios + str(self.files[3].text())
        if not xlsx:
            data_file += 'x'
        self.progressbar.setValue(2)
        headers = ['Facility', 'Capacity (MW)', 'Subtotal (MWh)', 'CF', 'Cost ($)', 'LCOE ($/MWh)', 'Emissions (tCO2e)']
        ds = oxl.Workbook()
        ns = ds.active
        ns.title = 'Detail'
        ss = ds.create_sheet('Summary', 0)
        cap_row = 1
        ns.cell(row=cap_row, column=2).value = headers[1]
        ss.cell(row=3, column=1).value = headers[0]
        ss.cell(row=3, column=2).value = headers[1]
        ini_row = 2
        ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
        sum_row = 3
        ns.cell(row=sum_row, column=2).value = headers[2]
        ss.cell(row=3, column=3).value = headers[2]
        cf_row = 4
        ns.cell(row=cf_row, column=2).value = headers[3]
        ss.cell(row=3, column=4).value = headers[3]
        cost_row = 5
        ns.cell(row=cost_row, column=2).value = headers[4]
        ss.cell(row=3, column=5).value = headers[4]
        lcoe_row = 6
        ns.cell(row=lcoe_row, column=2).value = headers[5]
        ss.cell(row=3, column=6).value = headers[5]
        emi_row = 7
        ns.cell(row=emi_row, column=2).value = headers[6]
        ss.cell(row=3, column=7).value = headers[6]
        ss_row = 3
        fall_row = 8
        ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
        what_row = 9
        hrows = 10
        ns.cell(row=what_row, column=1).value = 'Hour'
        ns.cell(row=what_row, column=2).value = 'Period'
        ns.cell(row=what_row, column=3).value = 'Load'
        o = 4
        col = 3
        if details:
            for i in cols:
                if tech_names[i] == 'Load':
                    continue
                col += 1
                ss_row += 1
                ns.cell(row=what_row, column=col).value = tech_names[i]
                ss.cell(row=ss_row, column=1).value = '=Detail!' + ss_col(col) + str(what_row)
                ns.cell(row=cap_row, column=col).value = re_capacities[i]
                ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=2).value = '=Detail!' + ss_col(col) + str(cap_row)
                ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) \
                        + str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ss.cell(row=ss_row, column=3).value = '=Detail!' + ss_col(col) + str(sum_row)
                ss.cell(row=ss_row, column=3).number_format = '#,##0'
                ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' \
                        + ss_col(col) + '1/8760'
                ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=4).value = '=Detail!' + ss_col(col) + str(cf_row)
                ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
                if generators[tech_names[i]].lcoe > 0:
                    ns.cell(row=cost_row, column=col).value = generators[tech_names[i]].lcoe * generators[tech_names[i]].lcoe_cf \
                            * 8760 * re_capacities[i]
                    ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                    ss.cell(row=ss_row, column=5).value = '=Detail!' + ss_col(col) + str(cost_row)
                    ss.cell(row=ss_row, column=5).number_format = '$#,##0'
                    ns.cell(row=lcoe_row, column=col).value = '=' + ss_col(col) + str(cost_row) + '/8760/' \
                            + ss_col(col) + str(cf_row) +'/' + ss_col(col) + str(cap_row)
                    ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
                    ss.cell(row=ss_row, column=6).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                    ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
                if generators[tech_names[i]].emissions > 0:
                    ns.cell(row=emi_row, column=col).value = '=' + ss_col(col) + str(sum_row) \
                            + '*' + str(generators[tech_names[i]].emissions)
                    ns.cell(row=emi_row, column=col).number_format = '#,##0'
                    ss.cell(row=ss_row, column=7).value = '=Detail!' + ss_col(col) + str(emi_row)
                    ss.cell(row=ss_row, column=7).number_format = '#,##0'
            shrt_col = col + 1
        else:
            shrt_col = 5
            ns.cell(row=what_row, column=4).value = 'Renewable'
            ns.cell(row=cap_row, column=4).value = re_capacity
            ns.cell(row=cap_row, column=4).number_format = '#,##0.00'
            if xlsx:
                ns.cell(row=sum_row, column=4).value = '=SUM(D' + str(hrows) + ':D' + str(hrows + 8759) + ')'
            else:
                ns.cell(row=sum_row, column=4).value = re_generation
            ns.cell(row=sum_row, column=4).number_format = '#,##0'
            ns.cell(row=cf_row, column=4).value = '=D3/D1/8760'
            ns.cell(row=cf_row, column=4).number_format = '#,##0.00'
            ss_row += 1
            ss.cell(row=ss_row, column=1).value = '=Detail!D' + str(what_row)
            ss.cell(row=ss_row, column=2).value = '=Detail!D' + str(cap_row)
            ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=3).value = '=Detail!D' + str(sum_row)
            ss.cell(row=ss_row, column=3).number_format = '#,##0'
            ss.cell(row=ss_row, column=4).value = '=Detail!D' + str(cf_row)
            ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
        ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ss_col(shrt_col) \
                        + str(hrows) + ':' + ss_col(shrt_col) + str(hrows + 8759) + ',">0")'
        ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
        ns.cell(row=what_row, column=shrt_col).value = 'Shortfall'
        for col in range(3, shrt_col + 1):
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
        for row in range(hrows, 8760 + hrows):
            if xlsx:
                ns.cell(row=row, column=1).value = ws.cell(row=top_row + row - hrows + 1, column=1).value
                ns.cell(row=row, column=2).value = ws.cell(row=top_row + row - hrows + 1, column=2).value
                ns.cell(row=row, column=3).value = round(data[load_col][row - hrows], 2)
                if not details:
                    ns.cell(row=row, column=4).value = round(data[load_col][row - hrows] - shortfall[row - hrows], 2)
            else:
                ns.cell(row=row, column=1).value = ws.cell_value(row - hrows + 2, 0)
                ns.cell(row=row, column=2).value = ws.cell_value(row - hrows + 2, 1)
            ns.cell(row=row, column=shrt_col).value = round(shortfall[row - hrows], 2)
            for col in range(3, shrt_col + 1):
                ns.cell(row=row, column=col).number_format = '#,##0.00'
        if details:
            col = 3
            for i in range(len(data)):
                if cols[i] == load_col:
                    continue
                col += 1
                for row in range(hrows, 8760 + hrows):
                    ns.cell(row=row, column=col).value = data[i][row - hrows]
        order = [] #'Storage', 'Biomass', 'PHS', 'Gas', 'CCG1', 'Other', 'Coal']
        for itm in range(self.order.count()):
            order.append(str(self.order.item(itm).text()))
        col = shrt_col + 1
        #storage? = []
        self.progressbar.setValue(3)
        for gen in order:
            if constraints[generators[gen].constraint].category == 'Storage': # storage
                storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                storage[0] = generators[gen].capacity
                ns.cell(row=cap_row, column=col).value = round(generators[gen].capacity, 2)
                ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                storage[1] = generators[gen].initial
                if constraints[generators[gen].constraint].capacity_min > 0:
                    storage[2] = generators[gen].capacity * constraints[generators[gen].constraint].capacity_min
                if constraints[generators[gen].constraint].capacity_max > 0:
                    storage[3] = generators[gen].capacity * constraints[generators[gen].constraint].capacity_max
                else:
                    storage[3] = generators[gen].capacity
                recharge = [0., 0.] # cap, loss
                if constraints[generators[gen].constraint].recharge_max > 0:
                    recharge[0] = generators[gen].capacity * constraints[generators[gen].constraint].recharge_max
                else:
                    recharge[0] = generators[gen].capacity
                if constraints[generators[gen].constraint].recharge_loss > 0:
                    recharge[1] = constraints[generators[gen].constraint].recharge_loss
                discharge = [0., 0.] # cap, loss
                if constraints[generators[gen].constraint].discharge_max > 0:
                    discharge[0] = generators[gen].capacity * constraints[generators[gen].constraint].discharge_max
                if constraints[generators[gen].constraint].discharge_loss > 0:
                    discharge[1] = constraints[generators[gen].constraint].discharge_loss
                if constraints[generators[gen].constraint].parasitic_loss > 0:
                    parasite = constraints[generators[gen].constraint].parasitic_loss / 24.
                else:
                    parasite = 0.
                storage_carry = storage[1] # generators[gen].initial
                ns.cell(row=ini_row, column=col).value = round(storage_carry, 2)
                ns.cell(row=ini_row, column=col).number_format = '#,##0.00'
                storage_bal = []
                for row in range(8760):
                    if storage_carry > 0:
                        storage_carry = storage_carry * (1 - parasite)
                    storage_loss = 0.
                    if shortfall[row] < 0:  # excess generation
                        can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                        if can_use < 0: # can use some
                            if shortfall[row] > can_use:
                                can_use = shortfall[row]
                            if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                can_use = - recharge[0]
                        else:
                            can_use = 0.
                        storage_carry -= (can_use * (1 - recharge[1]))
                        shortfall[row] -= can_use
                    else: # shortfall
                        can_use = shortfall[row] * (1 / (1 - discharge[1]))
                        if can_use > storage_carry - storage[2]:
                            can_use = storage_carry - storage[2]
                        if can_use > 0:
                            storage_loss = can_use * discharge[1]
                            storage_carry -= can_use
                            shortfall[row] -= (can_use - storage_loss)
                            if storage_carry < 0:
                                storage_carry = 0
                        else:
                            can_use = 0.
                    storage_bal.append(storage_carry)
                    ns.cell(row=row + hrows, column=col).value = round(can_use, 2)
                    ns.cell(row=row + hrows, column=col + 1).value = round(storage_carry, 2)
                    ns.cell(row=row + hrows, column=col + 2).value = round(shortfall[row], 2)
                    ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                    ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=row + hrows, column=col + 2).number_format = '#,##0.00'
                ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ss_col(col) + \
                        str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' + ss_col(col) + '1/8760'
                ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                col += 3
            else:
                ns.cell(row=cap_row, column=col).value = round(generators[gen].capacity, 2)
                ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                for row in range(8760):
                    if shortfall[row] >= 0: # shortfall?
                        if shortfall[row] >= generators[gen].capacity:
                            shortfall[row] = shortfall[row] - generators[gen].capacity
                            ns.cell(row=row + hrows, column=col).value = round(generators[gen].capacity, 2)
                        else:
                            ns.cell(row=row + hrows, column=col).value = round(shortfall[row], 2)
                            shortfall[row] = 0
                    else:
                        ns.cell(row=row + hrows, column=col).value = 0
                    ns.cell(row=row + hrows, column=col + 1).value = round(shortfall[row], 2)
                    ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                    ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) + str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
                ns.cell(row=sum_row, column=col).number_format = '#,##0'
                ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' + ss_col(col) + '1/8760'
                ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                col += 2
        self.progressbar.setValue(4)
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            for cell in column_cells:
                if cell.row >= 0:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                        value = cell.value
                        row = cell.row
            ns.column_dimensions[column_cells[0].column].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        col = shrt_col + 1
        for gen in order:
            ss_row += 1
            ns.cell(row=what_row, column=col).value = gen
            ss.cell(row=ss_row, column=1).value = '=Detail!' + ss_col(col) + str(what_row)
            ss.cell(row=ss_row, column=2).value = '=Detail!' + ss_col(col) + str(cap_row)
            ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=3).value = '=Detail!' + ss_col(col) + str(sum_row)
            ss.cell(row=ss_row, column=3).number_format = '#,##0'
            ss.cell(row=ss_row, column=4).value = '=Detail!' + ss_col(col) + str(cf_row)
            ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
            if generators[gen].lcoe > 0:
                ns.cell(row=cost_row, column=col).value = generators[gen].lcoe * generators[gen].lcoe_cf * 8760 * generators[gen].capacity
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ss.cell(row=ss_row, column=5).value = '=Detail!' + ss_col(col) + str(cost_row)
                ss.cell(row=ss_row, column=5).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=' + ss_col(col) + str(cost_row) + '/8760/' \
                            + ss_col(col) + str(cf_row) +'/' + ss_col(col) + str(cap_row)
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=6).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
            if generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ss_col(col) + str(sum_row) \
                        + '*' + str(generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
                ss.cell(row=ss_row, column=7).value = '=Detail!' + ss_col(col) + str(emi_row)
                ss.cell(row=ss_row, column=7).number_format = '#,##0'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if constraints[generators[gen].constraint].category == 'Storage':
                ns.cell(row=what_row, column=col + 1).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 2).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 2).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 2).value = '=COUNTIF(' + ss_col(col + 2) \
                        + str(hrows) + ':' + ss_col(col + 2) + str(hrows + 8759) + ',">0")'
                ns.cell(row=fall_row, column=col + 2).number_format = '#,##0'
                col += 3
            else:
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ss_col(col + 1) \
                        + str(hrows) + ':' + ss_col(col + 1) + str(hrows + 8759) + ',">0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
        self.progressbar.setValue(6)
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        ss.cell(row=1, column=1).value = 'PowerBalance2 - Summary'
        bold = oxl.styles.Font(bold=True, name=ss.cell(row=1, column=1).font.name)
        ss.cell(row=1, column=1).font = bold
        ss_row +=1
        for col in range(1, 8):
            ss.cell(row=3, column=col).font = bold
            ss.cell(row=ss_row, column=col).font = bold
        ss.cell(row=ss_row, column=2).value = '=SUM(B4:B' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
        ss.cell(row=ss_row, column=3).value = '=SUM(C4:C' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/B' + str(ss_row) + '/8760'
        ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
        ss.cell(row=ss_row, column=5).value = '=SUM(E4:E' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=5).number_format = '$#,##0'
        ss.cell(row=ss_row, column=6).value = '=E' + str(ss_row) + '/C' + str(ss_row)
        ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
        ss.cell(row=ss_row, column=7).value = '=SUM(G4:G' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=7).number_format = '#,##0'
        ss.cell(row=ss_row, column=1).value = 'Total'
        for column_cells in ss.columns:
            length = 0
            value = ''
            row = 0
            for cell in column_cells:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
                    value = cell.value
                    row = cell.row
            ss.column_dimensions[column_cells[0].column].width = length * 1.15
        ss.column_dimensions['D'].width = 7
        ss.column_dimensions['E'].width = 18
        last_col = ss_col(ns.max_column)
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Load Analysis'
        ss.cell(row=ss_row, column=1).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Load met'
        ss.cell(row=ss_row, column=3).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',"<=0",Detail!C' + str(hrows) + ':Detail!C' \
            + str(hrows + 8759) + ')+SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',">0",Detail!C' + str(hrows) + ':Detail!C' \
            + str(hrows + 8759) + ')-C' + str(ss_row + 1)
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row + 2)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Shortfall'
        ss.cell(row=ss_row, column=3).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',">0",Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row + 1)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Total Load'
        ss.cell(row=ss_row, column=1).font = bold
        ss.cell(row=ss_row, column=3).value = '=SUM(C' + str(ss_row - 2) + ':C' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=3).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Surplus'
        ss.cell(row=ss_row, column=3).value = '=-SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',"<0",Detail!' + last_col + str(hrows) \
            + ':Detail!' + last_col + str(hrows + 8759) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row - 1)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Data sources:'
        ss.cell(row=ss_row, column=1).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Scenarios folder'
        ss.cell(row=ss_row, column=2).value = self.scenarios
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'PowerBalance data file'
        if pb_data_file[: len(self.scenarios)] == self.scenarios:
            pb_data_file = pb_data_file[len(self.scenarios):]
        ss.cell(row=ss_row, column=2).value = pb_data_file
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Constraints worksheet'
        ss.cell(row=ss_row, column=2).value = str(self.files[0].text()) \
               + '.' + str(self.sheets[0].currentText())
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Facility worksheet'
        ss.cell(row=ss_row, column=2).value = str(self.files[1].text()) \
               + '.' + str(self.sheets[1].currentText())
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        self.progressbar.setValue(7)
        try:
            if adjustby is not None:
                adjusted = ''
                for key, value in iter(sorted(adjustby.iteritems())):
                    if value != 1:
                        adjusted += key + ': ' + str(value) + '; '
                if len(adjusted) > 0:
                    ss_row += 1
                    ss.cell(row=ss_row, column=1).value = 'Renewable inputs adjusted'
                    ss.cell(row=ss_row, column=2).value = adjusted[:-2]
        except:
            pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        ds.save(data_file)
        self.progressbar.setValue(10)
        j = data_file.rfind('/')
        data_file = data_file[j + 1:]
        msg = '%s created (%.2f or %.2f seconds)' % (data_file, time.time() - start_time,
              time.clock() - clock_start)
        msg = '%s created.' % data_file
        self.log.setText(msg)
        self.progressbar.setHidden(True)

if "__main__" == __name__:
    app = QtGui.QApplication(sys.argv)
    ex = powerBalance()
    app.exec_()
    app.deleteLater()
    sys.exit()
