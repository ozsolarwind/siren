#!/usr/bin/python3
#
#  Copyright (C) 2016-2024 Sustainable Energy Now Inc., Angus King
#
#  indexweather.py - This file is part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#

import openpyxl as oxl
import os
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
import configparser   # decode .ini file
import xlwt

import displayobject
from credits import fileVersion
from getmodels import getModelFile
from senutils import ClickableQLabel, getParents, getUser, ssCol


class makeIndex():

    def close(self):
        return

    def getLog(self):
        return self.log

    def __init__(self, what, src_dir, tgt_fil):
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        self.log = ''
        files = []
        fils = os.listdir(src_dir)
        for fil in fils:
            if (what[0].lower() == 's' and (fil[-4:] == '.csv' or fil[-4:] == '.smw')) \
              or (what[0].lower() == 'w' and fil[-4:] == '.srw'):
                tf = open(src_dir + '/' + fil, 'r')
                lines = tf.readlines()
                tf.close()
                if fil[-4:] == '.smw':
                    bits = lines[0].split(',')
                    src_lat = float(bits[4])
                    src_lon = float(bits[5])
                elif fil[-4:] == '.srw':
                    bits = lines[0].split(',')
                    src_lat = float(bits[5])
                    src_lon = float(bits[6])
                elif fil[-10:] == '(TMY2).csv' or fil[-10:] == '(TMY3).csv' \
                  or fil[-10:] == '(INTL).csv' or fil[-4:] == '.csv':
                    fst_row = len(lines) - 8760
                    if fst_row < 3:
                        bits = lines[0].split(',')
                        src_lat = float(bits[4])
                        src_lon = float(bits[5])
                    else:
                        cols = lines[fst_row - 3].split(',')
                        bits = lines[fst_row - 2].split(',')
                        for i in range(len(cols)):
                            if cols[i].lower() in ['latitude', 'lat']:
                                src_lat = float(bits[i])
                            elif cols[i].lower() in ['longitude', 'lon', 'long', 'lng']:
                                src_lon = float(bits[i])
                else:
                    continue
                files.append([src_lat, src_lon, fil])
        if tgt_fil[-5:] == '.xlsx':
            wb = oxl.Workbook()
            normal = oxl.styles.Font(name='Arial', size='10')
            bold = oxl.styles.Font(name='Arial', bold=True, size='10')
            ws = wb.active
            ws.title = 'Index'
            ws.cell(row=1, column=1).value = 'Latitude'
            ws.cell(row=1, column=1).font = normal
            ws.cell(row=1, column=2).value = 'Longitude'
            ws.cell(row=1, column=2).font = normal
            ws.cell(row=1, column=3).value = 'Filename'
            ws.cell(row=1, column=3).font = normal
            lens = [8, 9, 8]
            for i in range(len(files)):
                for c in range(3):
                    ws.cell(row=i + 2, column=c + 1).value = files[i][c]
                    ws.cell(row=i + 2, column=c + 1).font = normal
                    lens[c] = max(len(str(files[i][c])), lens[c])
            for c in range(len(lens)):
                ws.column_dimensions[ssCol(c + 1)].width = lens[c]
            ws.freeze_panes = 'A2'
            wb.save(tgt_fil)
        elif tgt_fil[-4:]:
            wb = xlwt.Workbook()
            fnt = xlwt.Font()
            fnt.bold = True
            styleb = xlwt.XFStyle()
            styleb.font = fnt
            ws = wb.add_sheet('Index')
            ws.write(0, 0, 'Latitude')
            ws.write(0, 1, 'Longitude')
            ws.write(0, 2, 'Filename')
            lens = [8, 9, 8]
            for i in range(len(files)):
                for c in range(3):
                    ws.write(i + 1, c, files[i][c])
                    lens[c] = max(len(str(files[i][c])), lens[c])
            for c in range(len(lens)):
                if lens[c] * 275 > ws.col(c).width:
                    ws.col(c).width = lens[c] * 275
            ws.set_panes_frozen(True)   # frozen headings instead of split panes
            ws.set_horz_split_pos(1)   # in general, freeze after last heading row
            ws.set_remove_splits(True)   # if user does unfreeze, don't leave a split there
            wb.save(tgt_fil)
        else:
            tf = open(tgt_fil, 'w')
            hdr = 'Latitude,Longitude,Filename\n'
            tf.write(hdr)
            for i in range(len(files)):
                line = '%s,%s,"%s"\n' % (files[i][0], files[i][1], files[i][2])
                tf.write(line)
            tf.close()
        self.log += '%s created' % tgt_fil[tgt_fil.rfind('/') + 1:]


class getParms(QtWidgets.QWidget):

    def __init__(self, help='help.html'):
        super(getParms, self).__init__()
        self.help = help
        self.initUI()

    def initUI(self):
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = getModelFile('SIREN.ini')
        config.read(config_file)
        try:
            self.base_year = config.get('Base', 'year')
        except:
            self.base_year = '2012'
        self.parents = []
        try:
            self.parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            self.solarfiles = config.get('Files', 'solar_files')
            for key, value in self.parents:
                self.solarfiles = self.solarfiles.replace(key, value)
            self.solarfiles = self.solarfiles.replace('$USER$', getUser())
            self.solarfiles = self.solarfiles.replace('$YEAR$', self.base_year)
        except:
            self.solarfiles = ''
        try:
            self.solarindex = config.get('Files', 'solar_index')
            for key, value in self.parents:
                self.solarindex = self.solarindex.replace(key, value)
            self.solarindex = self.solarindex.replace('$USER$', getUser())
            self.solarindex = self.solarindex.replace('$YEAR$', self.base_year)
        except:
            self.solarindex = ''
        if self.solarindex == '':
            self.solarindex = self.solarfiles + '/solar_index.xlsx'
        try:
            self.windfiles = config.get('Files', 'wind_files')
            for key, value in self.parents:
                self.windfiles = self.windfiles.replace(key, value)
            self.windfiles = self.windfiles.replace('$USER$', getUser())
            self.windfiles = self.windfiles.replace('$YEAR$', self.base_year)
        except:
            self.windfiles = ''
        try:
            self.windindex = config.get('Files', 'wind_index')
            for key, value in self.parents:
                self.windindex = self.windindex.replace(key, value)
            self.windindex = self.windindex.replace('$USER$', getUser())
            self.windindex = self.windindex.replace('$YEAR$', self.base_year)
        except:
            self.windindex = ''
        if self.windindex == '':
            self.windindex = self.windfiles + '/wind_index.xlsx'
        self.grid = QtWidgets.QGridLayout()
        self.grid.addWidget(QtWidgets.QLabel('Solar Folder:'), 1, 0)
        self.ssource = ClickableQLabel()
        self.ssource.setText(self.solarfiles)
        self.ssource.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.ssource.clicked.connect(self.sdirChanged)
        self.grid.addWidget(self.ssource, 1, 1, 1, 4)
        self.grid.addWidget(QtWidgets.QLabel('Solar Index:'), 2, 0)
        self.starget = ClickableQLabel()
        self.starget.setText(self.solarindex)
        self.starget.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.starget.clicked.connect(self.stgtChanged)
        self.grid.addWidget(self.starget, 2, 1, 1, 4)
        self.grid.addWidget(QtWidgets.QLabel('Wind Folder:'), 3, 0)
        self.wsource = ClickableQLabel()
        self.wsource.setText(self.windfiles)
        self.wsource.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.wsource.clicked.connect(self.wdirChanged)
        self.grid.addWidget(self.wsource, 3, 1, 1, 4)
        self.grid.addWidget(QtWidgets.QLabel('Wind Index:'), 4, 0)
        self.wtarget = ClickableQLabel()
        self.wtarget.setText(self.windindex)
        self.wtarget.setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
        self.wtarget.clicked.connect(self.wtgtChanged)
        self.grid.addWidget(self.wtarget, 4, 1, 1, 4)
        self.grid.addWidget(QtWidgets.QLabel('Properties:'), 5, 0)
        self.properties = QtWidgets.QLineEdit()
        self.properties.setReadOnly(True)
        self.grid.addWidget(self.properties, 5, 1, 1, 4)
        self.log = QtWidgets.QLabel(' ')
        self.grid.addWidget(self.log, 6, 1, 1, 4)
        quit = QtWidgets.QPushButton('Quit', self)
        self.grid.addWidget(quit, 7, 0)
        quit.clicked.connect(self.quitClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        dosolar = QtWidgets.QPushButton('Produce Solar Index', self)
        wdth = dosolar.fontMetrics().boundingRect(dosolar.text()).width() + 9
        dosolar.setMaximumWidth(wdth)
        self.grid.addWidget(dosolar, 7, 1)
        dosolar.clicked.connect(self.dosolarClicked)
        dowind = QtWidgets.QPushButton('Produce Wind Index', self)
        dowind.setMaximumWidth(wdth)
        self.grid.addWidget(dowind, 7, 2)
        dowind.clicked.connect(self.dowindClicked)
        help = QtWidgets.QPushButton('Help', self)
        help.setMaximumWidth(wdth)
        self.grid.addWidget(help, 7, 3)
        help.clicked.connect(self.helpClicked)
        QtWidgets.QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
        self.grid.setColumnStretch(3, 5)
        frame = QtWidgets.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtWidgets.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - indexweather (' + fileVersion() + ') - Make resource grid file')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        self.center()
        self.resize(int(self.sizeHint().width()* 1.07), int(self.sizeHint().height() * 1.07))
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        centerPoint = QtWidgets.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def sdirChanged(self):
        curdir = self.ssource.text()
        newdir = str(QtWidgets.QFileDialog.getExistingDirectory(self, 'Choose Solar Folder',
                 curdir, QtWidgets.QFileDialog.ShowDirsOnly))
        if newdir != '':
            self.ssource.setText(newdir)

    def wdirChanged(self):
        curdir = self.wsource.text()
        newdir = str(QtWidgets.QFileDialog.getExistingDirectory(self, 'Choose Wind Folder',
                 curdir, QtWidgets.QFileDialog.ShowDirsOnly))
        if newdir != '':
            self.wsource.setText(newdir)

    def stgtChanged(self):
        curtgt = self.starget.text()
        newtgt = QtWidgets.QFileDialog.getSaveFileName(self, 'Choose Solar Index',
                 curtgt, 'Excel Files (*.xls*)')[0]
        if newtgt != '':
            self.starget.setText(newtgt)

    def wtgtChanged(self):
        curtgt = self.starget.text()
        newtgt = QtWidgets.QFileDialog.getSaveFileName(self, 'Choose Wind Index',
                 curtgt, 'Excel Files (*.xls*)')[0]
        if newtgt != '':
            self.wtarget.setText(newtgt)

    def helpClicked(self):
        dialog = displayobject.AnObject(QtWidgets.QDialog(), self.help,
                 title='Help for indexweather (' + fileVersion() + ')', section='index')
        dialog.exec_()

    def quitClicked(self):
        self.close()

    def dosolarClicked(self):
        solar = makeIndex('Solar', str(self.ssource.text()), str(self.starget.text()))
        self.log.setText(solar.getLog())
        prop = 'solar_index=' + str(self.starget.text())
        self.properties.setText(prop)

    def dowindClicked(self):
        wind = makeIndex('Wind', str(self.wsource.text()), str(self.wtarget.text()))
        self.log.setText(wind.getLog())
        prop = 'wind_index=' + str(self.wtarget.text())
        self.properties.setText(prop)


if "__main__" == __name__:
    app = QtWidgets.QApplication(sys.argv)
    if len(sys.argv) > 2:   # arguments
        src_dir_s = ''
        src_dir_w = ''
        tgt_fil = ''
        for i in range(1, len(sys.argv)):
            if sys.argv[i][:6] == 'solar=':
                src_dir_s = sys.argv[i][6:]
            elif sys.argv[i][:5] == 'wind=':
                src_dir_w = sys.argv[i][5:]
            elif sys.argv[i][:7] == 'target=' or sys.argv[i][:7] == 'tgtfil=':
                tgt_fil = sys.argv[i][7:]
        if src_dir_s != '':
            files = makeIndex('Solar', src_dir_s, tgt_fil)
        elif src_dir_w != '':
            files = makeIndex('Wind', src_dir_w, tgt_fil)
        else:
            print('No source directory specified')
    else:
        ex = getParms()
        app.exec_()
        app.deleteLater()
        sys.exit()
